from __future__ import annotations

import calendar
import os
import random
import re
import threading
import time
import unicodedata
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Dict

import customtkinter as ctk
import pandas as pd
from openpyxl import load_workbook
from PIL import Image
from playwright.sync_api import sync_playwright
from tkinter import filedialog, messagebox


APP_TITLE = "Robô de Lançamentos de Caixa"
APP_GEOMETRY = "1180x760"
TOTVS_URL = "https://focoaluguel162907.protheus.cloudtotvs.com.br:1453/webapp/"

MAIN_BG = "#f6f4f1"
CARD_BG = "#ffffff"
CARD_BORDER = "#eadfdb"
PRIMARY_TEXT = "#d81919"
MUTED_TEXT = "#5c5c5c"
BUTTON_BG = "#ef1a14"
BUTTON_ACTIVE_BG = "#c91410"
SOFT_RED = "#fff1ef"
SUCCESS_GREEN = "#0f8a4b"
WARNING_ORANGE = "#b96a10"

MONTH_NAMES_PT = {
    1: "JANEIRO",
    2: "FEVEREIRO",
    3: "MARCO",
    4: "ABRIL",
    5: "MAIO",
    6: "JUNHO",
    7: "JULHO",
    8: "AGOSTO",
    9: "SETEMBRO",
    10: "OUTUBRO",
    11: "NOVEMBRO",
    12: "DEZEMBRO",
}


def _format_period_text(start_date: datetime, end_date: datetime) -> str:
    if end_date < start_date:
        start_date, end_date = end_date, start_date

    start_month = MONTH_NAMES_PT[start_date.month]
    end_month = MONTH_NAMES_PT[end_date.month]

    if start_date.month == end_date.month and start_date.year == end_date.year:
        return f"{start_date.day} A {end_date.day} DE {start_month}"
    if start_date.year == end_date.year:
        return f"{start_date.day} DE {start_month} A {end_date.day} DE {end_month}"
    return f"{start_date.day} DE {start_month} DE {start_date.year} A {end_date.day} DE {end_month} DE {end_date.year}"


FLOW_SPECS = {
    "cash": {
        "label": "Cash",
        "sheet_name": "CASH",
        "header_row": 1,
        "pending_column": "BAIXADO",
        "required_columns": [
            "CONTRATO",
            "LOJA",
            "VALOR",
            "DATA PAGAMENTO",
            "BAIXADO",
        ],
        "aliases": {
            "CONTRATO": {"CONTRATO"},
            "LOJA": {"LOJA", "LOJA "},
            "VALOR": {"VALOR", "VALOR ", "VALOR PAGO"},
            "DATA PAGAMENTO": {"DATA PAGAMENTO", "DATA_PAGAMENTO"},
            "BAIXADO": {"BAIXADO", "STATUS"},
        },
        "group_hint": "Agrupamento previsto: somar cash por loja.",
    },
    "despesas": {
        "label": "Despesas",
        "sheet_name": "DESPESAS",
        "header_row": 0,
        "pending_column": "STATUS",
        "required_columns": [
            "LOJA",
            "DATA DE DESPESA",
            "TIPO DA DESPESA",
            "DESCRICAO",
            "VALOR",
            "STATUS",
        ],
        "aliases": {
            "LOJA": {"LOJA"},
            "DATA DE DESPESA": {"DATA DE DESPESA", "DATA DESPESA"},
            "TIPO DA DESPESA": {"TIPO DA DESPESA", "TIPO DESPESA"},
            "DESCRICAO": {"DESCRICAO", "DESCRIÇÃO", "DESCRICAO ", "DESCRIÇÃO "},
            "VALOR": {"VALOR"},
            "STATUS": {"STATUS", "BAIXADO"},
        },
        "group_hint": "Agrupamento previsto: separar despesas por loja e tipo para rateio.",
    },
    "depositos": {
        "label": "Depósitos",
        "sheet_name": "DEPÓSITOS",
        "header_row": 1,
        "pending_column": "BAIXADO",
        "required_columns": [
            "DATA",
            "BANCO",
            "AGENCIA",
            "VALOR",
            "LOJA",
            "BAIXADO",
        ],
        "aliases": {
            "DATA": {"DATA"},
            "BANCO": {"BANCO"},
            "AGENCIA": {"AGENCIA", "AGÊNCIA"},
            "VALOR": {"VALOR"},
            "LOJA": {"LOJA"},
            "BAIXADO": {"BAIXADO", "STATUS"},
        },
        "group_hint": "Agrupamento previsto: consolidar depósitos por loja, dia e banco.",
    },
}


def _normalize_text(value: str) -> str:
    text = unicodedata.normalize("NFKD", str(value or "")).encode("ASCII", "ignore").decode("ASCII")
    return " ".join(text.strip().upper().split())


def _resolve_sheet_name(sheet_names: list[str], target_name: str) -> str | None:
    target_norm = _normalize_text(target_name)
    for sheet_name in sheet_names:
        if _normalize_text(sheet_name) == target_norm:
            return sheet_name
    return None


def _parse_date(value) -> datetime | None:
    if pd.isna(value):
        return None
    if isinstance(value, datetime):
        return value
    parsed = pd.to_datetime(value, dayfirst=True, errors="coerce")
    if pd.isna(parsed):
        return None
    return parsed.to_pydatetime()


def _format_brl(value: float) -> str:
    text = f"{float(value):,.2f}"
    return text.replace(",", "X").replace(".", ",").replace("X", ".")


def _format_totvs_money(value: float) -> str:
    return f"{float(value):.2f}".replace(".", ",")


def _generate_expense_title_number() -> str:
    return f"DESP{random.randint(0, 99999):05d}"


def _expense_account_for_type(tipo_despesa: str) -> str | None:
    tipo_norm = _normalize_text(tipo_despesa)
    if tipo_norm in EXPENSE_ACCOUNT_MAP:
        return EXPENSE_ACCOUNT_MAP[tipo_norm]
    for key, account in EXPENSE_ACCOUNT_MAP.items():
        if key in tipo_norm or tipo_norm in key:
            return account
    return None


def _normalize_store_code(value: str) -> str:
    store = _normalize_text(value)
    if store == "SAO11":
        return "CGH"
    if store == "SAO10":
        return "GRU"
    return store[:3]


def _resolve_cash_guide_path() -> Path:
    base_dir = Path(__file__).resolve().parent
    candidates = [
        base_dir.parent / "assets" / "CADASTRO_LOJA_CASH.xlsx",
        Path.cwd() / "DESENVOLVIMENTO" / "assets" / "CADASTRO_LOJA_CASH.xlsx",
        Path.cwd() / "assets" / "CADASTRO_LOJA_CASH.xlsx",
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return candidates[0]


def _resolve_expense_client_guide_path() -> Path:
    base_dir = Path(__file__).resolve().parent
    candidates = [
        base_dir.parent / "assets" / "CADASTRO_CLIENTE_DESPESAS.xlsx",
        Path.cwd() / "DESENVOLVIMENTO" / "assets" / "CADASTRO_CLIENTE_DESPESAS.xlsx",
        Path.cwd() / "assets" / "CADASTRO_CLIENTE_DESPESAS.xlsx",
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return candidates[0]


def _resolve_expense_rateio_path() -> Path:
    base_dir = Path(__file__).resolve().parent
    candidates = [
        base_dir.parent / "assets" / "DESPESAS.xlsx",
        Path.cwd() / "DESENVOLVIMENTO" / "assets" / "DESPESAS.xlsx",
        Path.cwd() / "assets" / "DESPESAS.xlsx",
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return candidates[0]


def _load_logo_candidates() -> list[Path]:
    candidates: list[Path] = []
    env_logo = os.environ.get("FOCO_LOGO_PNG", "").strip()
    env_assets = os.environ.get("FOCO_ASSETS_DIR", "").strip()
    if env_logo:
        candidates.append(Path(env_logo))
    if env_assets:
        candidates.append(Path(env_assets) / "logo.png")

    base_dir = Path(__file__).resolve().parent
    candidates.append(base_dir.parent / "assets" / "logo.png")
    candidates.append(Path.cwd() / "DESENVOLVIMENTO" / "assets" / "logo.png")
    candidates.append(Path.cwd() / "assets" / "logo.png")
    return candidates


@dataclass
class ValidationResult:
    flow_key: str
    flow_label: str
    sheet_name: str
    total_rows: int
    pending_rows: int
    launched_rows: int
    missing_columns: list[str]
    workbook_sheets: list[str]


@dataclass
class CashGroup:
    loja: str
    cliente_codigo: str
    valor_total: float
    periodo: str
    row_numbers: list[int]


@dataclass
class ExpenseGroup:
    loja: str
    tipo_despesa: str
    tipos_despesa: list[str]
    contas_por_tipo: dict[str, str]
    descricao: str
    cliente_codigo: str
    valor_total: float
    periodo: str
    row_numbers: list[int]

    @property
    def precisa_rateio(self) -> bool:
        return len(self.tipos_despesa) > 1


CASH_FIELD_LABELS = {
    "prefixo": "Prefixo",
    "tipo": "Tipo",
    "data_emissao": "Data de emissao",
    "vencimento": "Vencimento",
    "valor_titulo": "Valor do titulo",
    "centro_custo": "Centro de custo",
    "segregacao": "Segregacao",
    "historico": "Historico",
    "natureza": "Natureza",
    "cliente": "Cliente",
    "conta_baixa": "Conta da baixa",
}

CASH_FIELD_MAP = {
    "prefixo": {"protheus_name": "E1_PREFIXO", "tag": "wa-text-input"},
    "tipo_pg": {"protheus_name": "E1_XTIPOPG", "tag": "wa-combobox"},
    "tipo": {"protheus_name": "E1_TIPO", "tag": "wa-text-input"},
    "natureza": {"protheus_name": "E1_NATUREZ", "tag": "wa-text-input"},
    "cliente": {"protheus_name": "E1_CLIENTE", "tag": "wa-text-input"},
    "data_emissao": {"protheus_name": "E1_EMISSAO", "tag": "wa-text-input"},
    "vencimento": {"protheus_name": "E1_VENCTO", "tag": "wa-text-input"},
    "valor_titulo": {"protheus_name": "E1_VALOR", "tag": "wa-text-input"},
    "centro_custo": {"protheus_name": "E1_CCUSTO", "tag": "wa-text-input"},
    "negocio": {"protheus_name": "E1_NEGOCIO", "tag": "wa-combobox"},
    "segregacao": {"protheus_name": "E1_XSEGREG", "tag": "wa-text-input"},
    "motivo": {"protheus_name": "E1_MOTIVO2", "tag": "wa-combobox"},
    "historico": {"protheus_name": "E1_HIST", "tag": "wa-text-input"},
}

EXPENSE_FIELD_LABELS = {
    "numero_titulo": "Numero do Titulo",
    "tipo_titulo": "Tipo do Titulo",
    "natureza": "Codigo da natureza",
    "fornecedor": "Fornecedor",
    "data_emissao": "Data emissao",
    "vencimento": "Vencimento",
    "valor": "Valor",
    "historico": "Historico",
    "rateio": "Rateio",
    "conta_contabil": "Conta contabil",
    "centro_custo": "Centro de custo",
    "forma_pagamento": "Forma de pagamento",
}

EXPENSE_ACCOUNT_MAP = {
    "ESTORNOS REEMBOLSOS": "3102020003",
    "REEMBOLSOS": "3102020003",
    "COMBUSTIVEIS E LUBRIFICANTES": "4101010004",
    "ESTACIONAMENTO": "4101010016",
    "DESPESAS COM ESCRITORIOS": "4101010022",
    "DESPESAS DE ESCRITORIO": "4101010022",
    "EVENTOS E CONFRATERNIZACOES": "4101010025",
    "MANUTENCAO DE EQUIPAMENTO": "4101010051",
    "MATERIAL E SERVICOS DE LIMPEZA": "4101010056",
    "OUTROS GASTOS COM FROTA": "4101010061",
    "CORREIOS E MALOTES": "4101010050",
    "CORREIO E MALOTE": "4101010050",
    "VIAGENS AEREO HOSPEDAGEM ALIMENTACAO": "4101010024",
    "SERVICOS DE TERCEIROS": "4101010027",
}
DEFAULT_EXPENSE_ACCOUNT = "4101010022"
DEFAULT_EXPENSE_ACCOUNT_DESCRIPTION = "DESPESAS COM ESCRITORIOS"

HELP_FIELD_PATTERNS = {
    "tipo": ("E1_TIPO", "TIPO PROBLEMA", "TIPO DO TITULO"),
    "natureza": ("E1_NATUREZ", "NATUREZ", "NATUREZA PROBLEMA"),
    "centro_custo": ("NOCUSTO", "E1_CCUSTO", "C. CUSTO", "CENTRO DE CUSTO"),
    "segregacao": ("REGNOIS", "E1_XSEGREG", "SEGREGACAO", "SEGREGA"),
    "cliente": ("E1_CLIENTE", "CLIENTE"),
}


class TotvsCaixaLoginBot:
    def __init__(self, username: str, password: str, headless: bool, log_callback) -> None:
        self.username = username
        self.password = password
        self.headless = headless
        self.log = log_callback
        self.playwright = None
        self.browser = None
        self.context = None
        self.page = None

    def start(self) -> None:
        self.log("Inicializando navegador para acesso ao TOTVS...")
        self.playwright = sync_playwright().start()
        self.browser = self.playwright.chromium.launch(
            headless=self.headless,
            args=["--start-maximized"],
        )
        self.context = self.browser.new_context(viewport=None)
        self.page = self.context.new_page()
        self.page.goto(TOTVS_URL, wait_until="domcontentloaded", timeout=90000)
        self._login()

    def _login(self) -> None:
        if self.page is None:
            raise RuntimeError("Pagina do TOTVS nao foi inicializada.")

        self.log("Aguardando botao inicial Ok...")
        self._click_first_visible_locator(
            [
                self.page.get_by_role("button", name="Ok"),
                self.page.get_by_role("button", name="OK"),
                self.page.get_by_text("Ok", exact=True),
            ],
            timeout=60000,
            pause_ms=700,
        )

        frame = self.page.frame_locator("iframe")
        user_field = frame.get_by_role("textbox", name=re.compile("Insira seu usu.rio", re.I))
        password_field = frame.get_by_role("textbox", name=re.compile("Insira sua senha", re.I))

        self.log("Preenchendo credenciais...")
        user_field.click(timeout=60000)
        user_field.fill(self.username)
        password_field.click(timeout=60000)
        password_field.fill(self.password)

        self.log("Confirmando login...")
        frame.get_by_role("button", name="Entrar").click(timeout=60000)
        self.page.wait_for_timeout(1000)
        frame.get_by_role("button", name="Entrar").click(timeout=90000)

        self.log("Aguardando tela principal do TOTVS...")
        self._wait_main_menu_ready()
        self._click_first_visible_locator(
            [
                self.page.get_by_text(re.compile(r"Atualiza..es\s*\(\d+\)", re.I)),
                self.page.get_by_text("Atualizacoes", exact=False),
                self.page.get_by_text("Atualizações", exact=False),
            ],
            timeout=120000,
            pause_ms=1000,
        )
        self.log("Login concluido. Menu Atualizacoes aberto.")

    def _require_page(self):
        if self.page is None:
            raise RuntimeError("Pagina do TOTVS ainda nao foi inicializada.")
        return self.page

    def _click_when_clickable(self, locator, label: str, timeout: int = 60000, pause_ms: int = 500) -> None:
        page = self._require_page()
        locators = [locator]
        label_norm = _normalize_text(label)
        if "FUNCOES CONTAS A RECEBER" in label_norm:
            locators.extend(
                [
                    page.get_by_title("Funcoes Contas a Receber"),
                    page.get_by_title("Funções Contas a Receber"),
                    page.locator('wa-menu-item[caption*="Funcoes Contas a Receber"]'),
                    page.locator('wa-menu-item[caption*="Funções Contas a Receber"]'),
                    page.get_by_text("Funcoes Contas a Receber", exact=False),
                    page.get_by_text("Funções Contas a Receber", exact=False),
                ]
            )
        elif "FUNCOES CONTAS A PAGAR" in label_norm:
            locators.extend(
                [
                    page.get_by_title("Funcoes Contas a Pagar"),
                    page.get_by_title("Funções Contas a Pagar"),
                    page.get_by_title("FunÃ§Ãµes Contas a Pagar"),
                    page.locator('wa-menu-item[caption*="Funcoes Contas a Pagar"]'),
                    page.locator('wa-menu-item[caption*="Funções Contas a Pagar"]'),
                    page.locator('wa-menu-item[caption*="FunÃ§Ãµes Contas a Pagar"]'),
                    page.get_by_text("Funcoes Contas a Pagar", exact=False),
                    page.get_by_text("Funções Contas a Pagar", exact=False),
                    page.get_by_text("FunÃ§Ãµes Contas a Pagar", exact=False),
                ]
            )
        elif "OUTRAS ACOES" in label_norm:
            locators.extend(
                [
                    page.get_by_role("button", name=re.compile("Outras A..es|Outras Acoes", re.I)),
                    page.get_by_text("Outras Ações", exact=True),
                    page.get_by_text("Outras Acoes", exact=True),
                    page.locator('wa-button[caption*="Outras"]'),
                ]
            )
        elif label_norm == "BAIXAS":
            locators.extend(
                [
                    page.get_by_text("Baixas", exact=True),
                    page.locator('wa-menu-popup-item[caption="Baixas"]'),
                    page.locator('wa-menu-popup-item:has-text("Baixas")'),
                ]
            )
        self._click_first_visible_locator(locators, timeout=timeout, pause_ms=pause_ms, label=label)

    def _click_first_visible_locator(self, locators, timeout: int = 30000, pause_ms: int = 300, label: str = "elemento") -> None:
        page = self._require_page()
        end_time = page.evaluate("Date.now()") + timeout
        last_error = ""

        while page.evaluate("Date.now()") < end_time:
            for locator in locators:
                try:
                    count = locator.count()
                    if count == 0:
                        continue
                    for index in range(count):
                        candidate = locator.nth(index)
                        if not candidate.is_visible():
                            continue
                        candidate.scroll_into_view_if_needed(timeout=3000)
                        candidate.click(force=True)
                        page.wait_for_timeout(pause_ms)
                        self.log(f"Clique OK em {label}.")
                        return
                except Exception as exc:
                    last_error = str(exc)
                    continue
            page.wait_for_timeout(250)

        raise RuntimeError(f"Nao foi possivel clicar em {label}. Ultima falha: {last_error}")

    def _first_visible_locator(self, locators):
        for locator in locators:
            try:
                count = locator.count()
                if count == 0:
                    continue
                for index in range(count):
                    candidate = locator.nth(index)
                    if candidate.is_visible():
                        return candidate
            except Exception:
                continue
        return None

    def _wait_for_first_visible_locator(self, locators, timeout: int = 30000, pause_ms: int = 250):
        page = self._require_page()
        end_time = page.evaluate("Date.now()") + timeout

        while page.evaluate("Date.now()") < end_time:
            candidate = self._first_visible_locator(locators)
            if candidate is not None:
                return candidate
            page.wait_for_timeout(pause_ms)
        return None

    def _wait_any_visible(self, locators, timeout: int = 30000) -> None:
        page = self._require_page()
        end_time = page.evaluate("Date.now()") + timeout

        while page.evaluate("Date.now()") < end_time:
            if self._is_any_locator_visible(locators):
                return
            page.wait_for_timeout(250)
        raise RuntimeError("Nao foi possivel validar a tela esperada no Protheus.")

    def _is_any_locator_visible(self, locators) -> bool:
        for locator in locators:
            try:
                count = locator.count()
                if count == 0:
                    continue
                for index in range(count):
                    if locator.nth(index).is_visible():
                        return True
            except Exception:
                continue
        return False

    def _wait_main_menu_ready(self) -> None:
        page = self._require_page()
        candidates = [
            page.locator('wa-menu-item:has-text("Atualizacoes")'),
            page.locator('wa-menu-item:has-text("Atualizações")'),
            page.locator('wa-menu-item:has-text("Contas a Receber")'),
            page.get_by_text("Atualizacoes", exact=False),
            page.get_by_text("Atualizações", exact=False),
            page.get_by_text("Contas a Receber", exact=False),
        ]
        self._wait_any_visible(candidates, timeout=120000)

    def open_cash_entry_screen(self) -> None:
        if self.page is None:
            raise RuntimeError("Pagina do TOTVS nao foi inicializada.")

        self.log("Abrindo modulo de Cash em Contas a Receber...")
        self._navigate_to_contas_receber()
        self._click_when_clickable(self.page.get_by_title("Funções Contas a Receber"), "Funcoes Contas a Receber")
        self._click_ctas_receber_button()
        self._click_incluir_menu()
        self._finalize_incluir()
        self.wait_until_billing_screen()
        self.log("Tela de faturamento aberta para lancamento de Cash.")

    def open_expense_entry_screen(self) -> None:
        if self.page is None:
            raise RuntimeError("Pagina do TOTVS nao foi inicializada.")

        self.log("Abrindo modulo de Despesas em Contas a Pagar...")
        self._navigate_to_contas_pagar()
        self._click_when_clickable(self.page.get_by_title("Funções Contas a Pagar"), "Funcoes Contas a Pagar")
        self._click_cancelar_contas_pagar_menu()
        self._click_ctas_pagar_button()
        self._click_incluir_menu()
        self._finalize_incluir()
        self.wait_until_payable_screen()
        self.log("Tela de Contas a Pagar aberta para lancamento de Despesas.")

    def _navigate_to_contas_receber(self) -> None:
        page = self._require_page()
        contas_candidates = [
            page.get_by_text("Contas a Receber (27)", exact=True),
            page.get_by_text("Contas a Receber", exact=False),
        ]
        if self._is_any_locator_visible(contas_candidates):
            self._click_first_visible_locator(contas_candidates, timeout=10000, pause_ms=1200, label="Contas a Receber")
            return

        possible_steps = [
            [page.get_by_text("Atualiza", exact=False), page.get_by_text("Contas a Receber", exact=False)],
            [page.get_by_text("Atualizacoes (17)", exact=True), page.get_by_text("Contas a Receber (27)", exact=True)],
        ]

        for steps in possible_steps:
            try:
                for level, locator in enumerate(steps, 1):
                    self._click_first_visible_locator([locator], timeout=45000, pause_ms=1000 + level * 250, label=f"menu nivel {level}")
                return
            except Exception:
                page.wait_for_timeout(1000)

        raise RuntimeError("Nenhum caminho de navegacao para Contas a Receber funcionou no menu do Protheus.")

    def _navigate_to_contas_pagar(self) -> None:
        page = self._require_page()
        self._click_first_visible_locator(
            [
                page.get_by_text("Atualizações (17)", exact=True),
                page.get_by_text("Atualizacoes (17)", exact=True),
                page.get_by_text(re.compile(r"Atualiza..es\s*\(\d+\)", re.I)),
                page.get_by_text("Atualiza", exact=False),
            ],
            timeout=45000,
            pause_ms=1000,
            label="Atualizacoes",
        )
        self._click_first_visible_locator(
            [
                page.get_by_text("Contas a Pagar (22)", exact=True),
                page.get_by_text(re.compile(r"Contas a Pagar\s*\(\d+\)", re.I)),
                page.get_by_text("Contas a Pagar", exact=False),
            ],
            timeout=45000,
            pause_ms=1200,
            label="Contas a Pagar",
        )

    def _click_ctas_receber_button(self) -> None:
        page = self._require_page()
        self._click_first_visible_locator(
            [
                page.get_by_role("button", name="Ctas a Receber"),
                page.locator("#COMP4599"),
                page.locator('wa-button[caption*="Ctas a Receber"]'),
            ],
            timeout=45000,
            pause_ms=1200,
            label="Ctas a Receber",
        )

    def _click_ctas_pagar_button(self) -> None:
        page = self._require_page()
        self._click_first_visible_locator(
            [
                page.get_by_role("button", name="Ctas a Pagar"),
                page.get_by_role("button", name="Contas a Pagar"),
                page.locator('wa-button[caption*="Ctas a Pagar"]'),
                page.locator('wa-button[caption*="Contas a Pagar"]'),
            ],
            timeout=45000,
            pause_ms=1200,
            label="Ctas a Pagar",
        )

    def _click_cancelar_contas_pagar_menu(self) -> None:
        page = self._require_page()
        self._click_first_visible_locator(
            [
                page.get_by_role("button", name="Cancelar"),
                page.locator('wa-button[caption="Cancelar"]'),
                page.get_by_text("Cancelar", exact=True),
            ],
            timeout=45000,
            pause_ms=1200,
            label="Cancelar menu Contas a Pagar",
        )

    def _click_incluir_menu(self) -> None:
        page = self._require_page()
        self._click_first_visible_locator(
            [
                page.get_by_text("Incluir", exact=True),
                page.locator('wa-menu-popup-item[caption="Incluir"]'),
                page.locator('wa-menu-popup-item:has-text("Incluir")'),
                page.locator('wa-text-view[caption*="Incluir"]'),
                page.locator("wa-menu-popup-item#COMP4602"),
            ],
            timeout=45000,
            pause_ms=1200,
            label="Incluir",
        )

    def _finalize_incluir(self) -> None:
        page = self._require_page()
        self._click_first_visible_locator(
            [
                page.get_by_role("button", name="OK"),
                page.locator("#COMP6057"),
                page.get_by_text("OK", exact=True),
                page.locator('wa-button:has-text("OK")'),
                page.locator('button:has-text("OK")'),
            ],
            timeout=30000,
            pause_ms=700,
            label="OK de inclusao",
        )

    def wait_until_billing_screen(self) -> None:
        if self.page is None:
            raise RuntimeError("Pagina do TOTVS nao foi inicializada.")
        dialog = self._billing_dialog()
        self._wait_any_visible(
            [
                dialog.locator('wa-text-view[caption="Contas a Receber"]'),
                dialog.locator('wa-tab-button#BUTTON-COMP6003[active]'),
                dialog.get_by_role("button", name="D ados Gerais"),
            ],
            timeout=45000,
        )
        self._wait_any_visible(
            [
                dialog.get_by_role("button", name="Salvar"),
                dialog.locator('wa-button[caption="Salvar"]'),
                dialog.locator("#COMP6156"),
            ],
            timeout=45000,
        )
        self._wait_any_visible(
            [
                dialog.get_by_title(re.compile("Prefixo", re.I)).get_by_role("textbox"),
                dialog.locator("#COMP6021 > input"),
            ],
            timeout=45000,
        )
        self.page.wait_for_timeout(500)

    def wait_until_payable_screen(self) -> None:
        if self.page is None:
            raise RuntimeError("Pagina do TOTVS nao foi inicializada.")
        dialog = self._payable_dialog()
        self._wait_any_visible(
            [
                dialog.locator('wa-text-view[caption="Contas a Pagar"]'),
                dialog.get_by_text("Contas a Pagar", exact=False),
                dialog.get_by_role("button", name=re.compile("Dados Gerais|D ados Gerais", re.I)),
            ],
            timeout=60000,
        )
        self._wait_any_visible(
            [
                dialog.get_by_role("button", name="Salvar"),
                dialog.locator('wa-button[caption="Salvar"]'),
            ],
            timeout=45000,
        )
        self.page.wait_for_timeout(500)

    def _billing_dialog(self):
        page = self._require_page()
        dialog = page.locator('wa-dialog.dict-msdialog[title="Contas a Receber"][opened]').last
        dialog.wait_for(state="visible", timeout=45000)
        return dialog

    def _payable_dialog(self):
        page = self._require_page()
        candidates = [
            page.locator('wa-dialog.dict-msdialog[title="Contas a Pagar"][opened]').last,
            page.locator('wa-dialog[title="Contas a Pagar"][opened]').last,
            page.locator('wa-dialog.dict-msdialog[opened]').filter(has_text="Contas a Pagar").last,
        ]
        dialog = self._wait_for_first_visible_locator(candidates, timeout=60000)
        if dialog is None:
            raise RuntimeError("Tela de Contas a Pagar nao foi localizada.")
        return dialog

    def _fill_text_field(
        self,
        locator,
        label: str,
        value: str,
        timeout: int = 60000,
        field_key: str | None = None,
        settle_locator=None,
    ) -> None:
        self._click_and_fill_field(locator, value, field_key=field_key, attempts=3, timeout=timeout, settle_locator=settle_locator)
        self.log(f"Campo {label} preenchido e validado.")

    def _fill_first_available(self, locators: list, label: str, value: str) -> None:
        last_error = ""
        for locator in locators:
            try:
                field_key = "prefixo" if _normalize_text(label) == "PREFIXO" else None
                self._fill_text_field(locator, label, value, timeout=12000, field_key=field_key, settle_locator=locator)
                return
            except Exception as exc:
                last_error = str(exc)
        raise RuntimeError(f"Nao foi possivel preencher {label}. Ultima falha: {last_error}")

    def _select_option(self, locator, label: str, option: str) -> None:
        self._select_combo_option(locator, option, label=label)
        self.log(f"Campo {label} selecionado: {option}.")

    @staticmethod
    def _build_cash_field_selector(field_key: str) -> str:
        field = CASH_FIELD_MAP[field_key]
        return f'{field["tag"]}[name="M->{field["protheus_name"]}"]'

    def _cash_field_locator(self, dialog, field_key: str, fallbacks: list | None = None):
        selector = self._build_cash_field_selector(field_key)
        candidates = [dialog.locator(selector).last]
        if fallbacks:
            candidates.extend(fallbacks)
        locator = self._first_visible_locator(candidates)
        if locator is None:
            raise RuntimeError(
                f"Campo {CASH_FIELD_LABELS.get(field_key, field_key)} nao foi localizado pelo identificador interno do TOTVS ({selector})."
            )
        return locator

    def _expense_field_locator(self, dialog, field_key: str):
        page = self._require_page()
        locators = {
            "numero_titulo": [
                dialog.get_by_title("Numero do Titulo         ").get_by_role("textbox"),
                page.get_by_title("Numero do Titulo         ").get_by_role("textbox"),
            ],
            "tipo_titulo": [
                dialog.get_by_title("Tipo do Titulo           ").get_by_role("textbox"),
                page.get_by_title("Tipo do Titulo           ").get_by_role("textbox"),
            ],
            "natureza": [
                dialog.get_by_title("Codigo da natureza       ").get_by_role("textbox"),
                page.get_by_title("Codigo da natureza       ").get_by_role("textbox"),
            ],
            "fornecedor": [dialog.locator("#COMP6022 > input"), page.locator("#COMP6022 > input")],
            "data_emissao": [dialog.locator("#COMP6025 > input"), page.locator("#COMP6025 > input")],
            "vencimento": [dialog.locator("#COMP6026 > input"), page.locator("#COMP6026 > input")],
            "valor": [dialog.locator("#COMP6028 > input"), page.locator("#COMP6028 > input")],
            "historico": [dialog.locator("#COMP6029 > input"), page.locator("#COMP6029 > input")],
            "rateio": [dialog.get_by_role("combobox"), page.get_by_role("combobox")],
            "conta_contabil": [dialog.locator("#COMP6035 > input"), page.locator("#COMP6035 > input")],
            "centro_custo": [dialog.locator("#COMP6038 > input"), page.locator("#COMP6038 > input")],
            "forma_pagamento": [dialog.locator("#COMP6040 > input"), page.locator("#COMP6040 > input")],
        }
        locator = self._first_visible_locator(locators.get(field_key, []))
        if locator is None:
            raise RuntimeError(f"Campo {EXPENSE_FIELD_LABELS.get(field_key, field_key)} nao foi localizado.")
        return locator

    def _wait_first_visible_locator(self, locators: list, timeout_ms: int = 45000, interval_ms: int = 500):
        page = self._require_page()
        deadline = time.time() + (timeout_ms / 1000)
        last_error = ""
        while time.time() < deadline:
            for locator in locators:
                try:
                    if locator.count() > 0 and locator.first.is_visible():
                        return locator.first
                except Exception as exc:
                    last_error = str(exc)
                    continue
            page.wait_for_timeout(interval_ms)
        if last_error:
            self.log(f"Ultima falha ao aguardar campo: {last_error}")
        return None

    def _click_and_fill_field(
        self,
        locator,
        value: str,
        field_key: str | None = None,
        attempts: int = 3,
        timeout: int = 45000,
        settle_locator=None,
    ) -> None:
        if locator is None:
            raise RuntimeError("Campo de entrada nao foi localizado.")

        page = self._require_page()
        expected = str(value)
        last_value = ""
        for attempt in range(1, attempts + 1):
            locator.wait_for(state="visible", timeout=timeout)
            locator.scroll_into_view_if_needed(timeout=3000)
            if not self._focus_and_verify_host_input(locator):
                raise RuntimeError(f"Nao foi possivel confirmar foco no campo {CASH_FIELD_LABELS.get(field_key or '', field_key or 'campo')}.")
            page.wait_for_timeout(250)

            try:
                self._clipboard_fill_locator(locator, expected)
            except Exception:
                self._keyboard_fill_locator(locator, expected)

            if settle_locator is not None:
                self._stabilize_on_previous_field(settle_locator, field_key)
            else:
                page.wait_for_timeout(1200)
            error_field = self._dismiss_help_popup_if_present()
            if error_field:
                if field_key and error_field != "desconhecido" and error_field != field_key:
                    raise RuntimeError(
                        f"TOTVS apontou erro no campo {CASH_FIELD_LABELS.get(error_field, error_field)} "
                        f"enquanto o robo preenchia {CASH_FIELD_LABELS.get(field_key, field_key)}. "
                        "A loja sera interrompida para evitar preenchimento cruzado."
                    )
                self.log(f"Erro do TOTVS detectado no campo {error_field}. Tentando novamente ({attempt}/{attempts}).")
                continue

            last_value = self._read_from_locator(locator)
            if field_key is None:
                if last_value.strip():
                    return
            elif self._field_value_matches(field_key, expected, last_value):
                return

            self.log(f"Validacao do campo falhou ({attempt}/{attempts}). Esperado: '{expected}' | Lido: '{last_value}'")

        label = CASH_FIELD_LABELS.get(field_key or "", field_key or "campo")
        raise RuntimeError(f"Nao foi possivel preencher corretamente {label}. Esperado '{expected}', lido '{last_value}'.")

    def _clipboard_fill_locator(self, locator, value: str) -> None:
        page = self._require_page()
        locator.wait_for(state="visible", timeout=45000)
        locator.scroll_into_view_if_needed(timeout=3000)
        if not self._focus_and_verify_host_input(locator):
            raise RuntimeError("Nao foi possivel focar o campo correto antes de colar.")
        page.keyboard.press("Control+A")
        page.wait_for_timeout(120)
        self._set_clipboard_text(str(value))
        page.keyboard.press("Control+V")
        page.wait_for_timeout(650)

    def _set_clipboard_text(self, value: str) -> None:
        page = self._require_page()
        try:
            page.evaluate("async (text) => await navigator.clipboard.writeText(text)", value)
            return
        except Exception:
            pass
        try:
            import tkinter as _tk

            root = _tk.Tk()
            root.withdraw()
            root.clipboard_clear()
            root.clipboard_append(value)
            root.update()
            root.destroy()
            return
        except Exception:
            pass
        raise RuntimeError("Nao foi possivel gravar o valor na area de transferencia.")

    def _keyboard_fill_locator(self, locator, value: str) -> None:
        page = self._require_page()
        locator.wait_for(state="visible", timeout=45000)
        locator.scroll_into_view_if_needed(timeout=3000)
        if not self._focus_and_verify_host_input(locator):
            raise RuntimeError("Nao foi possivel focar o campo correto antes de preencher.")
        page.keyboard.press("Control+A")
        page.wait_for_timeout(120)
        page.keyboard.press("Backspace")
        page.wait_for_timeout(120)
        page.keyboard.insert_text(str(value))
        page.wait_for_timeout(500)

    def _stabilize_on_previous_field(self, locator, current_field_key: str | None = None) -> None:
        page = self._require_page()
        if locator is None:
            page.wait_for_timeout(1200)
            return

        label = CASH_FIELD_LABELS.get(current_field_key or "", current_field_key or "campo")
        try:
            locator.wait_for(state="visible", timeout=45000)
            locator.scroll_into_view_if_needed(timeout=3000)
            focused = self._focus_host_input(locator)
            if not focused:
                locator.click(force=True)
            page.wait_for_timeout(1200)
            self.log(f"Campo {label} estabilizado voltando ao campo anterior.")
        except Exception as exc:
            raise RuntimeError(f"Nao foi possivel estabilizar o TOTVS apos preencher {label}: {exc}") from exc

    def _focus_host_input(self, locator) -> bool:
        try:
            locator.click(force=True)
            return bool(
                locator.evaluate(
                    """
                    (host) => {
                        const root = host.shadowRoot || host;
                        const input =
                            root.querySelector('input') ||
                            root.querySelector('textarea') ||
                            host.querySelector('input') ||
                            host.querySelector('textarea');
                        if (!input) return false;
                        input.focus();
                        const active = root.activeElement || document.activeElement;
                        return active === input || document.activeElement === host || document.activeElement === input;
                    }
                    """
                )
            )
        except Exception:
            return False

    def _focus_and_verify_host_input(self, locator) -> bool:
        if locator is None:
            return False
        for _ in range(3):
            try:
                locator.scroll_into_view_if_needed(timeout=3000)
            except Exception:
                pass
            try:
                locator.click(timeout=3000)
            except Exception:
                try:
                    locator.click(force=True, timeout=3000)
                except Exception:
                    pass
            try:
                focused = bool(
                    locator.evaluate(
                        """
                        (host) => {
                            const root = host.shadowRoot || host;
                            const input =
                                root.querySelector('input') ||
                                root.querySelector('textarea') ||
                                host.querySelector('input') ||
                                host.querySelector('textarea') ||
                                (host.tagName && host.tagName.toLowerCase() === 'input' ? host : null);
                            if (!input) return false;
                            input.focus();
                            const active = root.activeElement || document.activeElement;
                            return active === input || document.activeElement === host || document.activeElement === input;
                        }
                        """
                    )
                )
                if focused:
                    return True
            except Exception:
                pass
            self._require_page().wait_for_timeout(200)
        return False

    def _fill_host_input(self, locator, value: str, dismiss_popup: bool = True) -> None:
        page = self._require_page()
        locator.wait_for(state="visible", timeout=45000)
        locator.scroll_into_view_if_needed(timeout=3000)
        locator.click(force=True)

        filled = False
        try:
            filled = bool(
                locator.evaluate(
                    """
                    (host, value) => {
                        const root = host.shadowRoot || host;
                        const input =
                            root.querySelector('input') ||
                            root.querySelector('textarea') ||
                            host.querySelector('input') ||
                            host.querySelector('textarea');
                        if (!input) return false;
                        input.focus();
                        input.value = '';
                        input.dispatchEvent(new Event('input', { bubbles: true, composed: true }));
                        input.value = String(value);
                        input.dispatchEvent(new Event('input', { bubbles: true, composed: true }));
                        input.dispatchEvent(new Event('change', { bubbles: true, composed: true }));
                        return true;
                    }
                    """,
                    value,
                )
            )
        except Exception:
            filled = False

        if not filled:
            locator.click(force=True)
            page.keyboard.press("Control+A")
            page.keyboard.press("Backspace")
            page.keyboard.type(str(value), delay=100)

        page.wait_for_timeout(500)
        if dismiss_popup:
            self._dismiss_help_popup_if_present()

    def _select_combo_option(self, locator, value: str, label: str = "") -> None:
        if locator is None:
            raise RuntimeError("Combobox nao localizado.")
        page = self._require_page()
        locator.wait_for(state="visible", timeout=45000)
        locator.scroll_into_view_if_needed(timeout=3000)

        try:
            locator.select_option(str(value))
            page.wait_for_timeout(800)
            self._dismiss_help_popup_if_present()
            return
        except Exception:
            pass

        if self._set_totvs_combobox_value(locator, str(value)):
            page.wait_for_timeout(800)
            self._dismiss_help_popup_if_present()
            return

        locator.click(force=True)
        page.wait_for_timeout(250)
        page.keyboard.press("Control+A")
        page.wait_for_timeout(100)
        page.keyboard.press("Backspace")
        page.wait_for_timeout(100)
        try:
            self._set_clipboard_text(str(value))
            page.keyboard.press("Control+V")
        except Exception:
            page.keyboard.insert_text(str(value))
        page.wait_for_timeout(200)
        page.keyboard.press("Enter")
        page.wait_for_timeout(300)
        page.keyboard.press("Tab")
        page.wait_for_timeout(800)
        self._dismiss_help_popup_if_present()

    def _set_totvs_combobox_value(self, locator, value: str) -> bool:
        try:
            return bool(
                locator.evaluate(
                    """
                    (node, rawValue) => {
                        const value = String(rawValue);
                        const selectedIndex = Number.parseInt(value, 10);
                        const combo = node.closest && node.closest('wa-combobox')
                            ? node.closest('wa-combobox')
                            : node;
                        if (!combo) return false;

                        const applyValue = (target) => {
                            if (!target) return;
                            try { target.value = value; } catch (error) {}
                            try { target.selectedIndex = selectedIndex; } catch (error) {}
                            try { target.selectedindex = selectedIndex; } catch (error) {}
                            try { target.setAttribute('value', value); } catch (error) {}
                            try { target.setAttribute('selectedindex', value); } catch (error) {}
                            try { target.setAttribute('selectedIndex', value); } catch (error) {}
                            try {
                                target.dispatchEvent(new Event('input', { bubbles: true, composed: true }));
                                target.dispatchEvent(new Event('change', { bubbles: true, composed: true }));
                                target.dispatchEvent(
                                    new CustomEvent('change', {
                                        bubbles: true,
                                        composed: true,
                                        detail: { value, selectedIndex },
                                    })
                                );
                            } catch (error) {}
                        };

                        applyValue(combo);

                        const root = combo.shadowRoot || combo;
                        const input =
                            root.querySelector('input') ||
                            root.querySelector('[contenteditable="true"]') ||
                            combo.querySelector('input');
                        if (input) {
                            input.focus();
                            applyValue(input);
                        }

                        const button =
                            root.querySelector('button') ||
                            root.querySelector('[role="button"]') ||
                            combo.querySelector('button');
                        if (button) {
                            try { button.dispatchEvent(new Event('change', { bubbles: true, composed: true })); } catch (error) {}
                        }

                        return true;
                    }
                    """,
                    value,
                )
            )
        except Exception:
            return False

    def _dismiss_help_popup_if_present(self) -> str | None:
        page = self._require_page()
        try:
            help_dialog = self._find_visible_totvs_help_dialog()
            if help_dialog is None:
                return None
            field_key = self._identify_help_field(help_dialog.inner_text() or "")
            self.log(f"Aviso do TOTVS detectado. Campo identificado: {field_key}.")
            close_button = self._first_visible_locator(
                [
                    help_dialog.get_by_role("button", name="Fechar"),
                    help_dialog.get_by_text("Fechar", exact=True),
                    help_dialog.locator('wa-button[caption="Fechar"]'),
                    help_dialog.locator('button:has-text("Fechar")'),
                    page.get_by_role("button", name="Fechar"),
                    page.get_by_text("Fechar", exact=True),
                ]
            )
            if close_button is not None:
                close_button.click(force=True)
                page.wait_for_timeout(1000)
                self.log("Aviso do TOTVS fechado.")
                return field_key
            self.log("Aviso do TOTVS detectado, mas o botao Fechar nao foi localizado.")
        except Exception:
            return None
        return None

    def _find_visible_totvs_help_dialog(self):
        page = self._require_page()
        dialog_selectors = [
            "wa-dialog.dict-msdialog",
            "wa-dialog",
            "div[role='dialog']",
        ]
        help_markers = ("PROBLEMA", "HELP:", "CAMPO", "E1_", "NOCUSTO", "REGNOIS")

        for selector in dialog_selectors:
            try:
                dialogs = page.locator(selector)
                count = dialogs.count()
            except Exception:
                continue

            for index in range(count - 1, -1, -1):
                try:
                    dialog = dialogs.nth(index)
                    if not dialog.is_visible():
                        continue
                    text = _normalize_text(dialog.inner_text() or "")
                    if any(marker in text for marker in help_markers):
                        return dialog
                except Exception:
                    continue
        return None

    def _raise_if_totvs_error_present(self) -> None:
        page = self._require_page()
        help_dialog = self._find_visible_totvs_help_dialog()
        if help_dialog is not None:
            text = _normalize_text(help_dialog.inner_text() or "")
            raise RuntimeError(f"TOTVS apresentou aviso antes de concluir o salvamento: {text[:350]}")

        try:
            dialogs = page.locator("wa-dialog.dict-msdialog")
            count = dialogs.count()
        except Exception:
            return

        error_markers = ("PROBLEMA:", "ERRO", "ERROR", "INCONSIST", "ATENCAO", "ATENCAO")
        for index in range(count):
            try:
                dialog = dialogs.nth(index)
                if not dialog.is_visible():
                    continue
                text = _normalize_text(dialog.inner_text() or "")
                if any(marker in text for marker in error_markers):
                    raise RuntimeError(f"TOTVS apresentou erro antes de concluir o salvamento: {text[:350]}")
            except RuntimeError:
                raise
            except Exception:
                continue

    def _identify_help_field(self, text: str) -> str:
        normalized = _normalize_text(text)
        for field_key, patterns in HELP_FIELD_PATTERNS.items():
            if any(_normalize_text(pattern) in normalized for pattern in patterns):
                return field_key
        return "desconhecido"

    def _field_value_matches(self, field_key: str, expected: str, actual: str) -> bool:
        expected = str(expected).strip()
        actual = str(actual).strip()
        if field_key == "valor_titulo":
            expected_value = self._parse_money_for_compare(expected)
            actual_value = self._parse_money_for_compare(actual)
            return expected_value is not None and actual_value is not None and abs(expected_value - actual_value) < 0.01
        if field_key in {"data_emissao", "vencimento"}:
            return "".join(char for char in expected if char.isdigit()) == "".join(char for char in actual if char.isdigit())

        expected_norm = _normalize_text(expected)
        actual_norm = _normalize_text(actual)
        return actual_norm == expected_norm or actual_norm.startswith(expected_norm)

    @staticmethod
    def _parse_money_for_compare(value: str) -> float | None:
        text = str(value).strip().replace("R$", "").replace(" ", "")
        if not text:
            return None
        if "," in text and "." in text:
            text = text.replace(".", "").replace(",", ".")
        elif "," in text:
            text = text.replace(",", ".")
        try:
            return float(text)
        except Exception:
            return None

    def _read_from_locator(self, locator) -> str:
        if locator is None:
            return ""
        try:
            if locator.evaluate("el => el.tagName && el.tagName.toLowerCase() === 'input'"):
                return str(locator.input_value()).strip()
        except Exception:
            pass
        try:
            return str(
                locator.evaluate(
                    """
                    (host) => {
                        const root = host.shadowRoot || host;
                        const input =
                            root.querySelector('input') ||
                            root.querySelector('textarea') ||
                            host.querySelector('input') ||
                            host.querySelector('textarea');
                        if (input) return input.value || '';
                        return host.value || host.textContent || '';
                    }
                    """
                )
                or ""
            ).strip()
        except Exception:
            return ""

    def fill_cash_title(self, group: CashGroup, issue_date: str) -> None:
        if self.page is None:
            raise RuntimeError("Pagina do TOTVS nao foi inicializada.")

        self.wait_until_billing_screen()
        dialog = self._billing_dialog()
        self.log(f"Preenchendo Cash da loja {group.loja} no valor R$ {_format_brl(group.valor_total)}.")

        prefixo_input = self._cash_field_locator(dialog, "prefixo")
        tipo_pg_combo = self._first_visible_locator(
            [
                dialog.get_by_title("Tipo de pagamento        ").get_by_role("combobox"),
                self.page.get_by_title("Tipo de pagamento        ").get_by_role("combobox"),
                self._cash_field_locator(dialog, "tipo_pg"),
            ]
        )
        if tipo_pg_combo is None:
            raise RuntimeError("Campo Tipo de pagamento nao foi localizado.")
        tipo_input = self._cash_field_locator(dialog, "tipo")
        data_emissao_input = self._cash_field_locator(dialog, "data_emissao")
        vencimento_input = self._cash_field_locator(dialog, "vencimento")
        valor_input = self._cash_field_locator(dialog, "valor_titulo")
        centro_custo_input = self._cash_field_locator(dialog, "centro_custo")
        negocio_combo = self._first_visible_locator(
            [
                dialog.get_by_title("Negocio                  ").get_by_role("combobox"),
                self.page.get_by_title("Negocio                  ").get_by_role("combobox"),
                self._cash_field_locator(dialog, "negocio"),
            ]
        )
        if negocio_combo is None:
            raise RuntimeError("Campo Negocio nao foi localizado.")
        segregacao_input = self._cash_field_locator(dialog, "segregacao")
        motivo_combo = self._first_visible_locator(
            [
                dialog.get_by_title("Motivo                   ").get_by_role("combobox"),
                self.page.get_by_title("Motivo                   ").get_by_role("combobox"),
                self._cash_field_locator(dialog, "motivo"),
            ]
        )
        if motivo_combo is None:
            raise RuntimeError("Campo Motivo nao foi localizado.")
        historico_input = self._cash_field_locator(dialog, "historico")
        natureza_input = self._cash_field_locator(dialog, "natureza")
        cliente_input = self._cash_field_locator(
            dialog,
            "cliente",
            fallbacks=[
                dialog.locator(
                    "xpath=/html/body/wa-dialog/wa-panel/wa-panel[2]/wa-tab-view/wa-tab-page/wa-dialog[2]/wa-tab-view/wa-tab-page[1]/wa-panel/wa-panel/wa-text-input[7]"
                ),
                dialog.locator(
                    "xpath=/html/body/wa-dialog/wa-panel/wa-panel[2]/wa-tab-view/wa-tab-page/wa-dialog[2]/wa-tab-view/wa-tab-page[1]/wa-panel/wa-panel/wa-text-input[7]//input"
                ),
                self.page.locator(
                    "xpath=/html/body/wa-dialog/wa-panel/wa-panel[2]/wa-tab-view/wa-tab-page/wa-dialog[2]/wa-tab-view/wa-tab-page[1]/wa-panel/wa-panel/wa-text-input[7]"
                ),
                self.page.locator(
                    "xpath=/html/body/wa-dialog/wa-panel/wa-panel[2]/wa-tab-view/wa-tab-page/wa-dialog[2]/wa-tab-view/wa-tab-page[1]/wa-panel/wa-panel/wa-text-input[7]//input"
                ),
            ],
        )

        self._fill_text_field(prefixo_input, "Prefixo", "DH", field_key="prefixo", settle_locator=prefixo_input)
        self._fill_text_field(tipo_input, "Tipo", "R$", field_key="tipo", settle_locator=prefixo_input)
        self._fill_text_field(data_emissao_input, "Data de emissao", issue_date, field_key="data_emissao", settle_locator=tipo_input)
        self._fill_text_field(vencimento_input, "Vencimento", issue_date, field_key="vencimento", settle_locator=data_emissao_input)
        self._fill_text_field(valor_input, "Valor do titulo", _format_totvs_money(group.valor_total), field_key="valor_titulo", settle_locator=vencimento_input)
        self._fill_text_field(centro_custo_input, "Centro de custo", group.loja, field_key="centro_custo", settle_locator=valor_input)
        self._select_option(negocio_combo, "Negocio", "3")
        self._stabilize_on_previous_field(centro_custo_input, "negocio")
        self._fill_text_field(segregacao_input, "Segregacao", "FIL", field_key="segregacao", settle_locator=centro_custo_input)
        self._select_option(motivo_combo, "Motivo", "6")
        self._stabilize_on_previous_field(segregacao_input, "motivo")
        self._fill_text_field(historico_input, "Historico", f"RECEITA DE {group.periodo} - {group.loja}", field_key="historico", settle_locator=segregacao_input)
        self._select_option(tipo_pg_combo, "Tipo de pagamento", "2")
        self._stabilize_on_previous_field(historico_input, "tipo_pagamento")
        self._fill_text_field(natureza_input, "Natureza", "RECEITA LJ", field_key="natureza", settle_locator=historico_input)
        self.log(f"Cliente guia para loja {group.loja}: {group.cliente_codigo}.")
        self._fill_text_field(cliente_input, "Cliente", 
        cliente_codigo, field_key="cliente", settle_locator=natureza_input)

    def prepare_expense_title_fields(self, group: ExpenseGroup, issue_date: str) -> dict[str, object]:
        if self.page is None:
            raise RuntimeError("Pagina do TOTVS nao foi inicializada.")

        self.wait_until_payable_screen()
        dialog = self._payable_dialog()
        self.log(f"Preparando campos de Despesa da loja {group.loja} no valor R$ {_format_brl(group.valor_total)}.")
        return {
            "numero_titulo": self._expense_field_locator(dialog, "numero_titulo"),
            "tipo_titulo": self._expense_field_locator(dialog, "tipo_titulo"),
            "natureza": self._expense_field_locator(dialog, "natureza"),
            "fornecedor": self._expense_field_locator(dialog, "fornecedor"),
            "data_emissao": self._expense_field_locator(dialog, "data_emissao"),
            "vencimento": self._expense_field_locator(dialog, "vencimento"),
            "valor": self._expense_field_locator(dialog, "valor"),
            "historico": self._expense_field_locator(dialog, "historico"),
            "rateio": self._expense_field_locator(dialog, "rateio"),
            "conta_contabil": self._expense_field_locator(dialog, "conta_contabil"),
            "centro_custo": self._expense_field_locator(dialog, "centro_custo"),
            "forma_pagamento": self._expense_field_locator(dialog, "forma_pagamento"),
        }

    def fill_expense_title_basic(self, group: ExpenseGroup, issue_date: str) -> None:
        fields = self.prepare_expense_title_fields(group, issue_date)
        title_number = _generate_expense_title_number()
        historico = f"DESPESAS {group.periodo} - {group.loja}"
        if group.descricao:
            historico = f"{historico} - {group.tipo_despesa}"

        self.log(f"Numero do titulo gerado para despesa: {title_number}.")
        self._fill_text_field(fields["numero_titulo"], "Numero do Titulo", title_number, settle_locator=fields["numero_titulo"])
        self._fill_text_field(fields["tipo_titulo"], "Tipo do Titulo", "R$", settle_locator=fields["numero_titulo"])
        self._fill_text_field(fields["natureza"], "Codigo da natureza", "DESPESA LJ", settle_locator=fields["numero_titulo"])

        if group.cliente_codigo:
            self._fill_text_field(fields["fornecedor"], "Fornecedor", group.cliente_codigo, settle_locator=fields["natureza"])

        self._fill_text_field(fields["data_emissao"], "Data emissao", issue_date, settle_locator=fields["fornecedor"])
        self._fill_text_field(fields["vencimento"], "Vencimento", issue_date, settle_locator=fields["data_emissao"])
        self._fill_text_field(fields["valor"], "Valor", _format_totvs_money(group.valor_total), settle_locator=fields["vencimento"])
        self._fill_text_field(fields["historico"], "Historico", historico, settle_locator=fields["valor"])
        rateio_option = "0" if group.precisa_rateio else "1"
        self._select_option(fields["rateio"], "Rateio", rateio_option)
        if group.precisa_rateio:
            self._confirm_expense_rateio_mode()
        self._stabilize_on_previous_field(fields["historico"], "rateio")
        if not group.precisa_rateio:
            conta_contabil = next(iter(group.contas_por_tipo.values()), DEFAULT_EXPENSE_ACCOUNT)
            self._fill_text_field(fields["conta_contabil"], "Conta contabil", conta_contabil, settle_locator=fields["historico"])
        self._fill_text_field(fields["centro_custo"], "Centro de custo", group.loja, settle_locator=fields["historico"])
        self._fill_text_field(fields["forma_pagamento"], "Forma de pagamento", "99", settle_locator=fields["centro_custo"])
        self.log(
            f"Rateio da loja {group.loja}: {'SIM' if group.precisa_rateio else 'NAO'} "
            f"({len(group.tipos_despesa)} tipo(s) de despesa)."
        )
        if group.precisa_rateio:
            contas = ", ".join(f"{tipo}: {conta}" for tipo, conta in group.contas_por_tipo.items())
            self.log(f"Contas contabeis previstas para rateio: {contas}.")

    def _confirm_expense_rateio_mode(self) -> None:
        page = self._require_page()
        self.log("Confirmando configuracao de rateio digitado...")

        pre_configurado = self._wait_for_first_visible_locator(
            [
                page.get_by_role("radio", name="Pre-Configurado"),
                page.get_by_text("Pre-Configurado", exact=False),
            ],
            timeout=45000,
        )
        if pre_configurado is None:
            raise RuntimeError("Opcao Pre-Configurado do rateio nao apareceu.")
        try:
            pre_configurado.check(timeout=5000)
        except Exception:
            pre_configurado.click(force=True, timeout=5000)
        page.wait_for_timeout(500)

        digitado = self._wait_for_first_visible_locator(
            [
                page.get_by_role("radio", name="Digitado"),
                page.get_by_text("Digitado", exact=False),
            ],
            timeout=45000,
        )
        if digitado is None:
            raise RuntimeError("Opcao Digitado do rateio nao apareceu.")
        try:
            digitado.check(timeout=5000)
        except Exception:
            digitado.click(force=True, timeout=5000)
        page.wait_for_timeout(500)

        self._click_first_visible_locator(
            [
                page.get_by_role("button", name="Ok"),
                page.get_by_role("button", name="OK"),
                page.get_by_text("Ok", exact=True),
                page.get_by_text("OK", exact=True),
            ],
            timeout=45000,
            pause_ms=1200,
            label="Ok do rateio",
        )

        self.log("Aguardando botao Salvar do rateio...")
        self._click_first_visible_locator(
            [
                page.locator("#COMP7518").get_by_role("button", name="Salvar"),
                page.locator("#COMP7518"),
                page.get_by_role("button", name="Salvar"),
            ],
            timeout=90000,
            pause_ms=1500,
            label="Salvar configuracao do rateio",
        )

    def save_and_lower_expense_title(self, group: ExpenseGroup) -> None:
        if self.page is None:
            raise RuntimeError("Pagina do TOTVS nao foi inicializada.")

        self.log(f"Salvando titulo de Despesa da loja {group.loja}...")
        dialog = self._payable_dialog()
        self._click_first_visible_locator(
            [dialog.get_by_role("button", name="Salvar"), dialog.get_by_text("Salvar", exact=True)],
            timeout=45000,
            pause_ms=1200,
            label="Salvar despesa",
        )
        self._raise_if_totvs_error_present()
        self._click_first_visible_locator(
            [self.page.get_by_role("button", name="Salvar"), self.page.get_by_text("Salvar", exact=True)],
            timeout=90000,
            pause_ms=1500,
            label="Salvar confirmacao despesa",
        )
        self._raise_if_totvs_error_present()
        self._click_first_visible_locator(
            [self.page.get_by_role("button", name="Cancelar"), self.page.get_by_text("Cancelar", exact=True)],
            timeout=90000,
            pause_ms=1200,
            label="Cancelar despesa",
        )

        self.log("Abrindo baixa manual da despesa...")
        self._click_when_clickable(self.page.get_by_role("button", name="Outras Ações"), "Outras Acoes")
        self._click_first_visible_locator(
            [
                self.page.get_by_text("Baixa Manual", exact=True),
                self.page.locator('wa-menu-popup-item[caption="Baixa Manual"]'),
                self.page.locator('wa-menu-popup-item:has-text("Baixa Manual")'),
            ],
            timeout=45000,
            pause_ms=1200,
            label="Baixa Manual",
        )
        self._click_first_visible_locator(
            [
                self.page.get_by_text("Baixar", exact=True),
                self.page.locator('wa-menu-popup-item[caption="Baixar"]'),
                self.page.locator('wa-menu-popup-item:has-text("Baixar")'),
            ],
            timeout=45000,
            pause_ms=1200,
            label="Baixar",
        )
        forma_baixa_combo = self._wait_for_first_visible_locator(
            [self.page.get_by_role("combobox")],
            timeout=60000,
        )
        if forma_baixa_combo is None:
            raise RuntimeError("Combobox da baixa manual nao foi localizado.")
        self._select_option(forma_baixa_combo, "Forma da baixa manual", "3")

        baixa_codigo = _normalize_store_code(group.loja)
        banco_baixa_input = self._wait_first_visible_locator(
            [
                self.page.locator("#COMP6028 > input"),
                self.page.locator("#COMP6028"),
            ],
            timeout_ms=60000,
            interval_ms=700,
        )
        if banco_baixa_input is None:
            raise RuntimeError("Campo Banco da baixa manual nao foi localizado.")
        self._fill_text_field(banco_baixa_input, "Banco da baixa manual", baixa_codigo)

        self._click_first_visible_locator(
            [self.page.get_by_role("button", name="Salvar"), self.page.get_by_text("Salvar", exact=True)],
            timeout=45000,
            pause_ms=1200,
            label="Salvar baixa manual",
        )
        self._raise_if_totvs_error_present()
        self._click_first_visible_locator(
            [self.page.get_by_role("button", name="Salvar"), self.page.get_by_text("Salvar", exact=True)],
            timeout=90000,
            pause_ms=1500,
            label="Salvar confirmacao baixa manual",
        )
        self._raise_if_totvs_error_present()
        self.log(f"Despesa da loja {group.loja} salva e baixada.")

    def save_and_lower_cash_title(self, group: CashGroup) -> None:
        if self.page is None:
            raise RuntimeError("Pagina do TOTVS nao foi inicializada.")

        self.log(f"Salvando titulo de Cash da loja {group.loja}...")
        dialog = self._billing_dialog()
        self._click_first_visible_locator(
            [dialog.get_by_role("button", name="Salvar"), dialog.get_by_text("Salvar", exact=True)],
            timeout=45000,
            pause_ms=1200,
            label="Salvar",
        )
        self._raise_if_totvs_error_present()
        self._click_first_visible_locator(
            [self.page.get_by_role("button", name="Salvar"), self.page.get_by_text("Salvar", exact=True)],
            timeout=90000,
            pause_ms=1500,
            label="Salvar confirmacao",
        )
        self._raise_if_totvs_error_present()
        self._click_first_visible_locator(
            [self.page.get_by_role("button", name="Cancelar"), self.page.get_by_text("Cancelar", exact=True)],
            timeout=90000,
            pause_ms=1200,
            label="Cancelar",
        )

        self.log("Abrindo baixa do titulo...")
        self._click_when_clickable(self.page.get_by_role("button", name="Outras Ações"), "Outras Acoes")
        self._click_when_clickable(self.page.get_by_text("Baixas", exact=True), "Baixas")
        self._click_when_clickable(self.page.locator("#COMP4631 > .img"), "icone de baixa", timeout=90000)

        baixa_codigo = _normalize_store_code(group.loja)
        self.log("Aguardando campo Banco da baixa carregar...")
        banco_baixa_input = self._wait_first_visible_locator(
            [
                self.page.locator(
                    "xpath=/html/body/wa-dialog/wa-panel/wa-panel[2]/wa-tab-view/wa-tab-page/wa-dialog[2]/wa-panel[1]/wa-panel[2]/wa-text-input[3]//input"
                ),
                self.page.locator(
                    "xpath=/html/body/wa-dialog/wa-panel/wa-panel[2]/wa-tab-view/wa-tab-page/wa-dialog[2]/wa-panel[1]/wa-panel[2]/wa-text-input[3]"
                ),
                self.page.locator("#COMP6031 > input"),
                self.page.locator("#COMP6031"),
            ],
            timeout_ms=60000,
            interval_ms=700,
        )
        if banco_baixa_input is None:
            raise RuntimeError("Campo Banco/Conta da baixa nao foi localizado.")
        self.log(f"Preenchendo banco da baixa com {baixa_codigo} para a loja {group.loja}.")
        self._fill_text_field(banco_baixa_input, "Conta da baixa", baixa_codigo, field_key="conta_baixa")
        self._click_first_visible_locator(
            [self.page.get_by_role("button", name="Salvar"), self.page.get_by_text("Salvar", exact=True)],
            timeout=45000,
            pause_ms=1200,
            label="Salvar baixa",
        )
        self._raise_if_totvs_error_present()
        self._click_first_visible_locator(
            [self.page.get_by_role("button", name="Salvar"), self.page.get_by_text("Salvar", exact=True)],
            timeout=90000,
            pause_ms=1500,
            label="Salvar confirmacao da baixa",
        )
        self._raise_if_totvs_error_present()
        self.log(f"Cash da loja {group.loja} salvo e baixado.")

    def close(self) -> None:
        for resource in (self.context, self.browser):
            try:
                if resource is not None:
                    resource.close()
            except Exception:
                pass
        try:
            if self.playwright is not None:
                self.playwright.stop()
        except Exception:
            pass


class RoboLancamentosCaixaApp(ctk.CTk):
    PRIMARY_TEXT = PRIMARY_TEXT
    CARD_BG = CARD_BG
    CARD_BORDER = CARD_BORDER
    BUTTON_BG = BUTTON_BG
    BUTTON_ACTIVE_BG = BUTTON_ACTIVE_BG

    def __init__(self) -> None:
        super().__init__()
        ctk.set_appearance_mode("light")
        ctk.set_default_color_theme("blue")

        self.title(APP_TITLE)
        self.geometry(APP_GEOMETRY)
        self.minsize(1080, 720)
        self.configure(fg_color=MAIN_BG)

        self.username_var = ctk.StringVar()
        self.password_var = ctk.StringVar()
        self.file_path_var = ctk.StringVar()
        self.issue_date_var = ctk.StringVar(value=datetime.now().strftime("%d/%m/%Y"))
        self.period_start_var = ctk.StringVar(value="")
        self.period_end_var = ctk.StringVar(value="")
        self.period_var = ctk.StringVar(value="")
        self.flow_var = ctk.StringVar(value="cash")
        self.headless_var = ctk.BooleanVar(value=False)

        self.total_var = ctk.StringVar(value="0")
        self.pending_var = ctk.StringVar(value="0")
        self.ready_var = ctk.StringVar(value="0")
        self.error_var = ctk.StringVar(value="0")
        self.sheet_var = ctk.StringVar(value="-")
        self.status_var = ctk.StringVar(value="Aguardando validação")

        self.validation_result: ValidationResult | None = None
        self.processing_thread: threading.Thread | None = None
        self.stop_requested = False
        self.totvs_bot: TotvsCaixaLoginBot | None = None
        self.logo_image = self._load_logo()

        self._build_layout()
        self._update_flow_description()
        self._update_action_buttons()

    def _load_logo(self):
        for candidate in _load_logo_candidates():
            try:
                if candidate.exists():
                    image = Image.open(candidate)
                    return ctk.CTkImage(light_image=image, dark_image=image, size=(86, 52))
            except Exception:
                continue
        return None

    def _build_layout(self) -> None:
        container = ctk.CTkScrollableFrame(self, fg_color="transparent", corner_radius=0)
        container.pack(fill="both", expand=True, padx=22, pady=22)
        container.grid_columnconfigure(0, weight=1)

        self._build_header(container)
        self._build_access_section(container)
        self._build_plan_section(container)
        self._build_indicator_section(container)
        self._build_execution_section(container)

        footer = ctk.CTkLabel(
            container,
            text="Criado por Diogo Medeiros © 2026",
            text_color="#ef574f",
            font=("Segoe UI", 12),
        )
        footer.grid(row=5, column=0, sticky="w", padx=6, pady=(6, 0))

    def _build_header(self, parent) -> None:
        header = ctk.CTkFrame(parent, fg_color=CARD_BG, corner_radius=26, border_width=1, border_color=CARD_BORDER)
        header.grid(row=0, column=0, sticky="ew", pady=(0, 18))
        header.grid_columnconfigure(1, weight=1)

        brand = ctk.CTkFrame(header, fg_color="transparent")
        brand.grid(row=0, column=0, padx=(22, 14), pady=22, sticky="nw")
        if self.logo_image is not None:
            ctk.CTkLabel(brand, text="", image=self.logo_image).pack(anchor="w")
        else:
            ctk.CTkLabel(brand, text="foco,", text_color=PRIMARY_TEXT, font=("Segoe UI", 24, "bold")).pack(anchor="w")
            ctk.CTkLabel(brand, text="aluguel de carros", text_color="#c7463f", font=("Segoe UI", 10, "bold")).pack(anchor="w")

        texts = ctk.CTkFrame(header, fg_color="transparent")
        texts.grid(row=0, column=1, sticky="ew", padx=(0, 22), pady=22)
        ctk.CTkLabel(texts, text="Lançamentos de Caixa", text_color=PRIMARY_TEXT, font=("Segoe UI", 30, "bold")).pack(anchor="w")
        ctk.CTkLabel(
            texts,
            text="Automação de cash, despesas e depósitos com leitura da planilha padrão e validação prévia.",
            text_color="#4b5563",
            font=("Segoe UI", 15),
        ).pack(anchor="w", pady=(8, 0))
        ctk.CTkLabel(
            texts,
            text="ESCOP0 1: INTERFACE + VALIDAÇÃO",
            text_color="#b75b2d",
            font=("Segoe UI", 13, "bold"),
        ).pack(anchor="w", pady=(16, 0))

    def _build_access_section(self, parent) -> None:
        section = self._create_section(parent, 1, "Credenciais do TOTVS")
        grid = ctk.CTkFrame(section, fg_color="transparent")
        grid.pack(fill="x", padx=18, pady=(0, 18))
        grid.grid_columnconfigure((0, 1), weight=1)

        self._form_label(grid, "Usuário").grid(row=0, column=0, sticky="w", padx=(0, 10), pady=(0, 6))
        self._form_label(grid, "Senha").grid(row=0, column=1, sticky="w", padx=(10, 0), pady=(0, 6))
        self._entry(grid, self.username_var).grid(row=1, column=0, sticky="ew", padx=(0, 10))
        self._entry(grid, self.password_var, show="*").grid(row=1, column=1, sticky="ew", padx=(10, 0))

    def _build_plan_section(self, parent) -> None:
        section = self._create_section(parent, 2, "Planilha e Fluxo")
        content = ctk.CTkFrame(section, fg_color="transparent")
        content.pack(fill="x", padx=18, pady=(0, 18))
        content.grid_columnconfigure(0, weight=1)

        self._form_label(content, "Planilha padrão").grid(row=0, column=0, sticky="w", pady=(0, 6))
        self._entry(content, self.file_path_var).grid(row=1, column=0, sticky="ew", padx=(0, 10), pady=(0, 10))
        self._secondary_button(content, "Selecionar", self.select_excel_file, width=150).grid(row=1, column=1, sticky="e", pady=(0, 10))
        date_grid = ctk.CTkFrame(content, fg_color="transparent")
        date_grid.grid(row=2, column=0, columnspan=2, sticky="ew", pady=(4, 10))
        date_grid.grid_columnconfigure((0, 1), weight=1)

        venc_frame = ctk.CTkFrame(date_grid, fg_color="transparent")
        venc_frame.grid(row=0, column=0, sticky="ew", padx=(0, 10))
        venc_frame.grid_columnconfigure(0, weight=1)
        self._form_label(venc_frame, "Data de emissao/vencimento").grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, 6))
        self._entry(venc_frame, self.issue_date_var).grid(row=1, column=0, sticky="ew", padx=(0, 8))
        self._secondary_button(venc_frame, "Calendario", lambda: self._open_calendar(self.issue_date_var), width=130).grid(row=1, column=1, sticky="e")

        period_frame = ctk.CTkFrame(date_grid, fg_color="transparent")
        period_frame.grid(row=0, column=1, sticky="ew", padx=(10, 0))
        period_frame.grid_columnconfigure(0, weight=1)
        self._form_label(period_frame, "Periodo do historico").grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, 6))
        self._entry(period_frame, self.period_var).grid(row=1, column=0, sticky="ew", padx=(0, 8))
        self._secondary_button(period_frame, "Selecionar periodo", self._open_period_calendar, width=170).grid(row=1, column=1, sticky="e")

        self._form_label(content, "Fluxo de lançamento").grid(row=3, column=0, sticky="w", pady=(4, 6))
        self.flow_selector = ctk.CTkSegmentedButton(
            content,
            values=[FLOW_SPECS[key]["label"] for key in FLOW_SPECS],
            fg_color="#f2e7e3",
            selected_color=BUTTON_BG,
            selected_hover_color=BUTTON_ACTIVE_BG,
            unselected_color="#fff9f8",
            unselected_hover_color="#fff0ec",
            text_color="#8e2d2d",
            command=self._on_flow_segment_change,
        )
        self.flow_selector.grid(row=4, column=0, columnspan=2, sticky="ew")
        self.flow_selector.set(FLOW_SPECS[self.flow_var.get()]["label"])

        self.flow_description = ctk.CTkTextbox(
            content,
            height=92,
            corner_radius=14,
            fg_color=SOFT_RED,
            border_width=1,
            border_color=CARD_BORDER,
            text_color="#5a4137",
        )
        self.flow_description.grid(row=5, column=0, columnspan=2, sticky="ew", pady=(12, 10))
        self.flow_description.configure(state="disabled")

        ctk.CTkCheckBox(
            content,
            text="Executar em modo invisível",
            variable=self.headless_var,
            text_color="#303030",
            fg_color=BUTTON_BG,
            hover_color=BUTTON_ACTIVE_BG,
            border_color="#d9c9c3",
        ).grid(row=6, column=0, sticky="w", pady=(0, 6))

    def _build_indicator_section(self, parent) -> None:
        section = self._create_section(parent, 3, "Resumo da Validação")
        cards = ctk.CTkFrame(section, fg_color="transparent")
        cards.pack(fill="x", padx=18, pady=(0, 18))
        for idx in range(5):
            cards.grid_columnconfigure(idx, weight=1)

        self._counter_card(cards, 0, "Linhas na aba", self.total_var)
        self._counter_card(cards, 1, "Pendentes", self.pending_var)
        self._counter_card(cards, 2, "Prontos para lançar", self.ready_var)
        self._counter_card(cards, 3, "Erros de validação", self.error_var)
        self._counter_card(cards, 4, "Aba selecionada", self.sheet_var)

        status_frame = ctk.CTkFrame(section, fg_color=SOFT_RED, corner_radius=18, border_width=1, border_color=CARD_BORDER)
        status_frame.pack(fill="x", padx=18, pady=(0, 18))
        ctk.CTkLabel(status_frame, text="Status", text_color=PRIMARY_TEXT, font=("Segoe UI", 14, "bold")).pack(anchor="w", padx=16, pady=(12, 4))
        ctk.CTkLabel(status_frame, textvariable=self.status_var, text_color="#5c5c5c", font=("Segoe UI", 14)).pack(anchor="w", padx=16, pady=(0, 12))

    def _build_execution_section(self, parent) -> None:
        section = self._create_section(parent, 4, "Execução")
        content = ctk.CTkFrame(section, fg_color="transparent")
        content.pack(fill="both", expand=True, padx=18, pady=(0, 18))
        content.grid_columnconfigure(0, weight=1)
        content.grid_rowconfigure(1, weight=1)

        self.progress = ctk.CTkProgressBar(
            content,
            progress_color=BUTTON_BG,
            fg_color="#f1dfdb",
            corner_radius=20,
            height=18,
        )
        self.progress.set(0)
        self.progress.grid(row=0, column=0, sticky="ew", padx=18, pady=(0, 12))

        self.log_box = ctk.CTkTextbox(
            content,
            corner_radius=14,
            fg_color="#fffdfc",
            border_width=1,
            border_color=CARD_BORDER,
            text_color="#303030",
            height=240,
        )
        self.log_box.grid(row=1, column=0, sticky="nsew", padx=18, pady=(0, 12))

        actions = ctk.CTkFrame(content, fg_color="transparent")
        actions.grid(row=2, column=0, sticky="ew", padx=18)
        for idx in range(4):
            actions.grid_columnconfigure(idx, weight=1)

        self.validate_button = self._primary_button(actions, "Validar planilha", self.validate_workbook)
        self.validate_button.grid(row=0, column=0, padx=(0, 8), sticky="ew")
        self.start_button = self._primary_button(actions, "Iniciar", self.start_processing)
        self.start_button.grid(row=0, column=1, padx=8, sticky="ew")
        self.pause_button = self._secondary_button(actions, "Pausar", self.pause_processing)
        self.pause_button.grid(row=0, column=2, padx=8, sticky="ew")
        self.stop_button = self._secondary_button(actions, "Parar", self.stop_processing)
        self.stop_button.grid(row=0, column=3, padx=(8, 0), sticky="ew")

    def _create_section(self, parent, row: int, title: str) -> ctk.CTkFrame:
        section = ctk.CTkFrame(parent, fg_color=CARD_BG, corner_radius=24, border_width=1, border_color=CARD_BORDER)
        section.grid(row=row, column=0, sticky="ew", pady=(0, 16))
        ctk.CTkLabel(section, text=title, text_color=PRIMARY_TEXT, font=("Segoe UI", 16, "bold")).pack(anchor="w", padx=18, pady=(16, 12))
        return section

    def _entry(self, parent, variable, show: str | None = None) -> ctk.CTkEntry:
        return ctk.CTkEntry(
            parent,
            textvariable=variable,
            height=42,
            corner_radius=14,
            border_color=CARD_BORDER,
            fg_color="#fffdfc",
            text_color="#303030",
            placeholder_text="",
            show=show,
        )

    def _form_label(self, parent, text: str) -> ctk.CTkLabel:
        return ctk.CTkLabel(parent, text=text, text_color="#1f1f1f", font=("Segoe UI", 14, "bold"))

    def _primary_button(self, parent, text: str, command):
        return ctk.CTkButton(
            parent,
            text=text,
            command=command,
            height=42,
            corner_radius=16,
            fg_color=BUTTON_BG,
            hover_color=BUTTON_ACTIVE_BG,
            text_color="#ffffff",
            font=("Segoe UI", 15, "bold"),
        )

    def _secondary_button(self, parent, text: str, command, width: int | None = None):
        kwargs = {
            "text": text,
            "command": command,
            "height": 42,
            "corner_radius": 16,
            "fg_color": "#fff9f8",
            "hover_color": "#fff0ec",
            "border_width": 1,
            "border_color": CARD_BORDER,
            "text_color": PRIMARY_TEXT,
            "font": ("Segoe UI", 14, "bold"),
        }
        if width is not None:
            kwargs["width"] = width
        return ctk.CTkButton(parent, **kwargs)

    def _counter_card(self, parent, column: int, title: str, value_var: ctk.StringVar) -> None:
        card = ctk.CTkFrame(parent, fg_color="#fffdfc", corner_radius=18, border_width=1, border_color=CARD_BORDER)
        card.grid(row=0, column=column, padx=6, sticky="ew")
        ctk.CTkLabel(card, text=title, text_color="#5a4137", font=("Segoe UI", 13, "bold")).pack(anchor="w", padx=14, pady=(14, 4))
        ctk.CTkLabel(card, textvariable=value_var, text_color=PRIMARY_TEXT, font=("Segoe UI", 18, "bold")).pack(anchor="w", padx=14, pady=(0, 14))

    def _open_calendar(self, target_var: ctk.StringVar) -> None:
        try:
            selected = datetime.strptime(target_var.get().strip(), "%d/%m/%Y")
        except ValueError:
            selected = datetime.now()

        current_year = selected.year
        current_month = selected.month

        popup = ctk.CTkToplevel(self)
        popup.title("Selecionar data")
        popup.geometry("340x360")
        popup.resizable(False, False)
        popup.configure(fg_color=MAIN_BG)
        popup.transient(self)
        popup.grab_set()

        header = ctk.CTkFrame(popup, fg_color="transparent")
        header.pack(fill="x", padx=16, pady=(16, 8))
        header.grid_columnconfigure(1, weight=1)

        body = ctk.CTkFrame(popup, fg_color=CARD_BG, corner_radius=16, border_width=1, border_color=CARD_BORDER)
        body.pack(fill="both", expand=True, padx=16, pady=(0, 16))

        title_var = ctk.StringVar()

        def choose_day(day: int) -> None:
            target_var.set(f"{day:02d}/{current_month:02d}/{current_year}")
            popup.destroy()

        def render_calendar() -> None:
            for widget in body.winfo_children():
                widget.destroy()

            title_var.set(f"{MONTH_NAMES_PT[current_month]} {current_year}")
            weekdays = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sab", "Dom"]
            for column, label in enumerate(weekdays):
                ctk.CTkLabel(body, text=label, text_color=MUTED_TEXT, font=("Segoe UI", 12, "bold")).grid(
                    row=0,
                    column=column,
                    padx=4,
                    pady=(12, 6),
                )

            for row_index, week in enumerate(calendar.monthcalendar(current_year, current_month), start=1):
                for column, day in enumerate(week):
                    if day == 0:
                        ctk.CTkLabel(body, text="", width=34, height=30).grid(row=row_index, column=column, padx=4, pady=4)
                        continue
                    ctk.CTkButton(
                        body,
                        text=str(day),
                        width=34,
                        height=30,
                        corner_radius=8,
                        fg_color=BUTTON_BG if day == selected.day and current_month == selected.month and current_year == selected.year else "#fff9f8",
                        hover_color=BUTTON_ACTIVE_BG,
                        text_color="#ffffff" if day == selected.day and current_month == selected.month and current_year == selected.year else PRIMARY_TEXT,
                        command=lambda value=day: choose_day(value),
                    ).grid(row=row_index, column=column, padx=4, pady=4)

        def change_month(delta: int) -> None:
            nonlocal current_month, current_year
            current_month += delta
            if current_month < 1:
                current_month = 12
                current_year -= 1
            elif current_month > 12:
                current_month = 1
                current_year += 1
            render_calendar()

        self._secondary_button(header, "<", lambda: change_month(-1), width=44).grid(row=0, column=0, sticky="w")
        ctk.CTkLabel(header, textvariable=title_var, text_color=PRIMARY_TEXT, font=("Segoe UI", 18, "bold")).grid(
            row=0,
            column=1,
            sticky="ew",
        )
        self._secondary_button(header, ">", lambda: change_month(1), width=44).grid(row=0, column=2, sticky="e")

        render_calendar()

    def _open_period_calendar(self) -> None:
        today = datetime.now()
        try:
            selected_start = datetime.strptime(self.period_start_var.get().strip(), "%d/%m/%Y")
        except ValueError:
            selected_start = None

        current_year = selected_start.year if selected_start else today.year
        current_month = selected_start.month if selected_start else today.month
        selected_end = None

        popup = ctk.CTkToplevel(self)
        popup.title("Selecionar periodo")
        popup.geometry("380x430")
        popup.resizable(False, False)
        popup.configure(fg_color=MAIN_BG)
        popup.transient(self)
        popup.grab_set()

        header = ctk.CTkFrame(popup, fg_color="transparent")
        header.pack(fill="x", padx=16, pady=(16, 8))
        header.grid_columnconfigure(1, weight=1)

        instruction_var = ctk.StringVar(value="Selecione a data inicial")
        ctk.CTkLabel(
            popup,
            textvariable=instruction_var,
            text_color=MUTED_TEXT,
            font=("Segoe UI", 13, "bold"),
        ).pack(fill="x", padx=16, pady=(0, 8))

        body = ctk.CTkFrame(popup, fg_color=CARD_BG, corner_radius=16, border_width=1, border_color=CARD_BORDER)
        body.pack(fill="both", expand=True, padx=16, pady=(0, 12))

        footer = ctk.CTkFrame(popup, fg_color="transparent")
        footer.pack(fill="x", padx=16, pady=(0, 16))
        footer.grid_columnconfigure((0, 1), weight=1)

        title_var = ctk.StringVar()
        preview_var = ctk.StringVar(value="Periodo: -")

        def same_day(left: datetime | None, day: int) -> bool:
            return bool(left and left.day == day and left.month == current_month and left.year == current_year)

        def choose_day(day: int) -> None:
            nonlocal selected_start, selected_end
            chosen = datetime(current_year, current_month, day)
            if selected_start is None or selected_end is not None:
                selected_start = chosen
                selected_end = None
                self.period_start_var.set(chosen.strftime("%d/%m/%Y"))
                self.period_end_var.set("")
                instruction_var.set("Selecione a data final")
                preview_var.set(f"Inicio: {chosen.strftime('%d/%m/%Y')} | Fim: -")
                render_calendar()
                return

            selected_end = chosen
            if selected_end < selected_start:
                selected_start, selected_end = selected_end, selected_start

            self.period_start_var.set(selected_start.strftime("%d/%m/%Y"))
            self.period_end_var.set(selected_end.strftime("%d/%m/%Y"))
            self.period_var.set(_format_period_text(selected_start, selected_end))
            popup.destroy()

        def render_calendar() -> None:
            for widget in body.winfo_children():
                widget.destroy()

            title_var.set(f"{MONTH_NAMES_PT[current_month]} {current_year}")
            weekdays = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sab", "Dom"]
            for column, label in enumerate(weekdays):
                ctk.CTkLabel(body, text=label, text_color=MUTED_TEXT, font=("Segoe UI", 12, "bold")).grid(
                    row=0,
                    column=column,
                    padx=5,
                    pady=(12, 6),
                )

            for row_index, week in enumerate(calendar.monthcalendar(current_year, current_month), start=1):
                for column, day in enumerate(week):
                    if day == 0:
                        ctk.CTkLabel(body, text="", width=36, height=32).grid(row=row_index, column=column, padx=5, pady=4)
                        continue

                    is_start = same_day(selected_start, day)
                    is_end = same_day(selected_end, day)
                    is_selected = is_start or is_end
                    ctk.CTkButton(
                        body,
                        text=str(day),
                        width=36,
                        height=32,
                        corner_radius=8,
                        fg_color=BUTTON_BG if is_selected else "#fff9f8",
                        hover_color=BUTTON_ACTIVE_BG,
                        text_color="#ffffff" if is_selected else PRIMARY_TEXT,
                        command=lambda value=day: choose_day(value),
                    ).grid(row=row_index, column=column, padx=5, pady=4)

        def change_month(delta: int) -> None:
            nonlocal current_month, current_year
            current_month += delta
            if current_month < 1:
                current_month = 12
                current_year -= 1
            elif current_month > 12:
                current_month = 1
                current_year += 1
            render_calendar()

        self._secondary_button(header, "<", lambda: change_month(-1), width=44).grid(row=0, column=0, sticky="w")
        ctk.CTkLabel(header, textvariable=title_var, text_color=PRIMARY_TEXT, font=("Segoe UI", 18, "bold")).grid(
            row=0,
            column=1,
            sticky="ew",
        )
        self._secondary_button(header, ">", lambda: change_month(1), width=44).grid(row=0, column=2, sticky="e")

        ctk.CTkLabel(footer, textvariable=preview_var, text_color=MUTED_TEXT, font=("Segoe UI", 12)).grid(
            row=0,
            column=0,
            sticky="w",
        )
        self._secondary_button(footer, "Limpar", lambda: self._clear_period_popup(popup), width=90).grid(
            row=0,
            column=1,
            sticky="e",
        )

        render_calendar()

    def _clear_period_popup(self, popup: ctk.CTkToplevel) -> None:
        self.period_start_var.set("")
        self.period_end_var.set("")
        self.period_var.set("")
        popup.destroy()

    def _on_flow_segment_change(self, label: str) -> None:
        for key, spec in FLOW_SPECS.items():
            if spec["label"] == label:
                self.flow_var.set(key)
                break
        self._update_flow_description()
        self.validation_result = None
        self.status_var.set("Fluxo alterado. Valide a planilha novamente.")
        self.progress.set(0)
        self._update_action_buttons()

    def _update_flow_description(self) -> None:
        flow = FLOW_SPECS[self.flow_var.get()]
        description = [
            f"Fluxo selecionado: {flow['label']}",
            f"Aba esperada: {flow['sheet_name']}",
            flow["group_hint"],
            f"Regra de pendência: coluna {flow['pending_column']} em branco.",
        ]
        self.flow_description.configure(state="normal")
        self.flow_description.delete("1.0", "end")
        self.flow_description.insert("1.0", "\n".join(description))
        self.flow_description.configure(state="disabled")

    def log(self, message: str) -> None:
        if threading.current_thread() is not threading.main_thread():
            self.after(0, lambda: self.log(message))
            return
        timestamp = pd.Timestamp.now().strftime("%H:%M:%S")
        self.log_box.insert("end", f"[{timestamp}] {message}\n")
        self.log_box.see("end")
        self.update_idletasks()

    def select_excel_file(self) -> None:
        file_path = filedialog.askopenfilename(
            title="Selecionar planilha de caixa",
            filetypes=[("Excel", "*.xlsx *.xlsm *.xls")],
        )
        if not file_path:
            return
        self.file_path_var.set(file_path)
        self.validation_result = None
        self.status_var.set("Planilha selecionada. Pronta para validação.")
        self.log(f"Planilha selecionada: {file_path}")
        self.progress.set(0)
        self._update_action_buttons()

    def validate_workbook(self) -> None:
        path_text = self.file_path_var.get().strip()
        if not path_text:
            messagebox.showwarning("Planilha", "Selecione a planilha antes de validar.")
            return

        workbook_path = Path(path_text)
        if not workbook_path.exists():
            messagebox.showerror("Arquivo não encontrado", f"A planilha não foi encontrada:\n{workbook_path}")
            return

        flow = FLOW_SPECS[self.flow_var.get()]
        self.log(f"Iniciando validação da aba {flow['sheet_name']}...")
        self.progress.set(0.15)

        try:
            excel_file = pd.ExcelFile(workbook_path)
            workbook_sheets = list(excel_file.sheet_names)
            actual_sheet_name = _resolve_sheet_name(workbook_sheets, flow["sheet_name"])
            if not actual_sheet_name:
                raise RuntimeError(f"A aba {flow['sheet_name']} não existe nesta planilha.")

            dataframe = pd.read_excel(
                workbook_path,
                sheet_name=actual_sheet_name,
                header=flow.get("header_row", 0),
            )
            renamed, missing_columns = self._normalize_flow_dataframe(dataframe, flow)

            total_rows = len(renamed.index)
            pending_column = flow["pending_column"]
            pending_rows = 0
            launched_rows = 0
            if pending_column in renamed.columns:
                status_series = renamed[pending_column].fillna("").astype(str).str.strip()
                pending_rows = int(status_series.eq("").sum())
                launched_rows = int((~status_series.eq("")).sum())

            self.validation_result = ValidationResult(
                flow_key=self.flow_var.get(),
                flow_label=flow["label"],
                sheet_name=actual_sheet_name,
                total_rows=total_rows,
                pending_rows=pending_rows,
                launched_rows=launched_rows,
                missing_columns=missing_columns,
                workbook_sheets=workbook_sheets,
            )

            self.total_var.set(str(total_rows))
            self.pending_var.set(str(pending_rows))
            self.ready_var.set(str(pending_rows if not missing_columns else 0))
            self.error_var.set(str(len(missing_columns)))
            self.sheet_var.set(actual_sheet_name)

            if missing_columns:
                self.status_var.set("Validação concluída com pendências de estrutura")
                self.log(f"Colunas obrigatórias ausentes: {', '.join(missing_columns)}")
                messagebox.showwarning(
                    "Validação com pendências",
                    "A planilha foi lida, mas faltam colunas obrigatórias:\n\n- " + "\n- ".join(missing_columns),
                )
            else:
                self.status_var.set(f"Validação concluída. {pending_rows} lançamento(s) pendente(s).")
                self.log(f"Aba {actual_sheet_name} validada com sucesso.")
                self.log(f"Linhas totais: {total_rows} | Pendentes: {pending_rows} | Já lançadas: {launched_rows}")

            self.progress.set(1)
        except Exception as exc:
            self.validation_result = None
            self.total_var.set("0")
            self.pending_var.set("0")
            self.ready_var.set("0")
            self.error_var.set("1")
            self.sheet_var.set("-")
            self.status_var.set("Falha na validação")
            self.progress.set(0)
            self.log(f"Erro de validação: {exc}")
            messagebox.showerror("Falha na validação", str(exc))
        finally:
            self._update_action_buttons()

    def _normalize_flow_dataframe(self, dataframe: pd.DataFrame, flow: Dict) -> tuple[pd.DataFrame, list[str]]:
        rename_map: Dict[str, str] = {}
        normalized_columns = {_normalize_text(column): column for column in dataframe.columns}

        for canonical, aliases in flow["aliases"].items():
            for alias in aliases:
                alias_norm = _normalize_text(alias)
                if alias_norm in normalized_columns:
                    rename_map[normalized_columns[alias_norm]] = canonical
                    break

        renamed = dataframe.rename(columns=rename_map).copy()
        missing_columns = [column for column in flow["required_columns"] if column not in renamed.columns]
        return renamed, missing_columns

    def _load_store_guide(self) -> dict[str, str]:
        guide_path = _resolve_cash_guide_path()
        if not guide_path.exists():
            raise RuntimeError(f"Planilha guia padrao nao encontrada: {guide_path}")

        self.log(f"Usando guia de lojas padrao: {guide_path}")
        guide_df = pd.read_excel(guide_path)
        normalized_columns = {_normalize_text(column): column for column in guide_df.columns}
        loja_col = normalized_columns.get("LOJA")
        codigo_col = normalized_columns.get("CODIGO")
        if not loja_col or not codigo_col:
            raise RuntimeError("A planilha guia precisa ter as colunas loja e codigo.")

        guide: dict[str, str] = {}
        for _, row in guide_df.iterrows():
            loja = _normalize_text(row.get(loja_col, ""))
            codigo = str(row.get(codigo_col, "")).strip()
            if loja and codigo and codigo.lower() != "nan":
                guide[loja] = codigo
        return guide

    def _load_expense_client_guide(self) -> dict[str, str]:
        guide_path = _resolve_expense_client_guide_path()
        if not guide_path.exists():
            self.log(f"Planilha guia de clientes de despesas ainda nao encontrada: {guide_path}")
            return {}

        self.log(f"Usando guia de clientes de despesas: {guide_path}")
        guide_df = pd.read_excel(guide_path)
        normalized_columns = {_normalize_text(column): column for column in guide_df.columns}
        nome_col = normalized_columns.get("NOME")
        codigo_col = normalized_columns.get("CODIGO")
        if not nome_col or not codigo_col:
            raise RuntimeError("A planilha guia de despesas precisa ter as colunas Nome e Codigo.")

        guide: dict[str, str] = {}
        for _, row in guide_df.iterrows():
            nome = _normalize_text(row.get(nome_col, ""))
            codigo = str(row.get(codigo_col, "")).strip()
            if nome and codigo and codigo.lower() != "nan":
                guide[nome] = codigo
        return guide

    def _load_expense_rateio_guide(self) -> pd.DataFrame:
        guide_path = _resolve_expense_rateio_path()
        if not guide_path.exists():
            self.log(f"Planilha de rateio de despesas ainda nao encontrada: {guide_path}")
            return pd.DataFrame()

        self.log(f"Usando planilha de rateio de despesas: {guide_path}")
        dataframe = pd.read_excel(guide_path, dtype=str).fillna("")
        dataframe.columns = [str(column).strip() for column in dataframe.columns]
        if "Conta Contábil Debito" in dataframe.columns:
            dataframe["Conta Contábil Debito"] = dataframe["Conta Contábil Debito"].astype(str).str.strip()
        return dataframe

    def _read_flow_dataframe(self) -> tuple[pd.DataFrame, Dict]:
        if self.validation_result is None:
            raise RuntimeError("Planilha ainda nao validada.")
        flow = FLOW_SPECS[self.validation_result.flow_key]
        workbook_path = Path(self.file_path_var.get().strip())
        dataframe = pd.read_excel(
            workbook_path,
            sheet_name=self.validation_result.sheet_name,
            header=flow.get("header_row", 0),
        )
        renamed, missing_columns = self._normalize_flow_dataframe(dataframe, flow)
        if missing_columns:
            raise RuntimeError("A planilha ainda tem colunas obrigatorias ausentes.")
        return renamed, flow

    def _build_cash_groups(self) -> list[CashGroup]:
        guide = self._load_store_guide()
        dataframe, flow = self._read_flow_dataframe()
        pending_column = flow["pending_column"]
        pending_mask = dataframe[pending_column].fillna("").astype(str).str.strip().eq("")
        pending = dataframe[pending_mask].copy()
        if pending.empty:
            raise RuntimeError("Nao ha linhas pendentes de Cash para lancar.")

        groups: list[CashGroup] = []
        header_row = int(flow.get("header_row", 0))
        for loja_raw, group_df in pending.groupby("LOJA", dropna=True):
            loja = _normalize_text(loja_raw)
            if not loja:
                continue
            cliente_codigo = guide.get(loja)
            if not cliente_codigo:
                raise RuntimeError(f"Loja {loja} nao encontrada na planilha guia.")

            valor_total = float(pd.to_numeric(group_df["VALOR"], errors="coerce").fillna(0).sum())
            periodo = _normalize_text(self.period_var.get())

            row_numbers = [int(index) + header_row + 2 for index in group_df.index]
            groups.append(
                CashGroup(
                    loja=loja,
                    cliente_codigo=cliente_codigo,
                    valor_total=valor_total,
                    periodo=periodo,
                    row_numbers=row_numbers,
                )
            )
        return groups

    def _build_expense_groups(self) -> list[ExpenseGroup]:
        guide = self._load_expense_client_guide()
        rateio_guide = self._load_expense_rateio_guide()
        if not rateio_guide.empty:
            self.log(f"Base de rateio carregada com {len(rateio_guide)} linha(s).")
        dataframe, flow = self._read_flow_dataframe()
        pending_column = flow["pending_column"]
        pending_mask = dataframe[pending_column].fillna("").astype(str).str.strip().eq("")
        pending = dataframe[pending_mask].copy()
        if pending.empty:
            raise RuntimeError("Nao ha linhas pendentes de Despesas para lancar.")

        pending["LOJA_NORMALIZADA"] = pending["LOJA"].map(_normalize_text)
        pending["TIPO_DESPESA_NORMALIZADO"] = pending["TIPO DA DESPESA"].map(_normalize_text)

        groups: list[ExpenseGroup] = []
        header_row = int(flow.get("header_row", 0))
        for loja_raw, group_df in pending.groupby("LOJA_NORMALIZADA", dropna=True):
            loja = _normalize_text(loja_raw)
            if not loja:
                continue

            valor_total = float(pd.to_numeric(group_df["VALOR"], errors="coerce").fillna(0).sum())
            tipos_despesa = [
                tipo
                for tipo in dict.fromkeys(group_df["TIPO_DESPESA_NORMALIZADO"].fillna("").map(_normalize_text).tolist())
                if tipo
            ]
            tipo_despesa = tipos_despesa[0] if len(tipos_despesa) == 1 else "MULTIPLAS DESPESAS"
            contas_por_tipo = {}
            for tipo in tipos_despesa:
                conta = _expense_account_for_type(tipo)
                if not conta:
                    self.log(
                        f"Tipo de despesa sem conta mapeada: {tipo}. "
                        f"Usando conta padrao {DEFAULT_EXPENSE_ACCOUNT} - {DEFAULT_EXPENSE_ACCOUNT_DESCRIPTION}."
                    )
                    conta = DEFAULT_EXPENSE_ACCOUNT
                contas_por_tipo[tipo] = conta
            descricoes = [
                str(value).strip()
                for value in group_df["DESCRICAO"].fillna("").tolist()
                if str(value).strip() and str(value).strip().lower() != "nan"
            ]
            descricao = " | ".join(dict.fromkeys(descricoes[:3]))
            periodo = _normalize_text(self.period_var.get())
            row_numbers = [int(index) + header_row + 2 for index in group_df.index]
            cliente_nome = f"FUNDO FIXO - {loja}"
            cliente_codigo = guide.get(_normalize_text(cliente_nome), "")
            if guide and not cliente_codigo:
                raise RuntimeError(f"Cliente de despesa nao encontrado na guia: {cliente_nome}")

            groups.append(
                ExpenseGroup(
                    loja=loja,
                    tipo_despesa=tipo_despesa,
                    tipos_despesa=tipos_despesa,
                    contas_por_tipo=contas_por_tipo,
                    descricao=descricao,
                    cliente_codigo=cliente_codigo,
                    valor_total=valor_total,
                    periodo=periodo,
                    row_numbers=row_numbers,
                )
            )
        return groups

    def _mark_cash_group_as_lowered(self, group: CashGroup) -> None:
        if self.validation_result is None:
            raise RuntimeError("Planilha ainda nao validada.")
        workbook_path = Path(self.file_path_var.get().strip())
        flow = FLOW_SPECS[self.validation_result.flow_key]
        header_excel_row = int(flow.get("header_row", 0)) + 1

        workbook = load_workbook(workbook_path)
        sheet = workbook[self.validation_result.sheet_name]
        status_col = None
        for cell in sheet[header_excel_row]:
            if _normalize_text(cell.value) == _normalize_text(flow["pending_column"]):
                status_col = cell.column
                break
        if status_col is None:
            raise RuntimeError(f"Coluna {flow['pending_column']} nao encontrada para atualizar a planilha.")

        for row_number in group.row_numbers:
            sheet.cell(row=row_number, column=status_col).value = "BAIXADO"
        workbook.save(workbook_path)
        self.log(f"Planilha atualizada: {len(group.row_numbers)} linha(s) da loja {group.loja} marcadas como BAIXADO.")

    def _mark_expense_group_as_lowered(self, group: ExpenseGroup) -> None:
        if self.validation_result is None:
            raise RuntimeError("Planilha ainda nao validada.")
        workbook_path = Path(self.file_path_var.get().strip())
        flow = FLOW_SPECS[self.validation_result.flow_key]
        header_excel_row = int(flow.get("header_row", 0)) + 1

        workbook = load_workbook(workbook_path)
        sheet = workbook[self.validation_result.sheet_name]
        status_col = None
        for cell in sheet[header_excel_row]:
            if _normalize_text(cell.value) == _normalize_text(flow["pending_column"]):
                status_col = cell.column
                break
        if status_col is None:
            raise RuntimeError(f"Coluna {flow['pending_column']} nao encontrada para atualizar a planilha.")

        for row_number in group.row_numbers:
            sheet.cell(row=row_number, column=status_col).value = "BAIXADO"
        workbook.save(workbook_path)
        self.log(f"Planilha atualizada: {len(group.row_numbers)} despesa(s) da loja {group.loja} marcadas como BAIXADO.")

    def start_processing(self) -> None:
        if self.validation_result is None:
            messagebox.showwarning("Validação", "Valide a planilha antes de iniciar.")
            return
        if self.validation_result.missing_columns:
            messagebox.showwarning("Validação", "Corrija as colunas obrigatórias antes de iniciar.")
            return
        messagebox.showinfo(
            "Escopo 1",
            "A interface e a validação já estão prontas.\n\n"
            "A automação operacional do fluxo selecionado será implementada no próximo passo.",
        )
        self.log(f"Início bloqueado: automação do fluxo {self.validation_result.flow_label} ainda não foi implementada neste escopo.")

    def pause_processing(self) -> None:
        messagebox.showinfo("Escopo 1", "O controle de pausa será habilitado junto com a automação operacional.")

    def stop_processing(self) -> None:
        messagebox.showinfo("Escopo 1", "O controle de parada será habilitado junto com a automação operacional.")

    def _update_action_buttons(self) -> None:
        validated = self.validation_result is not None and not self.validation_result.missing_columns
        self.start_button.configure(state="normal" if validated else "disabled")
        self.pause_button.configure(state="disabled")
        self.stop_button.configure(state="disabled")

    def start_processing(self) -> None:
        if self.validation_result is None:
            messagebox.showwarning("Validacao", "Valide a planilha antes de iniciar.")
            return
        if self.validation_result.missing_columns:
            messagebox.showwarning("Validacao", "Corrija as colunas obrigatorias antes de iniciar.")
            return
        if not self.username_var.get().strip() or not self.password_var.get().strip():
            messagebox.showwarning("Credenciais", "Informe usuario e senha do TOTVS antes de iniciar.")
            return
        if self.validation_result.flow_key in {"cash", "despesas"}:
            try:
                datetime.strptime(self.issue_date_var.get().strip(), "%d/%m/%Y")
            except ValueError:
                messagebox.showwarning("Data", "Informe a data no formato dd/mm/aaaa.")
                return
            if not self.period_var.get().strip():
                messagebox.showwarning("Periodo", "Informe o periodo que sera usado no historico.")
                return
        if self.processing_thread is not None and self.processing_thread.is_alive():
            messagebox.showinfo("Execucao", "O login no TOTVS ja esta em andamento.")
            return

        self.stop_requested = False
        self.progress.set(0.05)
        self.status_var.set("Iniciando login no TOTVS...")
        self.start_button.configure(state="disabled")
        self.pause_button.configure(state="disabled")
        self.stop_button.configure(state="normal")
        self.log(f"Iniciando preparacao do fluxo {self.validation_result.flow_label}.")

        self.processing_thread = threading.Thread(target=self._run_login_flow, daemon=True)
        self.processing_thread.start()

    def _run_login_flow(self) -> None:
        try:
            self.totvs_bot = TotvsCaixaLoginBot(
                username=self.username_var.get().strip(),
                password=self.password_var.get().strip(),
                headless=self.headless_var.get(),
                log_callback=self.log,
            )
            self.totvs_bot.start()
            if self.validation_result is not None and self.validation_result.flow_key == "cash":
                cash_groups = self._build_cash_groups()
                total_groups = len(cash_groups)
                self.log(f"Cash agrupado em {total_groups} loja(s) para lancamento.")
                self.after(0, lambda: self.status_var.set("TOTVS logado. Iniciando lancamentos de Cash..."))

                for index, group in enumerate(cash_groups, start=1):
                    if self.stop_requested:
                        break
                    self.log(f"Processando Cash {index}/{total_groups}: loja {group.loja}.")
                    self.totvs_bot.open_cash_entry_screen()
                    self.totvs_bot.fill_cash_title(group, self.issue_date_var.get().strip())
                    self.totvs_bot.save_and_lower_cash_title(group)
                    self._mark_cash_group_as_lowered(group)
                    progress = 0.35 + (index / total_groups) * 0.65
                    self.after(0, lambda value=progress: self.progress.set(value))

                self.after(0, lambda: self.status_var.set("Fluxo Cash concluido."))
                self.log("Fluxo Cash concluido.")
            elif self.validation_result is not None and self.validation_result.flow_key == "despesas":
                expense_groups = self._build_expense_groups()
                total_groups = len(expense_groups)
                self.log(f"Despesas agrupadas em {total_groups} loja(s).")
                self.after(0, lambda: self.status_var.set("TOTVS logado. Iniciando lancamentos de Despesas..."))

                if not expense_groups:
                    raise RuntimeError("Nenhum grupo de despesas foi montado.")

                for index, group in enumerate(expense_groups, start=1):
                    if self.stop_requested:
                        break
                    self.log(
                        f"Processando Despesa {index}/{total_groups}: loja {group.loja} | "
                        f"valor R$ {_format_brl(group.valor_total)}."
                    )
                    self.totvs_bot.open_expense_entry_screen()
                    self.totvs_bot.fill_expense_title_basic(group, self.issue_date_var.get().strip())
                    self.totvs_bot.save_and_lower_expense_title(group)
                    self._mark_expense_group_as_lowered(group)
                    progress = 0.35 + (index / total_groups) * 0.65
                    self.after(0, lambda value=progress: self.progress.set(value))

                self.after(0, lambda: self.status_var.set("Fluxo Despesas concluido."))
                self.log("Fluxo Despesas concluido.")
            else:
                self.after(0, lambda: self.progress.set(0.35))
                self.after(0, lambda: self.status_var.set("TOTVS logado. Pronto para entrar no fluxo operacional."))
                self.log("Base comum de login finalizada para Cash, Despesas e Depositos.")
        except Exception as exc:
            self.log(f"Erro ao iniciar/login TOTVS: {exc}")
            self.after(0, lambda: self.status_var.set("Falha no login do TOTVS"))
            self.after(0, lambda: messagebox.showerror("Falha no login TOTVS", str(exc)))
            self.log("Navegador mantido aberto para diagnostico. Use Parar para fechar manualmente.")
        finally:
            self.after(0, self._update_action_buttons)

    def stop_processing(self) -> None:
        self.stop_requested = True
        if self.totvs_bot is not None:
            self.totvs_bot.close()
            self.totvs_bot = None
        self.status_var.set("Execucao interrompida.")
        self.log("Execucao interrompida pelo usuario.")
        self._update_action_buttons()

    def _update_action_buttons(self) -> None:
        running = self.processing_thread is not None and self.processing_thread.is_alive()
        validated = self.validation_result is not None and not self.validation_result.missing_columns
        self.start_button.configure(state="normal" if validated and not running else "disabled")
        self.pause_button.configure(state="disabled")
        self.stop_button.configure(state="normal" if running else "disabled")


if __name__ == "__main__":
    app = RoboLancamentosCaixaApp()
    app.mainloop()
