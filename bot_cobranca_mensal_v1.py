import time
import queue
import threading
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import os
import re
import requests
import getpass
import unicodedata
from datetime import datetime
from urllib.parse import quote

import pythoncom
import win32com.client as win32
from openpyxl import Workbook, load_workbook

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager


class RevisaoManualObrigatoria(Exception):
    pass


class ExecucaoInterrompida(Exception):
    pass


class RoboCobrancaMensalApp:
    MAIN_BG = "#f6f4f1"
    CARD_BG = "#ffffff"
    PRIMARY_TEXT = "#d81919"
    MUTED_TEXT = "#5c5c5c"
    BUTTON_BG = "#ef1a14"
    BUTTON_ACTIVE_BG = "#c91410"
    SUCCESS_TEXT = "#187a2f"

    URL_BASE = "https://coral.aluguefoco.com.br/"
    URL_VALIDACAO = "https://raw.githubusercontent.com/diogodiasyt-blip/validacaofoco/refs/heads/main/chave"
    URL_WHATSAPP_WEB = "https://web.whatsapp.com/"

    STATUS_AGUARDANDO_DEVOLUCAO = "aguardando devolução"
    STATUS_VENCIDO = "vencido"

    TOLERANCIA_VALOR = 0.01

    TIMEOUT_PADRAO = 30
    TIMEOUT_LONGO = 60
    TIMEOUT_CURTO = 5
    TIMEOUT_POPUP = 8

    MAX_TENTATIVAS_ITEM = 2
    MAX_TENTATIVAS_BUSCA = 2
    MAX_TENTATIVAS_TELA_CRITICA = 2

    ASSUNTO_EMAIL_LINK = "Link de Pagamento - Foco Aluguel de Carros"
    EMAIL_REMETENTE_CORPORATIVO = "cobranca@aluguefoco.com.br"
    MODALIDADE_MENSAL = "Mensal"
    MODALIDADE_QUINZENAL = "Quinzenal"

    # LOGIN
    XPATH_USUARIO = "/html/body/foco-app/div[1]/foco-login/div/div/div/div/div[2]/form/div[1]/input"
    XPATH_SENHA = "/html/body/foco-app/div[1]/foco-login/div/div/div/div/div[2]/form/div[2]/input"
    XPATH_BOTAO_ENTRAR = "/html/body/foco-app/div[1]/foco-login/div/div/div/div/div[2]/form/button"

    # CONTRATOS
    XPATH_MENU_CONTRATOS = "/html/body/foco-app/div[1]/div/ul/li[5]/a/i"
    XPATH_ABA_CONTRATOS = "/html/body/foco-app/div[1]/foco-rent-agreement-home/div/ngb-tabset/ul/li[3]/a"
    XPATH_CAMPO_BUSCA = "/html/body/foco-app/div[1]/foco-rent-agreement-home/div/div/div[2]/input"
    XPATH_TD_CONTRATO_RESULTADO = "/html/body/foco-app/div[1]/foco-rent-agreement-home/div/ngb-tabset/div/div/div/div/foco-rent-agreement-list/div/div/div[3]/table/tbody/tr/td[2]"
    XPATH_TD_STATUS_RESULTADO = "/html/body/foco-app/div[1]/foco-rent-agreement-home/div/ngb-tabset/div/div/div/div/foco-rent-agreement-list/div/div/div[3]/table/tbody/tr/td[7]"
    XPATH_MAIS_OPCOES = "/html/body/foco-app/div[1]/foco-rent-agreement-home/div/ngb-tabset/div/div/div/div/foco-rent-agreement-list/div/div/div[3]/table/tbody/tr/td[8]/div/div"
    XPATH_EDITAR = "/html/body/foco-app/div[1]/foco-rent-agreement-home/div/ngb-tabset/div/div/div/div/foco-rent-agreement-list/div/div/div[3]/table/tbody/tr/td[8]/div/div/div/button[1]"

    # EDIÇÃO / PAGAMENTOS
    XPATH_ABA_DADOS_CLIENTE = "/html/body/foco-app/div[1]/foco-rent-agreement-edit/div/div[1]/div/div/div[2]/div[1]/button"
    XPATH_ABA_PAGAMENTOS_RAPIDA = "/html/body/foco-app/div[1]/foco-rent-agreement-edit/div/div[1]/div/div/div[2]/div[11]/button"
    XPATH_ABA_PAGAMENTOS_RAPIDA_ICONE = "/html/body/foco-app/div[1]/foco-rent-agreement-edit/div/div[1]/div/div/div[2]/div[11]/button/i"
    XPATH_POPUP_SIM = "/html/body/ngb-modal-window/div/div/foco-confirm-modal/div[3]/button[2]"
    XPATH_AVANCAR_1 = "/html/body/foco-app/div[1]/foco-rent-agreement-edit/div/div[3]/div/div/div[2]/button"
    XPATH_AVANCAR_2 = "/html/body/foco-app/div[1]/foco-rent-agreement-edit/div/div[3]/div/div/div[2]/button[3]"
    XPATH_AVANCAR_3 = "/html/body/foco-app/div[1]/foco-rent-agreement-edit/div/div[3]/div/div/div[2]/button[2]"
    XPATH_AVANCAR_4 = "/html/body/foco-app/div[1]/foco-rent-agreement-edit/div/div[3]/div/div/div[2]/button[2]"
    XPATH_TEXTO_PAGAMENTO = "/html/body/foco-app/div[1]/foco-rent-agreement-edit/div/div[2]/div[6]/foco-rent-agreement-payment/div/div[2]/div/div[1]"
    XPATH_BOTAO_CARTEIRA = "/html/body/foco-app/div[1]/foco-rent-agreement-edit/div/div[2]/div[6]/foco-rent-agreement-payment/div/div[2]/div/div[2]/div[1]/div[1]/button[6]"

    # CARTEIRA / COBRANÇA
    XPATH_CONTAINER_CARTOES = "//div[contains(@class,'payment-area')]//form"
    XPATH_CAMPO_VALOR_PAGAMENTO = "/html/body/foco-app/div[1]/foco-rent-agreement-edit/div/div[2]/div[6]/foco-rent-agreement-payment/div/div[2]/div/div[2]/div[11]/div/div[2]/foco-form-input/div/div[1]/input"
    XPATH_LISTA_PAGAMENTOS = "/html/body/foco-app/div[1]/foco-rent-agreement-edit/div/div[2]/div[6]/foco-rent-agreement-payment/div/div[2]/div/div[2]/div[11]/div/div[3]/foco-form-dropdown/div/foco-dropdown/div/button/div"
    XPATH_PARCELAMENTO_1X = "/html/body/foco-app/div[1]/foco-rent-agreement-edit/div/div[2]/div[6]/foco-rent-agreement-payment/div/div[2]/div/div[2]/div[11]/div/div[3]/foco-form-dropdown/div/foco-dropdown/div/div/div[2]/perfect-scrollbar/div/div[1]/button[1]/div"
    XPATH_BOTAO_EFETUAR_PAGAMENTO = "/html/body/foco-app/div[1]/foco-rent-agreement-edit/div/div[2]/div[6]/foco-rent-agreement-payment/div/div[2]/div/div[2]/div[19]/button"

    # RESULTADO PAGAMENTO
    XPATH_POPUP_ERRO_FECHAR = "/html/body/ngb-modal-window/div/div//button[contains(normalize-space(), 'Fechar')]"
    XPATH_HISTORICO_PAGAMENTO = "/html/body/foco-app/div[1]/foco-rent-agreement-edit/div/div[2]/div[6]/foco-rent-agreement-payment/div/div[1]/div/div[3]"
    XPATH_BOTAO_CONCLUIR = "/html/body/foco-app/div[1]/foco-rent-agreement-edit/div/div[3]/div/div/div[2]/button[2]"
    XPATH_BOTAO_ATUALIZAR_CONTRATO = "/html/body/ngb-modal-window/div/div/foco-confirm-modal/div[3]/button[2]"
    XPATH_BOTAO_FECHAR_FINAL = "/html/body/ngb-modal-window/div/div/foco-reservation-created/div[3]/button"

    # DADOS DO CLIENTE / LINK
    XPATH_INPUT_EMAIL_CLIENTE = "/html/body/foco-app/div[1]/foco-rent-agreement-edit/div/div[2]/div[1]/div[2]/foco-rent-agreement-customer-form/div/div[2]/div[6]/div/div[1]/foco-form-input/div/div[1]/input"
    XPATH_BOTAO_LINK = "/html/body/foco-app/div[1]/foco-rent-agreement-edit/div/div[2]/div[6]/foco-rent-agreement-payment/div/div[2]/div/div[2]/div[1]/div[1]/button[5]"
    XPATH_CAMPO_VALOR_LINK = "/html/body/foco-app/div[1]/foco-rent-agreement-edit/div/div[2]/div[6]/foco-rent-agreement-payment/div/div[2]/div/div[2]/div[8]/div/div[2]/foco-form-input/div/div[1]/input"
    XPATH_MODALIDADE_A_VENCER = "/html/body/foco-app/div[1]/foco-rent-agreement-edit/div/div[2]/div[6]/foco-rent-agreement-payment/div/div[2]/div/div[2]/div[8]/div/div[3]/foco-form-button-group/div/div/label[1]"
    XPATH_BOTAO_EFETUAR_PAGAMENTO_LINK = "/html/body/foco-app/div[1]/foco-rent-agreement-edit/div/div[2]/div[6]/foco-rent-agreement-payment/div/div[2]/div/div[2]/div[15]/button"
    XPATH_BOTAO_COPIAR_LINK = "/html/body/ngb-modal-window/div/div/foco-pbl-modal/div[2]/div/div[1]/div/button"

    def __init__(self, root):
        self.root = root
        self.root.title("Robô de Cobrança Mensal - Desenvolvido por Diogo Medeiros © 2026")
        self.root.geometry("980x680")
        self.root.minsize(880, 620)
        self.root.configure(bg=self.MAIN_BG)

        self.planilha_path = ""
        self.report_path = None
        self.report_rows = []
        self.report_columns = []
        self.whatsapp_report_input_path = ""
        self.whatsapp_report_output_path = None
        self.whatsapp_report_rows = []
        self.whatsapp_report_columns = []

        self.log_queue = queue.Queue()
        self.whatsapp_log_queue = queue.Queue()
        self.driver = None
        self.whatsapp_driver = None
        self.df_base = []
        self.df_aptos = []
        self.whatsapp_df_base = []
        self.whatsapp_df_aptos = []
        self.total_aptos = 0
        self.whatsapp_total_aptos = 0
        self.cobrancas_concluidas = 0
        self.links_gerados = 0
        self.whatsapp_enviados = 0
        self.pausado = False
        self.parar_solicitado = False
        self.contrato_em_andamento = ""
        self.whatsapp_pausado = False
        self.whatsapp_parar_solicitado = False
        self.whatsapp_item_em_andamento = ""

        self.configurar_estilo()
        self.criar_interface()
        self.atualizar_logs_periodicamente()

    def configurar_estilo(self):
        style = ttk.Style()
        try:
            style.theme_use("clam")
        except Exception:
            pass
        style.configure("App.TFrame", background=self.MAIN_BG)
        style.configure("Card.TFrame", background=self.CARD_BG)
        style.configure("TLabelframe", background=self.CARD_BG, borderwidth=1, relief="solid")
        style.configure("TLabelframe.Label", background=self.CARD_BG, foreground=self.PRIMARY_TEXT, font=("Segoe UI", 10, "bold"))
        style.configure("TLabel", background=self.CARD_BG, foreground="#303030", font=("Segoe UI", 10))
        style.configure("Muted.TLabel", background=self.MAIN_BG, foreground=self.MUTED_TEXT, font=("Segoe UI", 10))
        style.configure("Status.TLabel", background=self.CARD_BG, foreground=self.SUCCESS_TEXT, font=("Segoe UI", 10, "bold"))
        style.configure("TEntry", padding=6)
        style.configure("TCombobox", padding=4)
        style.configure("TCheckbutton", background=self.CARD_BG, font=("Segoe UI", 10))
        style.configure("Primary.TButton", background=self.BUTTON_BG, foreground="#ffffff", padding=(14, 8), font=("Segoe UI", 10, "bold"), borderwidth=0)
        style.map("Primary.TButton", background=[("active", self.BUTTON_ACTIVE_BG), ("pressed", self.BUTTON_ACTIVE_BG)])
        style.configure("Secondary.TButton", background="#ffffff", foreground=self.PRIMARY_TEXT, padding=(12, 8), font=("Segoe UI", 10, "bold"), borderwidth=1)
        style.map("Secondary.TButton", background=[("active", "#fff3f2")], foreground=[("active", self.PRIMARY_TEXT)])
        style.configure("TNotebook", background=self.MAIN_BG, borderwidth=0)
        style.configure("TNotebook.Tab", padding=(18, 8), font=("Segoe UI", 10, "bold"))

    # =========================
    # VALIDAÇÃO REMOTA + PING
    # =========================
    def registrar_abertura(self):
        try:
            url = "https://docs.google.com/forms/d/e/1FAIpQLScmxNbTO-vXw0LEOKIyEhSpIl9aTbw8x5hnEI5VY2eVMRh5gQ/formResponse"
            data = {
                "entry.846583903": getpass.getuser(),
                "entry.1509395143": datetime.now().strftime("%d/%m/%Y %H:%M:%S")
            }
            requests.post(url, data=data, timeout=5)
            self.adicionar_log("Abertura liberada.")
        except Exception as e:
            self.adicionar_log(f"Falha {e}")

    def verificar_chave(self):
        try:
            r = requests.get(self.URL_VALIDACAO, timeout=10)
            ativo = r.text.strip().upper() == "ATIVO"
            self.adicionar_log(f"STATUS DO ROBÔ: {r.text.strip()}")
            return ativo
        except Exception as e:
            self.adicionar_log(f"Falha ao validar chave remota: {e}. Mantendo execução liberada.")
            return True

    # =========================
    # RELATÓRIO EM TEMPO REAL
    # =========================
    def criar_relatorio(self):
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        if self.planilha_path:
            pasta = os.path.dirname(self.planilha_path)
            nome_base = os.path.splitext(os.path.basename(self.planilha_path))[0]
            self.report_path = os.path.join(pasta, f"Relatorio_Cobranca_{nome_base}_{timestamp}.xlsx")
        else:
            self.report_path = f"Relatorio_Cobranca_{timestamp}.xlsx"

        self.report_columns = [
            "Contrato", "Nome", "Telefone", "Valor a Cobrar", "Descricao Cobranca",
            "Tipo", "Status Cobranca", "Tentativas", "Link", "Email Cliente",
            "Status Email", "Status Final", "Data/Hora", "Erro"
        ]
        self.report_rows = []
        self.salvar_relatorio()
        self.adicionar_log(f"Relatório criado: {os.path.basename(self.report_path)}")

    def atualizar_relatorio(self, dados):
        self.report_rows.append(dados)
        self.salvar_relatorio()
        self.adicionar_log(f"Relatório atualizado (Contrato {dados.get('Contrato', '')})")

    def salvar_relatorio(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório"
        ws.append(self.report_columns)
        for linha in self.report_rows:
            ws.append([linha.get(coluna, "") for coluna in self.report_columns])
        wb.save(self.report_path)

    def criar_relatorio_whatsapp(self):
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        if self.whatsapp_report_input_path:
            pasta = os.path.dirname(self.whatsapp_report_input_path)
            nome_base = os.path.splitext(os.path.basename(self.whatsapp_report_input_path))[0]
            self.whatsapp_report_output_path = os.path.join(pasta, f"Relatorio_WhatsApp_{nome_base}_{timestamp}.xlsx")
        else:
            self.whatsapp_report_output_path = f"Relatorio_WhatsApp_{timestamp}.xlsx"

        self.whatsapp_report_columns = [
            "Contrato", "Nome", "Telefone", "Valor a Cobrar", "Descricao Cobranca",
            "Link", "Status WhatsApp", "Status Final", "Data/Hora", "Erro"
        ]
        self.whatsapp_report_rows = []
        self.salvar_relatorio_whatsapp()
        self.adicionar_log_whatsapp(f"Relatório do WhatsApp criado: {os.path.basename(self.whatsapp_report_output_path)}")

    def atualizar_relatorio_whatsapp(self, dados):
        self.whatsapp_report_rows.append(dados)
        self.salvar_relatorio_whatsapp()
        self.adicionar_log_whatsapp(f"Relatório do WhatsApp atualizado (Contrato {dados.get('Contrato', '')})")

    def salvar_relatorio_whatsapp(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Relatorio WhatsApp"
        ws.append(self.whatsapp_report_columns)
        for linha in self.whatsapp_report_rows:
            ws.append([linha.get(coluna, "") for coluna in self.whatsapp_report_columns])
        wb.save(self.whatsapp_report_output_path)

    # =========================
    # INTERFACE
    # =========================
    def criar_interface(self):
        main_frame = ttk.Frame(self.root, padding=12, style="App.TFrame")
        main_frame.pack(fill="both", expand=True)

        header = tk.Frame(main_frame, bg=self.MAIN_BG)
        header.pack(fill="x", pady=(0, 10))
        tk.Label(
            header,
            text="Cobrança FOCO",
            bg=self.MAIN_BG,
            fg=self.PRIMARY_TEXT,
            font=("Segoe UI", 22, "bold")
        ).pack()
        tk.Label(
            header,
            text="Cobrança, link, e-mail e WhatsApp em um único fluxo operacional.",
            bg=self.MAIN_BG,
            fg=self.MUTED_TEXT,
            font=("Segoe UI", 10)
        ).pack(pady=(4, 0))
        self.notebook = ttk.Notebook(main_frame)
        self.notebook.pack(fill="both", expand=True)

        aba_cobranca = ttk.Frame(self.notebook, padding=10)
        aba_whatsapp = ttk.Frame(self.notebook, padding=10)
        self.notebook.add(aba_cobranca, text="Cobrança")
        self.notebook.add(aba_whatsapp, text="WhatsApp")

        topo_cobranca = ttk.Frame(aba_cobranca)
        topo_cobranca.pack(fill="x", pady=(0, 5))

        frame_login = ttk.LabelFrame(topo_cobranca, text="🔑 Login do Sistema", padding=8)
        frame_login.pack(side="left", fill="x", expand=True, padx=(0, 5))

        ttk.Label(frame_login, text="Usuário:").grid(row=0, column=0, sticky="w", padx=5, pady=5)
        self.entry_usuario = ttk.Entry(frame_login, width=35)
        self.entry_usuario.grid(row=0, column=1, sticky="w", padx=5, pady=5)

        ttk.Label(frame_login, text="Senha:").grid(row=1, column=0, sticky="w", padx=5, pady=5)
        self.entry_senha = ttk.Entry(frame_login, width=35, show="*")
        self.entry_senha.grid(row=1, column=1, sticky="w", padx=5, pady=5)

        frame_config = ttk.LabelFrame(topo_cobranca, text="Configurações", padding=8)
        frame_config.pack(side="left", fill="x", expand=True, padx=(5, 0))

        self.var_headless = tk.BooleanVar(value=True)
        self.check_headless = ttk.Checkbutton(
            frame_config,
            text="Executar em modo invisível (sem abrir janela do Chrome)",
            variable=self.var_headless
        )
        self.check_headless.pack(anchor="w", padx=5, pady=5)

        ttk.Label(frame_config, text="E-mail em cópia (opcional):").pack(anchor="w", padx=5, pady=(8, 2))
        self.entry_email_copia = ttk.Entry(frame_config, width=45)
        self.entry_email_copia.pack(anchor="w", padx=5, pady=(0, 5))
        ttk.Label(frame_config, text="Modalidade para validar antes de iniciar:").pack(anchor="w", padx=5, pady=(8, 2))
        self.var_modalidade = tk.StringVar(value=self.MODALIDADE_MENSAL)
        self.combo_modalidade = ttk.Combobox(
            frame_config,
            textvariable=self.var_modalidade,
            values=[self.MODALIDADE_MENSAL, self.MODALIDADE_QUINZENAL],
            state="readonly",
            width=27
        )
        self.combo_modalidade.pack(anchor="w", padx=5, pady=(0, 5))
        self.combo_modalidade.bind("<<ComboboxSelected>>", self.ao_alterar_modalidade)

        meio_cobranca = ttk.Frame(aba_cobranca)
        meio_cobranca.pack(fill="x", pady=5)

        frame_planilha = ttk.LabelFrame(meio_cobranca, text="Planilha Base", padding=8)
        frame_planilha.pack(side="left", fill="both", expand=True, padx=(0, 5))

        self.btn_planilha = ttk.Button(
            frame_planilha,
            text="Selecionar Planilha Excel",
            command=self.selecionar_planilha,
            style="Secondary.TButton"
        )
        self.btn_planilha.pack(pady=(2, 5))

        self.label_planilha = ttk.Label(
            frame_planilha,
            text="Nenhuma planilha selecionada",
            foreground="blue"
        )
        self.label_planilha.pack(anchor="w", padx=5, pady=5)

        self.label_aptos = ttk.Label(
            frame_planilha,
            text=f"Contratos aptos para cobrança ({self.var_modalidade.get()}): 0",
            font=("Arial", 10, "bold")
        )
        self.label_aptos.pack(anchor="w", padx=5, pady=5)

        lateral_execucao = ttk.LabelFrame(meio_cobranca, text="Execução", padding=8)
        lateral_execucao.pack(side="left", fill="both", padx=(5, 0))

        frame_progresso = ttk.Frame(lateral_execucao)
        frame_progresso.pack(fill="x")

        self.progress = ttk.Progressbar(frame_progresso, orient="horizontal", mode="determinate")
        self.progress.pack(fill="x", padx=5, pady=5)

        self.label_progresso = ttk.Label(frame_progresso, text="0/0 - Aguardando início...")
        self.label_progresso.pack(pady=5)

        frame_botao = ttk.Frame(lateral_execucao)
        frame_botao.pack(fill="x", pady=(8, 0))

        self.btn_iniciar = ttk.Button(frame_botao, text="INICIAR ROBÔ", command=self.iniciar_robo, style="Primary.TButton")
        self.btn_iniciar.pack(fill="x", pady=(0, 4))

        self.btn_pausar = ttk.Button(frame_botao, text="Pausar", command=self.alternar_pausa, state="disabled", style="Secondary.TButton")
        self.btn_pausar.pack(fill="x", pady=4)

        self.btn_parar = ttk.Button(frame_botao, text="Parar", command=self.solicitar_parada, state="disabled", style="Secondary.TButton")
        self.btn_parar.pack(fill="x", pady=(4, 0))

        frame_logs = ttk.LabelFrame(aba_cobranca, text="Logs em Tempo Real", padding=8)
        frame_logs.pack(fill="both", expand=True, pady=(5, 0))

        self.txt_logs = scrolledtext.ScrolledText(frame_logs, height=20, wrap="word", state="disabled")
        self.txt_logs.pack(fill="both", expand=True)

        frame_info_whatsapp = ttk.LabelFrame(aba_whatsapp, text="Autenticação do WhatsApp Web", padding=10)
        frame_info_whatsapp.pack(fill="x", pady=5)

        ttk.Label(
            frame_info_whatsapp,
            text=(
                "Selecione o relatório gerado pela cobrança. O robô abrirá o WhatsApp Web, "
                "aguardará a autenticação e enviará as mensagens com o mesmo texto do e-mail."
            ),
            wraplength=880,
            justify="left"
        ).pack(anchor="w", padx=5, pady=5)

        frame_relatorio_whatsapp = ttk.LabelFrame(aba_whatsapp, text="Relatório Base da Cobrança", padding=10)
        frame_relatorio_whatsapp.pack(fill="x", pady=5)

        self.btn_relatorio_whatsapp = ttk.Button(
            frame_relatorio_whatsapp,
            text="Selecionar Relatório da Cobrança",
            command=self.selecionar_relatorio_whatsapp,
            style="Secondary.TButton"
        )
        self.btn_relatorio_whatsapp.pack(pady=5)

        self.label_relatorio_whatsapp = ttk.Label(
            frame_relatorio_whatsapp,
            text="Nenhum relatório selecionado",
            foreground="blue"
        )
        self.label_relatorio_whatsapp.pack(anchor="w", padx=5, pady=5)

        self.label_whatsapp_aptos = ttk.Label(
            frame_relatorio_whatsapp,
            text="Contratos aptos para envio no WhatsApp: 0",
            font=("Arial", 10, "bold")
        )
        self.label_whatsapp_aptos.pack(anchor="w", padx=5, pady=5)

        frame_progresso_whatsapp = ttk.LabelFrame(aba_whatsapp, text="Progresso do WhatsApp", padding=10)
        frame_progresso_whatsapp.pack(fill="x", pady=5)

        self.progress_whatsapp = ttk.Progressbar(frame_progresso_whatsapp, orient="horizontal", mode="determinate")
        self.progress_whatsapp.pack(fill="x", padx=5, pady=5)

        self.label_progresso_whatsapp = ttk.Label(frame_progresso_whatsapp, text="0/0 - Aguardando início...")
        self.label_progresso_whatsapp.pack(pady=5)

        frame_botoes_whatsapp = ttk.Frame(aba_whatsapp)
        frame_botoes_whatsapp.pack(fill="x", pady=10)

        self.btn_iniciar_whatsapp = ttk.Button(
            frame_botoes_whatsapp,
            text="INICIAR WHATSAPP",
            command=self.iniciar_robo_whatsapp,
            style="Primary.TButton"
        )
        self.btn_iniciar_whatsapp.pack(side="left", padx=5)

        self.btn_pausar_whatsapp = ttk.Button(
            frame_botoes_whatsapp,
            text="Pausar",
            command=self.alternar_pausa_whatsapp,
            state="disabled",
            style="Secondary.TButton"
        )
        self.btn_pausar_whatsapp.pack(side="left", padx=5)

        self.btn_parar_whatsapp = ttk.Button(
            frame_botoes_whatsapp,
            text="Parar",
            command=self.solicitar_parada_whatsapp,
            state="disabled",
            style="Secondary.TButton"
        )
        self.btn_parar_whatsapp.pack(side="left", padx=5)

        frame_logs_whatsapp = ttk.LabelFrame(aba_whatsapp, text="Logs do WhatsApp", padding=10)
        frame_logs_whatsapp.pack(fill="both", expand=True, pady=5)

        self.txt_logs_whatsapp = scrolledtext.ScrolledText(frame_logs_whatsapp, height=18, wrap="word", state="disabled")
        self.txt_logs_whatsapp.pack(fill="both", expand=True)

    # =========================
    # LOGS / PROGRESSO
    # =========================
    def adicionar_log(self, mensagem):
        timestamp = time.strftime("%H:%M:%S")
        self.log_queue.put(f"[{timestamp}] {mensagem}")

    def adicionar_log_whatsapp(self, mensagem):
        timestamp = time.strftime("%H:%M:%S")
        self.whatsapp_log_queue.put(f"[{timestamp}] {mensagem}")

    def processar_fila_logs(self, fila, widget):
        try:
            while True:
                mensagem = fila.get_nowait()
                widget.configure(state="normal")
                widget.insert("end", mensagem + "\n")
                widget.see("end")
                widget.configure(state="disabled")
        except queue.Empty:
            pass

    def atualizar_logs_periodicamente(self):
        self.processar_fila_logs(self.log_queue, self.txt_logs)
        self.processar_fila_logs(self.whatsapp_log_queue, self.txt_logs_whatsapp)
        self.root.after(200, self.atualizar_logs_periodicamente)

    def atualizar_progresso(self, atual, total, texto_extra=""):
        percentual = (atual / total) * 100 if total > 0 else 0
        self.progress["value"] = percentual
        texto = f"{atual}/{total}"
        if texto_extra:
            texto += f" — {texto_extra}"
        self.label_progresso.config(text=texto)
        self.root.update_idletasks()

    def atualizar_progresso_whatsapp(self, atual, total, texto_extra=""):
        percentual = (atual / total) * 100 if total > 0 else 0
        self.progress_whatsapp["value"] = percentual
        texto = f"{atual}/{total}"
        if texto_extra:
            texto += f" - {texto_extra}"
        self.label_progresso_whatsapp.config(text=texto)
        self.root.update_idletasks()

    # =========================
    # CONTROLES DE EXECUÇÃO
    # =========================
    def alternar_pausa(self):
        self.pausado = not self.pausado
        if self.pausado:
            self.btn_pausar.config(text="Retomar")
            self.adicionar_log("Pausa solicitada. O robô vai pausar no próximo ponto seguro.")
        else:
            self.btn_pausar.config(text="Pausar")
            self.adicionar_log("Execução retomada.")

    def solicitar_parada(self):
        self.parar_solicitado = True
        self.pausado = False
        self.btn_pausar.config(text="Pausar")
        self.adicionar_log("Parada solicitada. O robô vai encerrar no próximo ponto seguro.")

    def verificar_controle_execucao(self):
        if self.parar_solicitado:
            raise ExecucaoInterrompida("Execução interrompida pelo usuário.")
        while self.pausado:
            self.atualizar_progresso(
                self.progress["value"],
                100,
                f"Pausado no contrato {self.contrato_em_andamento}".strip()
            )
            time.sleep(0.5)
            if self.parar_solicitado:
                raise ExecucaoInterrompida("Execução interrompida pelo usuário durante pausa.")

    def alternar_pausa_whatsapp(self):
        self.whatsapp_pausado = not self.whatsapp_pausado
        if self.whatsapp_pausado:
            self.btn_pausar_whatsapp.config(text="Retomar")
            self.adicionar_log_whatsapp("Pausa solicitada. O envio vai pausar no próximo ponto seguro.")
        else:
            self.btn_pausar_whatsapp.config(text="Pausar")
            self.adicionar_log_whatsapp("Envio retomado.")

    def solicitar_parada_whatsapp(self):
        self.whatsapp_parar_solicitado = True
        self.whatsapp_pausado = False
        self.btn_pausar_whatsapp.config(text="Pausar")
        self.adicionar_log_whatsapp("Parada solicitada. O envio vai encerrar no próximo ponto seguro.")

    def verificar_controle_execucao_whatsapp(self):
        if self.whatsapp_parar_solicitado:
            raise ExecucaoInterrompida("Execução do WhatsApp interrompida pelo usuário.")
        while self.whatsapp_pausado:
            self.atualizar_progresso_whatsapp(
                self.progress_whatsapp["value"],
                100,
                f"Pausado no contrato {self.whatsapp_item_em_andamento}".strip()
            )
            time.sleep(0.5)
            if self.whatsapp_parar_solicitado:
                raise ExecucaoInterrompida("Execução do WhatsApp interrompida pelo usuário durante pausa.")

    # =========================
    # PLANILHA
    # =========================
    def valor_vazio(self, valor):
        if valor is None:
            return True
        if isinstance(valor, str):
            texto = valor.strip()
            return texto == "" or texto.lower() in ("nan", "none", "nat")
        return False

    def normalizar_contrato(self, valor):
        if self.valor_vazio(valor):
            return ""
        if isinstance(valor, float) and valor.is_integer():
            return str(int(valor))
        texto = str(valor).strip()
        if texto.endswith(".0"):
            texto = texto[:-2]
        return texto

    def normalizar_texto(self, valor):
        if self.valor_vazio(valor):
            return ""
        texto = str(valor).strip().lower()
        texto = unicodedata.normalize("NFKD", texto).encode("ascii", "ignore").decode("ascii")
        return " ".join(texto.split())

    def identificar_modalidade_periodo(self, periodo):
        texto = self.normalizar_texto(periodo)
        if not texto:
            return "nao_informado"
        if "quinzen" in texto or "15 em 15" in texto or "15 dias" in texto:
            return "quinzenal"
        if "mens" in texto or "30 dias" in texto:
            return "mensal"
        return "outro"

    def obter_modalidades_selecionadas(self):
        modalidade = self.var_modalidade.get()
        if modalidade == self.MODALIDADE_QUINZENAL:
            return {"quinzenal"}
        return {"mensal"}

    def obter_valor_coluna(self, registro, *nomes):
        for nome in nomes:
            if nome in registro:
                return registro.get(nome)
        return None

    def converter_valor_monetario(self, valor):
        if self.valor_vazio(valor):
            return None
        if isinstance(valor, (int, float)):
            return round(float(valor), 2)
        texto = str(valor).strip()
        if not texto:
            return None
        texto = texto.replace("R$", "").replace(" ", "")
        if "," in texto and "." in texto:
            texto = texto.replace(".", "").replace(",", ".")
        elif "," in texto:
            texto = texto.replace(",", ".")
        try:
            return round(float(texto), 2)
        except Exception:
            return None

    def formatar_valor_pagamento(self, valor):
        valor_float = self.converter_valor_monetario(valor)
        if valor_float is None:
            raise Exception(f"Valor inválido para pagamento: {valor}")
        valor_float = round(valor_float, 2)
        return f"{valor_float:.2f}".replace(".", ",")

    def valores_iguais(self, a, b, tolerancia=None):
        if tolerancia is None:
            tolerancia = self.TOLERANCIA_VALOR
        if a is None or b is None:
            return False
        return abs(a - b) <= tolerancia

    def calcular_metade_valor(self, valor):
        valor_convertido = self.converter_valor_monetario(valor)
        if valor_convertido is None:
            return None
        return round(valor_convertido / 2, 2)

    def obter_descricao_modalidade_cobranca(self):
        if self.var_modalidade.get() == self.MODALIDADE_QUINZENAL:
            return "QUINZENA"
        return "MENSALIDADE"

    def validar_colunas_planilha(self, colunas):
        colunas_obrigatorias = [
            "Contratos", "Total", "Pago", "Período", "Meses Utilizados",
            "Mês", "Mensalidade", "R$ Devido de Pagamento", "R$ a Cobrar",
            "Vencimento", "Nome", "Telefone"
        ]
        faltantes = [col for col in colunas_obrigatorias if col not in colunas]
        if faltantes:
            raise Exception("Colunas obrigatórias ausentes:\n- " + "\n- ".join(faltantes))

    def validar_dados_planilha(self, linhas):
        contratos_vazios = [linha for linha in linhas if not linha.get("Contratos")]
        if contratos_vazios:
            raise Exception(f"Existem {len(contratos_vazios)} linhas com contrato vazio.")

        valores_invalidos = [
            linha for linha in linhas
            if linha.get("R$ a Cobrar_num") is None or linha.get("R$ a Cobrar_num") <= 0
        ]
        if valores_invalidos:
            exemplos = ", ".join(str(linha.get("Contratos", "")) for linha in valores_invalidos[:5])
            raise Exception(f"Existem valores inválidos ou zerados em 'R$ a Cobrar'. Exemplos: {exemplos}")

        vistos = set()
        duplicados = []
        for linha in linhas:
            contrato = linha.get("Contratos", "")
            if contrato in vistos and contrato not in duplicados:
                duplicados.append(contrato)
            vistos.add(contrato)
        if duplicados:
            exemplos = ", ".join(map(str, duplicados[:10]))
            raise Exception(f"Existem contratos duplicados na planilha. Exemplos: {exemplos}")

    def ler_planilha_excel(self, caminho):
        if not caminho.lower().endswith(".xlsx"):
            raise Exception("Use uma planilha .xlsx. O formato .xls depende de bibliotecas antigas e foi desativado para evitar o erro do NumPy.")

        wb = load_workbook(caminho, data_only=True)
        ws = wb.active
        linhas = list(ws.iter_rows(values_only=True))
        if not linhas:
            raise Exception("A planilha está vazia.")

        colunas = [str(col).strip() if col is not None else "" for col in linhas[0]]
        self.validar_colunas_planilha(colunas)

        registros = []
        for valores in linhas[1:]:
            registro = {colunas[i]: valores[i] if i < len(valores) else None for i in range(len(colunas))}
            if self.valor_vazio(registro.get("Contratos")):
                continue
            registros.append(registro)
        return registros

    def preparar_dataframe(self, linhas):
        registros = []
        for linha in linhas:
            registro = dict(linha)
            registro["Contratos"] = self.normalizar_contrato(registro.get("Contratos"))
            registro["Nome"] = "" if self.valor_vazio(registro.get("Nome")) else str(registro.get("Nome")).strip()
            registro["Vencimento"] = "" if self.valor_vazio(registro.get("Vencimento")) else str(registro.get("Vencimento")).strip().lower()
            periodo_original = self.obter_valor_coluna(registro, "Período", "Período", "Período", "Periodo")
            registro["Periodo_norm"] = self.normalizar_texto(periodo_original)
            registro["Modalidade"] = self.identificar_modalidade_periodo(periodo_original)
            registro["Mensalidade_num"] = self.converter_valor_monetario(registro.get("Mensalidade"))
            registro["R$ a Cobrar_num"] = self.converter_valor_monetario(registro.get("R$ a Cobrar"))
            registro["Quinzenalidade_num"] = self.calcular_metade_valor(registro.get("Mensalidade"))
            registro["mensalidade_igual_cobrar"] = self.valores_iguais(
                registro["Mensalidade_num"],
                registro["R$ a Cobrar_num"]
            )
            registro["quinzenalidade_igual_cobrar"] = self.valores_iguais(
                registro["Quinzenalidade_num"],
                registro["R$ a Cobrar_num"]
            )
            registros.append(registro)
        self.validar_dados_planilha(registros)
        return registros

    def contrato_esta_apto(self, linha):
        if linha.get("Vencimento") != self.STATUS_VENCIDO:
            return False
        modalidade = self.var_modalidade.get()
        if modalidade == self.MODALIDADE_MENSAL:
            return linha.get("mensalidade_igual_cobrar") is True
        return linha.get("quinzenalidade_igual_cobrar") is True

    def filtrar_contratos_aptos(self, linhas):
        return [linha for linha in linhas if self.contrato_esta_apto(linha)]

    def carregar_e_validar_planilha(self, caminho):
        self.adicionar_log("Lendo planilha Excel...")
        linhas = self.ler_planilha_excel(caminho)
        registros = self.preparar_dataframe(linhas)
        registros_aptos = self.filtrar_contratos_aptos(registros)
        return registros, registros_aptos

    def atualizar_interface_planilha(self):
        self.label_planilha.config(text=self.planilha_path or "Nenhuma planilha selecionada")
        self.label_aptos.config(text=f"Contratos aptos para cobrança ({self.var_modalidade.get()}): {self.total_aptos}")
        self.atualizar_progresso(0, self.total_aptos, "Planilha carregada" if self.total_aptos > 0 else "Sem contratos aptos")

    def registrar_resumo_planilha(self):
        qtd_vencidos = len([linha for linha in self.df_base if linha.get("Vencimento") == self.STATUS_VENCIDO])
        qtd_mes_igual = len([linha for linha in self.df_base if linha.get("mensalidade_igual_cobrar") is True])
        qtd_quinzenal = len([linha for linha in self.df_base if linha.get("quinzenalidade_igual_cobrar") is True])
        self.adicionar_log(f"Planilha selecionada: {self.planilha_path}")
        self.adicionar_log(f"Total de registros na base: {len(self.df_base)}")
        self.adicionar_log(f"Contratos com vencimento '{self.STATUS_VENCIDO}': {qtd_vencidos}")
        self.adicionar_log(f"Contratos aptos pela regra mensal: {qtd_mes_igual}")
        self.adicionar_log(f"Contratos aptos pela regra quinzenal: {qtd_quinzenal}")
        self.adicionar_log(f"Filtro selecionado antes da execução: {self.var_modalidade.get()}")
        self.adicionar_log(f"Contratos aptos para cobrança ({self.var_modalidade.get()}): {self.total_aptos}")

    def recalcular_contratos_aptos(self):
        if not self.df_base:
            self.df_aptos = []
            self.total_aptos = 0
        else:
            self.df_aptos = self.filtrar_contratos_aptos(self.df_base)
            self.total_aptos = len(self.df_aptos)
        self.atualizar_interface_planilha()

    def ao_alterar_modalidade(self, event=None):
        self.adicionar_log(f"Modalidade selecionada para validação: {self.var_modalidade.get()}")
        self.recalcular_contratos_aptos()

    def limpar_estado_planilha(self):
        self.planilha_path = ""
        self.df_base = []
        self.df_aptos = []
        self.total_aptos = 0
        self.label_planilha.config(text="Nenhuma planilha selecionada")
        self.label_aptos.config(text=f"Contratos aptos para cobrança ({self.var_modalidade.get()}): 0")
        self.atualizar_progresso(0, 0, "Erro na leitura da planilha")

    def selecionar_planilha(self):
        caminho = filedialog.askopenfilename(
            title="Selecionar planilha",
            filetypes=[("Arquivos Excel", "*.xlsx")]
        )
        if not caminho:
            return
        try:
            self.planilha_path = caminho
            self.df_base, self.df_aptos = self.carregar_e_validar_planilha(caminho)
            self.total_aptos = len(self.df_aptos)
            self.atualizar_interface_planilha()
            self.registrar_resumo_planilha()
        except Exception as e:
            self.limpar_estado_planilha()
            messagebox.showerror("Erro", f"Falha ao carregar planilha:\n\n{str(e)}")

    def limpar_estado_whatsapp(self):
        self.whatsapp_report_input_path = ""
        self.whatsapp_df_base = []
        self.whatsapp_df_aptos = []
        self.whatsapp_total_aptos = 0
        self.label_relatorio_whatsapp.config(text="Nenhum relatório selecionado")
        self.label_whatsapp_aptos.config(text="Contratos aptos para envio no WhatsApp: 0")
        self.atualizar_progresso_whatsapp(0, 0, "Erro na leitura do relatório")

    def normalizar_telefone_whatsapp(self, valor):
        if self.valor_vazio(valor):
            return ""
        telefone = re.sub(r"\D", "", str(valor))
        if not telefone:
            return ""
        if telefone.startswith("55"):
            telefone = telefone[2:]
        if len(telefone) == 10:
            telefone = telefone[:2] + "9" + telefone[2:]
        if len(telefone) != 11:
            return ""
        return f"55{telefone}"

    def ler_relatorio_whatsapp(self, caminho):
        wb = load_workbook(caminho, data_only=True)
        ws = wb.active
        linhas = list(ws.iter_rows(values_only=True))
        if not linhas:
            raise Exception("O relatório selecionado está vazio.")

        colunas = [str(col).strip() if col is not None else "" for col in linhas[0]]
        registros = []
        for valores in linhas[1:]:
            if not valores or all(self.valor_vazio(valor) for valor in valores):
                continue
            registro = {colunas[i]: valores[i] if i < len(valores) else None for i in range(len(colunas))}
            registro["Contrato"] = self.normalizar_contrato(registro.get("Contrato"))
            registro["Nome"] = "" if self.valor_vazio(registro.get("Nome")) else str(registro.get("Nome")).strip()
            registro["Telefone"] = "" if self.valor_vazio(registro.get("Telefone")) else str(registro.get("Telefone")).strip()
            registro["Telefone_whatsapp"] = self.normalizar_telefone_whatsapp(registro.get("Telefone"))
            registro["Link"] = "" if self.valor_vazio(registro.get("Link")) else str(registro.get("Link")).strip()
            registro["Descricao Cobranca"] = (
                "" if self.valor_vazio(registro.get("Descricao Cobranca"))
                else str(registro.get("Descricao Cobranca")).strip().upper()
            )
            registros.append(registro)
        return registros

    def contrato_apto_whatsapp(self, linha):
        return bool(linha.get("Contrato")) and bool(linha.get("Link")) and bool(linha.get("Telefone_whatsapp"))

    def carregar_relatorio_whatsapp(self, caminho):
        self.adicionar_log_whatsapp("Lendo relatório da cobrança para envio no WhatsApp...")
        linhas = self.ler_relatorio_whatsapp(caminho)
        aptos = [linha for linha in linhas if self.contrato_apto_whatsapp(linha)]
        return linhas, aptos

    def atualizar_interface_whatsapp(self):
        self.label_relatorio_whatsapp.config(text=self.whatsapp_report_input_path or "Nenhum relatório selecionado")
        self.label_whatsapp_aptos.config(text=f"Contratos aptos para envio no WhatsApp: {self.whatsapp_total_aptos}")
        self.atualizar_progresso_whatsapp(
            0,
            self.whatsapp_total_aptos,
            "Relatório carregado" if self.whatsapp_total_aptos > 0 else "Sem contratos aptos"
        )

    def selecionar_relatorio_whatsapp(self):
        caminho = filedialog.askopenfilename(
            title="Selecionar relatório da cobrança",
            filetypes=[("Arquivos Excel", "*.xlsx")]
        )
        if not caminho:
            return
        try:
            self.whatsapp_report_input_path = caminho
            self.whatsapp_df_base, self.whatsapp_df_aptos = self.carregar_relatorio_whatsapp(caminho)
            self.whatsapp_total_aptos = len(self.whatsapp_df_aptos)
            self.atualizar_interface_whatsapp()
            self.adicionar_log_whatsapp(f"Relatório selecionado: {self.whatsapp_report_input_path}")
            self.adicionar_log_whatsapp(f"Total de registros na base: {len(self.whatsapp_df_base)}")
            self.adicionar_log_whatsapp(f"Contratos aptos para envio no WhatsApp: {self.whatsapp_total_aptos}")
        except Exception as e:
            self.limpar_estado_whatsapp()
            messagebox.showerror("Erro", f"Falha ao carregar relatório:\n\n{str(e)}")

    # =========================
    # VALIDAÇÕES
    # =========================
    def validar_campos(self):
        usuario = self.entry_usuario.get().strip()
        senha = self.entry_senha.get().strip()
        if not usuario:
            messagebox.showwarning("Atenção", "Preencha o usuário.")
            return False
        if not senha:
            messagebox.showwarning("Atenção", "Preencha a senha.")
            return False
        if not self.planilha_path:
            messagebox.showwarning("Atenção", "Selecione a planilha base.")
            return False
        if self.total_aptos <= 0:
            messagebox.showwarning("Atenção", "Não há contratos aptos para cobrança.")
            return False
        email_copia = self.entry_email_copia.get().strip()
        if email_copia and not re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", email_copia):
            messagebox.showwarning("Atenção", "E-mail em cópia inválido.")
            return False
        return True

    # =========================
    # DRIVER / SESSÃO
    # =========================
    def configurar_driver(self):
        options = Options()
        if self.var_headless.get():
            options.add_argument("--headless=new")
        options.add_argument("--start-maximized")
        options.add_argument("--disable-gpu")
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--log-level=3")
        driver = webdriver.Chrome(
            service=Service(ChromeDriverManager().install()),
            options=options
        )
        driver.set_page_load_timeout(self.TIMEOUT_LONGO)
        return driver

    def criar_driver_se_necessario(self):
        if self.driver is None:
            self.adicionar_log("Criando nova sessão do navegador...")
            self.driver = self.configurar_driver()

    def fechar_driver(self):
        try:
            if self.driver:
                self.driver.quit()
        except Exception:
            pass
        finally:
            self.driver = None

    def reset_completo_navegador(self, motivo=""):
        self.adicionar_log(f"Reset completo do navegador. Motivo: {motivo}")
        self.fechar_driver()
        self.criar_driver_se_necessario()

    def configurar_driver_whatsapp(self):
        options = Options()
        options.add_argument("--start-maximized")
        options.add_argument("--disable-gpu")
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--log-level=3")
        driver = webdriver.Chrome(
            service=Service(ChromeDriverManager().install()),
            options=options
        )
        driver.set_page_load_timeout(self.TIMEOUT_LONGO)
        return driver

    def criar_driver_whatsapp_se_necessario(self):
        if self.whatsapp_driver is None:
            self.adicionar_log_whatsapp("Criando nova sessão do navegador para o WhatsApp...")
            self.whatsapp_driver = self.configurar_driver_whatsapp()

    def fechar_driver_whatsapp(self):
        try:
            if self.whatsapp_driver:
                self.whatsapp_driver.quit()
        except Exception:
            pass
        finally:
            self.whatsapp_driver = None

    # =========================
    # HELPERS SELENIUM
    # =========================
    def esperar_elemento(self, by, locator, timeout=None):
        if timeout is None:
            timeout = self.TIMEOUT_PADRAO
        return WebDriverWait(self.driver, timeout).until(
            EC.presence_of_element_located((by, locator))
        )

    def esperar_visivel(self, by, locator, timeout=None):
        if timeout is None:
            timeout = self.TIMEOUT_PADRAO
        return WebDriverWait(self.driver, timeout).until(
            EC.visibility_of_element_located((by, locator))
        )

    def esperar_clicavel(self, by, locator, timeout=None):
        if timeout is None:
            timeout = self.TIMEOUT_PADRAO
        return WebDriverWait(self.driver, timeout).until(
            EC.element_to_be_clickable((by, locator))
        )

    def elemento_existe(self, by, locator, timeout=None):
        if timeout is None:
            timeout = self.TIMEOUT_CURTO
        try:
            self.esperar_elemento(by, locator, timeout=timeout)
            return True
        except Exception:
            return False

    def obter_texto(self, by, locator, descricao="", timeout=None):
        el = self.esperar_visivel(by, locator, timeout=timeout)
        texto = el.text.strip()
        if descricao:
            self.adicionar_log(f"{descricao}: {texto}")
        return texto

    def obter_valor_input(self, by, locator, descricao="", timeout=None):
        el = self.esperar_visivel(by, locator, timeout=timeout)
        valor = el.get_attribute("value") or ""
        valor = valor.strip()
        if descricao:
            self.adicionar_log(f"{descricao}: {valor}")
        return valor

    def clicar_seguro(self, by, locator, descricao="", timeout=None):
        if timeout is None:
            timeout = self.TIMEOUT_PADRAO
        try:
            el = self.esperar_clicavel(by, locator, timeout=timeout)
            self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", el)
            time.sleep(0.3)
            el.click()
            self.adicionar_log(f"Clique OK em {descricao}.")
            return True
        except Exception as erro_click:
            self.adicionar_log(f"Clique normal falhou em {descricao}: {erro_click}")
            el = self.esperar_elemento(by, locator, timeout=timeout)
            self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", el)
            time.sleep(0.3)
            self.driver.execute_script("arguments[0].click();", el)
            self.adicionar_log(f"Clique via JavaScript OK em {descricao}.")
            return True

    def digitar_seguro(self, by, locator, texto, descricao="", timeout=None, limpar=True, enter=False):
        if timeout is None:
            timeout = self.TIMEOUT_PADRAO
        el = self.esperar_visivel(by, locator, timeout=timeout)
        self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", el)
        if limpar:
            try:
                el.click()
                el.send_keys(Keys.CONTROL, "a")
                el.send_keys(Keys.DELETE)
            except Exception:
                pass
        el.send_keys(str(texto))
        texto_log = texto
        descricao_normalizada = self.normalizar_texto(descricao)
        if any(chave in descricao_normalizada for chave in ["senha", "password"]):
            texto_log = "*" * max(len(str(texto)), 8)
        self.adicionar_log(f"Texto digitado em {descricao}: {texto_log}")
        if enter:
            el.send_keys(Keys.ENTER)
            self.adicionar_log(f"ENTER enviado em {descricao}.")
        return el

    def preencher_input_valor(self, xpath_input, valor_formatado, descricao):
        campo = self.esperar_visivel(By.XPATH, xpath_input, timeout=self.TIMEOUT_PADRAO)
        self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", campo)
        time.sleep(0.3)
        try:
            campo.click()
            time.sleep(0.2)
            campo.send_keys(Keys.CONTROL, "a")
            campo.send_keys(Keys.DELETE)
            time.sleep(0.2)
            campo.send_keys(valor_formatado)
            time.sleep(0.5)
            valor_lido = (campo.get_attribute("value") or "").strip()
            if valor_lido:
                self.adicionar_log(f"{descricao} preenchido com sucesso: {valor_lido}")
                return True
        except Exception as e:
            self.adicionar_log(f"Falha na digitação normal em {descricao}: {e}")

        self.driver.execute_script("arguments[0].value = '';", campo)
        self.driver.execute_script("arguments[0].dispatchEvent(new Event('input', {bubbles:true}));", campo)
        self.driver.execute_script("arguments[0].value = arguments[1];", campo, valor_formatado)
        self.driver.execute_script("arguments[0].dispatchEvent(new Event('input', {bubbles:true}));", campo)
        self.driver.execute_script("arguments[0].dispatchEvent(new Event('change', {bubbles:true}));", campo)
        time.sleep(0.5)
        valor_lido = (campo.get_attribute("value") or "").strip()
        if valor_lido:
            self.adicionar_log(f"{descricao} preenchido via JavaScript: {valor_lido}")
            return True
        raise Exception(f"Não foi possível preencher {descricao}")

    def tentar_ler_clipboard(self):
        try:
            self.root.update()
            texto = self.root.clipboard_get()
            return texto.strip()
        except Exception:
            return ""

    # =========================
    # LOGIN / SESSÃO
    # =========================
    def esta_logado(self):
        return self.elemento_existe(By.XPATH, self.XPATH_MENU_CONTRATOS, timeout=self.TIMEOUT_CURTO)

    def fazer_login(self, usuario, senha):
        self.adicionar_log("Executando login...")
        self.digitar_seguro(By.XPATH, self.XPATH_USUARIO, usuario, "campo usuário", timeout=self.TIMEOUT_PADRAO)
        self.digitar_seguro(By.XPATH, self.XPATH_SENHA, senha, "campo senha", timeout=self.TIMEOUT_PADRAO)
        self.clicar_seguro(By.XPATH, self.XPATH_BOTAO_ENTRAR, "botão Entrar", timeout=self.TIMEOUT_PADRAO)
        WebDriverWait(self.driver, self.TIMEOUT_LONGO).until(lambda d: self.esta_logado())
        self.adicionar_log("Login concluído com sucesso.")

    def abrir_sistema(self, relogin=False):
        self.criar_driver_se_necessario()
        self.adicionar_log("Acessando sistema...")
        self.driver.get(self.URL_BASE)
        if self.esta_logado():
            self.adicionar_log("Sessão já autenticada reaproveitada.")
            return True
        self.adicionar_log("Sessão não autenticada. Login necessário.")
        usuario = self.entry_usuario.get().strip()
        senha = self.entry_senha.get().strip()
        self.esperar_visivel(By.XPATH, self.XPATH_USUARIO, timeout=self.TIMEOUT_LONGO)
        self.fazer_login(usuario, senha)
        return True

    def recuperar_sessao(self):
        self.adicionar_log("Tentando recuperar sessão...")
        try:
            if self.esta_logado():
                self.adicionar_log("Sessão já está válida.")
                return True
            self.abrir_sistema(relogin=False)
            return True
        except Exception as e:
            self.adicionar_log(f"Recuperação simples falhou: {e}")
        self.reset_completo_navegador("Falha ao recuperar sessão")
        self.abrir_sistema(relogin=True)
        return True

    # =========================
    # TELA BASE / CONTRATOS
    # =========================
    def validar_tela_contratos(self):
        self.esperar_visivel(By.XPATH, self.XPATH_CAMPO_BUSCA, timeout=self.TIMEOUT_PADRAO)
        self.adicionar_log("Validação da tela de contratos OK.")
        return True

    def ir_para_contratos(self):
        self.adicionar_log("Indo para a tela base de contratos...")
        if not self.esta_logado():
            self.recuperar_sessao()
        self.clicar_seguro(By.XPATH, self.XPATH_MENU_CONTRATOS, "menu Contratos", timeout=self.TIMEOUT_PADRAO)
        self.esperar_visivel(By.XPATH, self.XPATH_ABA_CONTRATOS, timeout=self.TIMEOUT_PADRAO)
        self.clicar_seguro(By.XPATH, self.XPATH_ABA_CONTRATOS, "aba Contratos", timeout=self.TIMEOUT_PADRAO)
        self.validar_tela_contratos()
        return True

    def esperar_resultado_busca(self, texto, timeout=None):
        if timeout is None:
            timeout = self.TIMEOUT_PADRAO
        xpath = f'//td[contains(normalize-space(), "{texto}")]'
        self.esperar_elemento(By.XPATH, xpath, timeout=timeout)
        self.adicionar_log(f"Resultado da busca encontrado para contrato {texto}.")
        return True

    def buscar_contrato_seguro(self, numero_contrato, tentativas=None):
        if tentativas is None:
            tentativas = self.MAX_TENTATIVAS_BUSCA
        for tentativa in range(1, tentativas + 1):
            try:
                self.adicionar_log(f"Busca do contrato {numero_contrato} - tentativa {tentativa}/{tentativas}")
                self.digitar_seguro(
                    By.XPATH,
                    self.XPATH_CAMPO_BUSCA,
                    numero_contrato,
                    descricao="campo de busca de contratos",
                    timeout=self.TIMEOUT_PADRAO,
                    limpar=True,
                    enter=True
                )
                self.esperar_resultado_busca(str(numero_contrato), timeout=self.TIMEOUT_PADRAO)
                texto_contrato = self.obter_texto(
                    By.XPATH,
                    self.XPATH_TD_CONTRATO_RESULTADO,
                    "Contrato retornado na pesquisa",
                    timeout=self.TIMEOUT_PADRAO
                )
                if str(texto_contrato).strip() != str(numero_contrato).strip():
                    raise Exception(
                        f"Contrato pesquisado não confere. Planilha: {numero_contrato} | Tela: {texto_contrato}"
                    )
                self.adicionar_log(f"Busca validada com sucesso para o contrato {numero_contrato}.")
                return True
            except Exception as e:
                self.adicionar_log(f"Falha na busca do contrato {numero_contrato}: {e}")
                if tentativa < tentativas:
                    self.reset_completo_navegador(f"Falha na busca do contrato {numero_contrato}")
                    self.abrir_sistema(relogin=True)
                    self.ir_para_contratos()
                else:
                    raise

    # =========================
    # CONTRATO / EDIÇÃO
    # =========================
    def validar_status_aguardando_devolucao(self):
        status = self.obter_texto(
            By.XPATH,
            self.XPATH_TD_STATUS_RESULTADO,
            "Status do contrato",
            timeout=self.TIMEOUT_PADRAO
        ).strip().lower()
        if status != self.STATUS_AGUARDANDO_DEVOLUCAO:
            raise Exception(f"Status do contrato diferente de '{self.STATUS_AGUARDANDO_DEVOLUCAO}': {status}")
        self.adicionar_log("Status validado: Aguardando devolução.")
        return True

    def validar_tela_edicao(self, timeout=None):
        if timeout is None:
            timeout = self.TIMEOUT_PADRAO
        self.esperar_elemento(By.XPATH, self.XPATH_ABA_PAGAMENTOS_RAPIDA_ICONE, timeout=timeout)
        self.adicionar_log("Tela de edição validada com sucesso.")
        return True

    def abrir_edicao_contrato(self):
        self.clicar_seguro(By.XPATH, self.XPATH_MAIS_OPCOES, "Mais opções", timeout=self.TIMEOUT_PADRAO)
        self.esperar_visivel(By.XPATH, self.XPATH_EDITAR, timeout=self.TIMEOUT_PADRAO)
        self.clicar_seguro(By.XPATH, self.XPATH_EDITAR, "Editar", timeout=self.TIMEOUT_PADRAO)
        self.validar_tela_edicao(timeout=self.TIMEOUT_PADRAO)
        return True

    # =========================
    # PAGAMENTOS
    # =========================
    def esta_na_aba_pagamentos(self, timeout=None):
        if timeout is None:
            timeout = self.TIMEOUT_CURTO
        try:
            texto = self.obter_texto(By.XPATH, self.XPATH_TEXTO_PAGAMENTO, timeout=timeout)
            return "pagamento" in texto.strip().lower()
        except Exception:
            return False

    def validar_aba_pagamentos(self, timeout=None):
        if timeout is None:
            timeout = self.TIMEOUT_PADRAO
        WebDriverWait(self.driver, timeout).until(
            lambda d: self.esta_na_aba_pagamentos(timeout=self.TIMEOUT_CURTO)
        )
        self.adicionar_log("Aba Pagamentos validada com sucesso.")
        return True

    def validar_aba_pagamentos_com_retry(self):
        for tentativa in range(1, self.MAX_TENTATIVAS_TELA_CRITICA + 1):
            try:
                self.validar_aba_pagamentos(timeout=self.TIMEOUT_PADRAO)
                return True
            except Exception as e:
                self.adicionar_log(f"Falha na validação da aba Pagamentos na tentativa {tentativa}: {e}")
                if tentativa < self.MAX_TENTATIVAS_TELA_CRITICA:
                    self.adicionar_log("Tentando refresh para validar aba Pagamentos novamente...")
                    self.driver.refresh()
                    time.sleep(2)
        raise Exception("Não foi possível validar a aba Pagamentos após as tentativas permitidas.")

    def popup_sim_apareceu(self, timeout=None):
        if timeout is None:
            timeout = self.TIMEOUT_POPUP
        try:
            self.esperar_visivel(By.XPATH, self.XPATH_POPUP_SIM, timeout=timeout)
            return True
        except Exception:
            return False

    def aguardar_aba_pagamentos_disponivel(self, timeout=None):
        if timeout is None:
            timeout = self.TIMEOUT_LONGO
        self.adicionar_log("Aguardando aba Pagamentos ficar realmente pronta para clique...")
        fim = time.time() + timeout
        ultima_falha = ""
        while time.time() < fim:
            try:
                botao = self.driver.find_element(By.XPATH, self.XPATH_ABA_PAGAMENTOS_RAPIDA)
                visivel = botao.is_displayed()
                habilitado = botao.is_enabled()
                if visivel and habilitado:
                    try:
                        WebDriverWait(self.driver, 2).until(
                            EC.element_to_be_clickable((By.XPATH, self.XPATH_ABA_PAGAMENTOS_RAPIDA))
                        )
                        self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", botao)
                        time.sleep(0.5)
                        self.adicionar_log("Aba Pagamentos liberada para clique.")
                        return True
                    except Exception as e:
                        ultima_falha = str(e)
                else:
                    ultima_falha = f"visível={visivel}, habilitado={habilitado}"
            except Exception as e:
                ultima_falha = str(e)
            time.sleep(1)
        raise Exception(f"A aba Pagamentos não ficou disponível dentro do tempo limite. Última falha: {ultima_falha}")

    def tentar_ir_pagamentos_via_rapida(self):
        self.adicionar_log("Tentando método rápido para ir à aba Pagamentos...")
        self.aguardar_aba_pagamentos_disponivel(timeout=self.TIMEOUT_LONGO)
        self.clicar_seguro(
            By.XPATH,
            self.XPATH_ABA_PAGAMENTOS_RAPIDA,
            "aba Pagamentos (método rápido)",
            timeout=self.TIMEOUT_LONGO
        )
        self.validar_aba_pagamentos_com_retry()
        self.adicionar_log("Método rápido funcionou. Já estamos em Pagamentos.")
        return True

    def tentar_fluxo_popup_para_pagamentos(self):
        self.adicionar_log("Popup detectado. Iniciando fluxo alternativo até Pagamentos...")
        self.clicar_seguro(By.XPATH, self.XPATH_POPUP_SIM, "botão Sim do popup", timeout=self.TIMEOUT_PADRAO)
        time.sleep(1)
        self.clicar_seguro(By.XPATH, self.XPATH_AVANCAR_1, "Avançar 1", timeout=self.TIMEOUT_PADRAO)
        time.sleep(1)
        self.clicar_seguro(By.XPATH, self.XPATH_AVANCAR_2, "Avançar 2", timeout=self.TIMEOUT_PADRAO)
        time.sleep(1)
        self.clicar_seguro(By.XPATH, self.XPATH_AVANCAR_3, "Avançar 3", timeout=self.TIMEOUT_PADRAO)
        time.sleep(1)
        self.clicar_seguro(By.XPATH, self.XPATH_AVANCAR_4, "Avançar 4", timeout=self.TIMEOUT_PADRAO)
        time.sleep(1)
        self.validar_aba_pagamentos_com_retry()
        self.adicionar_log("Fluxo alternativo concluído. Já estamos em Pagamentos.")
        return True

    def clicar_carteira(self):
        self.adicionar_log("Clicando em Carteira...")
        self.clicar_seguro(By.XPATH, self.XPATH_BOTAO_CARTEIRA, "botão Carteira", timeout=self.TIMEOUT_PADRAO)
        self.adicionar_log("Botão Carteira clicado com sucesso.")
        return True

    def ir_para_pagamentos(self):
        self.adicionar_log("Verificando se aparece popup antes de tentar ir para Pagamentos...")
        if self.esta_na_aba_pagamentos():
            self.adicionar_log("Já está na aba Pagamentos.")
            return True
        if self.popup_sim_apareceu(timeout=self.TIMEOUT_POPUP):
            self.adicionar_log("Popup apareceu após clicar em Editar.")
            self.tentar_fluxo_popup_para_pagamentos()
            return True
        self.adicionar_log("Popup não apareceu. Aguardando liberação da aba Pagamentos...")
        self.tentar_ir_pagamentos_via_rapida()
        return True

    # =========================
    # COBRANÇA POR CARTÃO
    # =========================
    def obter_texto_historico_pagamento(self):
        try:
            return self.obter_texto(By.XPATH, self.XPATH_HISTORICO_PAGAMENTO, timeout=self.TIMEOUT_CURTO)
        except Exception:
            return ""

    def listar_cartoes_disponiveis(self):
        self.adicionar_log("Localizando cartões disponíveis na aba Carteira...")
        WebDriverWait(self.driver, self.TIMEOUT_PADRAO).until(
            lambda d: len(d.find_elements(
                By.XPATH,
                "//div[contains(@class,'payment-area')]//div[contains(@class,'cardWallet')]"
            )) > 0
        )
        time.sleep(1)
        cartoes = self.driver.find_elements(
            By.XPATH,
            "//div[contains(@class,'payment-area')]//div[contains(@class,'cardWallet')]"
        )
        cartoes_validos = []
        for i, cartao in enumerate(cartoes, start=1):
            try:
                self.driver.execute_script("arguments[0].scrollIntoView({block:'center'});", cartao)
                time.sleep(0.2)
                radio = cartao.find_element(
                    By.XPATH,
                    ".//input[@type='radio' and @formcontrolname='cardSelected']"
                )
                try:
                    texto = cartao.find_element(
                        By.XPATH,
                        ".//span[contains(normalize-space(),'Número do cartão')]"
                    ).text.strip()
                except Exception:
                    texto = f"Cartão {i}"
                cartoes_validos.append({
                    "elemento": cartao,
                    "radio": radio,
                    "texto": texto
                })
                self.adicionar_log(f"Cartão detectado {len(cartoes_validos)}: {texto}")
            except Exception as e:
                self.adicionar_log(f"Falha ao mapear cartão {i}: {e}")
        self.adicionar_log(f"Total de cartões detectados: {len(cartoes_validos)}")
        if not cartoes_validos:
            raise Exception("Nenhum cartão foi encontrado na aba Carteira.")
        return cartoes_validos

    def selecionar_cartao_por_indice(self, indice):
        cartoes = self.listar_cartoes_disponiveis()
        if indice >= len(cartoes):
            raise Exception(f"Índice de cartão inválido: {indice}")
        cartao = cartoes[indice]
        texto_cartao = cartao["texto"]
        radio = cartao["radio"]
        bloco = cartao["elemento"]
        self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", bloco)
        time.sleep(0.5)
        try:
            if not radio.is_selected():
                radio.click()
        except Exception:
            try:
                self.driver.execute_script("arguments[0].click();", radio)
            except Exception:
                self.driver.execute_script("arguments[0].click();", bloco)
        time.sleep(0.5)
        try:
            selecionado = radio.is_selected()
        except Exception:
            selecionado = False
        self.adicionar_log(f"Cartão selecionado: {texto_cartao} | selected={selecionado}")
        return texto_cartao

    def preencher_valor_pagamento_cartao(self, valor):
        valor_formatado = self.formatar_valor_pagamento(valor)
        self.adicionar_log(f"Preenchendo valor do pagamento no cartão: {valor_formatado}")
        self.preencher_input_valor(
            self.XPATH_CAMPO_VALOR_PAGAMENTO,
            valor_formatado,
            "valor do pagamento no cartão"
        )

    def selecionar_parcela_1x(self):
        self.adicionar_log("Selecionando parcelamento 1x...")
        self.clicar_seguro(By.XPATH, self.XPATH_LISTA_PAGAMENTOS, "abrir lista de parcelamento", timeout=self.TIMEOUT_PADRAO)
        time.sleep(0.8)
        self.clicar_seguro(By.XPATH, self.XPATH_PARCELAMENTO_1X, "parcelamento 1x", timeout=self.TIMEOUT_PADRAO)
        time.sleep(0.5)

    def clicar_efetuar_pagamento_cartao(self):
        self.adicionar_log("Clicando em Efetuar pagamento no cartão...")
        self.clicar_seguro(By.XPATH, self.XPATH_BOTAO_EFETUAR_PAGAMENTO, "Efetuar pagamento cartão", timeout=self.TIMEOUT_PADRAO)

    def popup_erro_apareceu(self, timeout=None):
        if timeout is None:
            timeout = self.TIMEOUT_CURTO
        try:
            self.esperar_visivel(By.XPATH, self.XPATH_POPUP_ERRO_FECHAR, timeout=timeout)
            return True
        except Exception:
            return False

    def fechar_popup_erro(self):
        self.adicionar_log("Erro de cobrança detectado. Clicando em Fechar...")
        self.clicar_seguro(By.XPATH, self.XPATH_POPUP_ERRO_FECHAR, "Fechar popup de erro", timeout=self.TIMEOUT_PADRAO)
        time.sleep(1)

    def houve_atualizacao_historico(self, historico_antes, timeout=15):
        fim = time.time() + timeout
        while time.time() < fim:
            historico_depois = self.obter_texto_historico_pagamento()
            if historico_depois and historico_depois != historico_antes:
                self.adicionar_log("Histórico do pagamento foi atualizado.")
                return True
            time.sleep(1)
        return False

    def data_hoje_pagamento(self):
        return datetime.now().strftime("%d/%m/%Y")

    def extrair_final_cartao(self, texto_cartao):
        texto = str(texto_cartao or "")
        match = re.search(r"\*+\s*(\d{4})", texto)
        if match:
            return match.group(1)
        digitos = re.findall(r"\d", texto)
        if len(digitos) >= 4:
            return "".join(digitos[-4:])
        return ""

    def texto_para_valor_historico(self, texto):
        texto = str(texto or "").replace("\xa0", " ")
        match = re.search(r"(\d{1,3}(?:\.\d{3})*,\d{2}|\d+,\d{2}|\d+\.\d{2})", texto)
        if not match:
            return None
        return self.converter_valor_monetario(match.group(1))

    def obter_linhas_historico_pagamento(self):
        try:
            historico = self.esperar_elemento(By.XPATH, self.XPATH_HISTORICO_PAGAMENTO, timeout=self.TIMEOUT_PADRAO)
            linhas = historico.find_elements(By.XPATH, ".//table/tbody/tr")
        except Exception:
            return []

        registros = []
        for linha in linhas:
            try:
                colunas = linha.find_elements(By.XPATH, "./td")
                if len(colunas) < 3:
                    continue
                registros.append({
                    "data": colunas[0].text.strip(),
                    "descricao": colunas[1].text.strip(),
                    "valor": self.texto_para_valor_historico(colunas[2].text.strip())
                })
            except Exception:
                continue
        return registros

    def historico_tem_cobranca_cartao(self, valor_pagamento, final_cartao, data_pagamento=None):
        if data_pagamento is None:
            data_pagamento = self.data_hoje_pagamento()
        valor_esperado = self.converter_valor_monetario(valor_pagamento)
        final_cartao = str(final_cartao or "").strip()
        if valor_esperado is None or not final_cartao:
            return False

        for registro in self.obter_linhas_historico_pagamento():
            mesma_data = registro["data"] == data_pagamento
            mesmo_valor = self.valores_iguais(registro["valor"], valor_esperado)
            mesmo_cartao = final_cartao in registro["descricao"]
            descricao_cartao = "*" in registro["descricao"] or "TEF" in registro["descricao"].upper()
            if mesma_data and mesmo_valor and mesmo_cartao and descricao_cartao:
                self.adicionar_log(
                    f"Cobrança confirmada no histórico: data {registro['data']} | "
                    f"cartão final {final_cartao} | valor {self.formatar_valor_pagamento(valor_esperado)}"
                )
                return True
        return False

    def concluir_fluxo_pos_pagamento_cartao(self):
        self.adicionar_log("Clicando em Concluir...")
        self.clicar_seguro(By.XPATH, self.XPATH_BOTAO_CONCLUIR, "Concluir", timeout=self.TIMEOUT_PADRAO)
        time.sleep(1)
        self.adicionar_log("Clicando em Atualizar contrato...")
        self.clicar_seguro(By.XPATH, self.XPATH_BOTAO_ATUALIZAR_CONTRATO, "Atualizar contrato", timeout=self.TIMEOUT_PADRAO)
        time.sleep(1)
        self.adicionar_log("Clicando em Fechar...")
        self.clicar_seguro(By.XPATH, self.XPATH_BOTAO_FECHAR_FINAL, "Fechar final", timeout=self.TIMEOUT_PADRAO)
        time.sleep(2)
        self.cobrancas_concluidas += 1
        self.adicionar_log(f"Cobrança concluída com sucesso. Total concluídas: {self.cobrancas_concluidas}")

    def validar_resultado_pagamento_cartao(self, valor_pagamento, final_cartao):
        if self.popup_erro_apareceu(timeout=self.TIMEOUT_CURTO):
            self.fechar_popup_erro()
            return "ERRO"

        fim = time.time() + 20
        while time.time() < fim:
            self.verificar_controle_execucao()
            if self.historico_tem_cobranca_cartao(valor_pagamento, final_cartao):
                return "SUCESSO"
            if self.popup_erro_apareceu(timeout=1):
                self.fechar_popup_erro()
                return "ERRO"
            time.sleep(1)
        return "INDEFINIDO"

    def tentar_cobranca_em_todos_os_cartoes(self, valor_pagamento):
        cartoes = self.listar_cartoes_disponiveis()
        for i in range(len(cartoes)):
            try:
                self.verificar_controle_execucao()
                self.adicionar_log(f"Tentando cobrança no cartão {i + 1}/{len(cartoes)}...")
                texto_cartao = self.selecionar_cartao_por_indice(i)
                final_cartao = self.extrair_final_cartao(texto_cartao)
                if not final_cartao:
                    raise RevisaoManualObrigatoria(f"Não foi possível identificar os 4 últimos dígitos do {texto_cartao}.")
                time.sleep(1)

                if self.historico_tem_cobranca_cartao(valor_pagamento, final_cartao):
                    self.adicionar_log(
                        f"Já existe cobrança de hoje no histórico para o cartão final {final_cartao}. "
                        "Contrato será considerado cobrado para evitar duplicidade."
                    )
                    self.cobrancas_concluidas += 1
                    return True

                if i == 0:
                    self.preencher_valor_pagamento_cartao(valor_pagamento)
                    time.sleep(0.5)
                    self.selecionar_parcela_1x()
                    time.sleep(0.5)
                else:
                    self.adicionar_log("Valor e parcelamento já preenchidos anteriormente. Selecionando apenas o próximo cartão.")
                self.clicar_efetuar_pagamento_cartao()
                resultado = self.validar_resultado_pagamento_cartao(valor_pagamento, final_cartao)
                if resultado == "SUCESSO":
                    self.adicionar_log(f"Pagamento aprovado no {texto_cartao}.")
                    self.concluir_fluxo_pos_pagamento_cartao()
                    return True
                if resultado == "ERRO":
                    self.adicionar_log(f"Pagamento recusado no {texto_cartao}. Tentando próximo cartão...")
                    continue
                raise RevisaoManualObrigatoria(
                    f"Resultado indefinido no {texto_cartao}. Verifique manualmente antes de tentar outro cartão."
                )
            except RevisaoManualObrigatoria:
                raise
            except Exception as e:
                self.adicionar_log(f"Falha no cartão {i + 1}: {str(e)}")
        self.adicionar_log("Nenhum cartão conseguiu realizar a cobrança.")
        return False

    # =========================
    # FLUXO DE LINK
    # =========================
    def ir_para_dados_cliente(self):
        self.adicionar_log("Indo para a aba Dados do cliente...")
        self.clicar_seguro(By.XPATH, self.XPATH_ABA_DADOS_CLIENTE, "aba Dados do cliente", timeout=self.TIMEOUT_PADRAO)
        self.esperar_visivel(By.XPATH, self.XPATH_INPUT_EMAIL_CLIENTE, timeout=self.TIMEOUT_PADRAO)
        return True

    def capturar_email_cliente(self):
        self.adicionar_log("Capturando e-mail do cliente...")
        email = self.obter_valor_input(By.XPATH, self.XPATH_INPUT_EMAIL_CLIENTE, "E-mail do cliente", timeout=self.TIMEOUT_PADRAO)
        return email

    def voltar_para_pagamentos_pela_aba(self):
        self.adicionar_log("Voltando para a aba Pagamentos...")
        self.clicar_seguro(By.XPATH, self.XPATH_ABA_PAGAMENTOS_RAPIDA, "aba Pagamentos", timeout=self.TIMEOUT_LONGO)
        self.validar_aba_pagamentos_com_retry()
        return True

    def clicar_link(self):
        self.adicionar_log("Clicando em Link...")
        self.clicar_seguro(By.XPATH, self.XPATH_BOTAO_LINK, "botão Link", timeout=self.TIMEOUT_PADRAO)
        return True

    def preencher_valor_link(self, valor):
        valor_formatado = self.formatar_valor_pagamento(valor)
        self.adicionar_log(f"Preenchendo valor do link: {valor_formatado}")
        self.preencher_input_valor(
            self.XPATH_CAMPO_VALOR_LINK,
            valor_formatado,
            "valor do link"
        )

    def selecionar_modalidade_a_vencer(self):
        self.adicionar_log("Selecionando modalidade À Vencer...")
        self.clicar_seguro(By.XPATH, self.XPATH_MODALIDADE_A_VENCER, "modalidade À Vencer", timeout=self.TIMEOUT_PADRAO)

    def clicar_efetuar_pagamento_link(self):
        self.adicionar_log("Clicando em Efetuar pagamento do link...")
        self.clicar_seguro(By.XPATH, self.XPATH_BOTAO_EFETUAR_PAGAMENTO_LINK, "Efetuar pagamento link", timeout=self.TIMEOUT_PADRAO)

    def copiar_link_gerado(self, tentativas=3):
        for tentativa in range(1, tentativas + 1):
            self.adicionar_log(f"Tentando copiar link... {tentativa}/{tentativas}")
            antes = self.tentar_ler_clipboard()
            self.clicar_seguro(By.XPATH, self.XPATH_BOTAO_COPIAR_LINK, "Copiar link", timeout=self.TIMEOUT_PADRAO)
            time.sleep(1)
            depois = self.tentar_ler_clipboard()
            if depois and depois != antes and depois.lower().startswith("http"):
                self.adicionar_log(f"Link válido capturado: {depois}")
                return depois.strip()
            self.adicionar_log("Clipboard não mudou ou link inválido.")
        raise Exception("Não foi possível capturar o link corretamente.")

    def gerar_link_de_pagamento(self, valor_pagamento):
        self.clicar_link()
        time.sleep(0.5)
        self.preencher_valor_link(valor_pagamento)
        time.sleep(0.5)
        self.selecionar_modalidade_a_vencer()
        time.sleep(0.5)
        self.clicar_efetuar_pagamento_link()
        time.sleep(1)
        link = self.copiar_link_gerado()
        self.links_gerados += 1
        self.adicionar_log(f"Link de pagamento gerado com sucesso. Total de links gerados: {self.links_gerados}")
        return link

    def executar_fluxo_link(self, valor_pagamento):
        self.ir_para_dados_cliente()
        email = self.capturar_email_cliente()
        self.voltar_para_pagamentos_pela_aba()
        link = self.gerar_link_de_pagamento(valor_pagamento)
        return {
            "email": email,
            "link": link
        }

    # =========================
    # OUTLOOK / E-MAIL
    # =========================
    def montar_corpo_email_link(self, nome_cliente, valor_pagamento, link_pagamento, descricao_cobranca):
        nome_cliente = (nome_cliente or "").strip()
        valor_formatado = self.formatar_valor_pagamento(valor_pagamento)
        link_pagamento = (link_pagamento or "").strip()
        descricao_cobranca = (descricao_cobranca or "MENSALIDADE").strip().upper()

        corpo = f"""{nome_cliente}
 
Sou atendente do setor de Cobrança de Locações da FOCO Aluguel de Carros e estarei conduzindo o seu atendimento a partir deste momento.

Segue informações sobre o pagamento da sua {descricao_cobranca}. Tentamos realizar a cobrança no seu cartão cadastrado e não foi autorizado. Abaixo segue link para pagamento:

💵  Valor: R$ {valor_formatado}

Link de pagamento : {link_pagamento}
 
PRAZO DE 48H.

-> * O não pagamento acarretará em inclusão de juros e/ou busca/apreensão

Formas de pagamento disponíveis: 

💳 Cartão de crédito em até 6x sem juros, débito ou PIX

Agradecemos sua atenção e compreensão.

Atenciosamente,

FOCO ALUGUEL DE CARROS
Checkout - Foco Aluguel de Carros"""
        return corpo

    def montar_mensagem_whatsapp_link(self, nome_cliente, valor_pagamento, link_pagamento, descricao_cobranca):
        return self.montar_corpo_email_link(
            nome_cliente=nome_cliente,
            valor_pagamento=valor_pagamento,
            link_pagamento=link_pagamento,
            descricao_cobranca=descricao_cobranca
        )

    def whatsapp_esta_autenticado(self):
        if self.whatsapp_driver is None:
            return False
        seletores = [
            (By.ID, "pane-side"),
            (By.ID, "side"),
            (By.XPATH, "//div[@role='grid']"),
            (By.XPATH, "//div[@contenteditable='true' and @data-tab]")
        ]
        for by, locator in seletores:
            if self.whatsapp_driver.find_elements(by, locator):
                return True
        return False

    def aguardar_autenticacao_whatsapp(self, timeout=300):
        self.criar_driver_whatsapp_se_necessario()
        self.whatsapp_driver.get(self.URL_WHATSAPP_WEB)
        self.adicionar_log_whatsapp("WhatsApp Web aberto. Aguarde a autenticação da sessão.")
        limite = time.time() + timeout
        ultimo_log = 0
        while time.time() < limite:
            self.verificar_controle_execucao_whatsapp()
            if self.whatsapp_esta_autenticado():
                self.adicionar_log_whatsapp("Sessão do WhatsApp autenticada com sucesso.")
                return
            if time.time() - ultimo_log >= 15:
                restante = int(limite - time.time())
                self.adicionar_log_whatsapp(f"Aguardando autenticação do WhatsApp Web... tempo restante: {restante}s")
                ultimo_log = time.time()
            time.sleep(2)
        raise Exception("Não foi possível autenticar o WhatsApp Web dentro do tempo limite.")

    def aguardar_campo_mensagem_whatsapp(self, timeout=45):
        return WebDriverWait(self.whatsapp_driver, timeout).until(
            EC.presence_of_element_located((By.XPATH, "//footer//div[@contenteditable='true']"))
        )

    def validar_erro_whatsapp(self):
        textos_erro = [
            "número de telefone compartilhado por url é inválido",
            "phone number shared via url is invalid",
            "não foi possível encontrar o número de telefone"
        ]
        texto_pagina = self.whatsapp_driver.page_source.lower()
        return any(texto in texto_pagina for texto in textos_erro)

    def enviar_mensagem_whatsapp(self, telefone, mensagem):
        url_envio = f"{self.URL_WHATSAPP_WEB}send?phone={telefone}&text={quote(mensagem)}"
        self.whatsapp_driver.get(url_envio)
        time.sleep(3)
        if self.validar_erro_whatsapp():
            raise Exception("Número inválido ou não encontrado no WhatsApp.")
        campo = self.aguardar_campo_mensagem_whatsapp()
        campo.send_keys(Keys.ENTER)
        time.sleep(2)

    def enviar_email_outlook(self, destinatario, assunto, corpo, cc=None, bcc=None):
        destinatario = (destinatario or "").strip()
        assunto = (assunto or "").strip()
        corpo = (corpo or "").strip()
        if not destinatario:
            raise Exception("Destinatário do e-mail está vazio.")
        if not re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", destinatario):
            raise Exception(f"E-mail do cliente inválido: {destinatario}")
        if cc:
            cc = cc.strip()
            if not re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", cc):
                raise Exception(f"E-mail em cópia inválido: {cc}")
        self.adicionar_log(f"Preparando envio de e-mail para: {destinatario}")
        outlook = win32.Dispatch("Outlook.Application")
        mail = outlook.CreateItem(0)
        conta_envio = self.obter_conta_outlook(outlook, self.EMAIL_REMETENTE_CORPORATIVO)
        self.forcar_conta_envio_outlook(mail, conta_envio)
        mail.To = destinatario
        mail.Subject = assunto
        mail.Body = corpo
        if cc:
            mail.CC = cc
        if bcc:
            mail.BCC = bcc
        mail.Save()
        mail.Send()
        self.adicionar_log(f"E-mail enviado com sucesso para: {destinatario} pela conta {self.EMAIL_REMETENTE_CORPORATIVO}")
        return True

    def forcar_conta_envio_outlook(self, mail, conta_envio):
        mail.SendUsingAccount = conta_envio
        mail._oleobj_.Invoke(*(64209, 0, 8, 0, conta_envio))
        try:
            smtp = str(conta_envio.SmtpAddress or "").strip()
        except Exception:
            smtp = self.EMAIL_REMETENTE_CORPORATIVO
        self.adicionar_log(f"Remetente forçado no Outlook: {smtp}")

    def obter_conta_outlook(self, outlook, email_remetente):
        email_remetente = (email_remetente or "").strip().lower()
        namespace = outlook.GetNamespace("MAPI")
        contas_disponiveis = []

        for conta in namespace.Accounts:
            smtp = ""
            try:
                smtp = str(conta.SmtpAddress or "").strip()
            except Exception:
                smtp = ""
            nome = ""
            try:
                nome = str(conta.DisplayName or "").strip()
            except Exception:
                nome = ""

            if smtp:
                contas_disponiveis.append(smtp)
            elif nome:
                contas_disponiveis.append(nome)

            if smtp.lower() == email_remetente:
                self.adicionar_log(f"Conta de envio selecionada no Outlook: {smtp}")
                return conta

        raise Exception(
            f"A conta corporativa {email_remetente} não foi encontrada no Outlook. "
            f"Contas disponíveis: {', '.join(contas_disponiveis) or 'nenhuma'}"
        )

    def enviar_email_link_pagamento(self, email_cliente, nome_cliente, valor_pagamento, link_pagamento):
        if not email_cliente:
            raise Exception("Cliente sem e-mail cadastrado.")
        if not link_pagamento:
            raise Exception("Link de pagamento não foi capturado.")
        assunto = self.ASSUNTO_EMAIL_LINK
        descricao_cobranca = self.obter_descricao_modalidade_cobranca()
        corpo = self.montar_corpo_email_link(
            nome_cliente=nome_cliente,
            valor_pagamento=valor_pagamento,
            link_pagamento=link_pagamento,
            descricao_cobranca=descricao_cobranca
        )
        return self.enviar_email_outlook(
            destinatario=email_cliente,
            assunto=assunto,
            corpo=corpo,
            cc=self.entry_email_copia.get().strip()
        )

    # =========================
    # PROCESSAMENTO
    # =========================
    def processar_contrato_com_cobranca(self, numero_contrato, valor_pagamento):
        self.verificar_controle_execucao()
        self.ir_para_contratos()
        self.verificar_controle_execucao()
        self.buscar_contrato_seguro(numero_contrato, tentativas=self.MAX_TENTATIVAS_BUSCA)
        self.validar_status_aguardando_devolucao()
        self.abrir_edicao_contrato()
        self.ir_para_pagamentos()
        self.clicar_carteira()

        sucesso_cartao = self.tentar_cobranca_em_todos_os_cartoes(valor_pagamento)

        if sucesso_cartao:
            return {"tipo": "cartao"}

        self.adicionar_log("Iniciando fluxo alternativo por link de pagamento...")
        dados_link = self.executar_fluxo_link(valor_pagamento)
        return {
            "tipo": "link",
            "email": dados_link.get("email", ""),
            "link": dados_link.get("link", "")
        }

    def processar_item_com_isolamento(self, numero_contrato, valor_pagamento):
        for tentativa in range(1, self.MAX_TENTATIVAS_ITEM + 1):
            try:
                self.verificar_controle_execucao()
                self.adicionar_log(f"Processando contrato {numero_contrato} - tentativa {tentativa}/{self.MAX_TENTATIVAS_ITEM}")
                if tentativa > 1:
                    self.reset_completo_navegador(f"Nova tentativa do contrato {numero_contrato}")
                    self.abrir_sistema(relogin=True)
                resultado = self.processar_contrato_com_cobranca(numero_contrato, valor_pagamento)
                resultado["tentativas"] = tentativa
                return resultado
            except RevisaoManualObrigatoria:
                raise
            except ExecucaoInterrompida:
                raise
            except Exception as e:
                self.adicionar_log(f"Falha no contrato {numero_contrato} na tentativa {tentativa}: {e}")
                if tentativa < self.MAX_TENTATIVAS_ITEM:
                    self.adicionar_log("Como o Coral não recupera bem depois de falhar nessa etapa, o navegador será reiniciado antes da próxima tentativa.")
                    continue
                raise

    # =========================
    # EXECUÇÃO
    # =========================
    def iniciar_robo(self):
        if not self.validar_campos():
            return
        self.pausado = False
        self.parar_solicitado = False
        self.contrato_em_andamento = ""
        self.btn_iniciar.config(state="disabled")
        self.btn_pausar.config(state="normal", text="Pausar")
        self.btn_parar.config(state="normal")
        self.adicionar_log("Iniciando robô...")
        thread = threading.Thread(target=self.executar_robo, daemon=True)
        thread.start()

    def executar_robo(self):
        try:
            pythoncom.CoInitialize()

            # validação remota e ping logo na abertura
            self.registrar_abertura()
            if not self.verificar_chave():
                self.adicionar_log("Robô bloqueado por validação remota.")
                messagebox.showerror("Bloqueado", "Este robô está temporariamente desativado.")
                return

            self.cobrancas_concluidas = 0
            self.links_gerados = 0

            usuario = self.entry_usuario.get().strip()
            self.atualizar_progresso(0, self.total_aptos, "Abrindo sistema")
            self.adicionar_log(f"Usuário informado: {usuario}")
            self.adicionar_log("Modo invisível ativado." if self.var_headless.get() else "Modo visível ativado.")

            self.abrir_sistema(relogin=False)
            self.criar_relatorio()

            for i, linha in enumerate(self.df_aptos, start=1):
                self.verificar_controle_execucao()
                contrato = str(linha.get("Contratos", "")).strip()
                self.contrato_em_andamento = contrato
                nome = str(linha.get("Nome", "")).strip()
                valor_cobrar = linha.get("R$ a Cobrar", "")
                mensalidade = linha.get("Mensalidade", "")
                vencimento = str(linha.get("Vencimento", "")).strip()

                valor_cobrar_log = self.formatar_valor_pagamento(valor_cobrar)
                mensalidade_log = self.formatar_valor_pagamento(mensalidade)

                self.atualizar_progresso(i, self.total_aptos, f"Processando contrato {contrato}")
                self.adicionar_log(
                    f"[{i}/{self.total_aptos}] Contrato: {contrato} | "
                    f"Cliente: {nome} | "
                    f"Mensalidade: {mensalidade_log} | "
                    f"R$ a Cobrar: {valor_cobrar_log} | "
                    f"Vencimento: {vencimento}"
                )

                tipo = ""
                status_cobranca = ""
                link = ""
                email_cliente = ""
                status_email = "Não Aplicável"
                status_final = ""
                erro_msg = ""
                tentativas_item = 0
                resetar_navegador_apos_relatorio = False

                try:
                    resultado = self.processar_item_com_isolamento(contrato, valor_cobrar)
                    tentativas_item = resultado.get("tentativas", 1)

                    if resultado.get("tipo") == "cartao":
                        tipo = "Cartão"
                        status_cobranca = "Sucesso"
                        status_final = "Cobrado com Cartão"
                        self.adicionar_log(f"Contrato {contrato} cobrado com sucesso no cartão.")
                    else:
                        tipo = "Link"
                        resetar_navegador_apos_relatorio = True
                        status_cobranca = "Sucesso (Link gerado)"
                        email_cliente = resultado.get("email", "").strip()
                        link = resultado.get("link", "").strip()

                        self.adicionar_log(
                            f"Contrato {contrato} gerou link de pagamento. "
                            f"E-mail capturado: {email_cliente} | "
                            f"Link: {link}"
                        )

                        try:
                            self.enviar_email_link_pagamento(
                                email_cliente=email_cliente,
                                nome_cliente=nome,
                                valor_pagamento=valor_cobrar,
                                link_pagamento=link
                            )
                            status_email = "Enviado com Sucesso"
                            status_final = "Link Gerado e Enviado"
                            self.cobrancas_concluidas += 1
                            self.adicionar_log(f"Link enviado por e-mail com sucesso para o contrato {contrato}.")
                        except Exception as erro_email:
                            status_email = "Erro no Envio"
                            status_final = "Link Gerado (E-mail falhou)"
                            erro_msg = f"Erro email: {str(erro_email)}"
                            self.adicionar_log(f"Falha ao enviar e-mail do contrato {contrato}: {str(erro_email)}")

                except RevisaoManualObrigatoria as e:
                    tentativas_item = max(tentativas_item, 1)
                    tipo = "Revisão Manual"
                    status_cobranca = "Indefinido"
                    status_final = "Revisar Manualmente"
                    erro_msg = str(e)
                    self.adicionar_log(f"REVISÃO MANUAL no contrato {contrato}: {str(e)}")
                    self.reset_completo_navegador(f"Revisão manual no contrato {contrato}")
                    self.abrir_sistema(relogin=True)

                except ExecucaoInterrompida as e:
                    tentativas_item = max(tentativas_item, 1)
                    tipo = "Interrompido"
                    status_cobranca = "Interrompido"
                    status_final = "Execução Interrompida"
                    erro_msg = str(e)
                    self.adicionar_log(str(e))
                    self.parar_solicitado = True

                except Exception as e:
                    tentativas_item = self.MAX_TENTATIVAS_ITEM
                    tipo = "Erro"
                    status_cobranca = "Falha na Cobrança"
                    status_final = "Falha na Cobrança"
                    erro_msg = str(e)
                    self.adicionar_log(f"ERRO definitivo no contrato {contrato}: {str(e)}")
                    self.reset_completo_navegador(f"Erro definitivo no contrato {contrato}")
                    self.abrir_sistema(relogin=True)

                finally:
                    telefone = "" if self.valor_vazio(linha.get("Telefone")) else str(linha.get("Telefone")).strip()
                    descricao_cobranca = self.obter_descricao_modalidade_cobranca()
                    dados_relatorio = {
                        "Contrato": contrato,
                        "Nome": nome,
                        "Telefone": telefone,
                        "Valor a Cobrar": str(valor_cobrar),
                        "Descricao Cobranca": descricao_cobranca,
                        "Tipo": tipo,
                        "Status Cobranca": status_cobranca,
                        "Tentativas": tentativas_item,
                        "Link": link,
                        "Email Cliente": email_cliente,
                        "Status Email": status_email,
                        "Status Final": status_final,
                        "Data/Hora": time.strftime("%d/%m/%Y %H:%M:%S"),
                        "Erro": erro_msg
                    }
                    self.atualizar_relatorio(dados_relatorio)

                    if resetar_navegador_apos_relatorio and not self.parar_solicitado and i < self.total_aptos:
                        self.adicionar_log(
                            "Link gerado deixa o Coral em tela de aguardando pagamento. "
                            "Reiniciando navegador antes do próximo contrato..."
                        )
                        self.reset_completo_navegador(f"Link gerado no contrato {contrato}")
                        self.abrir_sistema(relogin=True)

                if self.parar_solicitado:
                    break

            self.atualizar_progresso(self.total_aptos, self.total_aptos, "Finalizado")
            self.adicionar_log(
                f"Execução finalizada. Cobranças concluídas: {self.cobrancas_concluidas} | "
                f"Links gerados: {self.links_gerados}"
            )
            self.adicionar_log(f"📊 Relatório salvo em: {self.report_path}")

            messagebox.showinfo(
                "Concluído",
                f"Processamento finalizado.\n\n"
                f"Cobranças concluídas: {self.cobrancas_concluidas}\n"
                f"Links gerados: {self.links_gerados}\n\n"
                f"Relatório salvo em:\n{self.report_path}"
            )

        except ExecucaoInterrompida as e:
            self.adicionar_log(str(e))
            messagebox.showinfo("Interrompido", "Execução interrompida pelo usuário.")

        except Exception as e:
            self.adicionar_log(f"ERRO GERAL: {str(e)}")
            messagebox.showerror("Erro", f"Ocorreu um erro:\n\n{str(e)}")

        finally:
            self.fechar_driver()
            self.adicionar_log("Navegador encerrado.")
            self.btn_iniciar.config(state="normal")
            self.btn_pausar.config(state="disabled", text="Pausar")
            self.btn_parar.config(state="disabled")
            self.pausado = False
            self.parar_solicitado = False
            self.contrato_em_andamento = ""
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass

    def validar_campos_whatsapp(self):
        if not self.whatsapp_report_input_path:
            messagebox.showwarning("Atenção", "Selecione o relatório da cobrança.")
            return False
        if self.whatsapp_total_aptos <= 0:
            messagebox.showwarning("Atenção", "Não há contratos aptos para envio no WhatsApp.")
            return False
        return True

    def iniciar_robo_whatsapp(self):
        if not self.validar_campos_whatsapp():
            return
        self.whatsapp_pausado = False
        self.whatsapp_parar_solicitado = False
        self.whatsapp_item_em_andamento = ""
        self.btn_iniciar_whatsapp.config(state="disabled")
        self.btn_pausar_whatsapp.config(state="normal", text="Pausar")
        self.btn_parar_whatsapp.config(state="normal")
        self.adicionar_log_whatsapp("Iniciando automação de WhatsApp...")
        thread = threading.Thread(target=self.executar_robo_whatsapp, daemon=True)
        thread.start()

    def executar_robo_whatsapp(self):
        try:
            self.whatsapp_enviados = 0
            self.atualizar_progresso_whatsapp(0, self.whatsapp_total_aptos, "Abrindo WhatsApp Web")
            self.criar_relatorio_whatsapp()
            self.aguardar_autenticacao_whatsapp()

            for i, linha in enumerate(self.whatsapp_df_aptos, start=1):
                self.verificar_controle_execucao_whatsapp()
                contrato = str(linha.get("Contrato", "")).strip()
                nome = str(linha.get("Nome", "")).strip()
                telefone = str(linha.get("Telefone", "")).strip()
                telefone_envio = str(linha.get("Telefone_whatsapp", "")).strip()
                valor_cobrar = linha.get("Valor a Cobrar", "")
                link = str(linha.get("Link", "")).strip()
                descricao_cobranca = (
                    str(linha.get("Descricao Cobranca", "")).strip().upper() or self.obter_descricao_modalidade_cobranca()
                )

                self.whatsapp_item_em_andamento = contrato
                self.atualizar_progresso_whatsapp(i, self.whatsapp_total_aptos, f"Enviando contrato {contrato}")
                self.adicionar_log_whatsapp(
                    f"[{i}/{self.whatsapp_total_aptos}] Contrato: {contrato} | Cliente: {nome} | Telefone: {telefone_envio}"
                )

                status_whatsapp = ""
                status_final = ""
                erro_msg = ""

                try:
                    mensagem = self.montar_mensagem_whatsapp_link(
                        nome_cliente=nome,
                        valor_pagamento=valor_cobrar,
                        link_pagamento=link,
                        descricao_cobranca=descricao_cobranca
                    )
                    self.enviar_mensagem_whatsapp(telefone_envio, mensagem)
                    self.whatsapp_enviados += 1
                    status_whatsapp = "Enviado com Sucesso"
                    status_final = "Mensagem Enviada"
                    self.adicionar_log_whatsapp(f"Mensagem enviada com sucesso para o contrato {contrato}.")
                except ExecucaoInterrompida:
                    raise
                except Exception as e:
                    status_whatsapp = "Erro no Envio"
                    status_final = "Falha no Envio"
                    erro_msg = str(e)
                    self.adicionar_log_whatsapp(f"Falha ao enviar mensagem do contrato {contrato}: {str(e)}")
                finally:
                    self.atualizar_relatorio_whatsapp({
                        "Contrato": contrato,
                        "Nome": nome,
                        "Telefone": telefone,
                        "Valor a Cobrar": str(valor_cobrar),
                        "Descricao Cobranca": descricao_cobranca,
                        "Link": link,
                        "Status WhatsApp": status_whatsapp,
                        "Status Final": status_final,
                        "Data/Hora": time.strftime("%d/%m/%Y %H:%M:%S"),
                        "Erro": erro_msg
                    })

                if self.whatsapp_parar_solicitado:
                    break

            self.atualizar_progresso_whatsapp(self.whatsapp_total_aptos, self.whatsapp_total_aptos, "Finalizado")
            self.adicionar_log_whatsapp(
                f"Envio finalizado. Mensagens enviadas: {self.whatsapp_enviados} | "
                f"Relatório salvo em: {self.whatsapp_report_output_path}"
            )
            messagebox.showinfo(
                "Concluído",
                f"Envio no WhatsApp finalizado.\n\n"
                f"Mensagens enviadas: {self.whatsapp_enviados}\n\n"
                f"Relatório salvo em:\n{self.whatsapp_report_output_path}"
            )
        except ExecucaoInterrompida as e:
            self.adicionar_log_whatsapp(str(e))
            messagebox.showinfo("Interrompido", "Envio no WhatsApp interrompido pelo usuário.")
        except Exception as e:
            self.adicionar_log_whatsapp(f"ERRO GERAL WHATSAPP: {str(e)}")
            messagebox.showerror("Erro", f"Ocorreu um erro no WhatsApp:\n\n{str(e)}")
        finally:
            self.fechar_driver_whatsapp()
            self.adicionar_log_whatsapp("Navegador do WhatsApp encerrado.")
            self.btn_iniciar_whatsapp.config(state="normal")
            self.btn_pausar_whatsapp.config(state="disabled", text="Pausar")
            self.btn_parar_whatsapp.config(state="disabled")
            self.whatsapp_pausado = False
            self.whatsapp_parar_solicitado = False
            self.whatsapp_item_em_andamento = ""


if __name__ == "__main__":
    root = tk.Tk()
    app = RoboCobrancaMensalApp(root)
    root.mainloop()
