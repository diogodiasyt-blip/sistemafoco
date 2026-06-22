from __future__ import annotations

import json
import hashlib
import os
import re
import shutil
import threading
import time
import unicodedata
import html
from dataclasses import dataclass, replace
from datetime import date
from datetime import datetime
from datetime import timedelta
from pathlib import Path
from tkinter import filedialog, messagebox
from urllib.parse import quote, unquote, urlparse
from urllib.request import Request, urlopen

import customtkinter as ctk
from openpyxl import load_workbook
from PIL import Image
from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager


APP_TITLE = "Robo de Cobranca de Pedagios"
APP_GEOMETRY = "1240x820"

MAIN_BG = "#f7f7f8"
CARD_BG = "#ffffff"
CARD_BORDER = "#e5e7eb"
PRIMARY_TEXT = "#e30613"
TITLE_TEXT = "#1f2937"
MUTED_TEXT = "#667085"
BUTTON_BG = "#e30613"
BUTTON_ACTIVE_BG = "#bd0010"
SOFT_RED = "#fff4f3"
SUCCESS_GREEN = "#0f8a4b"
WARNING_ORANGE = "#b96a10"
PANEL_BG = "#fafafa"
FIELD_BORDER = "#d0d5dd"
NAVY_TEXT = "#111827"

DEFAULT_EXCEL_URL = (
    "https://focoaluguel.sharepoint.com/Documentos Partilhados/Financeiro/Contas a Receber/"
    "Cobrança Pedágios/base_cobranca_pedagios_operacional.xlsx"
)
DEFAULT_OUTLOOK_ACCOUNT = os.environ.get("FOCO_OUTLOOK_ACCOUNT", "cobranca@aluguefoco.com.br")
DEFAULT_CANAL_COBRANCA = "COBRANCA AO CLIENTE"
CANAL_TODOS = "Todos os canais"
SHEET_CLIENTES = "CLIENTES"
SHEET_CONTRATOS = "CONTRATOS"
SHEET_HISTORICO = "HISTORICO"
SHEET_CONFIG = "CONFIG"
APP_DATA_DIR = Path(os.environ.get("LOCALAPPDATA") or Path.home()) / "SistemaFOCO" / "cobranca_pedagios"
UI_CONFIG_PATH = APP_DATA_DIR / "interface_config.json"
CORAL_CREDENTIAL_TARGET = "SistemaFOCO/RoboCobrancaPedagios/Coral"
PENDING_CONTROL_DIR = "_controle_robo_pedagios"
PENDING_D0_FILENAME = "pendencias_d0.jsonl"
PENDING_D0_SYNCED_FILENAME = "pendencias_d0_sincronizadas.jsonl"
PROCESSING_EVENTS_PREFIX = "registros_processamento"
PROCESSING_EVENTS_SYNCED_PREFIX = "registros_processamento_sincronizados"
PROCESSING_SYNC_BATCH_SIZE = 50
CORAL_OPERATION_MAX_ATTEMPTS = 3
DEFAULT_CORAL_HEADLESS = True
PEDAGIOS_REPORT_APP_URL = "https://consulta-pedagios.netlify.app/"
RELATORIOS_PDF_DIR = APP_DATA_DIR / "relatorios_pdf"
OUTLOOK_FOLDER_OUTBOX = 4
OUTLOOK_SEND_CONFIRM_TIMEOUT_SECONDS = 60
OUTLOOK_SEND_CONFIRM_POLL_SECONDS = 3

REQUIRED_CLIENT_COLUMNS = {"ID_CLIENTE", "CLIENTE", "DOCUMENTO", "EMAIL", "VALOR_TOTAL", "STATUS", "ETAPA", "PROXIMA_ACAO"}
REQUIRED_D0_CLIENT_COLUMNS = REQUIRED_CLIENT_COLUMNS
REQUIRED_CONTRACT_COLUMNS = {"ID_CLIENTE", "CONTRATO", "VALOR_A_COBRAR", "STATUS_CARTAO"}
REQUIRED_D0_CONTRACT_COLUMNS = REQUIRED_CONTRACT_COLUMNS | {"PLACA", "INICIO", "FIM", "QTD_PEDAGIOS"}
MAX_CONTRATOS_EMAIL_D0 = 5
LINK_EXPIRATION_DAYS = 4
WHATSAPP_PEDAGIOS_URL = (
    "https://wa.me/558008810102?text=Olá%2C%20Bem-vindo(a)%20à%20Foco%20Aluguel%20de%20Carros.%20"
    "Por%20favor%2C%20envie%20essa%20mensagem%20e%20após%20isso%2C%20selecione%20a%20opção%20"
    "%22Contratuais%20e%20Locações%22%20no%20menu%20inicial."
)
SIGNATURE_IMAGE_PATH = Path(__file__).resolve().parent.parent / "assets" / "assinatura_email_pedagio_whatsapp.png"
SIGNATURE_CONTENT_ID = "assinatura_pedagios@foco"
ACTION_UPDATE_QUEUE = "Atualizar fila"
ACTION_PROCESS_D0 = "Processar D0"
ACTION_PROCESS_D2 = "Processar D0+2"
ACTION_PROCESS_D45 = "Processar D0+4/D0+5"
ACTION_PROCESS_D7 = "Processar D0+7"
ACTION_PROCESS_ALL = "Processar todos aptos"
ACTION_KEYS = [
    ACTION_UPDATE_QUEUE,
    ACTION_PROCESS_D0,
    ACTION_PROCESS_D2,
    ACTION_PROCESS_D45,
    ACTION_PROCESS_D7,
    ACTION_PROCESS_ALL,
]
ACTION_LABELS = {
    ACTION_UPDATE_QUEUE: "Atualizar fila",
    ACTION_PROCESS_D0: "Etapa 1 - Aviso e link",
    ACTION_PROCESS_D2: "Etapa 2 - Cartao e link residual",
    ACTION_PROCESS_D45: "Etapa 3 - Nova tentativa e cobranca firme",
    ACTION_PROCESS_D7: "Etapa 4 - Aviso de negativacao",
    ACTION_PROCESS_ALL: "Processar todos aptos",
}
ACTION_OPTIONS = [ACTION_LABELS[key] for key in ACTION_KEYS]
ACTION_RULES = {
    ACTION_UPDATE_QUEUE: "Atualiza os indicadores lendo a base operacional.",
    ACTION_PROCESS_D0: "Etapa 1: envia o aviso inicial com link de pagamento. A proxima etapa fica apta apos 4 dias.",
    ACTION_PROCESS_D2: "Etapa 2: tenta os cartoes dos contratos; se sobrar saldo, envia link residual. A proxima etapa fica apta apos 4 dias.",
    ACTION_PROCESS_D45: "Etapa 3: repete tentativa de cartao e link residual com e-mail em tom mais firme.",
    ACTION_PROCESS_D7: "Etapa 4: envia o aviso de negativacao para pendencias nao regularizadas.",
    ACTION_PROCESS_ALL: "Processamento em lote ainda nao implementado.",
}
ACTION_KEY_BY_LABEL = {label: key for key, label in ACTION_LABELS.items()}

URL_CORAL_LOGIN = "https://coral.aluguefoco.com.br/login"
URL_CORAL_DASHBOARD = "https://coral.aluguefoco.com.br/precificacao/dashboard"
URL_CORAL_CONTRATOS = "https://coral.aluguefoco.com.br/contratos"
XPATH_CORAL_LOGIN = "/html/body/foco-app/div[1]/foco-login/div/div/div/div/div[2]/form/div[1]/input"
XPATH_CORAL_SENHA = "/html/body/foco-app/div[1]/foco-login/div/div/div/div/div[2]/form/div[2]/input"
XPATH_CORAL_ENTRAR = "/html/body/foco-app/div[1]/foco-login/div/div/div/div/div[2]/form/button"
XPATH_ABA_CONTRATOS = "/html/body/foco-app/div[1]/foco-rent-agreement-home/div/ngb-tabset/ul/li[3]/a"
XPATH_CAMPO_BUSCA_CONTRATOS = "/html/body/foco-app/div[1]/foco-rent-agreement-home/div/div/div[2]/input"
XPATH_TD_CONTRATO_RESULTADO = "/html/body/foco-app/div[1]/foco-rent-agreement-home/div/ngb-tabset/div/div/div/div/foco-rent-agreement-list/div/div/div[3]/table/tbody/tr/td[2]"
XPATH_MAIS_OPCOES_CONTRATO = "/html/body/foco-app/div[1]/foco-rent-agreement-home/div/ngb-tabset/div/div/div/div/foco-rent-agreement-list/div/div/div[3]/table/tbody/tr/td[8]/div/div/button"
XPATH_EDITAR_CONTRATO = "/html/body/foco-app/div[1]/foco-rent-agreement-home/div/ngb-tabset/div/div/div/div/foco-rent-agreement-list/div/div/div[3]/table/tbody/tr/td[8]/div/div/div/button[1]"
XPATH_ABA_PAGAMENTOS_RAPIDA = "/html/body/foco-app/div[1]/foco-rent-agreement-edit/div/div[1]/div/div/div[2]/div[11]/button"
XPATH_ABA_PAGAMENTOS_RAPIDA_ICONE = "/html/body/foco-app/div[1]/foco-rent-agreement-edit/div/div[1]/div/div/div[2]/div[11]/button/i"
XPATH_BOTAO_FECHAR_MODAL_CARREGAR_CLIENTE = "/html/body/ngb-modal-window/div/div/div[1]/button"
XPATH_BOTAO_SIM_MODAL_CARREGAR_CLIENTE = "(//ngb-modal-window//button[normalize-space()='Sim'])[last()]"
XPATH_BOTAO_AVANCAR_FLUXO_EDICAO = "/html/body/foco-app/div[1]/foco-rent-agreement-edit/div/div[3]/div/div/div[2]/button"
XPATH_BOTAO_AVANCAR_FLUXO_EDICAO_2 = "/html/body/foco-app/div[1]/foco-rent-agreement-edit/div/div[3]/div/div/div[2]/button[3]"
XPATH_BOTAO_AVANCAR_FLUXO_EDICAO_3 = "/html/body/foco-app/div[1]/foco-rent-agreement-edit/div/div[3]/div/div/div[2]/button[2]"
XPATH_BOTOES_RODAPE_EDICAO = "/html/body/foco-app/div[1]/foco-rent-agreement-edit/div/div[3]/div/div/div[2]//button"
XPATH_RESUMO_PAGAMENTO_TITULO = "/html/body/foco-app/div[1]/foco-rent-agreement-edit/div/div[2]/div[6]/foco-rent-agreement-payment/div/div[1]/div/div[1]"
XPATH_POPUP_SIM = "(//ngb-modal-window//foco-confirm-modal//button[normalize-space()='Sim'])[last()]"
XPATH_BOTAO_CARTEIRA = "/html/body/foco-app/div[1]/foco-rent-agreement-edit/div/div[2]/div[6]/foco-rent-agreement-payment/div/div[2]/div/div[2]/div[1]/div[1]/button[6]"
XPATH_BOTAO_LINK = "/html/body/foco-app/div[1]/foco-rent-agreement-edit/div/div[2]/div[6]/foco-rent-agreement-payment/div/div[2]/div/div[2]/div[1]/div[1]/button[5]"
XPATH_CAMPO_VALOR_PAGAMENTO_CARTAO = "/html/body/foco-app/div[1]/foco-rent-agreement-edit/div/div[2]/div[6]/foco-rent-agreement-payment/div/div[2]/div/div[2]/div[11]/div/div[2]/foco-form-input/div/div[1]/input"
XPATH_CAMPO_VALOR_LINK = "/html/body/foco-app/div[1]/foco-rent-agreement-edit/div/div[2]/div[6]/foco-rent-agreement-payment/div/div[2]/div/div[2]/div[8]/div/div[2]/foco-form-input/div/div[1]/input"
XPATH_MODALIDADE_A_VENCER_LINK = "/html/body/foco-app/div[1]/foco-rent-agreement-edit/div/div[2]/div[6]/foco-rent-agreement-payment/div/div[2]/div/div[2]/div[8]/div/div[3]/foco-form-button-group/div/div/label[1]"
XPATH_BOTAO_EFETUAR_PAGAMENTO_LINK = "/html/body/foco-app/div[1]/foco-rent-agreement-edit/div/div[2]/div[6]/foco-rent-agreement-payment/div/div[2]/div/div[2]/div[15]/button"
XPATH_BOTAO_COPIAR_LINK = "/html/body/ngb-modal-window/div/div/foco-pbl-modal/div[2]/div/div[1]/div/button"
XPATH_LISTA_PAGAMENTOS_CARTAO = "/html/body/foco-app/div[1]/foco-rent-agreement-edit/div/div[2]/div[6]/foco-rent-agreement-payment/div/div[2]/div/div[2]/div[11]/div/div[3]/foco-form-dropdown/div/foco-dropdown/div/button/div"
XPATH_PARCELAMENTO_1X_CARTAO = "/html/body/foco-app/div[1]/foco-rent-agreement-edit/div/div[2]/div[6]/foco-rent-agreement-payment/div/div[2]/div/div[2]/div[11]/div/div[3]/foco-form-dropdown/div/foco-dropdown/div/div/div[2]/perfect-scrollbar/div/div[1]/button[1]/div"
XPATH_BOTAO_EFETUAR_PAGAMENTO_CARTAO = "/html/body/foco-app/div[1]/foco-rent-agreement-edit/div/div[2]/div[6]/foco-rent-agreement-payment/div/div[2]/div/div[2]/div[19]/button"
XPATH_POPUP_ERRO_FECHAR = "/html/body/ngb-modal-window/div/div//button[contains(normalize-space(), 'Fechar')]"
XPATH_HISTORICO_PAGAMENTO = "/html/body/foco-app/div[1]/foco-rent-agreement-edit/div/div[2]/div[6]/foco-rent-agreement-payment/div/div[1]/div/div[3]"
XPATH_BOTAO_CONCLUIR_CONTRATO = "/html/body/foco-app/div[1]/foco-rent-agreement-edit/div/div[3]/div/div/div[2]/button[2]"
XPATH_BOTAO_ATUALIZAR_CONTRATO = "/html/body/ngb-modal-window/div/div/foco-confirm-modal/div[3]/button[2]"
XPATH_BOTAO_FECHAR_FINAL = "/html/body/ngb-modal-window/div/div/foco-reservation-created/div[3]/button"
XPATH_CARTEIRA_CARD = "//div[contains(@class,'payment-area')]//div[contains(@class,'cardWallet')]"
XPATH_TOAST_CONTAINER = "//*[@id='toast-container']"
TOASTS_CRITICOS_EDICAO = (
    "CLIENTE NAO PODE SER ADICIONADO COMO CONDUTOR ADICIONAL",
)


@dataclass(frozen=True)
class ResumoCobrancasExcel:
    total: int
    contratos_total: int
    contratos_pendentes: int
    aptos_d0: int
    aptos_d2: int
    aptos_d45: int
    aptos_d7: int
    em_processamento: int
    pagos: int
    erros: int
    valor_total: float
    exemplos_aptos: list[dict[str, object]]


@dataclass(frozen=True)
class EmailD0Pedagio:
    id_cliente: str
    nome: str
    destinatario: str
    assunto: str
    corpo: str
    total_contratos: int
    contratos_listados: int
    total_pedagios: int
    valor_total: float
    contrato_referencia: str
    link_pagamento: str = ""
    documento: str = ""


@dataclass(frozen=True)
class RelatorioPedagiosCliente:
    nome: str
    documento: str
    contratos: list[dict[str, object]]
    passagens: list[dict[str, object]]

    @property
    def total_contratos(self) -> int:
        return len(self.contratos)

    @property
    def total_passagens(self) -> int:
        return len(self.passagens)

    @property
    def valor_total(self) -> float:
        return round(sum(_format_money(item.get("total_valor")) for item in self.contratos), 2)

    @property
    def valor_cobrado(self) -> float:
        return round(sum(_format_money(item.get("total_cobrado")) for item in self.contratos), 2)

    @property
    def valor_pendente(self) -> float:
        return round(sum(_format_money(item.get("gap")) for item in self.contratos), 2)


@dataclass(frozen=True)
class ContratoD2Pedagio:
    id_cliente: str
    cliente: str
    documento: str
    contrato: str
    placa: str
    valor: float


@dataclass(frozen=True)
class ResultadoD2Pedagio:
    id_cliente: str
    contrato: str
    status: str
    cartoes_encontrados: int
    cartoes_tentados: int
    detalhe: str


def _normalizar_texto(value: object) -> str:
    text = unicodedata.normalize("NFKD", str(value or "")).encode("ASCII", "ignore").decode("ASCII")
    return " ".join(text.strip().upper().split())


def _parse_date(value: object) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y, %H:%M"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _is_active(status: object) -> bool:
    return _normalizar_texto(status) in {"ATIVO", "PENDENTE", "AGUARDANDO", "EMAIL_ENVIADO"}


def _is_processing(status: object) -> bool:
    return _normalizar_texto(status) in {"EM PROCESSAMENTO", "PROCESSANDO"}


def _is_paid(status: object) -> bool:
    return _normalizar_texto(status) in {"PAGO", "QUITADO", "COBRADO"}


def _is_error(status: object) -> bool:
    return _normalizar_texto(status) in {"ERRO", "FALHA"}


def _is_due(value: object, hoje: date) -> bool:
    parsed = _parse_date(value)
    return parsed is None or parsed <= hoje


def _format_money(value: object) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = str(value).replace("R$", "").replace(" ", "").strip()
    if "," in cleaned:
        cleaned = cleaned.replace(".", "").replace(",", ".")
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def _format_brl(value: float) -> str:
    text = f"{float(value):,.2f}"
    return text.replace(",", "X").replace(".", ",").replace("X", ".")


def _digits_only(value: object) -> str:
    return re.sub(r"\D+", "", str(value or ""))


def _documentos_equivalentes(left: object, right: object) -> bool:
    left_digits = _digits_only(left)
    right_digits = _digits_only(right)
    if not left_digits or not right_digits:
        return False
    return left_digits == right_digits or left_digits.lstrip("0") == right_digits.lstrip("0")


def _pdf_safe_text(value: object) -> str:
    text = str(value or "").replace("\r", " ").replace("\n", " ")
    text = unicodedata.normalize("NFKD", text).encode("ASCII", "ignore").decode("ASCII")
    return " ".join(text.split())


def _pdf_escape(value: object) -> str:
    text = _pdf_safe_text(value)
    return text.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")


def _slug_filename(value: object, fallback: str = "cliente") -> str:
    text = _normalizar_texto(value).lower()
    text = re.sub(r"[^a-z0-9]+", "_", text).strip("_")
    return text or fallback


def _parse_supabase_public_config(html_text: str) -> tuple[str, str]:
    url_match = re.search(r'SUPA_URL\s*=\s*"([^"]+)"', html_text)
    key_match = re.search(r'SUPA_KEY\s*=\s*"([^"]+)"', html_text)
    if not url_match or not key_match:
        raise RuntimeError("Nao foi possivel localizar configuracao publica do Supabase no site de pedagios.")
    return url_match.group(1).strip(), key_match.group(1).strip()


_SUPABASE_PUBLIC_CONFIG_CACHE: tuple[str, str] | None = None


def _carregar_supabase_public_config() -> tuple[str, str]:
    global _SUPABASE_PUBLIC_CONFIG_CACHE
    env_url = os.environ.get("FOCO_PEDAGIOS_SUPABASE_URL", "").strip()
    env_key = os.environ.get("FOCO_PEDAGIOS_SUPABASE_KEY", "").strip()
    if env_url and env_key:
        return env_url, env_key
    if _SUPABASE_PUBLIC_CONFIG_CACHE is not None:
        return _SUPABASE_PUBLIC_CONFIG_CACHE
    with urlopen(PEDAGIOS_REPORT_APP_URL, timeout=30) as response:
        html_text = response.read().decode("utf-8", errors="replace")
    _SUPABASE_PUBLIC_CONFIG_CACHE = _parse_supabase_public_config(html_text)
    return _SUPABASE_PUBLIC_CONFIG_CACHE


def _supabase_get_public(path: str) -> list[dict[str, object]]:
    base_url, api_key = _carregar_supabase_public_config()
    request = Request(
        f"{base_url.rstrip('/')}/rest/v1/{path}",
        headers={"apikey": api_key, "Authorization": f"Bearer {api_key}"},
    )
    with urlopen(request, timeout=45) as response:
        payload = response.read().decode("utf-8", errors="replace")
    data = json.loads(payload)
    if not isinstance(data, list):
        raise RuntimeError("Resposta inesperada do Supabase para relatorio de pedagios.")
    return data


def baixar_relatorio_pedagios_cliente(
    documento: str,
    nome: str = "",
    *,
    supabase_get=None,
) -> RelatorioPedagiosCliente:
    supabase_get = supabase_get or _supabase_get_public
    doc_digits = _digits_only(documento)
    filtros: list[str] = []
    if len(doc_digits) >= 6:
        filtros.append(f"doc_cliente.eq.{quote(doc_digits, safe='')}")
        doc_sem_zero = doc_digits.lstrip("0")
        if doc_sem_zero and doc_sem_zero != doc_digits and len(doc_sem_zero) >= 6:
            filtros.append(f"doc_cliente.eq.{quote(doc_sem_zero, safe='')}")
    if str(nome or "").strip():
        filtros.append(f"nome_cliente.ilike.*{quote(str(nome).strip(), safe='')}*")
    if not filtros:
        raise ValueError("Informe documento ou nome para buscar relatorio de pedagios.")

    contratos: list[dict[str, object]] = []
    for filtro in filtros:
        contratos = supabase_get(f"pedagios_contratos?or=({filtro})&order=id_contrato.asc&limit=500")
        if contratos:
            break
    if not contratos:
        raise RuntimeError("Nenhum contrato de pedagio encontrado para o cliente.")

    passagens: list[dict[str, object]] = []
    for contrato in contratos:
        id_contrato = str(contrato.get("id_contrato") or "").strip()
        if not id_contrato:
            continue
        itens = supabase_get(
            f"pedagios_itens?id_contrato=eq.{quote(id_contrato, safe='')}&order=data_hora.asc&limit=1000"
        )
        for item in itens:
            passagem = dict(item)
            passagem["contrato"] = id_contrato
            passagens.append(passagem)

    passagens.sort(key=lambda item: str(item.get("data_hora") or ""))
    first = contratos[0]
    return RelatorioPedagiosCliente(
        nome=str(first.get("nome_cliente") or nome or "").strip(),
        documento=_digits_only(first.get("doc_cliente") or documento),
        contratos=contratos,
        passagens=passagens,
    )


def _format_datetime_relatorio(value: object) -> str:
    text = str(value or "").strip()
    if not text:
        return "-"
    for fmt in ("%Y-%m-%dT%H:%M:%S.%fZ", "%Y-%m-%dT%H:%M:%S%z", "%Y-%m-%dT%H:%M:%S"):
        try:
            return datetime.strptime(text.replace("+00:00", "+0000"), fmt).strftime("%d/%m/%Y %H:%M")
        except ValueError:
            continue
    parsed = _parse_date(text[:10])
    if parsed is not None:
        return parsed.strftime("%d/%m/%Y")
    return text[:16]


def _pdf_stream(lines: list[str]) -> bytes:
    return "\n".join(lines).encode("latin-1", errors="replace")


def _pdf_page_stream(title: str, rows: list[list[str]], relatorio: RelatorioPedagiosCliente, first_page: bool) -> bytes:
    lines = ["BT"]
    if first_page:
        lines.extend(
            [
                "0.89 0.02 0.07 rg",
                "0 792 595 50 re f",
                "1 1 1 rg /F1 18 Tf 1 0 0 1 32 820 Tm (Historico de Pedagios) Tj",
                "0 0 0 rg /F2 9 Tf 1 0 0 1 410 820 Tm (Emitido em) Tj",
                f"/F1 10 Tf 1 0 0 1 410 806 Tm ({datetime.now().strftime('%d/%m/%Y')}) Tj",
                "0 0 0 rg /F1 12 Tf 1 0 0 1 32 760 Tm (LOCATARIO / CONTRATANTE) Tj",
                f"/F1 14 Tf 1 0 0 1 32 742 Tm ({_pdf_escape(relatorio.nome)}) Tj",
                f"/F2 10 Tf 1 0 0 1 32 726 Tm (CPF/CNPJ: {_pdf_escape(relatorio.documento)}) Tj",
                f"/F2 10 Tf 1 0 0 1 32 710 Tm ({relatorio.total_contratos} contratos - {relatorio.total_passagens} passagens) Tj",
                f"/F1 10 Tf 1 0 0 1 360 742 Tm (Total pedagios: R$ {_format_brl(relatorio.valor_total)}) Tj",
                f"/F1 10 Tf 1 0 0 1 360 726 Tm (Cobrado: R$ {_format_brl(relatorio.valor_cobrado)}) Tj",
                f"/F1 10 Tf 1 0 0 1 360 710 Tm (Pendente: R$ {_format_brl(relatorio.valor_pendente)}) Tj",
            ]
        )
        y = 675
    else:
        lines.extend(
            [
                "0.89 0.02 0.07 rg",
                "0 812 595 30 re f",
                f"1 1 1 rg /F1 12 Tf 1 0 0 1 32 824 Tm ({_pdf_escape(relatorio.nome)} - passagens) Tj",
            ]
        )
        y = 780

    lines.extend(
        [
            "0 0 0 rg",
            f"/F1 11 Tf 1 0 0 1 32 {y} Tm ({_pdf_escape(title)}) Tj",
            f"/F1 8 Tf 1 0 0 1 32 {y - 18} Tm (CONTRATO) Tj",
            f"/F1 8 Tf 1 0 0 1 120 {y - 18} Tm (PLACA) Tj",
            f"/F1 8 Tf 1 0 0 1 178 {y - 18} Tm (DATA / HORA) Tj",
            f"/F1 8 Tf 1 0 0 1 270 {y - 18} Tm (ENDERECO) Tj",
            f"/F1 8 Tf 1 0 0 1 520 {y - 18} Tm (VALOR) Tj",
        ]
    )
    y -= 34
    for row in rows:
        contrato, placa, data_hora, endereco, valor = row
        if len(endereco) > 42:
            endereco = endereco[:39] + "..."
        lines.extend(
            [
                f"/F2 7 Tf 1 0 0 1 32 {y} Tm ({_pdf_escape(contrato)}) Tj",
                f"/F2 8 Tf 1 0 0 1 120 {y} Tm ({_pdf_escape(placa)}) Tj",
                f"/F2 8 Tf 1 0 0 1 178 {y} Tm ({_pdf_escape(data_hora)}) Tj",
                f"/F2 8 Tf 1 0 0 1 270 {y} Tm ({_pdf_escape(endereco)}) Tj",
                f"/F2 8 Tf 1 0 0 1 520 {y} Tm ({_pdf_escape(valor)}) Tj",
            ]
        )
        y -= 15
    lines.append("ET")
    return _pdf_stream(lines)


def _write_simple_pdf(output_path: Path, page_streams: list[bytes]) -> None:
    objects: list[bytes] = [
        b"<< /Type /Catalog /Pages 2 0 R >>",
        b"<< /Type /Pages /Kids [" + b" ".join(f"{3 + i * 2} 0 R".encode("ascii") for i in range(len(page_streams))) + b"] /Count "
        + str(len(page_streams)).encode("ascii")
        + b" >>",
    ]
    for index, stream in enumerate(page_streams):
        page_obj = 3 + index * 2
        stream_obj = page_obj + 1
        objects.append(
            f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] /Resources << /Font << /F1 << /Type /Font /Subtype /Type1 /BaseFont /Helvetica-Bold >> /F2 << /Type /Font /Subtype /Type1 /BaseFont /Helvetica >> >> >> /Contents {stream_obj} 0 R >>".encode(
                "ascii"
            )
        )
        objects.append(b"<< /Length " + str(len(stream)).encode("ascii") + b" >>\nstream\n" + stream + b"\nendstream")

    output = bytearray(b"%PDF-1.4\n")
    offsets = [0]
    for obj_number, obj in enumerate(objects, start=1):
        offsets.append(len(output))
        output.extend(f"{obj_number} 0 obj\n".encode("ascii"))
        output.extend(obj)
        output.extend(b"\nendobj\n")
    xref_offset = len(output)
    output.extend(f"xref\n0 {len(objects) + 1}\n".encode("ascii"))
    output.extend(b"0000000000 65535 f \n")
    for offset in offsets[1:]:
        output.extend(f"{offset:010d} 00000 n \n".encode("ascii"))
    output.extend(
        f"trailer\n<< /Size {len(objects) + 1} /Root 1 0 R >>\nstartxref\n{xref_offset}\n%%EOF".encode("ascii")
    )
    output_path.write_bytes(bytes(output))


def gerar_pdf_relatorio_pedagios(relatorio: RelatorioPedagiosCliente, output_dir: str | Path = RELATORIOS_PDF_DIR) -> Path:
    output_root = Path(output_dir)
    output_root.mkdir(parents=True, exist_ok=True)
    output_path = output_root / f"Pedagios - {_slug_filename(relatorio.nome)}.pdf"
    rows = [
        [
            str(item.get("contrato") or item.get("id_contrato") or ""),
            str(item.get("placa") or "-"),
            _format_datetime_relatorio(item.get("data_hora")),
            str(item.get("praca") or item.get("endereco") or "-"),
            f"R$ {_format_brl(_format_money(item.get('valor_net') or item.get('valor')))}",
        ]
        for item in relatorio.passagens
    ]
    if not rows:
        rows = [["-", "-", "-", "Sem passagens registradas no Supabase.", "R$ 0,00"]]

    page_streams: list[bytes] = []
    first_page_limit = 32
    next_page_limit = 45
    first_rows = rows[:first_page_limit]
    page_streams.append(_pdf_page_stream(f"HISTORICO DE PASSAGENS ({relatorio.total_passagens})", first_rows, relatorio, True))
    remaining = rows[first_page_limit:]
    while remaining:
        chunk = remaining[:next_page_limit]
        page_streams.append(_pdf_page_stream("HISTORICO DE PASSAGENS", chunk, relatorio, False))
        remaining = remaining[next_page_limit:]
    _write_simple_pdf(output_path, page_streams)
    return output_path


def gerar_pdf_relatorio_pedagios_para_email(email: EmailD0Pedagio, output_dir: str | Path = RELATORIOS_PDF_DIR) -> Path:
    documento = _digits_only(email.documento)
    if not documento:
        raise RuntimeError(f"Documento ausente para gerar relatorio de pedagios de {email.nome}.")
    relatorio = baixar_relatorio_pedagios_cliente(documento)
    return gerar_pdf_relatorio_pedagios(relatorio, output_dir)


def _extrair_cobrancas_cartao_historico(historico: object) -> list[tuple[date, float, str]]:
    text = str(historico or "").replace("\xa0", " ")
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    cobrancas: list[tuple[date, float, str]] = []

    for index, line in enumerate(lines):
        if not re.match(r"^\d{2}/\d{2}/\d{4}\b", line):
            continue
        data_lancamento = _parse_date(line[:10])
        if data_lancamento is None:
            continue

        chunk = [line]
        for next_line in lines[index + 1 :]:
            if re.match(r"^\d{2}/\d{2}/\d{4}\b", next_line):
                break
            chunk.append(next_line)

        chunk_text = " ".join(chunk)
        normalized_chunk = _normalizar_texto(chunk_text)
        if "ORINT" in normalized_chunk or "A FATURAR" in normalized_chunk:
            continue
        if "CARTEIRA" not in normalized_chunk:
            continue
        value_matches = re.findall(r"R\$\s*([\d\.\,]+)", chunk_text)
        if not value_matches:
            continue
        cobrancas.append((data_lancamento, _format_money(value_matches[-1]), _normalizar_texto(chunk_text)))

    return cobrancas


def _historico_indica_nova_cobranca_cartao(
    historico_antes: object,
    historico_depois: object,
    valor: object,
    data_pagamento: date | None = None,
) -> bool:
    expected = _format_money(valor)
    expected_date = data_pagamento or date.today()
    cobrancas_antes = {_normalizar_texto(texto) for _, _, texto in _extrair_cobrancas_cartao_historico(historico_antes)}

    for data_lancamento, valor_cobrado, texto in _extrair_cobrancas_cartao_historico(historico_depois):
        if (
            data_lancamento == expected_date
            and abs(valor_cobrado - expected) < 0.01
            and _normalizar_texto(texto) not in cobrancas_antes
        ):
            return True
    return False


def _format_display_date(value: object) -> str:
    if value in (None, ""):
        return "Nao informado"
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")
    text = str(value).strip()
    parsed = _parse_date(text)
    if parsed is not None:
        return parsed.strftime("%d/%m/%Y")
    return text


def _is_valid_email(value: object) -> bool:
    email = str(value or "").strip()
    return bool(re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", email))


def _canal_filtro_ativo(canal: object) -> bool:
    normalized = _normalizar_texto(canal)
    return normalized not in {"", "TODOS", "TODOS OS CANAIS"}


def _canal_do_registro(record: dict[str, object]) -> object:
    return record.get("CANAL") or record.get("CANAL DA COBRANCA")


def _registro_pertence_ao_canal(record: dict[str, object], canal: object) -> bool:
    if not _canal_filtro_ativo(canal):
        return True
    return _normalizar_texto(_canal_do_registro(record)) == _normalizar_texto(canal)


def montar_url_edicao_contrato_coral(contrato: str) -> str:
    contrato_limpo = str(contrato or "").strip()
    if not contrato_limpo:
        raise ValueError("Contrato vazio para montar URL de edicao do Coral.")
    return f"{URL_CORAL_CONTRATOS}/editar/{quote(contrato_limpo, safe='')}"


def url_coral_corresponde_ao_contrato(url_atual: object, contrato: object) -> bool:
    return montar_url_edicao_contrato_coral(str(contrato or "")).lower() in str(url_atual or "").strip().lower()


def _sheet_records(workbook, sheet_name: str, required_columns: set[str]) -> list[dict[str, object]]:
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Aba obrigatoria ausente: {sheet_name}")
    sheet = workbook[sheet_name]
    rows = sheet.iter_rows(values_only=True)
    try:
        headers = [str(value or "").strip() for value in next(rows)]
    except StopIteration as exc:
        raise ValueError(f"Aba {sheet_name} esta vazia") from exc
    header_map = {_normalizar_texto(header): index for index, header in enumerate(headers)}
    missing = sorted(required_columns - set(header_map))
    if missing:
        raise ValueError(f"Colunas obrigatorias ausentes em {sheet_name}: {', '.join(missing)}")

    records: list[dict[str, object]] = []
    for row in rows:
        if not row or all(value in (None, "") for value in row):
            continue
        record = {}
        for column, index in header_map.items():
            record[column] = row[index] if index < len(row) else None
        records.append(record)
    return records


def carregar_canais_cobranca_excel(path: str | Path) -> list[str]:
    workbook_path = Path(path)
    if not workbook_path.exists():
        raise FileNotFoundError(f"Planilha nao encontrada: {workbook_path}")

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    canais: dict[str, str] = {}
    try:
        contratos = _sheet_records(workbook, SHEET_CONTRATOS, REQUIRED_CONTRACT_COLUMNS)
        for contrato in contratos:
            canal = str(_canal_do_registro(contrato) or "").strip()
            if not canal:
                continue
            canais.setdefault(_normalizar_texto(canal), canal)
    finally:
        workbook.close()

    ordered = sorted(canais.values(), key=_normalizar_texto)
    default_norm = _normalizar_texto(DEFAULT_CANAL_COBRANCA)
    default_value = canais.get(default_norm, DEFAULT_CANAL_COBRANCA)
    values = [default_value, CANAL_TODOS]
    values.extend(canal for canal in ordered if _normalizar_texto(canal) not in {default_norm, _normalizar_texto(CANAL_TODOS)})
    return values


def _montar_email_d0(cliente: dict[str, object], contratos: list[dict[str, object]]) -> EmailD0Pedagio:
    nome = str(cliente.get("CLIENTE") or "").strip()
    id_cliente = str(cliente.get("ID_CLIENTE") or cliente.get("DOCUMENTO") or "").strip()
    documento = _digits_only(cliente.get("DOCUMENTO"))
    destinatario = str(cliente.get("EMAIL") or "").strip()
    contratos_ordenados = sorted(contratos, key=lambda item: str(item.get("CONTRATO") or ""))
    contratos_listados = contratos_ordenados[:MAX_CONTRATOS_EMAIL_D0]
    total_contratos = len(contratos_ordenados)
    total_pedagios = int(sum(_format_money(item.get("QTD_PEDAGIOS")) for item in contratos_ordenados))
    valor_total = round(sum(_format_money(item.get("VALOR_A_COBRAR")) for item in contratos_ordenados), 2)
    contrato_referencia = str(contratos_ordenados[0].get("CONTRATO") or "").strip()
    link_pagamento = str(cliente.get("LINK_D0") or "").strip()
    assunto = "Alerta Pedágio: A FOCO identificou pedágios pendentes em seu contrato de locação"

    blocos: list[str] = []
    for contrato in contratos_listados:
        bloco = "\n".join(
            [
                f"Contrato: {contrato.get('CONTRATO') or 'Não informado'}",
                f"Placa do veículo: {contrato.get('PLACA') or 'Não informado'}",
                f"Loja: {contrato.get('LOJA') or 'Nao informado'}",
                f"Retirada: {_format_display_date(contrato.get('INICIO'))}",
                f"Devolução: {_format_display_date(contrato.get('FIM'))}",
                f"Quantidade de passagens em pedágio: {int(_format_money(contrato.get('QTD_PEDAGIOS')))}",
                f"Valor a regularizar: R$ {_format_brl(_format_money(contrato.get('VALOR_A_COBRAR')))}",
            ]
        )
        blocos.append(bloco)

    restante = total_contratos - len(contratos_listados)
    aviso_restante = ""
    if restante > 0:
        aviso_restante = f"\n\nAlém das locações listadas acima, identificamos mais {restante} contrato(s) vinculado(s) a esta cobrança."

    link_text = link_pagamento or "{LINK_PAGAMENTO}"
    corpo = f"""Olá, {nome}!

Identificamos, após o encerramento do seu contrato, valores referentes à utilização de pedágios durante sua locação que nos foram notificadas pelas operadoras de tag. Dessa forma, estamos realizando a regularização dos valores abaixo relacionados.

Dados da Locação e Utilização do pedágio:

{(chr(10) + chr(10)).join(blocos)}{aviso_restante}

Quantidade total de passagens em pedágio: {total_pedagios}
Valor total a regularizar: R$ {_format_brl(valor_total)}

Para sua comodidade, o link para regularização do pagamento já está disponível neste e-mail. Basta acessá-lo para concluir o pagamento de forma rápida e segura.

Link para pagamento: {link_text}

O link permanecerá ativo por 4 (quatro) dias corridos a partir do recebimento desta comunicação. Após esse período, caso não seja identificado o pagamento ou não haja manifestação para tratativa do débito, a pendência poderá ser encaminhada para os procedimentos de cobrança previstos em contrato, incluindo eventual registro junto aos órgãos de proteção ao crédito, observadas as notificações e exigências legais aplicáveis.

Caso tenha alguma dúvida, entre em contato conosco respondendo esse e-mail ou pelos nossos canais de atendimento, nosso objetivo é facilitar essa resolução, evitando qualquer transtorno futuro.: {WHATSAPP_PEDAGIOS_URL}
"""

    return EmailD0Pedagio(
        id_cliente=id_cliente,
        nome=nome,
        destinatario=destinatario,
        assunto=assunto,
        corpo=corpo,
        total_contratos=total_contratos,
        contratos_listados=len(contratos_listados),
        total_pedagios=total_pedagios,
        valor_total=valor_total,
        contrato_referencia=contrato_referencia,
        link_pagamento=link_pagamento,
        documento=documento,
    )


def preparar_email_d0_com_link(email: EmailD0Pedagio, link_pagamento: str) -> EmailD0Pedagio:
    link = str(link_pagamento or "").strip()
    if not link.lower().startswith("http"):
        raise ValueError("Link de pagamento D0 invalido.")
    corpo = email.corpo.replace("{LINK_PAGAMENTO}", link)
    return replace(email, corpo=corpo, link_pagamento=link)


def _local_pending_dir(local_root: str | Path | None = None) -> Path:
    return Path(local_root) if local_root is not None else APP_DATA_DIR


def _shared_pending_dir(workbook_path: str | Path) -> Path:
    return Path(workbook_path).resolve().parent / PENDING_CONTROL_DIR


def _pending_paths_d0(workbook_path: str | Path | None = None, local_root: str | Path | None = None) -> list[Path]:
    paths = [_local_pending_dir(local_root) / PENDING_D0_FILENAME]
    if workbook_path is not None:
        paths.append(_shared_pending_dir(workbook_path) / PENDING_D0_FILENAME)
    return paths


def _synced_paths_d0(workbook_path: str | Path | None = None, local_root: str | Path | None = None) -> list[Path]:
    paths = [_local_pending_dir(local_root) / PENDING_D0_SYNCED_FILENAME]
    if workbook_path is not None:
        paths.append(_shared_pending_dir(workbook_path) / PENDING_D0_SYNCED_FILENAME)
    return paths


def _append_jsonl(path: Path, payload: dict[str, object]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("a", encoding="utf-8") as file:
        file.write(json.dumps(payload, ensure_ascii=False, sort_keys=True) + "\n")
        file.flush()
        os.fsync(file.fileno())


def _read_jsonl(path: Path) -> list[dict[str, object]]:
    if not path.exists():
        return []
    records: list[dict[str, object]] = []
    with path.open("r", encoding="utf-8") as file:
        for line in file:
            line = line.strip()
            if not line:
                continue
            try:
                record = json.loads(line)
            except json.JSONDecodeError:
                continue
            if isinstance(record, dict):
                records.append(record)
    return records


def action_key_from_label(label_or_key: str) -> str:
    value = str(label_or_key or "").strip()
    return ACTION_KEY_BY_LABEL.get(value, value)


def action_label_from_key(key_or_label: str) -> str:
    key = action_key_from_label(key_or_label)
    return ACTION_LABELS.get(key, str(key_or_label or "").strip())


def carregar_config_interface(config_path: str | Path = UI_CONFIG_PATH) -> dict[str, object]:
    defaults: dict[str, object] = {
        "conta_envio": DEFAULT_OUTLOOK_ACCOUNT,
        "canal_cobranca": DEFAULT_CANAL_COBRANCA,
        "caminho_excel": DEFAULT_EXCEL_URL,
        "limite_execucao": "10",
        "usuario_coral": "",
        "salvar_login_coral": False,
        "coral_headless": DEFAULT_CORAL_HEADLESS,
    }
    path = Path(config_path)
    if not path.exists():
        return defaults
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return defaults
    if not isinstance(data, dict):
        return defaults
    merged = defaults | data
    merged["salvar_login_coral"] = bool(merged.get("salvar_login_coral"))
    merged["coral_headless"] = bool(merged.get("coral_headless", DEFAULT_CORAL_HEADLESS))
    return merged


def salvar_config_interface(
    *,
    conta_envio: str,
    limite_execucao: str,
    usuario_coral: str,
    salvar_login_coral: bool,
    coral_headless: bool = DEFAULT_CORAL_HEADLESS,
    caminho_excel: str = DEFAULT_EXCEL_URL,
    canal_cobranca: str = DEFAULT_CANAL_COBRANCA,
    config_path: str | Path = UI_CONFIG_PATH,
) -> None:
    path = Path(config_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "conta_envio": str(conta_envio or "").strip() or DEFAULT_OUTLOOK_ACCOUNT,
        "canal_cobranca": str(canal_cobranca or "").strip() or DEFAULT_CANAL_COBRANCA,
        "caminho_excel": str(caminho_excel or "").strip() or DEFAULT_EXCEL_URL,
        "limite_execucao": str(limite_execucao or "").strip() or "10",
        "usuario_coral": str(usuario_coral or "").strip(),
        "salvar_login_coral": bool(salvar_login_coral),
        "coral_headless": bool(coral_headless),
    }
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True), encoding="utf-8")


def salvar_senha_coral_windows(usuario: str, senha: str, target: str = CORAL_CREDENTIAL_TARGET) -> bool:
    usuario = str(usuario or "").strip()
    senha = str(senha or "")
    if not usuario or not senha:
        return False
    try:
        import win32cred

        win32cred.CredWrite(
            {
                "Type": win32cred.CRED_TYPE_GENERIC,
                "TargetName": target,
                "UserName": usuario,
                "CredentialBlob": senha,
                "Persist": win32cred.CRED_PERSIST_LOCAL_MACHINE,
            },
            0,
        )
        return True
    except Exception:
        return False


def carregar_senha_coral_windows(usuario: str, target: str = CORAL_CREDENTIAL_TARGET) -> str:
    usuario = str(usuario or "").strip()
    if not usuario:
        return ""
    try:
        import win32cred

        credential = win32cred.CredRead(target, win32cred.CRED_TYPE_GENERIC)
        if str(credential.get("UserName") or "").strip() != usuario:
            return ""
        blob = credential.get("CredentialBlob") or ""
        if isinstance(blob, bytes):
            return blob.decode("utf-16-le", errors="ignore").rstrip("\x00")
        return str(blob)
    except Exception:
        return ""


def apagar_senha_coral_windows(target: str = CORAL_CREDENTIAL_TARGET) -> None:
    try:
        import win32cred

        win32cred.CredDelete(target, win32cred.CRED_TYPE_GENERIC)
    except Exception:
        pass


def _email_d0_to_dict(email: EmailD0Pedagio) -> dict[str, object]:
    return {
        "id_cliente": email.id_cliente,
        "documento": email.documento,
        "nome": email.nome,
        "destinatario": email.destinatario,
        "assunto": email.assunto,
        "corpo": email.corpo,
        "total_contratos": email.total_contratos,
        "contratos_listados": email.contratos_listados,
        "total_pedagios": email.total_pedagios,
        "valor_total": email.valor_total,
        "contrato_referencia": email.contrato_referencia,
        "link_pagamento": email.link_pagamento,
    }


def _email_d0_from_dict(data: dict[str, object]) -> EmailD0Pedagio:
    return EmailD0Pedagio(
        id_cliente=str(data.get("id_cliente") or ""),
        documento=str(data.get("documento") or ""),
        nome=str(data.get("nome") or ""),
        destinatario=str(data.get("destinatario") or ""),
        assunto=str(data.get("assunto") or ""),
        corpo=str(data.get("corpo") or ""),
        total_contratos=int(data.get("total_contratos") or 0),
        contratos_listados=int(data.get("contratos_listados") or 0),
        total_pedagios=int(data.get("total_pedagios") or 0),
        valor_total=_format_money(data.get("valor_total") or 0),
        contrato_referencia=str(data.get("contrato_referencia") or ""),
        link_pagamento=str(data.get("link_pagamento") or ""),
    )


def _resultado_d2_to_dict(resultado: ResultadoD2Pedagio) -> dict[str, object]:
    return {
        "id_cliente": resultado.id_cliente,
        "contrato": resultado.contrato,
        "status": resultado.status,
        "cartoes_encontrados": resultado.cartoes_encontrados,
        "cartoes_tentados": resultado.cartoes_tentados,
        "detalhe": resultado.detalhe,
    }


def _resultado_d2_from_dict(data: dict[str, object]) -> ResultadoD2Pedagio:
    return ResultadoD2Pedagio(
        id_cliente=str(data.get("id_cliente") or ""),
        contrato=str(data.get("contrato") or ""),
        status=str(data.get("status") or ""),
        cartoes_encontrados=int(data.get("cartoes_encontrados") or 0),
        cartoes_tentados=int(data.get("cartoes_tentados") or 0),
        detalhe=str(data.get("detalhe") or ""),
    )


def _date_suffix(value: date | datetime | None = None) -> str:
    value = value or date.today()
    if isinstance(value, datetime):
        value = value.date()
    return value.strftime("%Y%m%d")


def _event_datetime_text(value: date | datetime | None = None) -> str:
    value = value or datetime.now()
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return f"{value:%Y-%m-%d} 00:00:00"


def _processing_events_filename(data_ref: date | datetime | None = None) -> str:
    return f"{PROCESSING_EVENTS_PREFIX}_{_date_suffix(data_ref)}.jsonl"


def _processing_synced_filename(data_ref: date | datetime | None = None) -> str:
    return f"{PROCESSING_EVENTS_SYNCED_PREFIX}_{_date_suffix(data_ref)}.jsonl"


def _processing_event_paths(
    workbook_path: str | Path | None = None,
    data_ref: date | datetime | None = None,
    local_root: str | Path | None = None,
) -> list[Path]:
    filename = _processing_events_filename(data_ref)
    paths = [_local_pending_dir(local_root) / filename]
    if workbook_path is not None:
        paths.append(_shared_pending_dir(workbook_path) / filename)
    return paths


def _processing_synced_paths(
    workbook_path: str | Path | None = None,
    data_ref: date | datetime | None = None,
    local_root: str | Path | None = None,
) -> list[Path]:
    filename = _processing_synced_filename(data_ref)
    paths = [_local_pending_dir(local_root) / filename]
    if workbook_path is not None:
        paths.append(_shared_pending_dir(workbook_path) / filename)
    return paths


def _workbook_identity(workbook_path: str | Path) -> str:
    return str(Path(workbook_path).expanduser().resolve(strict=False)).casefold()


def _event_hash(parts: list[object]) -> str:
    raw = json.dumps([str(part or "") for part in parts], ensure_ascii=False, sort_keys=True)
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def _deterministic_processing_event_id(tipo: str, payload: dict[str, object], processed_at: date | datetime) -> str:
    tipo_norm = str(tipo or "").strip()
    data = _event_datetime_text(processed_at)[:10]
    if tipo_norm == "EMAIL_D0_ENVIADO":
        email = payload.get("email")
        email_data = email if isinstance(email, dict) else {}
        return _event_hash([tipo_norm, email_data.get("id_cliente"), email_data.get("link_pagamento")])
    if tipo_norm == "RESULTADO_D2":
        resultado = payload.get("resultado")
        resultado_data = resultado if isinstance(resultado, dict) else {}
        return _event_hash(
            [
                tipo_norm,
                resultado_data.get("id_cliente"),
                resultado_data.get("contrato"),
                resultado_data.get("status"),
                data,
            ]
        )
    if tipo_norm == "LINK_D2_GERADO":
        contratos = payload.get("contratos")
        contratos_data = contratos if isinstance(contratos, list) else []
        return _event_hash([tipo_norm, payload.get("id_cliente"), "|".join(str(item) for item in contratos_data), payload.get("link_pagamento")])
    return _event_hash([tipo_norm, payload, data])


def registrar_evento_processamento_json(
    workbook_path: str | Path,
    *,
    tipo: str,
    payload: dict[str, object],
    processed_at: date | datetime | None = None,
    local_root: str | Path | None = None,
) -> Path:
    processed_at = processed_at or datetime.now()
    tipo = str(tipo or "").strip()
    event_id = str(payload.get("id_evento") or _deterministic_processing_event_id(tipo, payload, processed_at))
    record = {
        "id_evento": event_id,
        "tipo": tipo,
        "data_evento": _date_suffix(processed_at),
        "data_hora": _event_datetime_text(processed_at),
        "workbook": _workbook_identity(workbook_path),
        "payload": payload,
    }
    paths = _processing_event_paths(workbook_path, processed_at, local_root)
    local_path = paths[0]
    _append_jsonl(local_path, record)
    for path in paths[1:]:
        try:
            _append_jsonl(path, record)
        except Exception:
            pass
    return local_path


def _synced_event_times_processamento(
    workbook_path: str | Path,
    data_ref: date | datetime | None = None,
    local_root: str | Path | None = None,
) -> dict[str, str]:
    synced: dict[str, str] = {}
    for path in _processing_synced_paths(workbook_path, data_ref, local_root):
        for record in _read_jsonl(path):
            event_id = str(record.get("id_evento") or "").strip()
            synced_at = str(record.get("data_hora_sync") or "").strip()
            if event_id and synced_at and synced_at > synced.get(event_id, ""):
                synced[event_id] = synced_at
    return synced


def carregar_pendencias_processamento_json(
    workbook_path: str | Path,
    *,
    data_ref: date | datetime | None = None,
    local_root: str | Path | None = None,
) -> list[dict[str, object]]:
    data_ref = data_ref or date.today()
    workbook_alvo = _workbook_identity(workbook_path)
    synced_times = _synced_event_times_processamento(workbook_path, data_ref, local_root)
    pending_by_id: dict[str, dict[str, object]] = {}
    for path in _processing_event_paths(workbook_path, data_ref, local_root):
        for record in _read_jsonl(path):
            event_id = str(record.get("id_evento") or "").strip()
            if not event_id:
                continue
            if str(record.get("workbook") or "").casefold() != workbook_alvo:
                continue
            if str(record.get("data_evento") or "") != _date_suffix(data_ref):
                continue
            record_time = str(record.get("data_hora") or "").strip()
            synced_at = synced_times.get(event_id)
            if synced_at and record_time and record_time <= synced_at:
                continue
            if not isinstance(record.get("payload"), dict):
                continue
            pending_by_id.setdefault(event_id, record)
    return list(pending_by_id.values())


def contar_pendencias_processamento_recentes(
    workbook_path: str | Path,
    *,
    data_ref: date | datetime | None = None,
    dias: int = 2,
    local_root: str | Path | None = None,
) -> dict[str, int]:
    data_ref = data_ref or date.today()
    if isinstance(data_ref, datetime):
        data_ref = data_ref.date()
    resultado: dict[str, int] = {}
    for offset in range(max(dias, 0)):
        dia = data_ref - timedelta(days=offset)
        pendencias = carregar_pendencias_processamento_json(workbook_path, data_ref=dia, local_root=local_root)
        if pendencias:
            resultado[dia.isoformat()] = len(pendencias)
    return resultado


def _marcar_eventos_processamento_sincronizados(
    workbook_path: str | Path,
    event_ids: list[str],
    *,
    data_ref: date | datetime | None = None,
    local_root: str | Path | None = None,
) -> None:
    if not event_ids:
        return
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    records = [{"id_evento": event_id, "data_hora_sync": now} for event_id in dict.fromkeys(event_ids) if event_id]
    for path in _processing_synced_paths(workbook_path, data_ref, local_root):
        for record in records:
            try:
                _append_jsonl(path, record)
            except Exception:
                pass


def sincronizar_eventos_processamento_json(
    workbook_path: str | Path,
    *,
    data_ref: date | datetime | None = None,
    local_root: str | Path | None = None,
) -> dict[str, int]:
    data_ref = data_ref or date.today()
    eventos = carregar_pendencias_processamento_json(workbook_path, data_ref=data_ref, local_root=local_root)
    if not eventos:
        return {"eventos_pendentes": 0, "eventos_sincronizados": 0, "clientes_atualizados": 0, "contratos_atualizados": 0}

    d0_por_grupo: dict[tuple[str, bool, str], list[EmailD0Pedagio]] = {}
    d2_resultados: list[ResultadoD2Pedagio] = []
    d2_links: list[dict[str, object]] = []
    synced_ids: list[str] = []
    clientes_atualizados = 0
    contratos_atualizados = 0

    for event in eventos:
        tipo = str(event.get("tipo") or "").strip()
        payload = event.get("payload")
        if not isinstance(payload, dict):
            continue
        event_id = str(event.get("id_evento") or "").strip()
        if tipo == "EMAIL_D0_ENVIADO":
            email_data = payload.get("email")
            if not isinstance(email_data, dict):
                continue
            conta_envio = str(payload.get("conta_envio") or DEFAULT_OUTLOOK_ACCOUNT)
            registrar_link = bool(payload.get("registrar_link"))
            usuario = str(payload.get("usuario") or "")
            d0_por_grupo.setdefault((conta_envio, registrar_link, usuario), []).append(_email_d0_from_dict(email_data))
            synced_ids.append(event_id)
        elif tipo == "RESULTADO_D2":
            resultado_data = payload.get("resultado")
            if not isinstance(resultado_data, dict):
                continue
            d2_resultados.append(_resultado_d2_from_dict(resultado_data))
            synced_ids.append(event_id)
        elif tipo == "LINK_D2_GERADO":
            d2_links.append(payload)
            synced_ids.append(event_id)

    for (conta_envio, registrar_link, usuario), emails in d0_por_grupo.items():
        clientes_atualizados += registrar_processamento_d0_excel(
            workbook_path,
            emails,
            conta_envio=conta_envio,
            enviado=True,
            registrar_link=registrar_link,
            usuario=usuario,
        )

    if d2_resultados:
        usuario_d2 = next(
            (
                str(event.get("payload", {}).get("usuario") or "")
                for event in eventos
                if isinstance(event.get("payload"), dict) and str(event.get("tipo") or "") == "RESULTADO_D2"
            ),
            "",
        )
        contratos_atualizados += registrar_processamento_d2_excel(workbook_path, d2_resultados, usuario=usuario_d2)

    for payload in d2_links:
        contratos = payload.get("contratos") if isinstance(payload.get("contratos"), list) else []
        contratos_atualizados += registrar_link_d2_excel(
            workbook_path,
            id_cliente=str(payload.get("id_cliente") or ""),
            contratos=[str(item) for item in contratos],
            valor_link=_format_money(payload.get("valor_link") or 0),
            link_pagamento=str(payload.get("link_pagamento") or ""),
            usuario=str(payload.get("usuario") or ""),
        )

    _marcar_eventos_processamento_sincronizados(
        workbook_path,
        synced_ids,
        data_ref=data_ref,
        local_root=local_root,
    )
    return {
        "eventos_pendentes": len(eventos),
        "eventos_sincronizados": len(synced_ids),
        "clientes_atualizados": clientes_atualizados,
        "contratos_atualizados": contratos_atualizados,
    }


def _pending_event_id_d0(email: EmailD0Pedagio) -> str:
    parts = ["EMAIL_D0", email.id_cliente, email.destinatario, email.contrato_referencia, email.link_pagamento]
    return "::".join(_normalizar_texto(part).replace(" ", "_") for part in parts)


def _synced_event_ids_d0(workbook_path: str | Path | None = None, local_root: str | Path | None = None) -> set[str]:
    synced: set[str] = set()
    for path in _synced_paths_d0(workbook_path, local_root):
        for record in _read_jsonl(path):
            event_id = str(record.get("id_evento") or "").strip()
            if event_id:
                synced.add(event_id)
    return synced


def _synced_event_times_d0(workbook_path: str | Path | None = None, local_root: str | Path | None = None) -> dict[str, str]:
    synced: dict[str, str] = {}
    for path in _synced_paths_d0(workbook_path, local_root):
        for record in _read_jsonl(path):
            event_id = str(record.get("id_evento") or "").strip()
            synced_at = str(record.get("data_hora_sync") or "").strip()
            if event_id and synced_at and synced_at > synced.get(event_id, ""):
                synced[event_id] = synced_at
    return synced


def registrar_pendencia_d0_json(
    email: EmailD0Pedagio,
    workbook_path: str | Path,
    conta_envio: str,
    registrar_link: bool,
    usuario: str = "",
    local_root: str | Path | None = None,
) -> dict[str, Path]:
    if not email.link_pagamento:
        raise ValueError("Pendencia D0 exige e-mail com link de pagamento.")
    record = {
        "id_evento": _pending_event_id_d0(email),
        "tipo": "EMAIL_D0_ENVIADO",
        "data_hora": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "conta_envio": conta_envio,
        "registrar_link": bool(registrar_link),
        "usuario": usuario,
        "email": _email_d0_to_dict(email),
    }
    local_path = _local_pending_dir(local_root) / PENDING_D0_FILENAME
    shared_path = _shared_pending_dir(workbook_path) / PENDING_D0_FILENAME
    _append_jsonl(local_path, record)
    try:
        _append_jsonl(shared_path, record)
    except Exception:
        pass
    return {"local": local_path, "compartilhado": shared_path}


def carregar_pendencias_d0_json(
    workbook_path: str | Path,
    local_root: str | Path | None = None,
) -> list[dict[str, object]]:
    synced_times = _synced_event_times_d0(workbook_path, local_root)
    pending_by_id: dict[str, dict[str, object]] = {}
    for path in _pending_paths_d0(workbook_path, local_root):
        for record in _read_jsonl(path):
            event_id = str(record.get("id_evento") or "").strip()
            if not event_id:
                continue
            synced_at = synced_times.get(event_id)
            record_time = str(record.get("data_hora") or "").strip()
            if synced_at and record_time and record_time <= synced_at:
                continue
            if record.get("tipo") != "EMAIL_D0_ENVIADO":
                continue
            if not isinstance(record.get("email"), dict):
                continue
            pending_by_id.setdefault(event_id, record)
    return list(pending_by_id.values())


def _marcar_pendencias_d0_sincronizadas(
    event_ids: list[str],
    workbook_path: str | Path,
    local_root: str | Path | None = None,
) -> None:
    if not event_ids:
        return
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    records = [{"id_evento": event_id, "data_hora_sync": now} for event_id in dict.fromkeys(event_ids) if event_id]
    if not records:
        return
    for path in _synced_paths_d0(workbook_path, local_root):
        for record in records:
            try:
                _append_jsonl(path, record)
            except Exception:
                pass


def sincronizar_pendencias_d0_json(
    workbook_path: str | Path,
    local_root: str | Path | None = None,
) -> dict[str, int]:
    pendencias = carregar_pendencias_d0_json(workbook_path, local_root)
    if not pendencias:
        return {"eventos_pendentes": 0, "eventos_sincronizados": 0, "clientes_atualizados": 0}

    grupos: dict[tuple[str, bool], list[tuple[str, EmailD0Pedagio, str]]] = {}
    for record in pendencias:
        event_id = str(record.get("id_evento") or "").strip()
        email = _email_d0_from_dict(record["email"])
        conta_envio = str(record.get("conta_envio") or DEFAULT_OUTLOOK_ACCOUNT)
        registrar_link = bool(record.get("registrar_link"))
        usuario = str(record.get("usuario") or "")
        grupos.setdefault((conta_envio, registrar_link), []).append((event_id, email, usuario))

    synced_ids: list[str] = []
    clientes_atualizados = 0
    for (conta_envio, registrar_link), items in grupos.items():
        emails = [item[1] for item in items]
        usuario = next((item[2] for item in items if item[2]), "")
        clientes_atualizados += registrar_processamento_d0_excel(
            workbook_path,
            emails,
            conta_envio=conta_envio,
            enviado=True,
            registrar_link=registrar_link,
            usuario=usuario,
        )
        synced_ids.extend(item[0] for item in items)
    _marcar_pendencias_d0_sincronizadas(synced_ids, workbook_path, local_root)
    return {
        "eventos_pendentes": len(pendencias),
        "eventos_sincronizados": len(synced_ids),
        "clientes_atualizados": clientes_atualizados,
    }


def _default_onedrive_roots() -> list[Path]:
    user_home = Path.home()
    roots = [
        user_home / "OneDrive - Foco Aluguel de Carros",
        user_home / "OneDrive",
    ]
    for key, value in os.environ.items():
        if key.upper().startswith("ONEDRIVE") and value:
            roots.append(Path(value))
    return list(dict.fromkeys(roots))


def resolver_caminho_excel_compartilhado(value: str | Path, search_roots: list[Path] | None = None) -> Path | None:
    raw = str(value or "").strip().strip('"')
    if not raw:
        return None

    local_path = Path(raw)
    if local_path.exists():
        return local_path

    if not raw.lower().startswith(("http://", "https://")):
        return None

    parsed = urlparse(raw)
    filename = Path(unquote(parsed.path)).name
    if not filename:
        return None

    roots = search_roots or _default_onedrive_roots()
    normalized_filename = _normalizar_texto(filename)
    matches: list[Path] = []
    for root in roots:
        if not root.exists():
            continue
        try:
            for candidate in root.rglob(filename):
                if candidate.is_file():
                    matches.append(candidate)
        except Exception:
            continue
        if not matches:
            try:
                for candidate in root.rglob("*.xlsx"):
                    if candidate.is_file() and _normalizar_texto(candidate.name) == normalized_filename:
                        matches.append(candidate)
            except Exception:
                continue

    if not matches:
        return None

    decoded_parts = [_normalizar_texto(part) for part in Path(unquote(parsed.path)).parts if part]

    def score(path: Path) -> int:
        path_text = _normalizar_texto(str(path))
        return sum(1 for part in decoded_parts if part and part in path_text)

    matches.sort(key=score, reverse=True)
    return matches[0]


def carregar_resumo_cobrancas_excel(path: str | Path, hoje: date | None = None, canal: str | None = None) -> ResumoCobrancasExcel:
    workbook_path = Path(path)
    if not workbook_path.exists():
        raise FileNotFoundError(f"Planilha nao encontrada: {workbook_path}")

    hoje = hoje or date.today()
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if SHEET_CLIENTES not in workbook.sheetnames:
            raise ValueError(f"Aba obrigatoria ausente: {SHEET_CLIENTES}")
        if SHEET_CONTRATOS not in workbook.sheetnames:
            raise ValueError(f"Aba obrigatoria ausente: {SHEET_CONTRATOS}")

        sheet = workbook[SHEET_CLIENTES]
        rows = sheet.iter_rows(values_only=True)
        try:
            headers = [str(value or "").strip() for value in next(rows)]
        except StopIteration as exc:
            raise ValueError("Aba CLIENTES esta vazia") from exc

        header_map = {_normalizar_texto(header): index for index, header in enumerate(headers)}
        missing = sorted(REQUIRED_CLIENT_COLUMNS - set(header_map))
        if missing:
            raise ValueError(f"Colunas obrigatorias ausentes em CLIENTES: {', '.join(missing)}")

        contracts_sheet = workbook[SHEET_CONTRATOS]
        contract_rows = contracts_sheet.iter_rows(values_only=True)
        try:
            contract_headers = [str(value or "").strip() for value in next(contract_rows)]
        except StopIteration as exc:
            raise ValueError("Aba CONTRATOS esta vazia") from exc

        contract_header_map = {_normalizar_texto(header): index for index, header in enumerate(contract_headers)}
        missing_contracts = sorted(REQUIRED_CONTRACT_COLUMNS - set(contract_header_map))
        if missing_contracts:
            raise ValueError(f"Colunas obrigatorias ausentes em CONTRATOS: {', '.join(missing_contracts)}")

        def get(row: tuple[object, ...], column: str):
            index = header_map[column]
            return row[index] if index < len(row) else None

        def get_contract(row: tuple[object, ...], column: str):
            index = contract_header_map[column]
            return row[index] if index < len(row) else None

        def contract_record(row: tuple[object, ...]) -> dict[str, object]:
            return {
                column: row[index] if index < len(row) else None
                for column, index in contract_header_map.items()
            }

        total = aptos_d0 = aptos_d2 = aptos_d45 = aptos_d7 = 0
        em_processamento = pagos = erros = 0
        contratos_total = contratos_pendentes = 0
        valor_total = 0.0
        exemplos_aptos: list[dict[str, object]] = []
        canal_ativo = _canal_filtro_ativo(canal)
        valor_por_cliente: dict[str, float] = {}
        clientes_com_contrato_no_canal: set[str] = set()

        for row in contract_rows:
            if not row or all(value in (None, "") for value in row):
                continue
            record = contract_record(row)
            if not _registro_pertence_ao_canal(record, canal):
                continue
            id_cliente_contrato = str(record.get("ID_CLIENTE") or record.get("DOCUMENTO") or "").strip()
            if id_cliente_contrato:
                clientes_com_contrato_no_canal.add(id_cliente_contrato)
                valor_por_cliente[id_cliente_contrato] = round(
                    valor_por_cliente.get(id_cliente_contrato, 0.0) + _format_money(record.get("VALOR_A_COBRAR")),
                    2,
                )
            contratos_total += 1
            if _normalizar_texto(get_contract(row, "STATUS_CARTAO")) == "PENDENTE":
                contratos_pendentes += 1

        for row in rows:
            if not row or all(value in (None, "") for value in row):
                continue

            id_cliente = str(get(row, "ID_CLIENTE") or get(row, "DOCUMENTO") or "").strip()
            if canal_ativo and id_cliente not in clientes_com_contrato_no_canal:
                continue

            total += 1
            status = get(row, "STATUS")
            etapa = _normalizar_texto(get(row, "ETAPA"))
            proxima_acao = get(row, "PROXIMA_ACAO")
            valor = (
                valor_por_cliente.get(id_cliente, 0.0)
                if canal_ativo
                else _format_money(get(row, "VALOR_TOTAL"))
            )
            valor_total += valor

            if _is_processing(status):
                em_processamento += 1
            if _is_paid(status):
                pagos += 1
            if _is_error(status):
                erros += 1

            apta = _is_active(status) and _is_due(proxima_acao, hoje) and _is_valid_email(get(row, "EMAIL"))
            if apta:
                if etapa == "D0":
                    aptos_d0 += 1
                elif etapa == "D0+2":
                    aptos_d2 += 1
                elif etapa in {"D0+4/D0+5", "D0+4", "D0+5"}:
                    aptos_d45 += 1
                elif etapa == "D0+7":
                    aptos_d7 += 1

                if len(exemplos_aptos) < 20:
                    exemplos_aptos.append(
                        {
                            "id": get(row, "ID_CLIENTE"),
                            "cliente": get(row, "CLIENTE"),
                            "documento": get(row, "DOCUMENTO"),
                            "etapa": get(row, "ETAPA"),
                            "valor": valor,
                        }
                    )

    finally:
        workbook.close()
    return ResumoCobrancasExcel(
        total=total,
        contratos_total=contratos_total,
        contratos_pendentes=contratos_pendentes,
        aptos_d0=aptos_d0,
        aptos_d2=aptos_d2,
        aptos_d45=aptos_d45,
        aptos_d7=aptos_d7,
        em_processamento=em_processamento,
        pagos=pagos,
        erros=erros,
        valor_total=valor_total,
        exemplos_aptos=exemplos_aptos,
    )


def carregar_emails_d0_excel(
    path: str | Path,
    hoje: date | None = None,
    limite: int | None = None,
    canal: str | None = None,
) -> list[EmailD0Pedagio]:
    workbook_path = Path(path)
    if not workbook_path.exists():
        raise FileNotFoundError(f"Planilha nao encontrada: {workbook_path}")

    hoje = hoje or date.today()
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        clientes = _sheet_records(workbook, SHEET_CLIENTES, REQUIRED_D0_CLIENT_COLUMNS)
        contratos = _sheet_records(workbook, SHEET_CONTRATOS, REQUIRED_D0_CONTRACT_COLUMNS)
    finally:
        workbook.close()

    contratos_por_cliente: dict[str, list[dict[str, object]]] = {}
    for contrato in contratos:
        if _normalizar_texto(contrato.get("STATUS_CARTAO")) != "PENDENTE":
            continue
        if not _registro_pertence_ao_canal(contrato, canal):
            continue
        id_cliente = str(contrato.get("ID_CLIENTE") or "").strip()
        if not id_cliente:
            continue
        contratos_por_cliente.setdefault(id_cliente, []).append(contrato)

    emails: list[EmailD0Pedagio] = []
    for cliente in clientes:
        if _normalizar_texto(cliente.get("ETAPA")) != "D0":
            continue
        if not (_is_active(cliente.get("STATUS")) and _is_due(cliente.get("PROXIMA_ACAO"), hoje)):
            continue
        if not _is_valid_email(cliente.get("EMAIL")):
            continue
        id_cliente = str(cliente.get("ID_CLIENTE") or cliente.get("DOCUMENTO") or "").strip()
        documento_cliente = _digits_only(cliente.get("DOCUMENTO"))
        contratos_cliente = []
        for contrato in contratos_por_cliente.get(id_cliente, []):
            documento_contrato = _digits_only(contrato.get("DOCUMENTO"))
            if documento_cliente and documento_contrato and not _documentos_equivalentes(documento_cliente, documento_contrato):
                continue
            contratos_cliente.append(contrato)
        if not contratos_cliente:
            continue
        emails.append(_montar_email_d0(cliente, contratos_cliente))
        if limite is not None and len(emails) >= limite:
            break
    return emails


def carregar_contratos_d2_excel(
    path: str | Path,
    hoje: date | None = None,
    limite: int | None = None,
    canal: str | None = None,
) -> list[ContratoD2Pedagio]:
    workbook_path = Path(path)
    if not workbook_path.exists():
        raise FileNotFoundError(f"Planilha nao encontrada: {workbook_path}")

    hoje = hoje or date.today()
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        clientes = _sheet_records(workbook, SHEET_CLIENTES, REQUIRED_CLIENT_COLUMNS)
        contratos = _sheet_records(workbook, SHEET_CONTRATOS, REQUIRED_CONTRACT_COLUMNS | {"CLIENTE", "DOCUMENTO"})
    finally:
        workbook.close()

    clientes_aptos: dict[str, dict[str, object]] = {}
    for cliente in clientes:
        if _normalizar_texto(cliente.get("ETAPA")) != "D0+2":
            continue
        if not (_is_active(cliente.get("STATUS")) and _is_due(cliente.get("PROXIMA_ACAO"), hoje)):
            continue
        id_cliente = str(cliente.get("ID_CLIENTE") or cliente.get("DOCUMENTO") or "").strip()
        if id_cliente:
            clientes_aptos[id_cliente] = cliente

    contratos_pendentes_por_cliente: dict[str, list[dict[str, object]]] = {}
    for contrato in contratos:
        if _normalizar_texto(contrato.get("STATUS_CARTAO")) != "PENDENTE":
            continue
        if not _registro_pertence_ao_canal(contrato, canal):
            continue
        id_cliente = str(contrato.get("ID_CLIENTE") or "").strip()
        cliente = clientes_aptos.get(id_cliente)
        if cliente is None:
            continue
        numero_contrato = str(contrato.get("CONTRATO") or "").strip()
        if not numero_contrato:
            continue
        contratos_pendentes_por_cliente.setdefault(id_cliente, []).append(contrato)

    fila: list[ContratoD2Pedagio] = []
    for id_cliente, cliente in clientes_aptos.items():
        contratos_cliente = contratos_pendentes_por_cliente.get(id_cliente, [])
        if not contratos_cliente:
            continue
        if limite is not None and len(fila) >= limite:
            break
        for contrato in contratos_cliente:
            fila.append(
                ContratoD2Pedagio(
                    id_cliente=id_cliente,
                    cliente=str(contrato.get("CLIENTE") or cliente.get("CLIENTE") or "").strip(),
                    documento=str(contrato.get("DOCUMENTO") or cliente.get("DOCUMENTO") or "").strip(),
                    contrato=str(contrato.get("CONTRATO") or "").strip(),
                    placa=str(contrato.get("PLACA") or "").strip(),
                    valor=round(_format_money(contrato.get("VALOR_A_COBRAR")), 2),
                )
            )
    return fila


def _outlook_app():
    try:
        import win32com.client as win32
    except Exception as exc:
        raise RuntimeError("pywin32/win32com nao esta disponivel para acessar o Outlook Desktop.") from exc
    return win32.Dispatch("Outlook.Application")


def _inicializar_com_outlook():
    try:
        import pythoncom
    except Exception as exc:
        raise RuntimeError("pythoncom nao esta disponivel para inicializar o Outlook nesta thread.") from exc
    pythoncom.CoInitialize()
    return pythoncom


def listar_contas_outlook() -> list[str]:
    pythoncom = _inicializar_com_outlook()
    try:
        outlook = _outlook_app()
        namespace = outlook.GetNamespace("MAPI")
        contas: list[str] = []
        for conta in namespace.Accounts:
            smtp = ""
            nome = ""
            try:
                smtp = str(conta.SmtpAddress or "").strip()
            except Exception:
                smtp = ""
            try:
                nome = str(conta.DisplayName or "").strip()
            except Exception:
                nome = ""
            label = smtp or nome
            if label:
                contas.append(label)
        return contas
    finally:
        pythoncom.CoUninitialize()


def _obter_conta_outlook(outlook, conta_escolhida: str):
    conta_escolhida_norm = _normalizar_texto(conta_escolhida)
    namespace = outlook.GetNamespace("MAPI")
    contas_disponiveis: list[str] = []
    for conta in namespace.Accounts:
        smtp = ""
        nome = ""
        try:
            smtp = str(conta.SmtpAddress or "").strip()
        except Exception:
            smtp = ""
        try:
            nome = str(conta.DisplayName or "").strip()
        except Exception:
            nome = ""
        label = smtp or nome
        if label:
            contas_disponiveis.append(label)
        if conta_escolhida_norm and conta_escolhida_norm in {_normalizar_texto(smtp), _normalizar_texto(nome), _normalizar_texto(label)}:
            return conta
    raise RuntimeError(
        f"Conta de envio nao encontrada no Outlook: {conta_escolhida}. "
        f"Contas disponiveis: {', '.join(contas_disponiveis) or 'nenhuma'}"
    )


def _montar_html_email_d0(email: EmailD0Pedagio) -> str:
    corpo_html = html.escape(email.corpo).replace("\n", "<br>")
    nome_escaped = html.escape(email.nome)
    for saudacao in (f"OlÃ¡, {nome_escaped}!", f"Olá, {nome_escaped}!"):
        corpo_html = corpo_html.replace(saudacao, saudacao.replace(nome_escaped, f"<strong>{nome_escaped}</strong>"))
    for titulo in ("Dados da LocaÃ§Ã£o e UtilizaÃ§Ã£o do pedÃ¡gio:", "Dados da Locação e Utilização do pedágio:"):
        corpo_html = corpo_html.replace(titulo, f"<strong>{titulo}</strong>")
    for total_pedagios_line in (
        f"Quantidade total de passagens em pedÃ¡gio: {email.total_pedagios}",
        f"Quantidade total de passagens em pedágio: {email.total_pedagios}",
    ):
        corpo_html = corpo_html.replace(total_pedagios_line, f"<strong>{total_pedagios_line}</strong>")
    valor_total_line = f"Valor total a regularizar: R$ {_format_brl(email.valor_total)}"
    corpo_html = corpo_html.replace(valor_total_line, f"<strong>{valor_total_line}</strong>")
    if email.link_pagamento:
        link_escaped = html.escape(email.link_pagamento, quote=True)
        corpo_html = corpo_html.replace(
            f"Link para pagamento: {link_escaped}",
            f'<strong>Link para pagamento: <a href="{link_escaped}">{link_escaped}</a></strong>',
        )
    whatsapp_escaped = html.escape(WHATSAPP_PEDAGIOS_URL, quote=True)
    corpo_html = corpo_html.replace(f".: {whatsapp_escaped}", ".")
    corpo_html = corpo_html.replace(f": {whatsapp_escaped}", "")
    corpo_html = corpo_html.replace(whatsapp_escaped, "")
    corpo_html = re.sub(r"\bWhatsApp\b:?", "<strong>Whatsapp:</strong>", corpo_html, flags=re.IGNORECASE)
    return (
        '<html><body style="font-family:Arial,sans-serif;font-size:11pt;color:#222;line-height:1.45;">'
        f"{corpo_html}<br><br>"
        f'<a href="{whatsapp_escaped}" style="display:inline-block;">'
        f'<img src="cid:{SIGNATURE_CONTENT_ID}" alt="Foco Aluguel de Carros" width="320" '
        'style="width:320px;max-width:100%;height:auto;border:0;display:block;"></a>'
        "</body></html>"
    )


def _preparar_assinatura_outlook(
    origem: str | Path = SIGNATURE_IMAGE_PATH,
    cache_root: str | Path | None = None,
) -> Path:
    origem_path = Path(origem)
    if not origem_path.exists():
        raise FileNotFoundError(f"Assinatura de e-mail nao encontrada: {origem_path}")
    if cache_root is None:
        cache_root = Path(os.environ.get("LOCALAPPDATA") or Path.home()) / "SistemaFOCO" / "cache"
    cache_path = Path(cache_root)
    cache_path.mkdir(parents=True, exist_ok=True)
    destino = cache_path / "assinatura_email_pedagio_whatsapp.png"
    if not destino.exists() or destino.stat().st_size != origem_path.stat().st_size:
        shutil.copy2(origem_path, destino)
    return destino


def _anexar_imagem_inline_outlook(mail, image_path: str | Path, content_id: str = SIGNATURE_CONTENT_ID):
    attachment = mail.Attachments.Add(str(image_path))
    accessor = attachment.PropertyAccessor
    accessor.SetProperty("http://schemas.microsoft.com/mapi/proptag/0x3712001F", content_id)
    accessor.SetProperty("http://schemas.microsoft.com/mapi/proptag/0x370E001F", "image/png")
    accessor.SetProperty("http://schemas.microsoft.com/mapi/proptag/0x7FFE000B", True)
    return attachment


def _iter_outlook_items(items):
    try:
        count = int(items.Count)
        for index in range(1, count + 1):
            yield items.Item(index)
        return
    except Exception:
        pass
    try:
        for item in items:
            yield item
    except Exception:
        return


def _outlook_message_matches(item, subject: str, destinatario: str) -> bool:
    try:
        item_subject = str(getattr(item, "Subject", "") or "").strip()
    except Exception:
        item_subject = ""
    if item_subject != str(subject or "").strip():
        return False
    destinatario_norm = str(destinatario or "").strip().casefold()
    if not destinatario_norm:
        return False
    try:
        to_text = str(getattr(item, "To", "") or "").casefold()
    except Exception:
        to_text = ""
    if destinatario_norm in to_text:
        return True
    try:
        recipients = getattr(item, "Recipients", None)
        if recipients is None:
            return False
        for recipient in _iter_outlook_items(recipients):
            address = str(getattr(recipient, "Address", "") or "").casefold()
            name = str(getattr(recipient, "Name", "") or "").casefold()
            if destinatario_norm in {address, name} or destinatario_norm in address:
                return True
    except Exception:
        return False
    return False


def _outbox_tem_email(namespace, subject: str, destinatario: str) -> bool:
    outbox = namespace.GetDefaultFolder(OUTLOOK_FOLDER_OUTBOX)
    items = getattr(outbox, "Items", [])
    for item in _iter_outlook_items(items):
        if _outlook_message_matches(item, subject, destinatario):
            return True
    return False


def _aguardar_saida_caixa_outlook(
    namespace,
    subject: str,
    destinatario: str,
    *,
    timeout_seconds: int = OUTLOOK_SEND_CONFIRM_TIMEOUT_SECONDS,
    poll_seconds: int = OUTLOOK_SEND_CONFIRM_POLL_SECONDS,
    log_callback=None,
) -> None:
    try:
        namespace.SendAndReceive(False)
    except Exception as exc:
        if log_callback is not None:
            log_callback(f"Outlook: nao foi possivel forcar Enviar/Receber: {exc}")

    end_time = time.time() + max(timeout_seconds, 1)
    while time.time() < end_time:
        if not _outbox_tem_email(namespace, subject, destinatario):
            if log_callback is not None:
                log_callback("Outlook: mensagem saiu da Caixa de Saida.")
            return
        time.sleep(max(poll_seconds, 1))
        try:
            namespace.SendAndReceive(False)
        except Exception:
            pass
    raise RuntimeError(
        "Outlook manteve o e-mail na Caixa de Saida/Na fila. "
        "Verifique se o Outlook esta online, se a conta de envio esta conectada e execute Enviar/Receber."
    )


def criar_email_outlook(
    email: EmailD0Pedagio,
    conta_envio: str,
    log_callback=None,
    anexos: list[str | Path] | None = None,
) -> str:
    if not _is_valid_email(email.destinatario):
        raise RuntimeError(f"E-mail invalido para {email.nome}: {email.destinatario}")
    if not conta_envio:
        raise RuntimeError("Selecione a conta do Outlook antes de processar D0.")
    if not email.link_pagamento:
        raise RuntimeError(f"Link de pagamento ausente para {email.nome}.")

    def log_etapa(message: str, started_at: float) -> float:
        now = time.perf_counter()
        if log_callback is not None:
            log_callback(f"Outlook: {message} ({now - started_at:.1f}s)")
        return now

    pythoncom = _inicializar_com_outlook()
    etapa = time.perf_counter()
    try:
        outlook = _outlook_app()
        outlook.GetNamespace("MAPI")
        etapa = log_etapa("aplicacao conectada", etapa)
        mail = outlook.CreateItem(0)
        conta = _obter_conta_outlook(outlook, conta_envio)
        mail.SendUsingAccount = conta
        mail._oleobj_.Invoke(*(64209, 0, 8, 0, conta))
        etapa = log_etapa("mensagem e conta preparadas", etapa)
        mail.To = email.destinatario
        mail.Subject = email.assunto
        if SIGNATURE_IMAGE_PATH.exists():
            assinatura_local = _preparar_assinatura_outlook()
            _anexar_imagem_inline_outlook(mail, assinatura_local)
            etapa = log_etapa("assinatura local anexada", etapa)
        for anexo in anexos or []:
            anexo_path = Path(anexo)
            if not anexo_path.exists():
                raise FileNotFoundError(f"Anexo nao encontrado: {anexo_path}")
            mail.Attachments.Add(str(anexo_path))
            etapa = log_etapa(f"anexo incluido: {anexo_path.name}", etapa)
        mail.HTMLBody = _montar_html_email_d0(email)
        etapa = log_etapa("HTML montado", etapa)
        if log_callback is not None:
            log_callback(f"Outlook: destinatario={email.destinatario}")
        mail.Save()
        etapa = log_etapa("mensagem salva", etapa)
        mail.Send()
        etapa = log_etapa("comando Send concluido", etapa)
        if log_callback is not None:
            log_callback("Outlook: verificacao pos-envio desativada para este robo.")
        return "enviado"
    finally:
        pythoncom.CoUninitialize()


def _header_map_from_sheet(sheet) -> dict[str, int]:
    return {_normalizar_texto(cell.value): cell.column for cell in sheet[1] if cell.value not in (None, "")}


def _ensure_columns(sheet, columns: list[str]) -> dict[str, int]:
    header_map = _header_map_from_sheet(sheet)
    for column in columns:
        normalized = _normalizar_texto(column)
        if normalized in header_map:
            continue
        next_column = sheet.max_column + 1
        sheet.cell(row=1, column=next_column, value=column)
        header_map[normalized] = next_column
    return header_map


def _date_to_text(value: date | datetime) -> str:
    if isinstance(value, datetime):
        value = value.date()
    return value.strftime("%Y-%m-%d")


def _safe_save_workbook(workbook, workbook_path: str | Path) -> None:
    workbook_file = Path(workbook_path)
    temp_file = workbook_file.with_name(f".{workbook_file.stem}.tmp{workbook_file.suffix}")
    backup_file = workbook_file.with_name(f".{workbook_file.stem}.bak{workbook_file.suffix}")
    try:
        if temp_file.exists():
            temp_file.unlink()
        workbook.save(temp_file)
        if workbook_file.exists():
            shutil.copy2(workbook_file, backup_file)
        os.replace(temp_file, workbook_file)
    except Exception as exc:
        raise RuntimeError(
            f"Falha ao salvar a planilha de forma segura: {workbook_file}. "
            "Feche o arquivo no Excel/OneDrive e tente novamente."
        ) from exc
    finally:
        try:
            if temp_file.exists():
                temp_file.unlink()
        except Exception:
            pass
        try:
            if backup_file.exists():
                backup_file.unlink()
        except Exception:
            pass


def registrar_link_d0_excel(
    path: str | Path,
    id_cliente: str,
    contrato_referencia: str,
    valor_link: float,
    link_pagamento: str,
    processed_at: date | datetime | None = None,
    usuario: str = "",
) -> None:
    workbook_path = Path(path)
    if not workbook_path.exists():
        raise FileNotFoundError(f"Planilha nao encontrada: {workbook_path}")

    processed_at = processed_at or datetime.now()
    process_date = processed_at.date() if isinstance(processed_at, datetime) else processed_at
    process_datetime_text = processed_at.strftime("%Y-%m-%d %H:%M:%S") if isinstance(processed_at, datetime) else f"{processed_at:%Y-%m-%d} 00:00:00"
    workbook = load_workbook(workbook_path)
    try:
        if SHEET_CLIENTES not in workbook.sheetnames:
            raise ValueError(f"Aba obrigatoria ausente: {SHEET_CLIENTES}")
        if SHEET_HISTORICO not in workbook.sheetnames:
            workbook.create_sheet(SHEET_HISTORICO)

        clientes_sheet = workbook[SHEET_CLIENTES]
        historico_sheet = workbook[SHEET_HISTORICO]
        clientes_headers = _ensure_columns(
            clientes_sheet,
            ["ID_CLIENTE", "LINK_D0", "VALOR_LINK_D0", "DATA_LINK_D0", "CONTRATO_REFERENCIA_LINK_D0"],
        )
        historico_headers = _ensure_columns(
            historico_sheet,
            ["DATA_HORA", "ID_CLIENTE", "DOCUMENTO", "CLIENTE", "CONTRATO", "ETAPA", "ACAO", "RESULTADO", "DETALHE", "USUARIO"],
        )

        encontrado = False
        for row in range(2, clientes_sheet.max_row + 1):
            row_id_cliente = str(clientes_sheet.cell(row=row, column=clientes_headers["ID_CLIENTE"]).value or "").strip()
            if row_id_cliente != id_cliente:
                continue
            clientes_sheet.cell(row=row, column=clientes_headers["LINK_D0"], value=link_pagamento)
            clientes_sheet.cell(row=row, column=clientes_headers["VALOR_LINK_D0"], value=valor_link)
            clientes_sheet.cell(row=row, column=clientes_headers["DATA_LINK_D0"], value=_date_to_text(process_date))
            clientes_sheet.cell(row=row, column=clientes_headers["CONTRATO_REFERENCIA_LINK_D0"], value=contrato_referencia)
            encontrado = True
            break
        if not encontrado:
            raise ValueError(f"Cliente nao encontrado para registrar link D0: {id_cliente}")

        historico_row = historico_sheet.max_row + 1
        values = {
            "DATA_HORA": process_datetime_text,
            "ID_CLIENTE": id_cliente,
            "DOCUMENTO": "",
            "CLIENTE": "",
            "CONTRATO": contrato_referencia,
            "ETAPA": "D0",
            "ACAO": "LINK_D0",
            "RESULTADO": "LINK_GERADO",
            "DETALHE": f"Link D0 gerado. Valor R$ {_format_brl(valor_link)}. Link: {link_pagamento}",
            "USUARIO": usuario,
        }
        for column, value in values.items():
            historico_sheet.cell(row=historico_row, column=historico_headers[column], value=value)

        _safe_save_workbook(workbook, workbook_path)
    finally:
        workbook.close()


def registrar_processamento_d0_excel(
    path: str | Path,
    emails: list[EmailD0Pedagio],
    conta_envio: str,
    enviado: bool,
    processed_at: date | datetime | None = None,
    registrar_link: bool = False,
    usuario: str = "",
) -> int:
    if not emails:
        return 0

    workbook_path = Path(path)
    if not workbook_path.exists():
        raise FileNotFoundError(f"Planilha nao encontrada: {workbook_path}")

    processed_at = processed_at or datetime.now()
    process_date = processed_at.date() if isinstance(processed_at, datetime) else processed_at
    process_datetime_text = processed_at.strftime("%Y-%m-%d %H:%M:%S") if isinstance(processed_at, datetime) else f"{processed_at:%Y-%m-%d} 00:00:00"
    status_email = "EMAIL_ENVIADO" if enviado else "RASCUNHO_D0_CRIADO"
    next_action = process_date + timedelta(days=LINK_EXPIRATION_DAYS)
    emails_by_id = {email.id_cliente: email for email in emails}

    workbook = load_workbook(workbook_path)
    try:
        if SHEET_CLIENTES not in workbook.sheetnames:
            raise ValueError(f"Aba obrigatoria ausente: {SHEET_CLIENTES}")
        if SHEET_CONTRATOS not in workbook.sheetnames:
            raise ValueError(f"Aba obrigatoria ausente: {SHEET_CONTRATOS}")
        if SHEET_HISTORICO not in workbook.sheetnames:
            workbook.create_sheet(SHEET_HISTORICO)

        clientes_sheet = workbook[SHEET_CLIENTES]
        contratos_sheet = workbook[SHEET_CONTRATOS]
        historico_sheet = workbook[SHEET_HISTORICO]

        clientes_headers = _ensure_columns(
            clientes_sheet,
            [
                "STATUS",
                "ETAPA",
                "PROXIMA_ACAO",
                "OBS",
                "LINK_D0",
                "VALOR_LINK_D0",
                "DATA_LINK_D0",
                "CONTRATO_REFERENCIA_LINK_D0",
            ],
        )
        contratos_headers = _ensure_columns(contratos_sheet, ["STATUS_EMAIL_D0", "DATA_EMAIL_D0", "CONTA_EMAIL_D0"])
        historico_headers = _ensure_columns(
            historico_sheet,
            ["DATA_HORA", "ID_CLIENTE", "DOCUMENTO", "CLIENTE", "CONTRATO", "ETAPA", "ACAO", "RESULTADO", "DETALHE", "USUARIO"],
        )

        id_cliente_col = clientes_headers.get("ID_CLIENTE")
        if not id_cliente_col:
            raise ValueError("Coluna obrigatoria ausente em CLIENTES: ID_CLIENTE")
        contrato_id_col = contratos_headers.get("ID_CLIENTE")
        if not contrato_id_col:
            raise ValueError("Coluna obrigatoria ausente em CONTRATOS: ID_CLIENTE")

        atualizados = 0
        for row in range(2, clientes_sheet.max_row + 1):
            id_cliente = str(clientes_sheet.cell(row=row, column=id_cliente_col).value or "").strip()
            email = emails_by_id.get(id_cliente)
            if email is None:
                continue
            clientes_sheet.cell(row=row, column=clientes_headers["STATUS"], value=status_email)
            if enviado:
                clientes_sheet.cell(row=row, column=clientes_headers["ETAPA"], value="D0+2")
                clientes_sheet.cell(row=row, column=clientes_headers["PROXIMA_ACAO"], value=_date_to_text(next_action))
            else:
                clientes_sheet.cell(row=row, column=clientes_headers["OBS"], value=f"Rascunho D0 criado em {_date_to_text(process_date)}")
            if registrar_link and email.link_pagamento:
                clientes_sheet.cell(row=row, column=clientes_headers["LINK_D0"], value=email.link_pagamento)
                clientes_sheet.cell(row=row, column=clientes_headers["VALOR_LINK_D0"], value=email.valor_total)
                clientes_sheet.cell(row=row, column=clientes_headers["DATA_LINK_D0"], value=_date_to_text(process_date))
                clientes_sheet.cell(
                    row=row,
                    column=clientes_headers["CONTRATO_REFERENCIA_LINK_D0"],
                    value=email.contrato_referencia,
                )
            atualizados += 1

        contratos_por_cliente: dict[str, int] = {}
        for row in range(2, contratos_sheet.max_row + 1):
            id_cliente = str(contratos_sheet.cell(row=row, column=contrato_id_col).value or "").strip()
            if id_cliente not in emails_by_id:
                continue
            status_cartao_col = contratos_headers.get("STATUS_CARTAO")
            if status_cartao_col and _normalizar_texto(contratos_sheet.cell(row=row, column=status_cartao_col).value) != "PENDENTE":
                continue
            contratos_sheet.cell(row=row, column=contratos_headers["STATUS_EMAIL_D0"], value=status_email)
            contratos_sheet.cell(row=row, column=contratos_headers["DATA_EMAIL_D0"], value=_date_to_text(process_date))
            contratos_sheet.cell(row=row, column=contratos_headers["CONTA_EMAIL_D0"], value=conta_envio)
            contratos_por_cliente[id_cliente] = contratos_por_cliente.get(id_cliente, 0) + 1

        for email in emails:
            if registrar_link and email.link_pagamento:
                row = historico_sheet.max_row + 1
                link_values = {
                    "DATA_HORA": process_datetime_text,
                    "ID_CLIENTE": email.id_cliente,
                    "DOCUMENTO": "",
                    "CLIENTE": email.nome,
                    "CONTRATO": email.contrato_referencia,
                    "ETAPA": "D0",
                    "ACAO": "LINK_D0",
                    "RESULTADO": "LINK_GERADO",
                    "DETALHE": f"Link D0 gerado. Valor R$ {_format_brl(email.valor_total)}. Link: {email.link_pagamento}",
                    "USUARIO": usuario,
                }
                for column, value in link_values.items():
                    historico_sheet.cell(row=row, column=historico_headers[column], value=value)

            row = historico_sheet.max_row + 1
            detalhe = (
                f"{status_email}; destinatario={email.destinatario}; "
                f"contratos={contratos_por_cliente.get(email.id_cliente, email.total_contratos)}; "
                f"valor=R$ {_format_brl(email.valor_total)}"
            )
            values = {
                "DATA_HORA": process_datetime_text,
                "ID_CLIENTE": email.id_cliente,
                "DOCUMENTO": "",
                "CLIENTE": email.nome,
                "CONTRATO": "",
                "ETAPA": "D0",
                "ACAO": "EMAIL_D0",
                "RESULTADO": status_email,
                "DETALHE": detalhe,
                "USUARIO": conta_envio,
            }
            for column, value in values.items():
                historico_sheet.cell(row=row, column=historico_headers[column], value=value)

        _safe_save_workbook(workbook, workbook_path)
        return atualizados
    finally:
        workbook.close()


def _status_planilha_d2(status: str) -> str:
    normalized = _normalizar_texto(status)
    if normalized == "COBRADO":
        return "COBRADO"
    if normalized in {"NAO COBRADO", "NAO_COBRADO", "SEM CARTAO", "CARTAO_RECUSADO"}:
        return "LINK_PAGAMENTO_PENDENTE"
    if normalized == "INTERROMPIDO":
        return "INTERROMPIDO"
    return "ERRO_D0_2"


def registrar_processamento_d2_excel(
    path: str | Path,
    resultados: list[ResultadoD2Pedagio],
    processed_at: date | datetime | None = None,
    usuario: str = "",
) -> int:
    if not resultados:
        return 0

    workbook_path = Path(path)
    if not workbook_path.exists():
        raise FileNotFoundError(f"Planilha nao encontrada: {workbook_path}")

    processed_at = processed_at or datetime.now()
    process_date = processed_at.date() if isinstance(processed_at, datetime) else processed_at
    process_datetime_text = processed_at.strftime("%Y-%m-%d %H:%M:%S") if isinstance(processed_at, datetime) else f"{processed_at:%Y-%m-%d} 00:00:00"
    resultados_por_chave = {(item.id_cliente, item.contrato): item for item in resultados}

    workbook = load_workbook(workbook_path)
    try:
        if SHEET_CLIENTES not in workbook.sheetnames:
            raise ValueError(f"Aba obrigatoria ausente: {SHEET_CLIENTES}")
        if SHEET_CONTRATOS not in workbook.sheetnames:
            raise ValueError(f"Aba obrigatoria ausente: {SHEET_CONTRATOS}")
        if SHEET_HISTORICO not in workbook.sheetnames:
            workbook.create_sheet(SHEET_HISTORICO)

        clientes_sheet = workbook[SHEET_CLIENTES]
        contratos_sheet = workbook[SHEET_CONTRATOS]
        historico_sheet = workbook[SHEET_HISTORICO]

        clientes_headers = _ensure_columns(clientes_sheet, ["ID_CLIENTE", "STATUS", "ETAPA", "PROXIMA_ACAO", "OBS"])
        contratos_headers = _ensure_columns(
            contratos_sheet,
            [
                "ID_CLIENTE",
                "CONTRATO",
                "STATUS_CARTAO",
                "DATA_D0_2",
                "CARTOES_ENCONTRADOS_D0_2",
                "CARTOES_TENTADOS_D0_2",
                "DETALHE_D0_2",
            ],
        )
        historico_headers = _ensure_columns(
            historico_sheet,
            ["DATA_HORA", "ID_CLIENTE", "DOCUMENTO", "CLIENTE", "CONTRATO", "ETAPA", "ACAO", "RESULTADO", "DETALHE", "USUARIO"],
        )

        atualizados = 0
        clientes_afetados: set[str] = set()
        for row in range(2, contratos_sheet.max_row + 1):
            id_cliente = str(contratos_sheet.cell(row=row, column=contratos_headers["ID_CLIENTE"]).value or "").strip()
            contrato = str(contratos_sheet.cell(row=row, column=contratos_headers["CONTRATO"]).value or "").strip()
            resultado = resultados_por_chave.get((id_cliente, contrato))
            if resultado is None:
                continue

            status_final = _status_planilha_d2(resultado.status)
            contratos_sheet.cell(row=row, column=contratos_headers["STATUS_CARTAO"], value=status_final)
            contratos_sheet.cell(row=row, column=contratos_headers["DATA_D0_2"], value=_date_to_text(process_date))
            contratos_sheet.cell(row=row, column=contratos_headers["CARTOES_ENCONTRADOS_D0_2"], value=resultado.cartoes_encontrados)
            contratos_sheet.cell(row=row, column=contratos_headers["CARTOES_TENTADOS_D0_2"], value=resultado.cartoes_tentados)
            contratos_sheet.cell(row=row, column=contratos_headers["DETALHE_D0_2"], value=resultado.detalhe)
            clientes_afetados.add(id_cliente)
            atualizados += 1

            historico_row = historico_sheet.max_row + 1
            values = {
                "DATA_HORA": process_datetime_text,
                "ID_CLIENTE": id_cliente,
                "DOCUMENTO": "",
                "CLIENTE": "",
                "CONTRATO": contrato,
                "ETAPA": "D0+2",
                "ACAO": "COBRANCA_CARTAO",
                "RESULTADO": status_final,
                "DETALHE": resultado.detalhe,
                "USUARIO": usuario,
            }
            for column, value in values.items():
                historico_sheet.cell(row=historico_row, column=historico_headers[column], value=value)

        status_por_cliente: dict[str, list[str]] = {id_cliente: [] for id_cliente in clientes_afetados}
        for row in range(2, contratos_sheet.max_row + 1):
            id_cliente = str(contratos_sheet.cell(row=row, column=contratos_headers["ID_CLIENTE"]).value or "").strip()
            if id_cliente not in status_por_cliente:
                continue
            status_por_cliente[id_cliente].append(
                _normalizar_texto(contratos_sheet.cell(row=row, column=contratos_headers["STATUS_CARTAO"]).value)
            )

        for row in range(2, clientes_sheet.max_row + 1):
            id_cliente = str(clientes_sheet.cell(row=row, column=clientes_headers["ID_CLIENTE"]).value or "").strip()
            statuses = status_por_cliente.get(id_cliente)
            if not statuses:
                continue
            if any(status == "PENDENTE" for status in statuses):
                clientes_sheet.cell(row=row, column=clientes_headers["STATUS"], value="ATIVO")
                clientes_sheet.cell(row=row, column=clientes_headers["ETAPA"], value="D0+2")
                clientes_sheet.cell(row=row, column=clientes_headers["PROXIMA_ACAO"], value=_date_to_text(process_date))
            elif any(status == "LINK_PAGAMENTO_PENDENTE" for status in statuses):
                clientes_sheet.cell(row=row, column=clientes_headers["STATUS"], value="LINK_PAGAMENTO_PENDENTE")
                clientes_sheet.cell(row=row, column=clientes_headers["ETAPA"], value="D0+2_LINK")
                clientes_sheet.cell(row=row, column=clientes_headers["PROXIMA_ACAO"], value=_date_to_text(process_date))
            elif any(status in {"ERRO_D0_2", "INTERROMPIDO"} for status in statuses):
                clientes_sheet.cell(row=row, column=clientes_headers["STATUS"], value="ERRO_D0_2")
                clientes_sheet.cell(row=row, column=clientes_headers["ETAPA"], value="D0+2")
                clientes_sheet.cell(row=row, column=clientes_headers["PROXIMA_ACAO"], value=_date_to_text(process_date))
            elif statuses and all(status == "COBRADO" for status in statuses):
                clientes_sheet.cell(row=row, column=clientes_headers["STATUS"], value="COBRADO")
                clientes_sheet.cell(row=row, column=clientes_headers["ETAPA"], value="PAGO")
                clientes_sheet.cell(row=row, column=clientes_headers["PROXIMA_ACAO"], value="")

        _safe_save_workbook(workbook, workbook_path)
        return atualizados
    finally:
        workbook.close()


def registrar_link_d2_excel(
    path: str | Path,
    id_cliente: str,
    contratos: list[str],
    valor_link: float,
    link_pagamento: str,
    processed_at: date | datetime | None = None,
    usuario: str = "",
) -> int:
    if not contratos:
        return 0

    workbook_path = Path(path)
    if not workbook_path.exists():
        raise FileNotFoundError(f"Planilha nao encontrada: {workbook_path}")

    processed_at = processed_at or datetime.now()
    process_date = processed_at.date() if isinstance(processed_at, datetime) else processed_at
    process_datetime_text = processed_at.strftime("%Y-%m-%d %H:%M:%S") if isinstance(processed_at, datetime) else f"{processed_at:%Y-%m-%d} 00:00:00"
    contratos_residuais = {str(contrato or "").strip() for contrato in contratos if str(contrato or "").strip()}
    if not contratos_residuais:
        return 0

    workbook = load_workbook(workbook_path)
    try:
        if SHEET_CLIENTES not in workbook.sheetnames:
            raise ValueError(f"Aba obrigatoria ausente: {SHEET_CLIENTES}")
        if SHEET_CONTRATOS not in workbook.sheetnames:
            raise ValueError(f"Aba obrigatoria ausente: {SHEET_CONTRATOS}")
        if SHEET_HISTORICO not in workbook.sheetnames:
            workbook.create_sheet(SHEET_HISTORICO)

        clientes_sheet = workbook[SHEET_CLIENTES]
        contratos_sheet = workbook[SHEET_CONTRATOS]
        historico_sheet = workbook[SHEET_HISTORICO]

        clientes_headers = _ensure_columns(
            clientes_sheet,
            [
                "ID_CLIENTE",
                "STATUS",
                "ETAPA",
                "PROXIMA_ACAO",
                "LINK_D0_2",
                "VALOR_LINK_D0_2",
                "DATA_LINK_D0_2",
            ],
        )
        contratos_headers = _ensure_columns(
            contratos_sheet,
            ["ID_CLIENTE", "CONTRATO", "STATUS_CARTAO", "LINK_D0_2", "VALOR_LINK_D0_2", "DATA_LINK_D0_2"],
        )
        historico_headers = _ensure_columns(
            historico_sheet,
            ["DATA_HORA", "ID_CLIENTE", "DOCUMENTO", "CLIENTE", "CONTRATO", "ETAPA", "ACAO", "RESULTADO", "DETALHE", "USUARIO"],
        )

        atualizados = 0
        for row in range(2, contratos_sheet.max_row + 1):
            row_id_cliente = str(contratos_sheet.cell(row=row, column=contratos_headers["ID_CLIENTE"]).value or "").strip()
            contrato = str(contratos_sheet.cell(row=row, column=contratos_headers["CONTRATO"]).value or "").strip()
            if row_id_cliente != id_cliente or contrato not in contratos_residuais:
                continue
            contratos_sheet.cell(row=row, column=contratos_headers["STATUS_CARTAO"], value="LINK_GERADO")
            contratos_sheet.cell(row=row, column=contratos_headers["LINK_D0_2"], value=link_pagamento)
            contratos_sheet.cell(row=row, column=contratos_headers["VALOR_LINK_D0_2"], value=valor_link)
            contratos_sheet.cell(row=row, column=contratos_headers["DATA_LINK_D0_2"], value=_date_to_text(process_date))
            atualizados += 1

        for row in range(2, clientes_sheet.max_row + 1):
            row_id_cliente = str(clientes_sheet.cell(row=row, column=clientes_headers["ID_CLIENTE"]).value or "").strip()
            if row_id_cliente != id_cliente:
                continue
            clientes_sheet.cell(row=row, column=clientes_headers["STATUS"], value="LINK_GERADO_AGUARDANDO_EMAIL")
            clientes_sheet.cell(row=row, column=clientes_headers["ETAPA"], value="D0+2_LINK")
            clientes_sheet.cell(
                row=row,
                column=clientes_headers["PROXIMA_ACAO"],
                value=_date_to_text(process_date + timedelta(days=LINK_EXPIRATION_DAYS)),
            )
            clientes_sheet.cell(row=row, column=clientes_headers["LINK_D0_2"], value=link_pagamento)
            clientes_sheet.cell(row=row, column=clientes_headers["VALOR_LINK_D0_2"], value=valor_link)
            clientes_sheet.cell(row=row, column=clientes_headers["DATA_LINK_D0_2"], value=_date_to_text(process_date))
            break

        historico_row = historico_sheet.max_row + 1
        detalhe = (
            f"Link residual gerado para {len(contratos_residuais)} contrato(s). "
            f"Valor R$ {_format_brl(valor_link)}. Link: {link_pagamento}"
        )
        values = {
            "DATA_HORA": process_datetime_text,
            "ID_CLIENTE": id_cliente,
            "DOCUMENTO": "",
            "CLIENTE": "",
            "CONTRATO": ", ".join(sorted(contratos_residuais)),
            "ETAPA": "D0+2_LINK",
            "ACAO": "LINK_D0_2",
            "RESULTADO": "LINK_GERADO",
            "DETALHE": detalhe,
            "USUARIO": usuario,
        }
        for column, value in values.items():
            historico_sheet.cell(row=historico_row, column=historico_headers[column], value=value)

        _safe_save_workbook(workbook, workbook_path)
        return atualizados
    finally:
        workbook.close()


class CoralNavigationAbortError(RuntimeError):
    """Erro de navegação considerado irrecuperável para a unidade atual."""


class RoboCobrancaPedagiosApp(ctk.CTk):
    def __init__(self) -> None:
        super().__init__()
        ctk.set_appearance_mode("light")
        ctk.set_default_color_theme("blue")

        self.title(APP_TITLE)
        self.geometry(APP_GEOMETRY)
        self.minsize(1120, 760)
        self.configure(fg_color=MAIN_BG)

        saved_config = carregar_config_interface()
        saved_user = str(saved_config.get("usuario_coral") or "")
        saved_login = bool(saved_config.get("salvar_login_coral"))
        saved_password = carregar_senha_coral_windows(saved_user) if saved_login else ""

        self.status_var = ctk.StringVar(value="Selecione a planilha compartilhada para carregar a fila.")
        self.connection_var = ctk.StringVar(value="Excel nao conectado")
        self.file_path_var = ctk.StringVar(value=str(saved_config.get("caminho_excel") or DEFAULT_EXCEL_URL))
        self.outlook_account_var = ctk.StringVar(value=str(saved_config.get("conta_envio") or DEFAULT_OUTLOOK_ACCOUNT))
        self.canal_cobranca_var = ctk.StringVar(
            value=str(saved_config.get("canal_cobranca") or DEFAULT_CANAL_COBRANCA)
        )
        self.execution_limit_var = ctk.StringVar(value=str(saved_config.get("limite_execucao") or "10"))
        self.d0_limit_var = self.execution_limit_var
        self.d2_limit_var = self.execution_limit_var
        self.coral_user_var = ctk.StringVar(value=saved_user)
        self.coral_password_var = ctk.StringVar(value=saved_password)
        self.save_coral_login_var = ctk.BooleanVar(value=saved_login)
        self.coral_headless_var = ctk.BooleanVar(value=bool(saved_config.get("coral_headless", DEFAULT_CORAL_HEADLESS)))
        self.action_var = ctk.StringVar(value=ACTION_LABELS[ACTION_UPDATE_QUEUE])
        self.action_rule_var = ctk.StringVar(value=ACTION_RULES[ACTION_UPDATE_QUEUE])
        self.logo_image = self._load_logo()
        self.driver = None
        self.processing_thread: threading.Thread | None = None
        self.pause_requested = threading.Event()
        self.stop_requested = threading.Event()
        self.current_processing_workbook_path: Path | None = None
        self._main_thread = threading.current_thread()
        self.pending_local_root: Path | None = None
        self.popup_edicao_tratado = False

        self.metrics = {
            "total": ctk.StringVar(value="0"),
            "contracts": ctk.StringVar(value="0"),
            "contracts_pending": ctk.StringVar(value="0"),
            "d0": ctk.StringVar(value="0"),
            "d2": ctk.StringVar(value="0"),
            "d45": ctk.StringVar(value="0"),
            "d7": ctk.StringVar(value="0"),
            "processing": ctk.StringVar(value="0"),
            "paid": ctk.StringVar(value="0"),
            "errors": ctk.StringVar(value="0"),
        }

        self._build_layout()
        self.action_commands = self._build_action_commands()
        self.protocol("WM_DELETE_WINDOW", self._on_close)
        self._log("Interface criada. Base padrao de pedagios configurada.")
        self.after(500, self._start_initial_background_loads)

    def _load_logo(self):
        candidates: list[Path] = []
        env_logo = os.environ.get("FOCO_LOGO_PNG", "").strip()
        env_assets = os.environ.get("FOCO_ASSETS_DIR", "").strip()
        if env_logo:
            candidates.append(Path(env_logo))
        if env_assets:
            candidates.append(Path(env_assets) / "logo.png")

        base_dir = Path(__file__).resolve().parent
        candidates.append(base_dir.parent / "assets" / "logo.png")
        candidates.append(Path.cwd() / "DESENVOLVIMENTO" / "assets" / "logo.png")
        candidates.append(Path.cwd() / "assets" / "logo.png")

        for candidate in candidates:
            try:
                if candidate.exists():
                    image = Image.open(candidate)
                    return ctk.CTkImage(light_image=image, dark_image=image, size=(112, 68))
            except Exception:
                continue
        return None

    def _build_layout(self) -> None:
        container = ctk.CTkScrollableFrame(self, fg_color="transparent", corner_radius=0)
        container.pack(fill="both", expand=True, padx=18, pady=18)
        container.grid_columnconfigure(0, weight=1)

        self._build_header(container)
        self._build_workspace(container)
        self._build_metrics_section(container)
        self._build_settings_section(container)
        self._build_log_section(container)

    def _build_header(self, parent) -> None:
        header = ctk.CTkFrame(parent, fg_color=CARD_BG, corner_radius=14, border_width=1, border_color=CARD_BORDER)
        header.grid(row=0, column=0, sticky="ew", pady=(0, 12))
        header.grid_columnconfigure(1, weight=1)

        brand = ctk.CTkFrame(header, fg_color="transparent")
        brand.grid(row=0, column=0, padx=18, pady=16, sticky="nw")
        if self.logo_image is not None:
            ctk.CTkLabel(brand, text="", image=self.logo_image).pack(anchor="w")
        else:
            ctk.CTkLabel(brand, text="FOCO", text_color=PRIMARY_TEXT, font=("Segoe UI", 26, "bold")).pack(anchor="w")
            ctk.CTkLabel(brand, text="aluguel de carros", text_color="#b42318", font=("Segoe UI", 10, "bold")).pack(anchor="w")

        texts = ctk.CTkFrame(header, fg_color="transparent")
        texts.grid(row=0, column=1, sticky="ew", padx=(0, 18), pady=16)
        ctk.CTkLabel(
            texts,
            text="Cobranca de Pedagios",
            text_color=NAVY_TEXT,
            font=("Segoe UI", 28, "bold"),
        ).pack(anchor="w")
        ctk.CTkLabel(
            texts,
            text="Painel operacional para envio D0, tentativa no Coral e links de pagamento.",
            text_color=MUTED_TEXT,
            font=("Segoe UI", 14),
        ).pack(anchor="w", pady=(8, 0))

        status = ctk.CTkFrame(header, fg_color=SOFT_RED, corner_radius=10, border_width=1, border_color="#ffd0cc")
        status.grid(row=0, column=2, padx=18, pady=16, sticky="e")
        ctk.CTkLabel(status, text="BASE", text_color=PRIMARY_TEXT, font=("Segoe UI", 10, "bold")).pack(
            padx=16, pady=(10, 0)
        )
        ctk.CTkLabel(status, textvariable=self.connection_var, text_color=TITLE_TEXT, font=("Segoe UI", 13, "bold")).pack(
            padx=16, pady=(2, 10)
        )

    def _build_connection_section(self, parent) -> None:
        section = self._create_section(parent, 1, "Configuracao da base")
        body = ctk.CTkFrame(section, fg_color="transparent")
        body.pack(fill="x", padx=18, pady=(0, 18))
        body.grid_columnconfigure(0, weight=1)
        body.grid_columnconfigure(1, weight=0)
        body.grid_columnconfigure(2, weight=0)

        wrapper = ctk.CTkFrame(body, fg_color="transparent")
        wrapper.grid(row=0, column=0, sticky="ew", padx=(0, 10), pady=6)
        ctk.CTkLabel(wrapper, text="Planilha de trabalho", text_color=MUTED_TEXT, font=("Segoe UI", 12, "bold")).pack(
            anchor="w"
        )
        self.file_path_entry = ctk.CTkEntry(
            wrapper,
            textvariable=self.file_path_var,
            fg_color="#ffffff",
            text_color=TITLE_TEXT,
            border_color="#9ca3af",
            height=38,
            corner_radius=8,
        )
        self.file_path_entry.pack(anchor="w", fill="x", pady=(4, 0))
        ctk.CTkLabel(
            wrapper,
            text="Selecione uma base .xlsx",
            text_color=TITLE_TEXT,
            font=("Segoe UI", 12),
            wraplength=760,
            justify="left",
        ).pack(anchor="w", pady=(4, 0))

        buttons = ctk.CTkFrame(body, fg_color="transparent")
        buttons.grid(row=0, column=1, columnspan=2, sticky="e", padx=(0, 0), pady=(24, 6))
        ctk.CTkButton(
            buttons,
            text="Selecionar Excel",
            fg_color="#ffffff",
            hover_color="#fff1ef",
            border_width=1,
            border_color=BUTTON_BG,
            text_color=BUTTON_BG,
            height=38,
            command=self._select_excel_file,
        ).pack(side="left", padx=(0, 10))
        ctk.CTkButton(
            buttons,
            text="Atualizar fila",
            fg_color=BUTTON_BG,
            hover_color=BUTTON_ACTIVE_BG,
            text_color="#ffffff",
            height=38,
            command=self._refresh_queue,
        ).pack(side="left")

    def _select_excel_file(self) -> None:
        current = self.file_path_var.get().strip()
        initial_dir = str(Path.home())
        if current and not current.lower().startswith("http"):
            current_path = Path(current)
            if current_path.exists():
                initial_dir = str(current_path.parent)
            elif current_path.parent.exists():
                initial_dir = str(current_path.parent)

        selected = filedialog.askopenfilename(
            title="Selecionar planilha operacional",
            initialdir=initial_dir,
            filetypes=[("Planilhas Excel", "*.xlsx *.xlsm *.xls"), ("Todos os arquivos", "*.*")],
        )
        if not selected:
            return
        self.file_path_var.set(selected)
        self._persist_interface_config()
        self._refresh_queue()

    def _build_settings_section(self, parent) -> None:
        section = ctk.CTkFrame(parent, fg_color="transparent")
        section.grid(row=3, column=0, sticky="ew", pady=(0, 12))
        section.grid_columnconfigure(0, weight=1)
        section.grid_columnconfigure(1, weight=1)

        base_card = ctk.CTkFrame(section, fg_color=CARD_BG, corner_radius=14, border_width=1, border_color=CARD_BORDER)
        base_card.grid(row=0, column=0, sticky="nsew", padx=(0, 8))
        base_card.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(base_card, text="Base e Outlook", text_color=NAVY_TEXT, font=("Segoe UI", 17, "bold")).grid(
            row=0, column=0, columnspan=2, sticky="w", padx=16, pady=(14, 10)
        )

        ctk.CTkLabel(base_card, text="Planilha de trabalho", text_color=MUTED_TEXT, font=("Segoe UI", 11, "bold")).grid(
            row=1, column=0, columnspan=2, sticky="w", padx=16
        )
        self.file_path_entry = ctk.CTkEntry(
            base_card,
            textvariable=self.file_path_var,
            fg_color="#ffffff",
            text_color=TITLE_TEXT,
            border_color=FIELD_BORDER,
            height=36,
            corner_radius=8,
        )
        self.file_path_entry.grid(row=2, column=0, sticky="ew", padx=(16, 8), pady=(4, 10))
        ctk.CTkButton(
            base_card,
            text="Selecionar",
            fg_color="#ffffff",
            hover_color="#fff1ef",
            border_width=1,
            border_color=BUTTON_BG,
            text_color=BUTTON_BG,
            height=36,
            width=108,
            corner_radius=8,
            command=self._select_excel_file,
        ).grid(row=2, column=1, sticky="e", padx=(0, 16), pady=(4, 10))

        ctk.CTkLabel(base_card, text="Conta de envio", text_color=MUTED_TEXT, font=("Segoe UI", 11, "bold")).grid(
            row=3, column=0, columnspan=2, sticky="w", padx=16
        )
        self.outlook_account_menu = ctk.CTkComboBox(
            base_card,
            values=[DEFAULT_OUTLOOK_ACCOUNT],
            variable=self.outlook_account_var,
            fg_color="#ffffff",
            border_color=FIELD_BORDER,
            button_color="#f3f4f6",
            button_hover_color="#e5e7eb",
            text_color=TITLE_TEXT,
            dropdown_fg_color="#ffffff",
            dropdown_text_color=TITLE_TEXT,
            dropdown_hover_color="#fff1ef",
            height=36,
            corner_radius=8,
        )
        self.outlook_account_menu.grid(row=4, column=0, columnspan=2, sticky="ew", padx=16, pady=(4, 14))

        coral_card = ctk.CTkFrame(section, fg_color=CARD_BG, corner_radius=14, border_width=1, border_color=CARD_BORDER)
        coral_card.grid(row=0, column=1, sticky="nsew", padx=(8, 0))
        coral_card.grid_columnconfigure(0, weight=1)
        coral_card.grid_columnconfigure(1, weight=1)
        ctk.CTkLabel(coral_card, text="Coral", text_color=NAVY_TEXT, font=("Segoe UI", 17, "bold")).grid(
            row=0, column=0, columnspan=2, sticky="w", padx=16, pady=(14, 10)
        )

        ctk.CTkLabel(coral_card, text="Usuario", text_color=MUTED_TEXT, font=("Segoe UI", 11, "bold")).grid(
            row=1, column=0, sticky="w", padx=(16, 8)
        )
        ctk.CTkLabel(coral_card, text="Senha", text_color=MUTED_TEXT, font=("Segoe UI", 11, "bold")).grid(
            row=1, column=1, sticky="w", padx=(8, 16)
        )
        ctk.CTkEntry(
            coral_card,
            textvariable=self.coral_user_var,
            height=36,
            corner_radius=8,
            border_color=FIELD_BORDER,
        ).grid(row=2, column=0, sticky="ew", padx=(16, 8), pady=(4, 10))
        ctk.CTkEntry(
            coral_card,
            textvariable=self.coral_password_var,
            height=36,
            corner_radius=8,
            border_color=FIELD_BORDER,
            show="*",
        ).grid(row=2, column=1, sticky="ew", padx=(8, 16), pady=(4, 10))

        options = ctk.CTkFrame(coral_card, fg_color=PANEL_BG, corner_radius=8, border_width=1, border_color=CARD_BORDER)
        options.grid(row=3, column=0, columnspan=2, sticky="ew", padx=16, pady=(0, 14))
        options.grid_columnconfigure(0, weight=1)
        options.grid_columnconfigure(1, weight=1)
        ctk.CTkCheckBox(
            options,
            text="Salvar login",
            variable=self.save_coral_login_var,
            command=self._on_save_coral_login_changed,
            fg_color=BUTTON_BG,
            hover_color=BUTTON_ACTIVE_BG,
            text_color=TITLE_TEXT,
            checkbox_width=18,
            checkbox_height=18,
        ).grid(row=0, column=0, sticky="w", padx=12, pady=10)
        ctk.CTkCheckBox(
            options,
            text="Modo invisivel",
            variable=self.coral_headless_var,
            command=self._on_coral_headless_changed,
            fg_color=BUTTON_BG,
            hover_color=BUTTON_ACTIVE_BG,
            text_color=TITLE_TEXT,
            checkbox_width=18,
            checkbox_height=18,
        ).grid(row=0, column=1, sticky="w", padx=12, pady=10)

    def _build_outlook_section(self, parent) -> None:
        section = self._create_section(parent, 3, "Configuracoes de execucao")
        body = ctk.CTkFrame(section, fg_color="transparent")
        body.pack(fill="x", padx=18, pady=(0, 16))
        body.grid_columnconfigure(0, weight=2)
        body.grid_columnconfigure(1, weight=1)
        body.grid_columnconfigure(2, weight=1)

        account_box = ctk.CTkFrame(body, fg_color="transparent")
        account_box.grid(row=0, column=0, sticky="ew", padx=(0, 10))
        ctk.CTkLabel(account_box, text="Conta de envio", text_color=MUTED_TEXT, font=("Segoe UI", 12, "bold")).pack(anchor="w")
        self.outlook_account_menu = ctk.CTkComboBox(
            account_box,
            values=[DEFAULT_OUTLOOK_ACCOUNT],
            variable=self.outlook_account_var,
            fg_color="#ffffff",
            border_color="#9ca3af",
            button_color="#f3f4f6",
            button_hover_color="#e5e7eb",
            text_color=TITLE_TEXT,
            dropdown_fg_color="#ffffff",
            dropdown_text_color=TITLE_TEXT,
            dropdown_hover_color="#fff1ef",
            height=38,
            corner_radius=8,
        )
        self.outlook_account_menu.pack(fill="x", pady=(4, 0))

        channel_box = ctk.CTkFrame(body, fg_color="transparent")
        channel_box.grid(row=0, column=1, sticky="ew", padx=(0, 10))
        ctk.CTkLabel(channel_box, text="Canal da cobranca", text_color=MUTED_TEXT, font=("Segoe UI", 12, "bold")).pack(anchor="w")
        self.canal_cobranca_menu = ctk.CTkComboBox(
            channel_box,
            values=[DEFAULT_CANAL_COBRANCA, CANAL_TODOS],
            variable=self.canal_cobranca_var,
            fg_color="#ffffff",
            border_color="#9ca3af",
            button_color="#f3f4f6",
            button_hover_color="#e5e7eb",
            text_color=TITLE_TEXT,
            dropdown_fg_color="#ffffff",
            dropdown_text_color=TITLE_TEXT,
            dropdown_hover_color="#fff1ef",
            height=38,
            corner_radius=8,
            state="readonly",
            command=self._on_canal_selected,
        )
        self.canal_cobranca_menu.pack(fill="x", pady=(4, 0))

        limit_box = ctk.CTkFrame(body, fg_color="transparent")
        limit_box.grid(row=0, column=2, sticky="ew", padx=(0, 10))
        ctk.CTkLabel(limit_box, text="Limite por execucao", text_color=MUTED_TEXT, font=("Segoe UI", 12, "bold")).pack(anchor="w")
        ctk.CTkEntry(limit_box, textvariable=self.execution_limit_var, height=38, corner_radius=8).pack(fill="x", pady=(4, 0))

        ctk.CTkLabel(body, text="Login Coral", text_color=TITLE_TEXT, font=("Segoe UI", 13, "bold")).grid(
            row=1, column=0, columnspan=3, sticky="w", pady=(16, 8)
        )

        coral_user_box = ctk.CTkFrame(body, fg_color="transparent")
        coral_user_box.grid(row=2, column=0, sticky="ew", padx=(0, 10))
        ctk.CTkLabel(coral_user_box, text="Usuario Coral", text_color=MUTED_TEXT, font=("Segoe UI", 12, "bold")).pack(anchor="w")
        ctk.CTkEntry(coral_user_box, textvariable=self.coral_user_var, height=38, corner_radius=8).pack(fill="x", pady=(4, 0))

        coral_password_box = ctk.CTkFrame(body, fg_color="transparent")
        coral_password_box.grid(row=2, column=1, sticky="ew", padx=(0, 10))
        ctk.CTkLabel(coral_password_box, text="Senha Coral", text_color=MUTED_TEXT, font=("Segoe UI", 12, "bold")).pack(anchor="w")
        ctk.CTkEntry(coral_password_box, textvariable=self.coral_password_var, height=38, corner_radius=8, show="*").pack(fill="x", pady=(4, 0))

        options_box = ctk.CTkFrame(body, fg_color="transparent")
        options_box.grid(row=2, column=2, sticky="sw", pady=(24, 0))
        ctk.CTkCheckBox(
            options_box,
            text="Salvar login Coral",
            variable=self.save_coral_login_var,
            command=self._on_save_coral_login_changed,
            fg_color=BUTTON_BG,
            hover_color=BUTTON_ACTIVE_BG,
            text_color=TITLE_TEXT,
            checkbox_width=18,
            checkbox_height=18,
        ).pack(anchor="w")
        ctk.CTkCheckBox(
            options_box,
            text="Modo invisivel Coral",
            variable=self.coral_headless_var,
            command=self._on_coral_headless_changed,
            fg_color=BUTTON_BG,
            hover_color=BUTTON_ACTIVE_BG,
            text_color=TITLE_TEXT,
            checkbox_width=18,
            checkbox_height=18,
        ).pack(anchor="w", pady=(10, 0))

    def _build_metrics_section(self, parent) -> None:
        section = self._create_section(parent, 2, "Fila operacional")
        grid = ctk.CTkFrame(section, fg_color="transparent")
        grid.pack(fill="x", padx=16, pady=(0, 12))
        for index in range(5):
            grid.grid_columnconfigure(index, weight=1)

        cards = [
            ("Clientes", self.metrics["total"], "#ffffff", "#98a2b3"),
            ("Contratos", self.metrics["contracts"], "#ffffff", "#98a2b3"),
            ("Pendentes", self.metrics["contracts_pending"], "#fff7f6", PRIMARY_TEXT),
            ("Etapa 1", self.metrics["d0"], "#fff7f6", PRIMARY_TEXT),
            ("Etapa 2", self.metrics["d2"], "#fff7f6", PRIMARY_TEXT),
            ("Etapa 3", self.metrics["d45"], "#ffffff", WARNING_ORANGE),
            ("Etapa 4", self.metrics["d7"], "#ffffff", WARNING_ORANGE),
            ("Processando", self.metrics["processing"], "#ffffff", "#1e40af"),
            ("Pagos", self.metrics["paid"], "#ffffff", SUCCESS_GREEN),
            ("Erros", self.metrics["errors"], "#fff7f6", PRIMARY_TEXT),
        ]

        for index, (title, variable, bg_color, accent) in enumerate(cards):
            self._build_metric_card(grid, index // 5, index % 5, title, variable, bg_color, accent)

    def _build_workspace(self, parent) -> None:
        workspace = ctk.CTkFrame(parent, fg_color="transparent")
        workspace.grid(row=1, column=0, sticky="ew", pady=(0, 12))
        workspace.grid_columnconfigure(0, weight=1)

        actions = ctk.CTkFrame(workspace, fg_color=CARD_BG, corner_radius=14, border_width=1, border_color=CARD_BORDER)
        actions.grid(row=0, column=0, sticky="ew")
        actions.grid_columnconfigure(0, weight=1)

        top = ctk.CTkFrame(actions, fg_color="transparent")
        top.pack(fill="x", padx=16, pady=(14, 8))
        top.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(top, text="Execucao", text_color=NAVY_TEXT, font=("Segoe UI", 19, "bold")).grid(
            row=0, column=0, sticky="w"
        )
        ctk.CTkLabel(top, textvariable=self.status_var, text_color=MUTED_TEXT, font=("Segoe UI", 12)).grid(
            row=0, column=1, sticky="e"
        )

        action_row = ctk.CTkFrame(actions, fg_color="transparent")
        action_row.pack(fill="x", padx=16, pady=(0, 12))
        action_row.grid_columnconfigure(0, weight=2)
        action_row.grid_columnconfigure(1, weight=1)
        action_row.grid_columnconfigure(2, weight=0)

        action_box = ctk.CTkFrame(action_row, fg_color="transparent")
        action_box.grid(row=0, column=0, sticky="ew", padx=(0, 10))
        ctk.CTkLabel(action_box, text="Acao", text_color=MUTED_TEXT, font=("Segoe UI", 11, "bold")).pack(anchor="w")
        self.action_menu = ctk.CTkComboBox(
            action_box,
            values=ACTION_OPTIONS,
            variable=self.action_var,
            fg_color="#ffffff",
            border_color=FIELD_BORDER,
            button_color="#f3f4f6",
            button_hover_color="#e5e7eb",
            text_color=TITLE_TEXT,
            dropdown_fg_color="#ffffff",
            dropdown_text_color=TITLE_TEXT,
            dropdown_hover_color="#fff1ef",
            height=40,
            corner_radius=8,
            state="readonly",
            command=self._on_action_selected,
        )
        self.action_menu.pack(fill="x", pady=(4, 0))

        channel_box = ctk.CTkFrame(action_row, fg_color="transparent")
        channel_box.grid(row=0, column=1, sticky="ew", padx=(0, 10))
        ctk.CTkLabel(channel_box, text="Canal", text_color=MUTED_TEXT, font=("Segoe UI", 11, "bold")).pack(anchor="w")
        self.canal_cobranca_menu = ctk.CTkComboBox(
            channel_box,
            values=[DEFAULT_CANAL_COBRANCA, CANAL_TODOS],
            variable=self.canal_cobranca_var,
            fg_color="#ffffff",
            border_color=FIELD_BORDER,
            button_color="#f3f4f6",
            button_hover_color="#e5e7eb",
            text_color=TITLE_TEXT,
            dropdown_fg_color="#ffffff",
            dropdown_text_color=TITLE_TEXT,
            dropdown_hover_color="#fff1ef",
            height=40,
            corner_radius=8,
            state="readonly",
            command=self._on_canal_selected,
        )
        self.canal_cobranca_menu.pack(fill="x", pady=(4, 0))

        limit_box = ctk.CTkFrame(action_row, fg_color="transparent")
        limit_box.grid(row=0, column=2, sticky="ew", padx=(0, 10))
        ctk.CTkLabel(limit_box, text="Limite", text_color=MUTED_TEXT, font=("Segoe UI", 11, "bold")).pack(anchor="w")
        ctk.CTkEntry(
            limit_box,
            textvariable=self.execution_limit_var,
            height=40,
            width=82,
            corner_radius=8,
            border_color=FIELD_BORDER,
        ).pack(fill="x", pady=(4, 0))

        buttons = ctk.CTkFrame(action_row, fg_color="transparent")
        buttons.grid(row=0, column=3, sticky="sew", pady=(20, 0))
        self.execute_button = ctk.CTkButton(
            buttons,
            text="Executar",
            fg_color=BUTTON_BG,
            hover_color=BUTTON_ACTIVE_BG,
            text_color="#ffffff",
            height=40,
            width=132,
            corner_radius=8,
            command=self._execute_selected_action,
        )
        self.execute_button.pack(side="left", padx=(0, 8))
        self.pause_button = ctk.CTkButton(
            buttons,
            text="Pausar",
            fg_color="#ffffff",
            hover_color="#fff1ef",
            border_width=1,
            border_color=BUTTON_BG,
            text_color=BUTTON_BG,
            height=40,
            width=110,
            state="disabled",
            command=self._toggle_pause_processing,
        )
        self.pause_button.pack(side="left", padx=(0, 8))
        self.stop_button = ctk.CTkButton(
            buttons,
            text="Parar",
            fg_color="#ffffff",
            hover_color="#fff1ef",
            border_width=1,
            border_color=BUTTON_BG,
            text_color=BUTTON_BG,
            height=40,
            width=110,
            state="disabled",
            command=self._request_stop_processing,
        )
        self.stop_button.pack(side="left")

        rule_box = ctk.CTkFrame(actions, fg_color=PANEL_BG, corner_radius=8, border_width=1, border_color=CARD_BORDER)
        rule_box.pack(fill="x", padx=16, pady=(0, 14))
        ctk.CTkLabel(
            rule_box,
            textvariable=self.action_rule_var,
            text_color=TITLE_TEXT,
            font=("Segoe UI", 12),
            wraplength=980,
            justify="left",
        ).pack(anchor="w", padx=12, pady=8)

    def _build_log_section(self, parent) -> None:
        section = self._create_section(parent, 4, "Log operacional")
        ctk.CTkLabel(section, textvariable=self.status_var, text_color=MUTED_TEXT, font=("Segoe UI", 12)).pack(
            anchor="w", padx=16, pady=(0, 8)
        )
        self.log_box = ctk.CTkTextbox(
            section,
            height=190,
            fg_color="#101828",
            text_color="#f9fafb",
            corner_radius=10,
            border_width=1,
            border_color="#1f2937",
            font=("Consolas", 11),
        )
        self.log_box.pack(fill="x", padx=16, pady=(0, 16))
        self.log_box.configure(state="disabled")

    def _create_section(self, parent, row: int, title: str):
        section = ctk.CTkFrame(parent, fg_color=CARD_BG, corner_radius=14, border_width=1, border_color=CARD_BORDER)
        section.grid(row=row, column=0, sticky="ew", pady=(0, 12))
        section.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(section, text=title, text_color=NAVY_TEXT, font=("Segoe UI", 17, "bold")).pack(
            anchor="w", padx=16, pady=(14, 10)
        )
        return section

    def _build_readonly_field(self, parent, row: int, column: int, label: str, value: str) -> None:
        wrapper = ctk.CTkFrame(parent, fg_color="transparent")
        wrapper.grid(row=row, column=column, sticky="ew", padx=6, pady=6)
        ctk.CTkLabel(wrapper, text=label, text_color=MUTED_TEXT, font=("Segoe UI", 12, "bold")).pack(anchor="w")
        entry = ctk.CTkEntry(wrapper, height=38)
        entry.pack(fill="x", pady=(4, 0))
        entry.insert(0, value)
        entry.configure(state="disabled")

    def _build_info_label(self, parent, row: int, column: int, label: str, value: str) -> None:
        item = ctk.CTkFrame(parent, fg_color="transparent")
        item.grid(row=row, column=column, sticky="ew", padx=14, pady=10)
        ctk.CTkLabel(item, text=label, text_color=MUTED_TEXT, font=("Segoe UI", 11, "bold")).pack(anchor="w")
        ctk.CTkLabel(item, text=value, text_color=TITLE_TEXT, font=("Segoe UI", 13), wraplength=420).pack(anchor="w")

    def _build_metric_card(
        self,
        parent,
        row: int,
        column: int,
        title: str,
        variable,
        bg_color: str,
        accent_color: str = PRIMARY_TEXT,
    ) -> None:
        card = ctk.CTkFrame(parent, fg_color=bg_color, corner_radius=8, border_width=1, border_color=CARD_BORDER)
        card.grid(row=row, column=column, sticky="ew", padx=4, pady=4)
        ctk.CTkFrame(card, fg_color=accent_color, width=4, height=46, corner_radius=4).pack(
            side="left", fill="y", padx=(0, 0), pady=8
        )
        content = ctk.CTkFrame(card, fg_color="transparent")
        content.pack(side="left", fill="both", expand=True, padx=10, pady=8)
        ctk.CTkLabel(content, text=title.upper(), text_color=MUTED_TEXT, font=("Segoe UI", 9, "bold")).pack(
            anchor="w"
        )
        ctk.CTkLabel(content, textvariable=variable, text_color=TITLE_TEXT, font=("Segoe UI", 20, "bold")).pack(
            anchor="w", pady=(0, 0)
        )

    def _build_step_card(self, parent, etapa: str, texto: str) -> None:
        card = ctk.CTkFrame(parent, fg_color="#fafafa", corner_radius=16, border_width=1, border_color="#eeeeee")
        card.pack(fill="x", padx=18, pady=(0, 10))
        card.grid_columnconfigure(1, weight=1)
        ctk.CTkLabel(card, text=etapa, text_color=PRIMARY_TEXT, font=("Segoe UI", 14, "bold")).grid(
            row=0, column=0, padx=14, pady=14, sticky="nw"
        )
        ctk.CTkLabel(card, text=texto, text_color=TITLE_TEXT, font=("Segoe UI", 13), wraplength=330, justify="left").grid(
            row=0, column=1, padx=(0, 14), pady=14, sticky="ew"
        )

    def _not_ready(self, message: str) -> None:
        self.status_var.set(message)
        self._log(message)

    def _persist_interface_config(self) -> None:
        salvar_config_interface(
            conta_envio=self.outlook_account_var.get(),
            canal_cobranca=self.canal_cobranca_var.get(),
            caminho_excel=self.file_path_var.get(),
            limite_execucao=self.execution_limit_var.get(),
            usuario_coral=self.coral_user_var.get(),
            salvar_login_coral=bool(self.save_coral_login_var.get()),
            coral_headless=bool(self.coral_headless_var.get()),
        )
        if self.save_coral_login_var.get():
            senha_salva = salvar_senha_coral_windows(self.coral_user_var.get(), self.coral_password_var.get())
            if not senha_salva and self.coral_password_var.get().strip():
                self._log("Nao foi possivel salvar a senha Coral no Gerenciador de Credenciais do Windows.")
        else:
            apagar_senha_coral_windows()

    def _on_save_coral_login_changed(self) -> None:
        self._persist_interface_config()

    def _on_coral_headless_changed(self) -> None:
        self._persist_interface_config()

    def _selected_canal(self) -> str:
        return self.canal_cobranca_var.get().strip() or DEFAULT_CANAL_COBRANCA

    def _on_canal_selected(self, _value: str | None = None) -> None:
        self._persist_interface_config()
        self._refresh_queue()

    def _apply_canais_cobranca(self, canais: list[str]) -> None:
        if not canais:
            canais = [DEFAULT_CANAL_COBRANCA, CANAL_TODOS]
        current = self._selected_canal()
        normalized_values = {_normalizar_texto(item): item for item in canais}
        if _normalizar_texto(current) not in normalized_values:
            self.canal_cobranca_var.set(canais[0])
        self.canal_cobranca_menu.configure(values=canais)

    def _on_close(self) -> None:
        self._persist_interface_config()
        self.destroy()

    def _action_command(self, message: str | None):
        if message is None:
            return self._refresh_queue
        if message == "PROCESS_D0":
            return self._process_d0
        if message == "PROCESS_D2":
            return self._process_d2
        return lambda item=message: self._not_ready(item)

    def _build_action_commands(self) -> dict[str, object]:
        return {
            ACTION_UPDATE_QUEUE: self._refresh_queue,
            ACTION_PROCESS_D0: self._process_d0,
            ACTION_PROCESS_D2: self._process_d2,
            ACTION_PROCESS_D45: lambda: self._not_ready("Retry/lembrete ainda nao implementado."),
            ACTION_PROCESS_D7: lambda: self._not_ready("Ultima etapa ainda nao implementada."),
            ACTION_PROCESS_ALL: lambda: self._not_ready("Processamento em lote ainda nao implementado."),
        }

    def _on_action_selected(self, action: str) -> None:
        action_key = action_key_from_label(action)
        self.action_rule_var.set(ACTION_RULES.get(action_key, "Acao sem regra cadastrada."))

    def _execute_selected_action(self) -> None:
        action = action_key_from_label(self.action_var.get())
        command = self.action_commands.get(action)
        if command is None:
            self._not_ready(f"Acao nao configurada: {self.action_var.get().strip()}")
            return
        command()

    def _set_processing_controls(self, running: bool) -> None:
        main_thread = self.__dict__.get("_main_thread")
        if main_thread is not None and threading.current_thread() is not main_thread:
            self.after(0, self._set_processing_controls, running)
            return
        if "pause_button" not in self.__dict__ or "stop_button" not in self.__dict__ or "execute_button" not in self.__dict__:
            return
        state = "normal" if running else "disabled"
        self.pause_button.configure(state=state, text="Pausar")
        self.stop_button.configure(state=state)
        self.execute_button.configure(state="disabled" if running else "normal")
        if not running:
            self.__dict__.setdefault("pause_requested", threading.Event()).clear()

    def _start_processing_control(self, workbook_path: Path) -> None:
        self.__dict__.setdefault("pause_requested", threading.Event()).clear()
        self.__dict__.setdefault("stop_requested", threading.Event()).clear()
        self.current_processing_workbook_path = workbook_path
        self._set_processing_controls(True)

    def _finish_processing_control(self) -> None:
        self.__dict__.setdefault("pause_requested", threading.Event()).clear()
        self.__dict__.setdefault("stop_requested", threading.Event()).clear()
        self.current_processing_workbook_path = None
        self._set_processing_controls(False)

    def _toggle_pause_processing(self) -> None:
        if self.processing_thread is None or not self.processing_thread.is_alive():
            return
        if self.pause_requested.is_set():
            self.pause_requested.clear()
            self.pause_button.configure(text="Pausar")
            self._set_status("Processamento retomado.")
            self._log("Processamento retomado pelo usuario.")
        else:
            self.pause_requested.set()
            self.pause_button.configure(text="Retomar")
            self._set_status("Pausa solicitada. O robo vai pausar no proximo ponto seguro.")
            self._log("Pausa solicitada pelo usuario.")

    def _request_stop_processing(self) -> None:
        if self.processing_thread is None or not self.processing_thread.is_alive():
            workbook_path = self.current_processing_workbook_path
            if workbook_path:
                threading.Thread(
                    target=self._sync_processing_events_safe,
                    args=(workbook_path, "parada manual"),
                    daemon=True,
                ).start()
            return
        self.stop_requested.set()
        self.pause_requested.clear()
        self.pause_button.configure(text="Pausar")
        self._set_status("Parada solicitada. O robo vai sincronizar a planilha no proximo ponto seguro.")
        self._log("Parada solicitada pelo usuario; sincronizacao sera feita antes de encerrar.")

    def _wait_if_paused_or_stopped(self, workbook_path: Path, contexto: str) -> bool:
        pause_requested = self.__dict__.setdefault("pause_requested", threading.Event())
        stop_requested = self.__dict__.setdefault("stop_requested", threading.Event())
        if stop_requested.is_set():
            self._log(f"Parada detectada antes de {contexto}. Sincronizando pendencias.")
            self._sync_processing_events_safe(workbook_path, f"parada manual antes de {contexto}")
            return False
        if pause_requested.is_set():
            self._set_status(f"Pausado: {contexto}")
            self._log(f"Processamento pausado antes de {contexto}.")
        while pause_requested.is_set() and not stop_requested.is_set():
            time.sleep(0.5)
        if stop_requested.is_set():
            self._log(f"Parada detectada apos pausa em {contexto}. Sincronizando pendencias.")
            self._sync_processing_events_safe(workbook_path, f"parada manual em {contexto}")
            return False
        return True

    def _is_coral_login_xpath(self, xpath: str) -> bool:
        return xpath in {XPATH_CORAL_LOGIN, XPATH_CORAL_SENHA, XPATH_CORAL_ENTRAR}

    def _classificar_estado_coral(self) -> str:
        driver = self.__dict__.get("driver")
        if driver is None:
            return "SEM_DRIVER"

        try:
            current_url = str(getattr(driver, "current_url", "") or "").lower()
        except Exception:
            current_url = ""

        if "/login" in current_url:
            return "LOGIN"
        if "/contratos/editar/" in current_url:
            return "EDICAO"
        if "/contratos" in current_url:
            return "CONTRATOS"
        if "precificacao/dashboard" in current_url:
            return "DASHBOARD"

        try:
            if driver.find_elements(By.XPATH, XPATH_CORAL_LOGIN):
                return "LOGIN"
        except Exception:
            pass
        return "OUTRA"

    def _raise_if_coral_login_redirected(self, description: str, xpath: str | None = None) -> None:
        if xpath and self._is_coral_login_xpath(xpath):
            return
        if self._classificar_estado_coral() == "LOGIN":
            raise RuntimeError(f"Sessao Coral redirecionada para login enquanto aguardava {description}.")

    def _create_driver(self):
        options = webdriver.ChromeOptions()
        coral_headless = bool(self.__dict__.get("coral_headless_var").get()) if self.__dict__.get("coral_headless_var") else DEFAULT_CORAL_HEADLESS
        if coral_headless:
            options.add_argument("--headless=new")
            options.add_argument("--window-size=1600,1200")
        else:
            options.add_argument("--start-maximized")
        options.add_argument("--disable-gpu")
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--log-level=3")
        options.add_experimental_option("excludeSwitches", ["enable-logging"])
        self._log(
            "Criando sessao do Chrome para o Coral"
            + (" em modo invisivel..." if coral_headless else "...")
        )
        return webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

    def _wait_clickable(self, xpath: str, description: str, timeout: int = 30):
        self._log(f"Aguardando {description} ficar disponivel...")
        self._raise_if_coral_login_redirected(description, xpath)

        def condition(driver):
            self._raise_if_coral_login_redirected(description, xpath)
            return EC.element_to_be_clickable((By.XPATH, xpath))(driver)

        return WebDriverWait(self.driver, timeout, poll_frequency=0.5).until(condition)

    def _wait_visible(self, xpath: str, description: str, timeout: int = 30):
        self._log(f"Aguardando {description} aparecer...")
        self._raise_if_coral_login_redirected(description, xpath)

        def condition(driver):
            self._raise_if_coral_login_redirected(description, xpath)
            return EC.visibility_of_element_located((By.XPATH, xpath))(driver)

        return WebDriverWait(self.driver, timeout, poll_frequency=0.5).until(condition)

    def _safe_click(self, xpath: str, description: str, timeout: int = 30) -> None:
        last_error = None
        for attempt in range(1, 4):
            try:
                element = self._wait_clickable(xpath, description, timeout=timeout)
                self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", element)
                time.sleep(0.2)
                try:
                    element.click()
                except Exception:
                    self.driver.execute_script("arguments[0].click();", element)
                self._log(f"Clique OK em {description}.")
                return
            except Exception as exc:
                last_error = exc
                self._log(f"Clique falhou em {description} ({attempt}/3): {exc}")
                time.sleep(1)
        raise RuntimeError(f"Nao foi possivel clicar em {description} apos 3 tentativas: {last_error}")

    def _safe_type(self, xpath: str, value: str, description: str, timeout: int = 30, press_enter: bool = False) -> None:
        last_error = None
        for attempt in range(1, 4):
            try:
                element = self._wait_clickable(xpath, description, timeout=timeout)
                self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", element)
                time.sleep(0.2)
                element.click()
                element.send_keys(Keys.CONTROL, "a")
                element.send_keys(Keys.DELETE)
                element.send_keys(str(value))
                if press_enter:
                    element.send_keys(Keys.ENTER)
                shown = "********" if "senha" in _normalizar_texto(description).lower() else value
                self._log(f"Texto preenchido em {description}: {shown}")
                return
            except Exception as exc:
                last_error = exc
                self._log(f"Preenchimento falhou em {description} ({attempt}/3): {exc}")
                time.sleep(1)
        raise RuntimeError(f"Nao foi possivel preencher {description} apos 3 tentativas: {last_error}")

    def _clear_input_field(self, element) -> None:
        element.click()
        element.send_keys(Keys.CONTROL, "a")
        element.send_keys(Keys.DELETE)
        element.clear()
        self.driver.execute_script(
            """
            arguments[0].value = '';
            arguments[0].dispatchEvent(new Event('input', { bubbles: true }));
            arguments[0].dispatchEvent(new Event('change', { bubbles: true }));
            """,
            element,
        )
        time.sleep(0.2)

    def _read_input_value(self, xpath: str, description: str, timeout: int = 10) -> str:
        element = self._wait_visible(xpath, description, timeout=timeout)
        return (element.get_attribute("value") or "").strip()

    @staticmethod
    def _format_money_for_coral(value: object) -> str:
        return f"{_format_money(value):.2f}".replace(".", ",")

    @staticmethod
    def _money_values_match(expected: object, actual: object) -> bool:
        return abs(_format_money(expected) - _format_money(actual)) < 0.01

    def _fill_and_validate_money(self, xpath: str, value: object, description: str, timeout: int = 30) -> None:
        expected_text = self._format_money_for_coral(value)
        last_read = ""
        for attempt in range(1, 4):
            try:
                element = self._wait_clickable(xpath, description, timeout=timeout)
                self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", element)
                self._clear_input_field(element)
                element.send_keys(expected_text)
                time.sleep(0.4)
                last_read = self._read_input_value(xpath, description, timeout=10)
                if self._money_values_match(expected_text, last_read):
                    self._log(f"Valor validado em {description}: {expected_text}")
                    return
            except Exception as exc:
                last_read = f"erro: {exc}"
                self._log(f"Erro ao preencher valor em {description} ({attempt}/3): {exc}")
                time.sleep(1)
        raise RuntimeError(f"Valor divergente em {description}. Esperado {expected_text}, lido {last_read or '<vazio>'}.")

    def _login_coral(self, force_new_driver: bool = True) -> None:
        usuario = self.coral_user_var.get().strip()
        senha = self.coral_password_var.get().strip()
        if not usuario or not senha:
            raise RuntimeError("Informe usuario e senha do Coral antes de processar D0+2.")

        if force_new_driver or self.driver is None:
            self._fechar_driver_coral()
            self.driver = self._create_driver()
        self._set_status("Acessando Coral...")
        self._log("Abrindo login do Coral...")
        self.driver.get(URL_CORAL_LOGIN)
        self._safe_type(XPATH_CORAL_LOGIN, usuario, "campo de login Coral", timeout=45)
        self._safe_type(XPATH_CORAL_SENHA, senha, "campo de senha Coral", timeout=45)
        self._safe_click(XPATH_CORAL_ENTRAR, "botao Entrar Coral", timeout=45)
        WebDriverWait(self.driver, 60).until(lambda driver: URL_CORAL_DASHBOARD in driver.current_url)
        self._log("Login Coral confirmado pelo dashboard.")

    def _fechar_driver_coral(self) -> None:
        try:
            if self.driver is not None:
                self.driver.quit()
        except Exception as exc:
            self._log(f"Aviso ao fechar sessao anterior do Coral: {exc}")
        finally:
            self.driver = None

    def _reiniciar_sessao_coral(self, motivo: str) -> None:
        self._log(f"Recuperacao nivel 3: reiniciando navegador e login do Coral. Motivo: {motivo}")
        self._fechar_driver_coral()
        self._login_coral(force_new_driver=True)

    def _garantir_tela_base_coral(self) -> None:
        if self.driver is None:
            self._log("Sessao Coral inexistente. Criando nova sessao antes de voltar para contratos.")
            self._login_coral(force_new_driver=True)

        estado_atual = self._classificar_estado_coral()
        if estado_atual == "LOGIN":
            self._log("Sessao Coral caiu para login. Reautenticando antes de abrir contratos.")
            self._login_coral(force_new_driver=False)

        self._set_status("Abrindo tela base de contratos...")
        self._log("Indo para tela base de contratos no Coral...")
        self.driver.get(URL_CORAL_CONTRATOS)
        if self._classificar_estado_coral() == "LOGIN":
            self._log("Coral redirecionou para login ao abrir contratos. Reautenticando.")
            self._login_coral(force_new_driver=False)
            self.driver.get(URL_CORAL_CONTRATOS)

        self._click_popup_sim_if_visible(timeout=3)
        self._safe_click(XPATH_ABA_CONTRATOS, "aba Contratos", timeout=45)
        self._wait_visible(XPATH_CAMPO_BUSCA_CONTRATOS, "campo de busca de contratos", timeout=45)
        self._log("Tela base de contratos validada.")

    def _executar_unidade_com_tentativas(self, identificador: str, descricao: str, operacao):
        ultimo_erro = None
        for tentativa in range(1, CORAL_OPERATION_MAX_ATTEMPTS + 1):
            try:
                if tentativa > 1:
                    self._log(
                        f"Nova tentativa tecnica {tentativa}/{CORAL_OPERATION_MAX_ATTEMPTS} para "
                        f"{descricao} [{identificador}]."
                    )
                return operacao()
            except CoralNavigationAbortError:
                raise
            except Exception as exc:
                ultimo_erro = exc
                self._log(
                    f"Falha tecnica {tentativa}/{CORAL_OPERATION_MAX_ATTEMPTS} em "
                    f"{descricao} [{identificador}]: {exc}"
                )
                if tentativa >= CORAL_OPERATION_MAX_ATTEMPTS:
                    break
                try:
                    self._garantir_tela_base_coral()
                except Exception as recovery_exc:
                    self._log(f"Retorno a tela base falhou para {identificador}: {recovery_exc}")
                    self._reiniciar_sessao_coral(str(recovery_exc))
        raise RuntimeError(
            f"Nao foi possivel concluir {descricao} para {identificador} "
            f"apos {CORAL_OPERATION_MAX_ATTEMPTS} tentativas tecnicas: {ultimo_erro}"
        ) from ultimo_erro

    def _executar_com_recuperacao_coral(self, contrato: str, descricao: str, operacao):
        ultimo_erro = None
        for nivel in range(1, 4):
            try:
                if nivel == 2:
                    self._log(f"Recuperacao nivel 2 para {contrato}: retornando a tela base de contratos.")
                    self._garantir_tela_base_coral()
                elif nivel == 3:
                    self._reiniciar_sessao_coral(str(ultimo_erro or descricao))
                resultado = operacao()
                if nivel > 1:
                    self._log(f"Recuperacao concluida no nivel {nivel}: {descricao} | contrato {contrato}.")
                return resultado
            except CoralNavigationAbortError:
                raise
            except Exception as exc:
                ultimo_erro = exc
                self._log(f"Falha no nivel {nivel}/3 ao {descricao} para {contrato}: {exc}")
        raise RuntimeError(
            f"Nao foi possivel {descricao} para o contrato {contrato} apos recuperacao completa: {ultimo_erro}"
        ) from ultimo_erro

    def _ir_para_contratos_coral(self) -> None:
        self._garantir_tela_base_coral()

    def _buscar_contrato_coral(self, contrato: str) -> None:
        self._safe_type(XPATH_CAMPO_BUSCA_CONTRATOS, contrato, "campo de busca de contratos", timeout=30, press_enter=True)
        resultado = self._wait_visible(XPATH_TD_CONTRATO_RESULTADO, "contrato retornado na pesquisa", timeout=45)
        texto = (resultado.text or "").strip()
        if texto != contrato:
            raise RuntimeError(f"Contrato pesquisado nao confere. Planilha: {contrato} | Tela: {texto}")
        self._log(f"Busca validada para contrato {contrato}.")

    def _abrir_edicao_contrato_coral(self, contrato: str) -> None:
        url_edicao = montar_url_edicao_contrato_coral(contrato)
        self._log(f"Abrindo edicao direta do contrato {contrato}: {url_edicao}")
        try:
            self.driver.get(url_edicao)
            self._click_popup_sim_if_visible(timeout=5)
            self._wait_visible(XPATH_ABA_PAGAMENTOS_RAPIDA_ICONE, "tela de edicao do contrato", timeout=45)
            self._log(f"Tela de edicao aberta diretamente para contrato {contrato}.")
            return
        except Exception as exc:
            self._log(f"Abertura direta falhou para {contrato}. Usando busca como fallback: {exc}")

        self._ir_para_contratos_coral()
        self._buscar_contrato_coral(contrato)
        self._safe_click(XPATH_MAIS_OPCOES_CONTRATO, "mais opcoes do contrato", timeout=30)
        self._safe_click(XPATH_EDITAR_CONTRATO, "editar contrato", timeout=30)
        self._click_popup_sim_if_visible(timeout=5)
        self._wait_visible(XPATH_ABA_PAGAMENTOS_RAPIDA_ICONE, "tela de edicao do contrato", timeout=45)
        self._log(f"Tela de edicao aberta para contrato {contrato}.")

    def _clicar_botao_sim_repetidamente(self, xpath: str, description: str, timeout: int = 3, max_clicks: int = 5) -> int:
        clicks = 0
        for tentativa in range(1, max_clicks + 1):
            try:
                self._wait_clickable(xpath, description, timeout=timeout if clicks == 0 else 1)
                self._safe_click(xpath, description, timeout=5)
                clicks += 1
                self.popup_edicao_tratado = True
                self._log(f"{description} confirmado ({clicks}x).")
                time.sleep(0.5)
            except Exception:
                break
        return clicks

    def _click_popup_sim_if_visible(self, timeout: int = 3) -> bool:
        return self._clicar_botao_sim_repetidamente(
            XPATH_POPUP_SIM,
            "botao Sim do popup",
            timeout=timeout,
            max_clicks=5,
        ) > 0

    def _fechar_modal_carregar_cliente_se_visivel(self, timeout: int = 2) -> bool:
        try:
            cliques_sim = self._clicar_botao_sim_repetidamente(
                XPATH_BOTAO_SIM_MODAL_CARREGAR_CLIENTE,
                "botao Sim do modal Carregar Cliente",
                timeout=timeout,
                max_clicks=5,
            )
            if cliques_sim:
                return True

            botao = self._wait_clickable(
                XPATH_BOTAO_FECHAR_MODAL_CARREGAR_CLIENTE,
                "fechar modal Carregar Cliente",
                timeout=timeout,
            )
            botao.click()
            self.popup_edicao_tratado = True
            self._log("Modal Carregar Cliente fechado.")
            return True
        except Exception:
            return False

    def _resumo_pagamento_visivel(self, timeout: int = 2) -> bool:
        try:
            titulo = self._wait_visible(XPATH_RESUMO_PAGAMENTO_TITULO, "titulo RESUMO PAGAMENTO", timeout=timeout)
            return "RESUMO PAGAMENTO" in _normalizar_texto(titulo.text)
        except Exception:
            return False

    def _texto_toast_visivel(self, timeout: float = 0.8) -> str:
        driver = self.__dict__.get("driver")
        if driver is None:
            return ""
        end_time = time.time() + max(timeout, 0)
        while time.time() <= end_time:
            try:
                for toast in driver.find_elements(By.XPATH, XPATH_TOAST_CONTAINER):
                    texto = " ".join(str(getattr(toast, "text", "") or "").split()).strip()
                    if toast.is_displayed() and texto:
                        return texto
            except Exception:
                return ""
            time.sleep(0.15)
        return ""

    def _raise_if_toast_critico_edicao(self, contexto: str) -> None:
        texto = self._texto_toast_visivel(timeout=0.8)
        if not texto:
            return
        texto_normalizado = _normalizar_texto(texto)
        if any(indicador in texto_normalizado for indicador in TOASTS_CRITICOS_EDICAO):
            self._log(f"Toast critico detectado em {contexto}: {texto}")
            raise CoralNavigationAbortError(f"Toast critico do Coral em {contexto}: {texto}")

    @staticmethod
    def _falha_limite_cliques(exc: Exception, *descricoes: str) -> bool:
        texto = str(exc)
        return any(f"Nao foi possivel clicar em {descricao} apos 3 tentativas" in texto for descricao in descricoes)

    def _encontrar_botao_avancar_fluxo_edicao(self):
        driver = self.__dict__.get("driver")
        if driver is None:
            return None
        for botao in driver.find_elements(By.XPATH, XPATH_BOTOES_RODAPE_EDICAO):
            texto = _normalizar_texto(getattr(botao, "text", ""))
            if "AVANCAR" not in texto:
                continue
            if not botao.is_displayed() or not botao.is_enabled():
                continue
            return botao
        return None

    def _clicar_avancar_fluxo_edicao(self, fallback_xpath: str, tentativa: int) -> None:
        try:
            self._log(f"Procurando botao Avancar por texto ({tentativa})...")
            botao = WebDriverWait(self.driver, 5).until(lambda _driver: self._encontrar_botao_avancar_fluxo_edicao())
            self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", botao)
            time.sleep(0.2)
            try:
                botao.click()
            except Exception:
                self.driver.execute_script("arguments[0].click();", botao)
            self._log("Clique OK em Avancar do fluxo da edicao por texto.")
            return
        except Exception as exc:
            self._log(f"Nao foi possivel clicar em Avancar por texto ({tentativa}). Usando XPath fixo: {exc}")
            try:
                self._safe_click(fallback_xpath, "Avancar do fluxo da edicao", timeout=15)
            except Exception as click_exc:
                raise CoralNavigationAbortError(
                    f"Falha critica ao clicar em Avancar do fluxo da edicao apos 3 tentativas locais: {click_exc}"
                ) from click_exc

    def _avancar_fluxo_edicao_ate_resumo_pagamento(self, max_tentativas: int = 4) -> None:
        if self._resumo_pagamento_visivel(timeout=2):
            self._log("Resumo Pagamento ja esta visivel.")
            return

        sequencia_avancar = [
            XPATH_BOTAO_AVANCAR_FLUXO_EDICAO,
            XPATH_BOTAO_AVANCAR_FLUXO_EDICAO_2,
            XPATH_BOTAO_AVANCAR_FLUXO_EDICAO_3,
            XPATH_BOTAO_AVANCAR_FLUXO_EDICAO_3,
        ]
        for tentativa, xpath_avancar in enumerate(sequencia_avancar[:max_tentativas], start=1):
            self._log(f"Avancando fluxo da edicao ate Pagamentos ({tentativa}/{max_tentativas})...")
            self._clicar_avancar_fluxo_edicao(xpath_avancar, tentativa)
            self._raise_if_toast_critico_edicao(f"avancar fluxo da edicao ({tentativa}/{max_tentativas})")
            if self._resumo_pagamento_visivel(timeout=5):
                self._log("Resumo Pagamento validado apos Avancar.")
                return

        raise CoralNavigationAbortError("Nao foi possivel chegar ao RESUMO PAGAMENTO apos clicar em Avancar.")

    def _acessar_aba_pagamentos_seguro(self, alvo_xpath: str, alvo_descricao: str) -> None:
        erros = []
        popup_tratado_anteriormente = bool(self.__dict__.get("popup_edicao_tratado", False))
        popup_tratado = popup_tratado_anteriormente or self._click_popup_sim_if_visible(timeout=3)
        modal_tratado = self._fechar_modal_carregar_cliente_se_visivel(timeout=2)
        if popup_tratado or modal_tratado:
            try:
                self._log("Popup/modal tratado. Usando caminho secundario ate Pagamentos.")
                self._avancar_fluxo_edicao_ate_resumo_pagamento()
                self._wait_clickable(alvo_xpath, alvo_descricao, timeout=25)
                self._log(f"Pagamentos validado apos caminho secundario pelo alvo {alvo_descricao}.")
                self.popup_edicao_tratado = False
                return
            except CoralNavigationAbortError:
                self.popup_edicao_tratado = False
                raise
            except Exception as exc:
                erros.append(f"caminho secundario apos popup/modal: {exc}")
                self._log(f"Caminho secundario apos popup/modal falhou: {exc}")
                self.popup_edicao_tratado = False

        try:
            self._safe_click(XPATH_ABA_PAGAMENTOS_RAPIDA, "aba Pagamentos", timeout=25)
            self._raise_if_toast_critico_edicao("abrir aba Pagamentos")
            if not self._resumo_pagamento_visivel(timeout=5):
                self._avancar_fluxo_edicao_ate_resumo_pagamento()
            self._wait_clickable(alvo_xpath, alvo_descricao, timeout=25)
            self._log(f"Pagamentos validado pelo alvo {alvo_descricao}.")
            return
        except CoralNavigationAbortError:
            raise
        except Exception as exc:
            erros.append(f"rota rapida: {exc}")
            self._log(f"Rota rapida para Pagamentos falhou: {exc}")
            if self._falha_limite_cliques(exc, "aba Pagamentos"):
                raise CoralNavigationAbortError(
                    f"Falha critica ao clicar em aba Pagamentos apos 3 tentativas locais: {exc}"
                ) from exc

        try:
            self._log("Atualizando a edicao do contrato para uma ultima tentativa local de Pagamentos...")
            self.driver.refresh()
            self._click_popup_sim_if_visible(timeout=5)
            if self._fechar_modal_carregar_cliente_se_visivel(timeout=2):
                self._avancar_fluxo_edicao_ate_resumo_pagamento()
            self._safe_click(XPATH_ABA_PAGAMENTOS_RAPIDA, "aba Pagamentos apos refresh", timeout=30)
            self._raise_if_toast_critico_edicao("abrir aba Pagamentos apos refresh")
            if not self._resumo_pagamento_visivel(timeout=5):
                self._avancar_fluxo_edicao_ate_resumo_pagamento()
            self._wait_clickable(alvo_xpath, alvo_descricao, timeout=25)
            self._log(f"Pagamentos validado apos refresh pelo alvo {alvo_descricao}.")
            return
        except CoralNavigationAbortError:
            raise
        except Exception as exc:
            erros.append(f"refresh: {exc}")
            if self._falha_limite_cliques(exc, "aba Pagamentos apos refresh"):
                raise CoralNavigationAbortError(
                    f"Falha critica ao clicar em aba Pagamentos apos refresh depois de 3 tentativas locais: {exc}"
                ) from exc
        raise RuntimeError("Falha nas rotas locais para Pagamentos: " + " | ".join(erros))

    def _abrir_pagamentos_contrato_com_recuperacao(
        self,
        contrato: str,
        alvo_xpath: str,
        alvo_descricao: str,
    ) -> None:
        def abrir_e_validar():
            self._abrir_edicao_contrato_coral(contrato)
            url_atual = self.driver.current_url
            if not url_coral_corresponde_ao_contrato(url_atual, contrato):
                raise RuntimeError(f"Contrato incorreto. Esperado: {contrato} | URL atual: {url_atual}")
            self._log(f"Contrato validado pela URL antes de acessar Pagamentos: {contrato}.")
            self._acessar_aba_pagamentos_seguro(alvo_xpath, alvo_descricao)

        self._executar_com_recuperacao_coral(
            contrato,
            f"abrir Pagamentos e validar {alvo_descricao}",
            abrir_e_validar,
        )

    def _ir_para_pagamentos_carteira(self, contrato: str) -> None:
        self._abrir_pagamentos_contrato_com_recuperacao(contrato, XPATH_BOTAO_CARTEIRA, "botao Carteira")
        self._safe_click(XPATH_BOTAO_CARTEIRA, "botao Carteira", timeout=45)
        self._log("Aba Pagamentos/Carteira aberta.")

    def _ir_para_pagamentos_para_link(self, contrato: str) -> None:
        self._log(f"Preparando contrato {contrato} para gerar link. A navegacao para a edicao e obrigatoria.")
        self._abrir_pagamentos_contrato_com_recuperacao(contrato, XPATH_BOTAO_LINK, "botao Link")
        self._log(f"Aba Pagamentos aberta para gerar link no contrato {contrato}.")

    def _read_clipboard_text(self) -> str:
        try:
            import win32clipboard

            win32clipboard.OpenClipboard()
            try:
                return str(win32clipboard.GetClipboardData(win32clipboard.CF_UNICODETEXT) or "").strip()
            finally:
                win32clipboard.CloseClipboard()
        except Exception:
            try:
                return str(self.clipboard_get() or "").strip()
            except Exception:
                return ""

    def _ler_link_gerado_no_modal(self) -> str:
        try:
            campos = self.driver.find_elements(By.XPATH, "//ngb-modal-window//input[contains(@value,'http')]")
            for campo in campos:
                valor = str(campo.get_attribute("value") or "").strip()
                if valor.lower().startswith("http"):
                    return valor
        except Exception:
            return ""
        return ""

    def _copiar_link_gerado(self, tentativas: int = 3) -> str:
        for tentativa in range(1, tentativas + 1):
            link_modal = self._ler_link_gerado_no_modal()
            if link_modal:
                self._log(f"Link capturado diretamente no modal: {link_modal}")
                return link_modal
            antes = self._read_clipboard_text()
            self._log(f"Tentando copiar link gerado ({tentativa}/{tentativas})...")
            self._safe_click(XPATH_BOTAO_COPIAR_LINK, "botao Copiar link", timeout=30)
            time.sleep(1)
            depois = self._read_clipboard_text()
            if depois and depois != antes and depois.lower().startswith("http"):
                self._log(f"Link capturado: {depois}")
                return depois
            if depois and depois.lower().startswith("http"):
                self._log(f"Link capturado sem alteracao aparente do clipboard: {depois}")
                return depois
            link_modal = self._ler_link_gerado_no_modal()
            if link_modal:
                self._log(f"Link capturado diretamente no modal apos tentativa de copia: {link_modal}")
                return link_modal
        raise RuntimeError("Nao foi possivel capturar o link gerado pelo Coral.")

    def _gerar_link_pagamento_residual(self, valor: float) -> str:
        valor_formatado = self._format_money_for_coral(valor)
        self._log(f"Gerando link residual no Coral. Valor: R$ {valor_formatado}")
        self._safe_click(XPATH_BOTAO_LINK, "botao Link", timeout=45)
        time.sleep(0.5)
        self._fill_and_validate_money(XPATH_CAMPO_VALOR_LINK, valor, "valor do link", timeout=30)
        time.sleep(0.5)
        self._safe_click(XPATH_MODALIDADE_A_VENCER_LINK, "modalidade A vencer", timeout=30)
        time.sleep(0.5)
        self._safe_click(XPATH_BOTAO_EFETUAR_PAGAMENTO_LINK, "botao Efetuar pagamento link", timeout=45)
        time.sleep(1)
        return self._copiar_link_gerado()

    def _listar_cartoes_disponiveis(self, timeout: int = 20) -> list[dict[str, object]]:
        try:
            WebDriverWait(self.driver, timeout).until(
                lambda driver: len(driver.find_elements(By.XPATH, XPATH_CARTEIRA_CARD)) > 0
            )
        except TimeoutException:
            return []
        cartoes = []
        for index, bloco in enumerate(self.driver.find_elements(By.XPATH, XPATH_CARTEIRA_CARD), start=1):
            try:
                radio = bloco.find_element(By.XPATH, ".//input[@type='radio' and @formcontrolname='cardSelected']")
                try:
                    texto = bloco.find_element(By.XPATH, ".//span[contains(normalize-space(),'Numero do cartao') or contains(normalize-space(),'NÃºmero do cartÃ£o')]").text.strip()
                except Exception:
                    texto = f"Cartao {index}"
                cartoes.append({"elemento": bloco, "radio": radio, "texto": texto})
            except Exception as exc:
                self._log(f"Falha ao mapear cartao {index}: {exc}")
        return cartoes

    def _selecionar_cartao(self, cartao: dict[str, object]) -> str:
        bloco = cartao["elemento"]
        radio = cartao["radio"]
        texto = str(cartao.get("texto") or "Cartao")
        self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", bloco)
        time.sleep(0.3)
        try:
            if not radio.is_selected():
                radio.click()
        except Exception:
            self.driver.execute_script("arguments[0].click();", radio)
        self._log(f"Cartao selecionado: {texto}")
        return texto

    @staticmethod
    def _extrair_final_cartao(texto_cartao: str) -> str:
        match = re.search(r"\*+\s*(\d{4})", str(texto_cartao or ""))
        if match:
            return match.group(1)
        digits = re.findall(r"\d", str(texto_cartao or ""))
        return "".join(digits[-4:]) if len(digits) >= 4 else ""

    def _selecionar_parcela_1x(self) -> None:
        self._safe_click(XPATH_LISTA_PAGAMENTOS_CARTAO, "lista de parcelamento", timeout=30)
        time.sleep(0.5)
        self._safe_click(XPATH_PARCELAMENTO_1X_CARTAO, "parcelamento 1x", timeout=30)

    def _popup_erro_pagamento_visible(self, timeout: int = 2) -> bool:
        try:
            WebDriverWait(self.driver, timeout).until(EC.visibility_of_element_located((By.XPATH, XPATH_POPUP_ERRO_FECHAR)))
            return True
        except Exception:
            return False

    def _fechar_popup_erro_pagamento(self) -> None:
        self._safe_click(XPATH_POPUP_ERRO_FECHAR, "fechar popup de erro", timeout=10)
        time.sleep(1)

    def _historico_pagamento_text(self) -> str:
        try:
            return self._wait_visible(XPATH_HISTORICO_PAGAMENTO, "historico de pagamento", timeout=8).text or ""
        except Exception:
            return ""

    def _historico_tem_cobranca_cartao(
        self,
        valor: object,
        final_cartao: str,
        historico_antes: object = "",
    ) -> bool:
        expected = self._format_money_for_coral(valor)
        historico = self._historico_pagamento_text()
        if _historico_indica_nova_cobranca_cartao(historico_antes, historico, valor, data_pagamento=date.today()):
            self._log(f"Cobranca confirmada no historico: Carteira na data atual | valor R$ {expected}.")
            return True
        return False

    def _validar_resultado_pagamento_cartao(
        self,
        valor: object,
        final_cartao: str,
        historico_antes: object = "",
        timeout: int = 18,
        contrato: str = "",
    ) -> str:
        if self._popup_erro_pagamento_visible(timeout=1):
            self._fechar_popup_erro_pagamento()
            return "ERRO"

        end_time = time.time() + timeout
        while time.time() < end_time:
            estado_coral = self._classificar_estado_coral()
            if estado_coral == "LOGIN":
                raise RuntimeError(
                    f"Sessao Coral caiu para login apos tentativa de cobranca no contrato {contrato or '<desconhecido>'}."
                )
            if estado_coral in {"CONTRATOS", "DASHBOARD", "OUTRA"}:
                raise RuntimeError(
                    f"Coral saiu da tela de edicao apos tentativa de cobranca no contrato {contrato or '<desconhecido>'}."
                )
            if self._historico_tem_cobranca_cartao(valor, final_cartao, historico_antes=historico_antes):
                return "SUCESSO"
            if self._popup_erro_pagamento_visible(timeout=1):
                self._fechar_popup_erro_pagamento()
                return "ERRO"
            time.sleep(1)
        return "INDEFINIDO"

    def _concluir_fluxo_pos_pagamento(self) -> None:
        self._safe_click(XPATH_BOTAO_CONCLUIR_CONTRATO, "botao Concluir", timeout=30)
        time.sleep(1)
        self._safe_click(XPATH_BOTAO_ATUALIZAR_CONTRATO, "botao Atualizar contrato", timeout=30)
        time.sleep(1)
        self._safe_click(XPATH_BOTAO_FECHAR_FINAL, "botao Fechar final", timeout=30)
        time.sleep(1)

    def _tentar_cobranca_cartao_contrato(self, contrato: ContratoD2Pedagio) -> ResultadoD2Pedagio:
        self._ir_para_pagamentos_carteira(contrato.contrato)
        historico_inicial = self._historico_pagamento_text()
        if _historico_indica_nova_cobranca_cartao("", historico_inicial, contrato.valor, data_pagamento=date.today()):
            return ResultadoD2Pedagio(
                contrato.id_cliente,
                contrato.contrato,
                "COBRADO",
                0,
                0,
                "Pagamento Carteira ja identificado no historico antes de nova tentativa.",
            )
        cartoes = self._listar_cartoes_disponiveis()
        if not cartoes:
            return ResultadoD2Pedagio(
                contrato.id_cliente,
                contrato.contrato,
                "NAO_COBRADO",
                0,
                0,
                "Nenhum cartao localizado na carteira. Geracao de link pendente.",
            )

        for index, cartao in enumerate(cartoes):
            texto_cartao = self._selecionar_cartao(cartao)
            final_cartao = self._extrair_final_cartao(texto_cartao)
            if index == 0:
                self._fill_and_validate_money(XPATH_CAMPO_VALOR_PAGAMENTO_CARTAO, contrato.valor, "valor do pagamento no cartao")
                self._selecionar_parcela_1x()
            historico_antes = self._historico_pagamento_text()
            self._safe_click(XPATH_BOTAO_EFETUAR_PAGAMENTO_CARTAO, "botao Efetuar pagamento cartao", timeout=30)
            resultado_pagamento = self._validar_resultado_pagamento_cartao(
                contrato.valor,
                final_cartao,
                historico_antes=historico_antes,
                contrato=contrato.contrato,
            )
            if resultado_pagamento == "SUCESSO":
                return ResultadoD2Pedagio(
                    contrato.id_cliente,
                    contrato.contrato,
                    "COBRADO",
                    len(cartoes),
                    index + 1,
                    f"Pagamento aprovado no {texto_cartao}.",
                )
            if resultado_pagamento == "ERRO":
                self._log(f"Pagamento recusado no {texto_cartao}. Tentando proximo cartao...")
                continue
            self._log(f"Pagamento nao confirmado no {texto_cartao}. Tentando proximo cartao...")

        return ResultadoD2Pedagio(
            contrato.id_cliente,
            contrato.contrato,
            "NAO_COBRADO",
            len(cartoes),
            len(cartoes),
            "Todos os cartoes foram tentados sem confirmacao. Geracao de link pendente.",
        )

    def _refresh_outlook_accounts(self) -> None:
        threading.Thread(target=self._load_outlook_accounts_worker, args=(False,), daemon=True).start()

    def _refresh_outlook_accounts_silent(self) -> None:
        threading.Thread(target=self._load_outlook_accounts_worker, args=(True,), daemon=True).start()

    def _load_outlook_accounts_worker(self, silent: bool) -> None:
        try:
            contas = listar_contas_outlook()
        except Exception as exc:
            self._log(f"Nao foi possivel carregar contas do Outlook: {exc}")
            if not silent:
                self.after(0, lambda error=str(exc): messagebox.showerror("Erro no Outlook", error))
            return
        if not contas:
            self._log("Nenhuma conta encontrada no Outlook Desktop.")
            if not silent:
                self.after(0, lambda: messagebox.showwarning("Outlook", "Nenhuma conta encontrada no Outlook Desktop."))
            return
        self.after(0, self._apply_outlook_accounts, contas)

    def _apply_outlook_accounts(self, contas: list[str]) -> None:
        self.outlook_account_menu.configure(values=contas)
        if not self.outlook_account_var.get() or self.outlook_account_var.get() not in contas:
            self.outlook_account_var.set(contas[0])
        self._log(f"Contas Outlook carregadas automaticamente: {', '.join(contas)}")

    def _start_initial_background_loads(self) -> None:
        self._refresh_outlook_accounts_silent()
        path = self.file_path_var.get().strip() or DEFAULT_EXCEL_URL
        canal = self._selected_canal()
        threading.Thread(target=self._refresh_queue_worker, args=(path, canal), daemon=True).start()

    def _refresh_queue_worker(self, path: str, canal: str) -> None:
        resolved_path = resolver_caminho_excel_compartilhado(path)
        if resolved_path is None:
            self._log("Base operacional ainda nao localizada no carregamento inicial.")
            return
        self._log(f"Atualizacao automatica da fila usando planilha: {resolved_path}")
        try:
            sincronizados = sincronizar_eventos_processamento_json(resolved_path).get("eventos_sincronizados", 0)
            if sincronizados:
                self._log(f"Pendencias JSON do dia aplicadas antes de carregar a fila: {sincronizados} evento(s).")
        except Exception as exc:
            self._log(f"Nao foi possivel aplicar pendencias JSON antes de carregar a fila: {exc}")
        try:
            recentes = contar_pendencias_processamento_recentes(resolved_path)
            antigos = {dia: qtd for dia, qtd in recentes.items() if dia != date.today().isoformat()}
            if antigos:
                self._log(f"Pendencias JSON antigas encontradas e nao aplicadas automaticamente: {antigos}")
        except Exception as exc:
            self._log(f"Nao foi possivel verificar pendencias JSON recentes: {exc}")
        try:
            canais = carregar_canais_cobranca_excel(resolved_path)
            resumo = carregar_resumo_cobrancas_excel(resolved_path, canal=canal)
        except Exception as exc:
            self._log(f"Nao foi possivel carregar a fila automaticamente: {exc}")
            return
        self.after(0, self._apply_canais_cobranca, canais)
        self.after(0, self._apply_queue_summary, resolved_path, resumo)

    def _execution_limit(self) -> int:
        raw = self.execution_limit_var.get().strip()
        if not raw:
            return 10
        try:
            value = int(raw)
        except ValueError as exc:
            raise RuntimeError("Limite por execucao precisa ser um numero inteiro.") from exc
        if value <= 0:
            raise RuntimeError("Limite por execucao precisa ser maior que zero.")
        return value

    def _d0_limit(self) -> int:
        return self._execution_limit()

    def _d2_limit(self) -> int:
        return self._execution_limit()

    def _process_d0(self) -> None:
        if self.processing_thread is not None and self.processing_thread.is_alive():
            messagebox.showinfo("Etapa 1", "Ja existe um processamento em andamento.")
            return
        conta_envio = self.outlook_account_var.get().strip()
        if not conta_envio:
            messagebox.showwarning("Conta Outlook", "Informe a conta de envio antes de processar a etapa 1.")
            return
        if not self.coral_user_var.get().strip() or not self.coral_password_var.get().strip():
            messagebox.showwarning("Coral", "Informe usuario e senha do Coral antes de processar a etapa 1.")
            return
        self._persist_interface_config()

        path = self.file_path_var.get().strip() or DEFAULT_EXCEL_URL
        resolved_path = resolver_caminho_excel_compartilhado(path)
        if resolved_path is None:
            message = "Nao encontrei a base operacional sincronizada. Confirme a sincronizacao da pasta no OneDrive."
            self.status_var.set(message)
            self._log(message)
            messagebox.showwarning("Base nao encontrada", message)
            return

        try:
            limite = self._d0_limit()
            canal = self._selected_canal()
            emails = carregar_emails_d0_excel(resolved_path, limite=limite, canal=canal)
        except Exception as exc:
            self.status_var.set(str(exc))
            self._log(f"Erro ao preparar D0: {exc}")
            messagebox.showerror("Erro ao preparar etapa 1", str(exc))
            return

        if not emails:
            message = "Nenhum cliente apto para a etapa 1 com e-mail valido e contratos pendentes no canal selecionado."
            self.status_var.set(message)
            self._log(message)
            messagebox.showinfo("Etapa 1", message)
            return

        if not messagebox.askyesno(
            "Confirmar etapa 1",
            f"Gerar links no Coral e enviar diretamente {len(emails)} e-mail(s) pela conta {conta_envio}?\n\nCanal: {canal}",
        ):
            return

        self._start_processing_control(resolved_path)
        self.processing_thread = threading.Thread(
            target=self._run_d0_processing,
            args=(resolved_path, emails, conta_envio),
            daemon=True,
        )
        self.processing_thread.start()

    def _sync_processing_events_safe(self, workbook_path: Path, motivo: str) -> None:
        json_path = _processing_event_paths(workbook_path, date.today())[0]
        self._log(f"Sincronizacao JSON ({motivo}). Planilha alvo: {workbook_path}")
        self._log(f"Arquivo JSON do dia: {json_path}")
        try:
            resultado = sincronizar_eventos_processamento_json(workbook_path)
        except Exception as exc:
            self._log(f"Falha ao atualizar planilha no {motivo}. Eventos permanecem no JSON: {exc}")
            return
        sincronizados = resultado.get("eventos_sincronizados", 0)
        if sincronizados:
            self._log(
                f"Planilha atualizada no {motivo}: {sincronizados} evento(s), "
                f"{resultado.get('clientes_atualizados', 0)} cliente(s), "
                f"{resultado.get('contratos_atualizados', 0)} contrato(s)."
            )
        else:
            self._log(f"Nenhum evento JSON pendente para sincronizar no {motivo}.")

    def _run_d0_processing(self, workbook_path: Path, emails: list[EmailD0Pedagio], conta_envio: str) -> None:
        sucessos = 0
        falhas = 0
        eventos_desde_sync = 0
        try:
            self._login_coral(force_new_driver=True)
            total = len(emails)
            for index, email in enumerate(emails, start=1):
                if not self._wait_if_paused_or_stopped(workbook_path, f"D0 cliente {index}/{total}"):
                    break
                self._set_status(f"D0 {index}/{total}: {email.nome}")
                self._log(
                    f"[D0 {index}/{total}] Cliente {email.nome} | {email.total_contratos} contrato(s) | "
                    f"R$ {_format_brl(email.valor_total)}"
                )
                try:
                    contexto_envio = {
                        "link_pagamento": email.link_pagamento,
                        "link_gerado_nesta_execucao": False,
                    }

                    def processar_cliente_d0():
                        email_pronto = email
                        if contexto_envio["link_pagamento"]:
                            email_pronto = preparar_email_d0_com_link(email, str(contexto_envio["link_pagamento"]))
                            self._log(f"Reutilizando link D0 ja registrado para {email.nome}.")
                        else:
                            self._ir_para_pagamentos_para_link(email.contrato_referencia)
                            link_pagamento = self._gerar_link_pagamento_residual(email.valor_total)
                            contexto_envio["link_pagamento"] = link_pagamento
                            contexto_envio["link_gerado_nesta_execucao"] = True
                            email_pronto = preparar_email_d0_com_link(email, link_pagamento)
                            self._log(f"Link D0 gerado para {email.nome}: {link_pagamento}")

                        anexos_email: list[Path] = []
                        try:
                            pdf_relatorio = gerar_pdf_relatorio_pedagios_para_email(email_pronto)
                            anexos_email.append(pdf_relatorio)
                            self._log(f"Relatorio de pedagios anexado: {pdf_relatorio}")
                        except Exception as exc:
                            self._log(f"Nao foi possivel gerar/anexar relatorio de pedagios para {email.nome}: {exc}")

                        self._log(f"Enviando e-mail D0 pelo Outlook para {email.destinatario}...")
                        criar_email_outlook(
                            email_pronto,
                            conta_envio=conta_envio,
                            log_callback=self._log,
                            anexos=anexos_email,
                        )
                        payload_evento = {
                            "email": _email_d0_to_dict(email_pronto),
                            "conta_envio": conta_envio,
                            "registrar_link": bool(contexto_envio["link_gerado_nesta_execucao"]),
                            "usuario": self.coral_user_var.get().strip(),
                        }
                        event_id = _deterministic_processing_event_id("EMAIL_D0_ENVIADO", payload_evento, datetime.now())
                        payload_evento["id_evento"] = event_id
                        json_path = registrar_evento_processamento_json(
                            workbook_path,
                            tipo="EMAIL_D0_ENVIADO",
                            payload=payload_evento,
                        )
                        self._log(
                            f"Outlook confirmou o envio para {email.destinatario}. "
                            f"Registro salvo no JSON local: {json_path} | id_evento={event_id}."
                        )

                    self._executar_unidade_com_tentativas(email.id_cliente, "processar etapa 1", processar_cliente_d0)
                    sucessos += 1
                    eventos_desde_sync += 1
                    if eventos_desde_sync >= PROCESSING_SYNC_BATCH_SIZE:
                        self._sync_processing_events_safe(
                            workbook_path,
                            f"lote de {PROCESSING_SYNC_BATCH_SIZE} eventos D0",
                        )
                        eventos_desde_sync = 0
                    self._log(f"D0 enviado: {email.destinatario}")
                except Exception as exc:
                    falhas += 1
                    self._log(f"Falha no D0 para {email.destinatario}: {exc}")

            self._sync_processing_events_safe(workbook_path, "fechamento D0")
            self._set_status(f"D0 concluido. Sucessos: {sucessos} | Falhas: {falhas}")
            self._log(f"D0 finalizado. Sucessos={sucessos}; falhas={falhas}.")
        except Exception as exc:
            self._set_status(f"Erro no D0: {exc}")
            self._log(f"Erro geral no D0: {exc}")
            self.after(0, lambda error=str(exc): messagebox.showerror("Erro etapa 1", error))
        finally:
            self._finish_processing_control()
            self.after(0, self._refresh_queue)

    def _process_d2(self) -> None:
        if self.processing_thread is not None and self.processing_thread.is_alive():
            messagebox.showinfo("Etapa 2", "Ja existe um processamento em andamento.")
            return
        if not self.coral_user_var.get().strip() or not self.coral_password_var.get().strip():
            messagebox.showwarning("Coral", "Informe usuario e senha do Coral antes de processar a etapa 2.")
            return
        self._persist_interface_config()

        path = self.file_path_var.get().strip() or DEFAULT_EXCEL_URL
        resolved_path = resolver_caminho_excel_compartilhado(path)
        if resolved_path is None:
            message = "Nao encontrei a base operacional sincronizada. Confirme a sincronizacao da pasta no OneDrive."
            self.status_var.set(message)
            self._log(message)
            messagebox.showwarning("Base nao encontrada", message)
            return

        try:
            limite = self._d2_limit()
            canal = self._selected_canal()
            contratos = carregar_contratos_d2_excel(resolved_path, limite=limite, canal=canal)
        except Exception as exc:
            self.status_var.set(str(exc))
            self._log(f"Erro ao preparar D0+2: {exc}")
            messagebox.showerror("Erro ao preparar etapa 2", str(exc))
            return

        if not contratos:
            message = "Nenhum contrato apto para cobranca no cartao na etapa 2 no canal selecionado."
            self.status_var.set(message)
            self._log(message)
            messagebox.showinfo("Etapa 2", message)
            return

        if not messagebox.askyesno(
            "Confirmar etapa 2",
            f"Executar a etapa 2 para {len(contratos)} contrato(s) no Coral?\n\nCanal: {canal}\nA cobranca sera tentada no cartao contrato por contrato.",
        ):
            return

        self._start_processing_control(resolved_path)
        self.processing_thread = threading.Thread(
            target=self._run_d2_processing,
            args=(resolved_path, contratos),
            daemon=True,
        )
        self.processing_thread.start()

    def _run_d2_processing(self, workbook_path: Path, contratos: list[ContratoD2Pedagio]) -> None:
        cobrados = 0
        links_gerados = 0
        link_pendente = 0
        erros = 0
        eventos_desde_sync = 0
        try:
            self._login_coral(force_new_driver=True)
            total = len(contratos)
            contratos_por_cliente: dict[str, list[ContratoD2Pedagio]] = {}
            for contrato in contratos:
                contratos_por_cliente.setdefault(contrato.id_cliente, []).append(contrato)

            index_global = 0
            for cliente_index, (id_cliente, contratos_cliente) in enumerate(contratos_por_cliente.items(), start=1):
                if not self._wait_if_paused_or_stopped(
                    workbook_path,
                    f"D0+2 cliente {cliente_index}/{len(contratos_por_cliente)}",
                ):
                    break
                cliente_nome = contratos_cliente[0].cliente if contratos_cliente else id_cliente
                self._log(
                    f"[D0+2 cliente {cliente_index}/{len(contratos_por_cliente)}] "
                    f"{cliente_nome} | {len(contratos_cliente)} contrato(s)."
                )
                resultados_cliente: list[ResultadoD2Pedagio] = []
                for contrato in contratos_cliente:
                    index_global += 1
                    if not self._wait_if_paused_or_stopped(workbook_path, f"D0+2 contrato {index_global}/{total}"):
                        break
                    self._set_status(f"D0+2 {index_global}/{total}: {contrato.contrato}")
                    self._log(
                        f"[D0+2 {index_global}/{total}] Contrato {contrato.contrato} | Cliente {contrato.cliente} | "
                        f"Valor R$ {_format_brl(contrato.valor)}"
                    )
                    try:
                        resultado = self._executar_unidade_com_tentativas(
                            contrato.contrato,
                            "processar cobranca da etapa 2",
                            lambda contrato_atual=contrato: self._tentar_cobranca_cartao_contrato(contrato_atual),
                        )
                    except Exception as exc:
                        resultado = ResultadoD2Pedagio(
                            contrato.id_cliente,
                            contrato.contrato,
                            "ERRO_D0_2",
                            0,
                            0,
                            str(exc),
                        )
                        self._log(f"Erro D0+2 no contrato {contrato.contrato}: {exc}")

                    payload_evento = {
                        "resultado": _resultado_d2_to_dict(resultado),
                        "usuario": self.coral_user_var.get().strip(),
                    }
                    event_id = _deterministic_processing_event_id("RESULTADO_D2", payload_evento, datetime.now())
                    payload_evento["id_evento"] = event_id
                    json_path = registrar_evento_processamento_json(
                        workbook_path,
                        tipo="RESULTADO_D2",
                        payload=payload_evento,
                    )
                    resultados_cliente.append(resultado)
                    status_planilha = _status_planilha_d2(resultado.status)
                    if status_planilha == "COBRADO":
                        cobrados += 1
                    elif status_planilha == "LINK_PAGAMENTO_PENDENTE":
                        link_pendente += 1
                    else:
                        erros += 1
                    self._log(
                        f"Resultado D0+2 salvo no JSON: {resultado.contrato} | {status_planilha} | "
                        f"cartoes {resultado.cartoes_tentados}/{resultado.cartoes_encontrados} | "
                        f"{json_path} | id_evento={event_id}"
                    )
                    eventos_desde_sync += 1
                    if eventos_desde_sync >= PROCESSING_SYNC_BATCH_SIZE:
                        self._sync_processing_events_safe(
                            workbook_path,
                            f"lote de {PROCESSING_SYNC_BATCH_SIZE} eventos D0+2",
                        )
                        eventos_desde_sync = 0

                contratos_por_numero = {contrato.contrato: contrato for contrato in contratos_cliente}
                residuais = [
                    contratos_por_numero[resultado.contrato]
                    for resultado in resultados_cliente
                    if _status_planilha_d2(resultado.status) == "LINK_PAGAMENTO_PENDENTE"
                    and resultado.contrato in contratos_por_numero
                ]
                if not residuais:
                    continue
                if not self._wait_if_paused_or_stopped(workbook_path, f"link residual de {cliente_nome}"):
                    break

                valor_residual = sum(item.valor for item in residuais)
                self._set_status(f"Gerando link residual: {cliente_nome}")
                self._log(
                    f"Cliente {cliente_nome}: gerando link residual de R$ {_format_brl(valor_residual)} "
                    f"para {len(residuais)} contrato(s)."
                )
                try:
                    contexto_link = {"link_pagamento": ""}

                    def gerar_link_residual_cliente():
                        if not contexto_link["link_pagamento"]:
                            self._ir_para_pagamentos_para_link(residuais[-1].contrato)
                            contexto_link["link_pagamento"] = self._gerar_link_pagamento_residual(valor_residual)
                        payload_evento = {
                            "id_cliente": id_cliente,
                            "cliente": residuais[0].cliente,
                            "documento": residuais[0].documento,
                            "contratos": [item.contrato for item in residuais],
                            "valor_link": valor_residual,
                            "link_pagamento": contexto_link["link_pagamento"],
                            "contrato_referencia_link": residuais[-1].contrato,
                            "usuario": self.coral_user_var.get().strip(),
                        }
                        event_id = _deterministic_processing_event_id("LINK_D2_GERADO", payload_evento, datetime.now())
                        payload_evento["id_evento"] = event_id
                        json_path = registrar_evento_processamento_json(
                            workbook_path,
                            tipo="LINK_D2_GERADO",
                            payload=payload_evento,
                        )
                        return contexto_link["link_pagamento"], json_path, event_id

                    link_pagamento, json_path, event_id = self._executar_unidade_com_tentativas(
                        id_cliente,
                        "gerar link residual da etapa 2",
                        gerar_link_residual_cliente,
                    )
                    links_gerados += 1
                    link_pendente -= len(residuais)
                    self._log(
                        f"Link residual salvo no JSON para {cliente_nome}: {link_pagamento} | "
                        f"R$ {_format_brl(valor_residual)} | {json_path} | id_evento={event_id}"
                    )
                    eventos_desde_sync += 1
                    if eventos_desde_sync >= PROCESSING_SYNC_BATCH_SIZE:
                        self._sync_processing_events_safe(
                            workbook_path,
                            f"lote de {PROCESSING_SYNC_BATCH_SIZE} eventos D0+2",
                        )
                        eventos_desde_sync = 0
                except Exception as exc:
                    erros += 1
                    self._log(f"Falha ao gerar link residual para {cliente_nome}: {exc}")

            self._sync_processing_events_safe(workbook_path, "fechamento D0+2")
            self._set_status(
                f"D0+2 concluido. Cobrados: {cobrados} | Links gerados: {links_gerados} | "
                f"Links pendentes: {link_pendente} | Erros: {erros}"
            )
            self._log(
                f"D0+2 finalizado. Cobrados={cobrados}; links_gerados={links_gerados}; "
                f"links_pendentes={link_pendente}; erros={erros}."
            )
        except Exception as exc:
            self._set_status(f"Erro no D0+2: {exc}")
            self._log(f"Erro geral no D0+2: {exc}")
            self.after(0, lambda error=str(exc): messagebox.showerror("Erro etapa 2", error))
        finally:
            self._finish_processing_control()
            self.after(0, self._refresh_queue)

    def _refresh_queue(self) -> None:
        path = self.file_path_var.get().strip()
        if not path:
            self.file_path_var.set(DEFAULT_EXCEL_URL)
            path = DEFAULT_EXCEL_URL
            return

        resolved_path = resolver_caminho_excel_compartilhado(path)
        if resolved_path is None:
            if path.lower().startswith("http"):
                message = (
                    "Nao encontrei esse arquivo sincronizado no OneDrive desta maquina. "
                    "Abra o link no navegador, clique em Sincronizar ou Adicionar atalho ao OneDrive, "
                    "aguarde o arquivo aparecer no Explorador e tente novamente."
                )
            else:
                message = f"Planilha nao encontrada: {path}"
            self.connection_var.set("Arquivo nao encontrado")
            self.status_var.set(message)
            self._log(message)
            messagebox.showwarning("Arquivo nao encontrado", message)
            return

        if str(resolved_path) != path:
            self.file_path_var.set(str(resolved_path))
            self._log(f"Link resolvido para arquivo local: {resolved_path}")

        self._log(f"Atualizacao manual da fila usando planilha: {resolved_path}")
        try:
            sincronizados = sincronizar_eventos_processamento_json(resolved_path).get("eventos_sincronizados", 0)
            if sincronizados:
                self._log(f"Pendencias JSON do dia aplicadas antes de atualizar a fila: {sincronizados} evento(s).")
        except Exception as exc:
            self._log(f"Nao foi possivel aplicar pendencias JSON antes de atualizar a fila: {exc}")
        try:
            recentes = contar_pendencias_processamento_recentes(resolved_path)
            antigos = {dia: qtd for dia, qtd in recentes.items() if dia != date.today().isoformat()}
            if antigos:
                self._log(f"Pendencias JSON antigas encontradas e nao aplicadas automaticamente: {antigos}")
        except Exception as exc:
            self._log(f"Nao foi possivel verificar pendencias JSON recentes: {exc}")

        try:
            canais = carregar_canais_cobranca_excel(resolved_path)
            self._apply_canais_cobranca(canais)
            canal = self._selected_canal()
            resumo = carregar_resumo_cobrancas_excel(resolved_path, canal=canal)
        except Exception as exc:
            self.connection_var.set("Erro na base")
            self.status_var.set(str(exc))
            self._log(f"Erro ao carregar planilha: {exc}")
            messagebox.showerror("Erro ao carregar planilha", str(exc))
            return

        self._apply_queue_summary(resolved_path, resumo)

    def _apply_queue_summary(self, resolved_path: Path, resumo: ResumoCobrancasExcel) -> None:
        self.file_path_var.set(str(resolved_path))
        self.metrics["total"].set(str(resumo.total))
        self.metrics["contracts"].set(str(resumo.contratos_total))
        self.metrics["contracts_pending"].set(str(resumo.contratos_pendentes))
        self.metrics["d0"].set(str(resumo.aptos_d0))
        self.metrics["d2"].set(str(resumo.aptos_d2))
        self.metrics["d45"].set(str(resumo.aptos_d45))
        self.metrics["d7"].set(str(resumo.aptos_d7))
        self.metrics["processing"].set(str(resumo.em_processamento))
        self.metrics["paid"].set(str(resumo.pagos))
        self.metrics["errors"].set(str(resumo.erros))

        self.connection_var.set("Excel conectado")
        canal = self._selected_canal()
        self.status_var.set(
            f"Fila carregada. Canal: {canal} | Clientes: {resumo.total} | "
            f"Contratos: {resumo.contratos_total} | Valor: R$ {resumo.valor_total:,.2f}"
        )
        self._log(
            f"Fila atualizada ({canal}): {resumo.total} clientes | "
            f"{resumo.contratos_total} contratos | R$ {resumo.valor_total:,.2f}"
        )

    def _log(self, message: str) -> None:
        if threading.current_thread() is not self._main_thread:
            self.after(0, self._log, message)
            return
        timestamp = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        self.log_box.configure(state="normal")
        self.log_box.insert("end", f"[{timestamp}] {message}\n")
        self.log_box.see("end")
        self.log_box.configure(state="disabled")

    def _set_status(self, message: str) -> None:
        if threading.current_thread() is not self._main_thread:
            self.after(0, self._set_status, message)
            return
        self.status_var.set(message)


if __name__ == "__main__":
    app = RoboCobrancaPedagiosApp()
    app.mainloop()
